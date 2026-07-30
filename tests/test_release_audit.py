from __future__ import annotations

import json
from dataclasses import replace
from pathlib import Path

from openpyxl import load_workbook

import bus_schedule_engine.release_audit as release_audit
from bus_schedule_engine.application_pipeline import (
    WORKBOOK_IMPORT_INVALID,
    WORKBOOK_OPTIMIZATION_NOT_READY,
)
from bus_schedule_engine.excel_exporter import create_input_template
from bus_schedule_engine.optimization_service import SolverChoice
from bus_schedule_engine.release_audit import (
    RELEASE_AUDIT_BLOCKED,
    RELEASE_AUDIT_PASSED,
    RELEASE_AUDIT_SCHEMA_V1,
    main,
    run_release_audit_v1,
)


def _set_parameter(sheet, key: str, value: object | None) -> None:
    row = next(cell.row for cell in sheet["A"] if cell.value == key)
    sheet.cell(row, 2).value = value


def test_release_audit_is_deterministic_and_omits_raw_comparison_values(
    tmp_path: Path,
) -> None:
    workbook = create_input_template(tmp_path / "input.xlsx")
    first_output = tmp_path / "first.json"
    second_output = tmp_path / "second.json"

    first_exit = run_release_audit_v1(
        workbook,
        solver_choice=SolverChoice.HEURISTIC,
        output=first_output,
    )
    second_exit = run_release_audit_v1(
        workbook,
        solver_choice=SolverChoice.HEURISTIC,
        output=second_output,
    )

    assert first_exit == second_exit
    assert first_output.read_bytes() == second_output.read_bytes()
    payload = json.loads(first_output.read_text(encoding="utf-8"))
    assert payload["schema"] == RELEASE_AUDIT_SCHEMA_V1
    assert payload["status"] in {RELEASE_AUDIT_PASSED, RELEASE_AUDIT_BLOCKED}
    assert len(payload["source_sha256"]) == 64
    assert payload["solver_choice"] == "HEURISTIC"
    assert payload["input_readiness"]["optimization_ready"] is True
    assert first_exit == (1 if payload["blocking_discrepancy_codes"] else 0)
    assert payload["comparisons"]
    assert all(
        "legacy_value" not in comparison and "unified_value" not in comparison
        for comparison in payload["comparisons"]
    )
    serialized = first_output.read_text(encoding="utf-8")
    assert "passenger_demand" not in serialized
    assert "trips_b" not in serialized


def test_release_audit_stops_at_input_readiness(tmp_path: Path) -> None:
    workbook_path = create_input_template(tmp_path / "not-ready.xlsx")
    workbook = load_workbook(workbook_path)
    _set_parameter(
        workbook["THONG_SO_B"],
        "available_fleet_limit",
        None,
    )
    workbook.save(workbook_path)
    workbook.close()
    output = tmp_path / "not-ready.json"

    exit_code = run_release_audit_v1(
        workbook_path,
        solver_choice=SolverChoice.HEURISTIC,
        output=output,
    )

    payload = json.loads(output.read_text(encoding="utf-8"))
    assert exit_code == 2
    assert payload["status"] == WORKBOOK_OPTIMIZATION_NOT_READY
    assert payload["input_readiness"]["missing_optimization_authority_codes"]
    assert "comparisons" not in payload


def test_release_audit_records_import_invalid_without_raw_bytes(tmp_path: Path) -> None:
    workbook = tmp_path / "invalid.xlsx"
    workbook.write_bytes(b"not an xlsx secret-payload")
    output = tmp_path / "invalid.json"

    exit_code = run_release_audit_v1(
        workbook,
        solver_choice=SolverChoice.HEURISTIC,
        output=output,
    )

    payload = json.loads(output.read_text(encoding="utf-8"))
    assert exit_code == 2
    assert payload["status"] == WORKBOOK_IMPORT_INVALID
    assert "secret-payload" not in output.read_text(encoding="utf-8")


def test_release_audit_cli_accepts_required_arguments(tmp_path: Path) -> None:
    workbook = create_input_template(tmp_path / "cli.xlsx")
    output = tmp_path / "cli.json"

    exit_code = main(
        [
            "--workbook",
            str(workbook),
            "--solver",
            "HEURISTIC",
            "--output",
            str(output),
        ]
    )

    assert exit_code in {0, 1}
    assert output.is_file()


def test_release_audit_exits_nonzero_when_comparison_has_blocker(
    tmp_path: Path,
    monkeypatch,
) -> None:
    workbook = create_input_template(tmp_path / "blocked.xlsx")
    output = tmp_path / "blocked.json"
    real_runner = release_audit.run_side_by_side_validation_v1

    def blocked_runner(*args, **kwargs):
        report = real_runner(*args, **kwargs)
        return replace(
            report,
            blocking_discrepancy_codes=("SYNTHETIC_RELEASE_BLOCKER",),
        )

    monkeypatch.setattr(
        release_audit,
        "run_side_by_side_validation_v1",
        blocked_runner,
    )
    exit_code = run_release_audit_v1(
        workbook,
        solver_choice=SolverChoice.HEURISTIC,
        output=output,
    )

    payload = json.loads(output.read_text(encoding="utf-8"))
    assert exit_code == 1
    assert payload["status"] == RELEASE_AUDIT_BLOCKED
    assert payload["blocking_discrepancy_codes"] == ["SYNTHETIC_RELEASE_BLOCKER"]
