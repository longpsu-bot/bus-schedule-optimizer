from __future__ import annotations

from openpyxl import load_workbook

from bus_schedule_engine.excel_exporter import (
    CONDITIONAL_DEMAND_LABEL,
    CONDITIONAL_TRIP_RIDERSHIP_LABEL,
    OPTIONAL_LABEL,
    REQUIRED_FOR_OPTIMIZATION_LABEL,
    REQUIRED_LABEL,
    create_input_template,
)
from bus_schedule_engine.importer import import_workbook


def _rows_by_key(sheet) -> dict[str, tuple[object, object, object]]:
    return {
        str(sheet.cell(row, 1).value): (
            sheet.cell(row, 2).value,
            sheet.cell(row, 3).value,
            sheet.cell(row, 4).value,
        )
        for row in range(4, sheet.max_row + 1)
        if sheet.cell(row, 1).value is not None
    }


def _value_cell(sheet, key: str):
    row = next(cell.row for cell in sheet["A"] if cell.value == key)
    return sheet.cell(row, 2)


def _validation_for_cell(sheet, coordinate: str):
    return next(
        validation
        for validation in sheet.data_validations.dataValidation
        if coordinate in validation.cells
    )


def test_generated_template_displays_all_requirement_levels(tmp_path) -> None:
    path = create_input_template(tmp_path / "input.xlsx")
    workbook = load_workbook(path)
    parameters_b = _rows_by_key(workbook["THONG_SO_B"])
    metadata = _rows_by_key(workbook["THONG_TIN_DU_LIEU"])

    levels = {row[1] for row in parameters_b.values()} | {row[1] for row in metadata.values()}

    assert REQUIRED_LABEL in levels
    assert REQUIRED_FOR_OPTIMIZATION_LABEL in levels
    assert OPTIONAL_LABEL in levels
    assert CONDITIONAL_DEMAND_LABEL in levels
    assert workbook["THONG_SO_B"].cell(3, 3).value == "Mức độ"
    assert workbook["THONG_TIN_DU_LIEU"].cell(3, 3).value == "Mức độ"
    workbook.close()


def test_generated_template_classifies_authority_and_optional_fields(tmp_path) -> None:
    path = create_input_template(tmp_path / "input.xlsx")
    workbook = load_workbook(path)
    parameters_a = _rows_by_key(workbook["THONG_SO_A"])
    parameters_b = _rows_by_key(workbook["THONG_SO_B"])
    metadata = _rows_by_key(workbook["THONG_TIN_DU_LIEU"])

    for parameters in (parameters_a, parameters_b):
        assert parameters["available_fleet_limit"][1] == REQUIRED_FOR_OPTIMIZATION_LABEL
        assert parameters["operating_day_type"][1] == REQUIRED_FOR_OPTIMIZATION_LABEL
        assert parameters["approved_active_fleet"][1] == OPTIONAL_LABEL

    assert parameters_b["terminal_1_max_occupancy_vehicles"][1] == OPTIONAL_LABEL
    assert parameters_b["terminal_2_max_occupancy_vehicles"][1] == OPTIONAL_LABEL
    for key in (
        "demand_source_type",
        "demand_confidence",
        "demand_response_mode",
    ):
        assert metadata[key][1] == CONDITIONAL_DEMAND_LABEL
    workbook.close()


def test_requirement_levels_are_text_and_visually_distinct(tmp_path) -> None:
    path = create_input_template(tmp_path / "input.xlsx")
    workbook = load_workbook(path)
    sheet = workbook["THONG_SO_B"]
    level_cells = {
        sheet.cell(row, 3).value: sheet.cell(row, 3) for row in range(4, sheet.max_row + 1)
    }

    assert (
        level_cells[REQUIRED_LABEL].fill.fgColor.rgb
        != level_cells[REQUIRED_FOR_OPTIMIZATION_LABEL].fill.fgColor.rgb
    )
    assert (
        level_cells[REQUIRED_FOR_OPTIMIZATION_LABEL].fill.fgColor.rgb
        != level_cells[OPTIONAL_LABEL].fill.fgColor.rgb
    )
    assert all(cell.font.bold for cell in level_cells.values())
    workbook.close()


def test_required_and_blank_permitted_integer_cells_have_separate_validations(
    tmp_path,
) -> None:
    path = create_input_template(tmp_path / "input.xlsx")
    workbook = load_workbook(path)

    for sheet_name in ("THONG_SO_A", "THONG_SO_B"):
        sheet = workbook[sheet_name]
        for key in ("vehicle_capacity_passengers", "total_daily_trips"):
            validation = _validation_for_cell(sheet, _value_cell(sheet, key).coordinate)
            assert validation.type == "whole"
            assert validation.operator == "greaterThan"
            assert validation.formula1 == "0"
            assert validation.allow_blank is False
        for key in (
            "available_fleet_limit",
            "approved_active_fleet",
            "minimum_layover_minutes",
        ):
            validation = _validation_for_cell(sheet, _value_cell(sheet, key).coordinate)
            assert validation.type == "whole"
            assert validation.allow_blank is True

    sheet_b = workbook["THONG_SO_B"]
    for key in (
        "terminal_1_max_occupancy_vehicles",
        "terminal_2_max_occupancy_vehicles",
    ):
        validation = _validation_for_cell(sheet_b, _value_cell(sheet_b, key).coordinate)
        assert validation.type == "whole"
        assert validation.allow_blank is True
    workbook.close()


def test_template_explains_runtime_one_of_compatibility_rule(tmp_path) -> None:
    path = create_input_template(tmp_path / "input.xlsx")
    workbook = load_workbook(path)
    parameters = _rows_by_key(workbook["THONG_SO_B"])
    guide = " ".join(
        str(cell.value)
        for row in workbook["HUONG_DAN"].iter_rows(min_row=4)
        for cell in row
        if cell.value is not None
    )

    assert "Định dạng runtime ưu tiên" in parameters["allowed_trip_runtime_minutes"][2]
    assert "có thể dùng trip_runtime_minutes" in parameters["allowed_trip_runtime_minutes"][2]
    assert "Giá trị tương thích cũ" in parameters["trip_runtime_minutes"][2]
    assert "allowed_trip_runtime_minutes để trống" in parameters["trip_runtime_minutes"][2]
    assert "Phải khai báo allowed_trip_runtime_minutes hoặc trip_runtime_minutes" in guide
    assert "allowed_trip_runtime_minutes là định dạng được ưu tiên" in guide
    assert "trip_runtime_minutes chỉ dùng để tương thích file cũ" in guide

    allowed_runtime_validation = _validation_for_cell(
        workbook["THONG_SO_B"],
        _value_cell(workbook["THONG_SO_B"], "allowed_trip_runtime_minutes").coordinate,
    )
    assert allowed_runtime_validation.allow_blank is True
    workbook.close()


def test_generated_template_adds_separate_trip_ridership_sheets_and_guidance(
    tmp_path,
) -> None:
    path = create_input_template(tmp_path / "input.xlsx")
    workbook = load_workbook(path)

    assert "THONG_TIN_SAN_LUONG_CHUYEN" in workbook.sheetnames
    assert "SAN_LUONG_CHUYEN" in workbook.sheetnames
    metadata = _rows_by_key(workbook["THONG_TIN_SAN_LUONG_CHUYEN"])
    assert set(metadata) == {
        "trip_ridership_dataset_id",
        "trip_ridership_source_type",
        "trip_ridership_confidence",
        "observed_schedule_scenario",
        "operating_day_type",
        "match_tolerance_minutes",
        "source_notes",
    }
    assert all(
        row[1] == CONDITIONAL_TRIP_RIDERSHIP_LABEL
        for key, row in metadata.items()
        if key != "source_notes"
    )
    trip_sheet = workbook["SAN_LUONG_CHUYEN"]
    assert [trip_sheet.cell(3, column).value for column in range(1, 11)] == [
        "observation_id",
        "service_date",
        "source_trip_id",
        "scheduled_trip_id",
        "direction",
        "scheduled_departure_time",
        "actual_departure_time",
        "passenger_count",
        "vehicle_id",
        "notes",
    ]
    assert "DÒNG MẪU" in trip_sheet.cell(2, 1).value
    guide = " ".join(
        str(cell.value)
        for row in workbook["HUONG_DAN"].iter_rows(min_row=4)
        for cell in row
        if cell.value is not None
    )
    assert "chưa được dùng để sinh phương án C" in guide
    assert "không được hiểu là 0 hành khách" in guide
    assert "va chạm cùng chuyến-ngày" in guide
    workbook.close()

    imported = import_workbook(path)
    assert imported.trip_ridership_metadata is None
    assert imported.trip_ridership_observations == ()
