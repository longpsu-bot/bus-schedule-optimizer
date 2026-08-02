from __future__ import annotations

import hashlib
import json
from dataclasses import replace
from datetime import timedelta
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from presentation_support import (
    build_corpus_result_and_report,
    build_result_and_report,
    rejected_result_and_report,
)

import bus_schedule_engine.unified_presentation as unified_presentation
from bus_schedule_engine.optimization_service import SolverChoice
from bus_schedule_engine.unified_diagram import (
    build_unified_demand_supply_figure_v1,
    build_unified_departure_figure_v1,
)
from bus_schedule_engine.unified_presentation import (
    UnifiedPresentationConsistencyError,
    build_unified_presentation_v1,
)
from bus_schedule_engine.unified_result_exporter import (
    UnifiedExportMetadataV1,
    export_unified_result_workbook_v1,
    read_unified_export_metadata_bytes_v1,
    read_unified_export_metadata_v1,
)


@pytest.fixture(scope="module")
def accepted_presentation():
    return build_unified_presentation_v1(*build_result_and_report())


@pytest.fixture(scope="module")
def alpha_presentation():
    return build_unified_presentation_v1(*build_corpus_result_and_report("corpus_alpha_80.json"))


def _table_rows(workbook, sheet_name: str) -> list[dict[str, object]]:
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[1]]
    return [
        dict(zip(headers, values, strict=True))
        for values in sheet.iter_rows(min_row=2, values_only=True)
    ]


def _key_values(workbook, sheet_name: str) -> dict[str, object]:
    return {
        str(key): value
        for key, value in workbook[sheet_name].iter_rows(
            min_row=2,
            max_col=2,
            values_only=True,
        )
        if key is not None
    }


def _seconds(excel_time: float | timedelta | None) -> int | None:
    if excel_time is None:
        return None
    if isinstance(excel_time, timedelta):
        return round(excel_time.total_seconds())
    return round(excel_time * 86_400)


def _with_recomputed_fingerprint(presentation):
    without_fingerprint = replace(presentation, presentation_fingerprint="")
    return replace(
        without_fingerprint,
        presentation_fingerprint=unified_presentation._presentation_fingerprint(
            without_fingerprint
        ),
    )


def _with_changed_b_departure(presentation):
    scenario_b = presentation.scenario("B")
    assert scenario_b is not None
    changed_trip = replace(
        scenario_b.trips[0],
        departure_time_seconds=scenario_b.trips[0].departure_time_seconds + 60,
    )
    changed_b = replace(
        scenario_b,
        trips=(changed_trip, *scenario_b.trips[1:]),
    )
    return replace(
        presentation,
        scenarios=tuple(
            changed_b if item.scenario_id == "B" else item for item in presentation.scenarios
        ),
    )


def _with_changed_block(presentation):
    changed = replace(
        presentation.blocks[0],
        passenger_demand=presentation.blocks[0].passenger_demand + 1,
    )
    return replace(presentation, blocks=(changed, *presentation.blocks[1:]))


def _with_changed_outcome(presentation):
    return replace(
        presentation,
        outcome=replace(presentation.outcome, selected_action="NO_CHANGE"),
    )


def _with_changed_discrepancy(presentation):
    changed = replace(
        presentation.discrepancies[0],
        explanation=presentation.discrepancies[0].explanation + " changed",
    )
    return replace(
        presentation,
        discrepancies=(changed, *presentation.discrepancies[1:]),
    )


@pytest.mark.parametrize(
    "mutation",
    (
        _with_changed_b_departure,
        _with_changed_block,
        _with_changed_outcome,
        _with_changed_discrepancy,
        lambda presentation: replace(
            presentation,
            presentation_mode="AUTHORITATIVE",
        ),
        lambda presentation: replace(
            presentation,
            presentation_fingerprint="f" * 64,
        ),
    ),
    ids=(
        "b-departure",
        "block-value",
        "outcome-field",
        "discrepancy-record",
        "presentation-mode",
        "stored-fingerprint",
    ),
)
def test_all_artifact_builders_reject_modified_presentations(
    mutation,
    tmp_path: Path,
    accepted_presentation,
) -> None:
    changed = mutation(accepted_presentation)
    for builder in (
        build_unified_demand_supply_figure_v1,
        build_unified_departure_figure_v1,
    ):
        with pytest.raises(UnifiedPresentationConsistencyError):
            builder(changed)
    target = tmp_path / "must-not-exist.xlsx"
    with pytest.raises(UnifiedPresentationConsistencyError):
        export_unified_result_workbook_v1(changed, target)
    assert not target.exists()


def test_export_creates_only_new_path_by_default_and_refuses_existing(
    tmp_path: Path,
    accepted_presentation,
) -> None:
    target = tmp_path / "unified.xlsx"
    assert export_unified_result_workbook_v1(accepted_presentation, target) == target
    before = target.read_bytes()
    with pytest.raises(FileExistsError):
        export_unified_result_workbook_v1(accepted_presentation, target)
    assert target.read_bytes() == before


def test_required_accepted_c_sheets_exist(
    tmp_path: Path,
    accepted_presentation,
) -> None:
    target = export_unified_result_workbook_v1(
        accepted_presentation,
        tmp_path / "accepted.xlsx",
    )
    workbook = load_workbook(target, data_only=False)
    try:
        required = {
            "TONG_QUAN",
            "A_BIEU_DO",
            "B_BIEU_DO",
            "C_BIEU_DO",
            "SO_SANH_B_C",
            "CUNG_CAU_BLOCK",
            "DANH_GIA_B",
            "FLEET_C",
            "HEADWAY_C",
            "SOLVER",
            "DOI_CHIEU_5A1",
            "GIOI_HAN",
            "FINGERPRINTS",
        }
        assert required <= set(workbook.sheetnames)
        assert "C_TRANG_THAI" not in workbook.sheetnames
    finally:
        workbook.close()


def test_no_c_workbook_has_status_not_c_artifacts(
    tmp_path: Path,
    alpha_presentation,
) -> None:
    target = export_unified_result_workbook_v1(
        alpha_presentation,
        tmp_path / "alpha.xlsx",
    )
    workbook = load_workbook(target, data_only=False)
    try:
        assert "C_TRANG_THAI" in workbook.sheetnames
        assert {
            "C_BIEU_DO",
            "SO_SANH_B_C",
            "FLEET_C",
            "HEADWAY_C",
        }.isdisjoint(workbook.sheetnames)
        status = _key_values(workbook, "C_TRANG_THAI")
        assert status["authoritative_c_statement"] == ("Không có Scenario C có thẩm quyền.")
        assert status["solver_attempted"] is False
    finally:
        workbook.close()


def test_b_and_c_schedule_rows_match_presentation_exactly(
    tmp_path: Path,
    accepted_presentation,
) -> None:
    target = export_unified_result_workbook_v1(
        accepted_presentation,
        tmp_path / "schedules.xlsx",
    )
    workbook = load_workbook(target, data_only=False)
    try:
        b_rows = _table_rows(workbook, "B_BIEU_DO")
        c_rows = _table_rows(workbook, "C_BIEU_DO")
        scenario_b = accepted_presentation.scenario("B")
        scenario_c = accepted_presentation.scenario("C")
        assert scenario_b is not None
        assert scenario_c is not None
        assert [row["trip_id"] for row in b_rows] == [trip.trip_id for trip in scenario_b.trips]
        assert [_seconds(row["departure_time"]) for row in b_rows] == [
            trip.departure_time_seconds for trip in scenario_b.trips
        ]
        assert [row["c_trip_id"] for row in c_rows] == [trip.trip_id for trip in scenario_c.trips]
        assert [row["source_b_trip_id"] for row in c_rows] == [
            trip.source_b_trip_id for trip in scenario_c.trips
        ]
        assert [_seconds(row["c_departure_time"]) for row in c_rows] == [
            trip.departure_time_seconds for trip in scenario_c.trips
        ]
        assert [_seconds(row["arrival_time"]) for row in c_rows] == [
            trip.arrival_time_seconds for trip in scenario_c.trips
        ]
        assert [row["vehicle"] for row in c_rows] == [
            trip.vehicle_assignment for trip in scenario_c.trips
        ]
    finally:
        workbook.close()


def test_b_to_c_mapping_is_one_to_one_by_source_id(
    tmp_path: Path,
    accepted_presentation,
) -> None:
    target = export_unified_result_workbook_v1(
        accepted_presentation,
        tmp_path / "trace.xlsx",
    )
    workbook = load_workbook(target, data_only=False)
    try:
        rows = _table_rows(workbook, "SO_SANH_B_C")
        scenario_b = accepted_presentation.scenario("B")
        assert scenario_b is not None
        assert len({row["source_b_trip_id"] for row in rows}) == len(rows)
        assert {row["source_b_trip_id"] for row in rows} == {
            trip.trip_id for trip in scenario_b.trips
        }
    finally:
        workbook.close()


def test_block_rows_match_presentation_without_aggregation(
    tmp_path: Path,
    accepted_presentation,
) -> None:
    target = export_unified_result_workbook_v1(
        accepted_presentation,
        tmp_path / "blocks.xlsx",
    )
    workbook = load_workbook(target, data_only=False)
    try:
        rows = _table_rows(workbook, "CUNG_CAU_BLOCK")
        assert len(rows) == len(accepted_presentation.blocks)
        assert [
            (
                row["block_id"],
                row["direction"],
                _seconds(row["block_start"]),
                _seconds(row["block_end"]),
            )
            for row in rows
        ] == [
            (
                block.block_id,
                block.direction,
                block.block_start_seconds,
                block.block_end_seconds,
            )
            for block in accepted_presentation.blocks
        ]
        assert [row["b_trip_count"] for row in rows] == [
            block.b_trip_count for block in accepted_presentation.blocks
        ]
        assert [row["c_actual_trip_count"] for row in rows] == [
            block.c_actual_trip_count for block in accepted_presentation.blocks
        ]
    finally:
        workbook.close()


def test_solver_vectors_are_preserved_exactly_without_weighted_total(
    tmp_path: Path,
) -> None:
    presentation = build_unified_presentation_v1(
        *build_result_and_report(solver_choice=SolverChoice.BOTH)
    )
    target = export_unified_result_workbook_v1(
        presentation,
        tmp_path / "both.xlsx",
    )
    workbook = load_workbook(target, data_only=False)
    try:
        values = _key_values(workbook, "SOLVER")
        assert json.loads(values["comparison_objective_names"]) == list(
            presentation.outcome.comparison_objective_names
        )
        assert json.loads(values["heuristic_objective_vector"]) == list(
            presentation.outcome.heuristic_objective_vector
        )
        assert json.loads(values["ortools_objective_vector"]) == list(
            presentation.outcome.ortools_objective_vector
        )
        assert values["recommended_solver"] == presentation.outcome.recommended_solver
        assert "weighted_total" not in values
    finally:
        workbook.close()


def test_all_side_by_side_records_are_exported_complete(
    tmp_path: Path,
    accepted_presentation,
) -> None:
    target = export_unified_result_workbook_v1(
        accepted_presentation,
        tmp_path / "comparison.xlsx",
    )
    workbook = load_workbook(target, data_only=False)
    try:
        rows = _table_rows(workbook, "DOI_CHIEU_5A1")
        assert len(rows) == len(accepted_presentation.discrepancies)
        assert [row["fact_code"] for row in rows] == [
            item.fact_code for item in accepted_presentation.discrepancies
        ]
        assert [row["reason_code"] for row in rows] == [
            item.reason_code for item in accepted_presentation.discrepancies
        ]
    finally:
        workbook.close()


def test_terminal_occupancy_statuses_and_limits_are_preserved(
    tmp_path: Path,
) -> None:
    presentation = build_unified_presentation_v1(*build_result_and_report(terminal_1_occupancy=10))
    target = export_unified_result_workbook_v1(
        presentation,
        tmp_path / "occupancy.xlsx",
    )
    workbook = load_workbook(target, data_only=False)
    try:
        overview = _key_values(workbook, "TONG_QUAN")
        assert overview["terminal_occupancy_status"] == "PARTIALLY_EVALUATED"
        assert overview["terminal_1_occupancy_status"] == "PASS"
        assert overview["terminal_2_occupancy_status"] == "NOT_EVALUATED"
        assert overview["terminal_1_occupancy_limit"] == 10
        assert overview["terminal_2_occupancy_limit"] is None
        limitations = _table_rows(workbook, "GIOI_HAN")
        assert any(row["record_type"] == "TERMINAL_CAPACITY" for row in limitations)
    finally:
        workbook.close()


def test_beta_demand_gap_is_visible_in_limitations(tmp_path: Path) -> None:
    presentation = build_unified_presentation_v1(
        *build_corpus_result_and_report("corpus_beta_46.json")
    )
    target = export_unified_result_workbook_v1(
        presentation,
        tmp_path / "beta.xlsx",
    )
    workbook = load_workbook(target, data_only=False)
    try:
        gaps = [
            row for row in _table_rows(workbook, "GIOI_HAN") if row["record_type"] == "DEMAND_GAP"
        ]
        assert any(
            row["direction"] == "outbound" and row["range_seconds"] == f"{17 * 3600}-{18 * 3600}"
            for row in gaps
        )
    finally:
        workbook.close()


def test_fingerprint_metadata_reader_matches_presentation(
    tmp_path: Path,
    accepted_presentation,
) -> None:
    target = export_unified_result_workbook_v1(
        accepted_presentation,
        tmp_path / "metadata.xlsx",
    )
    metadata = read_unified_export_metadata_v1(target)

    assert UnifiedExportMetadataV1.__dataclass_params__.frozen
    assert "__slots__" in UnifiedExportMetadataV1.__dict__
    assert metadata.presentation_fingerprint == (accepted_presentation.presentation_fingerprint)
    assert metadata.b_fingerprint == accepted_presentation.source_b_fingerprint
    assert metadata.accepted_solution_fingerprint == (
        accepted_presentation.accepted_solution_fingerprint
    )
    assert metadata.source_id == accepted_presentation.source_id
    assert metadata.presentation_mode == "VALIDATION_ONLY"
    assert metadata.cutover_blocked == accepted_presentation.cutover_blocked


def test_byte_metadata_reader_matches_path_reader(
    tmp_path: Path,
    accepted_presentation,
) -> None:
    target = export_unified_result_workbook_v1(
        accepted_presentation,
        tmp_path / "metadata-bytes.xlsx",
    )

    assert read_unified_export_metadata_bytes_v1(
        target.read_bytes()
    ) == read_unified_export_metadata_v1(target)


def test_byte_metadata_reader_rejects_invalid_xlsx_bytes() -> None:
    with pytest.raises(ValueError, match="invalid"):
        read_unified_export_metadata_bytes_v1(b"not-an-xlsx")


def test_byte_metadata_reader_rejects_missing_fingerprints_sheet() -> None:
    workbook = Workbook()
    workbook.active.title = "TONG_QUAN"
    content = BytesIO()
    workbook.save(content)
    workbook.close()

    with pytest.raises(ValueError, match="FINGERPRINTS sheet is missing"):
        read_unified_export_metadata_bytes_v1(content.getvalue())


def test_byte_metadata_reader_aligns_none_accepted_c(
    tmp_path: Path,
    alpha_presentation,
) -> None:
    target = export_unified_result_workbook_v1(
        alpha_presentation,
        tmp_path / "alpha-metadata-bytes.xlsx",
    )
    metadata = read_unified_export_metadata_bytes_v1(target.read_bytes())

    assert metadata.accepted_solution_fingerprint is None


def test_workbook_is_unprotected_editable_and_has_no_formulas(
    tmp_path: Path,
    accepted_presentation,
) -> None:
    target = export_unified_result_workbook_v1(
        accepted_presentation,
        tmp_path / "plain.xlsx",
    )
    workbook = load_workbook(target, data_only=False)
    try:
        for sheet in workbook.worksheets:
            assert sheet.protection.sheet is False
            assert sheet.sheet_state == "visible"
            for row in sheet.iter_rows():
                for cell in row:
                    assert cell.data_type != "f"
    finally:
        workbook.close()


def test_rejected_candidate_timetable_is_absent_but_codes_are_visible(
    tmp_path: Path,
) -> None:
    presentation = build_unified_presentation_v1(*rejected_result_and_report())
    target = export_unified_result_workbook_v1(
        presentation,
        tmp_path / "rejected.xlsx",
    )
    workbook = load_workbook(target, data_only=False)
    try:
        assert "C_TRANG_THAI" in workbook.sheetnames
        assert "C_BIEU_DO" not in workbook.sheetnames
        codes = _key_values(workbook, "C_TRANG_THAI")["validator_rejection_codes"]
        assert json.loads(codes) == ["SYNTHETIC_DOMAIN_REJECTION"]
        all_text = " ".join(
            str(cell.value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None
        )
        assert "rejected-diagnostic-candidate" not in all_text
    finally:
        workbook.close()


def test_logical_source_id_is_not_interpreted_as_a_path(
    monkeypatch,
    tmp_path: Path,
    accepted_presentation,
) -> None:
    monkeypatch.chdir(tmp_path)
    presentation = _with_recomputed_fingerprint(
        replace(accepted_presentation, source_id="input.xlsx")
    )
    target = tmp_path / "input.xlsx"

    export_unified_result_workbook_v1(presentation, target)
    metadata = read_unified_export_metadata_v1(target)

    assert metadata.source_id == "input.xlsx"


def test_different_source_and_output_paths_export_successfully(
    tmp_path: Path,
    accepted_presentation,
) -> None:
    source = tmp_path / "source.xlsx"
    source.write_bytes(b"source workbook bytes stay untouched")
    target = tmp_path / "separate-output.xlsx"

    assert (
        export_unified_result_workbook_v1(
            accepted_presentation,
            target,
            source_workbook_path=source,
        )
        == target
    )
    assert target.exists()


def test_equal_source_and_output_paths_are_rejected_even_with_overwrite(
    tmp_path: Path,
    accepted_presentation,
) -> None:
    source = tmp_path / "source.xlsx"
    source.write_bytes(b"source")
    with pytest.raises(ValueError, match="source workbook"):
        export_unified_result_workbook_v1(
            accepted_presentation,
            source,
            overwrite=True,
            source_workbook_path=source,
        )
    assert source.read_bytes() == b"source"


def test_overwrite_requires_source_path_safety_metadata(
    tmp_path: Path,
    accepted_presentation,
) -> None:
    target = tmp_path / "existing.xlsx"
    target.write_bytes(b"existing output")

    with pytest.raises(
        ValueError,
        match="SOURCE_WORKBOOK_PATH_REQUIRED_FOR_OVERWRITE",
    ):
        export_unified_result_workbook_v1(
            accepted_presentation,
            target,
            overwrite=True,
        )
    assert target.read_bytes() == b"existing output"


def test_source_input_bytes_are_unchanged_when_different_output_is_overwritten(
    tmp_path: Path,
    accepted_presentation,
) -> None:
    source = tmp_path / "source.xlsx"
    source.write_bytes(b"source workbook bytes stay untouched")
    target = tmp_path / "existing-output.xlsx"
    target.write_bytes(b"replaceable output")
    before = hashlib.sha256(source.read_bytes()).hexdigest()

    export_unified_result_workbook_v1(
        accepted_presentation,
        target,
        overwrite=True,
        source_workbook_path=source,
    )

    assert hashlib.sha256(source.read_bytes()).hexdigest() == before
    assert target.read_bytes() != b"replaceable output"
