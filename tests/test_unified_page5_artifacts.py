from __future__ import annotations

from dataclasses import FrozenInstanceError, replace
from io import BytesIO
from pathlib import Path

import plotly.graph_objects as go
import pytest
from openpyxl import load_workbook
from presentation_support import (
    build_corpus_result_and_report,
    build_result_and_report,
    rejected_result_and_report,
)

import bus_schedule_engine.optimization_service as optimization_service
import bus_schedule_engine.service as service
import bus_schedule_engine.unified_page5_artifacts as page5_artifacts
import bus_schedule_engine.unified_presentation as unified_presentation
from bus_schedule_engine.unified_diagram import (
    available_unified_directions_v1,
    build_unified_demand_supply_figure_v1,
    build_unified_departure_figure_v1,
)
from bus_schedule_engine.unified_page5_artifacts import (
    UNIFIED_PAGE5_HTML_FILENAME,
    UNIFIED_PAGE5_PNG_FILENAME,
    UNIFIED_PAGE5_XLSX_FILENAME,
    UnifiedPage5ArtifactError,
    UnifiedPage5ArtifactsV1,
    build_unified_page5_artifacts_v1,
)
from bus_schedule_engine.unified_presentation import build_unified_presentation_v1
from bus_schedule_engine.unified_result_exporter import export_unified_result_workbook_v1


@pytest.fixture(scope="module")
def accepted_presentation():
    return build_unified_presentation_v1(*build_result_and_report())


@pytest.fixture(scope="module")
def alpha_presentation():
    return build_unified_presentation_v1(*build_corpus_result_and_report("corpus_alpha_80.json"))


@pytest.fixture
def fast_render(monkeypatch):
    monkeypatch.setattr(page5_artifacts, "_build_html_bytes", lambda *args, **kwargs: b"<html/>")
    monkeypatch.setattr(page5_artifacts, "_build_png_bytes", lambda figure: b"png")


def _xlsx_bytes(presentation, tmp_path: Path, name: str = "unified.xlsx") -> bytes:
    return export_unified_result_workbook_v1(presentation, tmp_path / name).read_bytes()


def _stored_figures(presentation):
    return (
        build_unified_demand_supply_figure_v1(presentation),
        build_unified_departure_figure_v1(presentation),
    )


def _build_artifacts(
    presentation,
    content: bytes,
    *,
    selected_direction: str | None = None,
    demand_figure=None,
    departure_figure=None,
):
    if demand_figure is None or departure_figure is None:
        stored_demand, stored_departure = _stored_figures(presentation)
        demand_figure = stored_demand if demand_figure is None else demand_figure
        departure_figure = stored_departure if departure_figure is None else departure_figure
    return build_unified_page5_artifacts_v1(
        presentation,
        demand_figure,
        departure_figure,
        content,
        selected_direction=(
            selected_direction
            if selected_direction is not None
            else available_unified_directions_v1(presentation)[0]
        ),
    )


def _tamper_metadata(content: bytes, key: str, value: object) -> bytes:
    workbook = load_workbook(BytesIO(content))
    try:
        sheet = workbook["FINGERPRINTS"]
        row_number = next(
            row for row in range(2, sheet.max_row + 1) if sheet.cell(row, 1).value == key
        )
        sheet.cell(row_number, 2, value)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()
    finally:
        workbook.close()


def _with_recomputed_fingerprint(presentation):
    provisional = replace(presentation, presentation_fingerprint="")
    return replace(
        provisional,
        presentation_fingerprint=unified_presentation._presentation_fingerprint(provisional),
    )


def test_artifact_model_is_frozen_and_slotted(
    fast_render,
    tmp_path: Path,
    accepted_presentation,
) -> None:
    artifacts = _build_artifacts(
        accepted_presentation,
        _xlsx_bytes(accepted_presentation, tmp_path),
    )

    assert UnifiedPage5ArtifactsV1.__dataclass_params__.frozen
    assert "__slots__" in UnifiedPage5ArtifactsV1.__dict__
    with pytest.raises(FrozenInstanceError):
        artifacts.selected_direction = "inbound"


@pytest.mark.parametrize(
    ("key", "value", "message"),
    (
        ("presentation_fingerprint", "stale-presentation", "presentation_fingerprint"),
        ("normalized_b_fingerprint", "stale-b", "b_fingerprint"),
        ("accepted_solution_fingerprint", "stale-c", "accepted_solution_fingerprint"),
        ("source_id", "another-source", "source_id"),
        ("presentation_mode", "OPERATIONAL", "presentation_mode"),
        ("cutover_blocked", True, "cutover_blocked"),
    ),
)
def test_tampered_xlsx_metadata_fails_closed(
    fast_render,
    tmp_path: Path,
    accepted_presentation,
    key: str,
    value: object,
    message: str,
) -> None:
    content = _tamper_metadata(
        _xlsx_bytes(accepted_presentation, tmp_path, f"{key}.xlsx"),
        key,
        value,
    )

    with pytest.raises(UnifiedPage5ArtifactError, match=message):
        _build_artifacts(accepted_presentation, content)


def test_semantic_presentation_mismatch_fails_before_artifact_exposure(
    fast_render,
    tmp_path: Path,
    accepted_presentation,
) -> None:
    changed_block = replace(
        accepted_presentation.blocks[0],
        passenger_demand=accepted_presentation.blocks[0].passenger_demand + 1,
    )
    stale = replace(
        accepted_presentation,
        blocks=(changed_block, *accepted_presentation.blocks[1:]),
    )
    demand, departure = _stored_figures(accepted_presentation)

    with pytest.raises(UnifiedPage5ArtifactError, match="fingerprint"):
        _build_artifacts(
            stale,
            _xlsx_bytes(accepted_presentation, tmp_path),
            selected_direction="outbound",
            demand_figure=demand,
            departure_figure=departure,
        )


@pytest.mark.parametrize("figure_name", ("demand", "departure"))
def test_stored_figure_metadata_mismatch_fails_closed(
    fast_render,
    tmp_path: Path,
    accepted_presentation,
    figure_name: str,
) -> None:
    demand, departure = _stored_figures(accepted_presentation)
    target = demand if figure_name == "demand" else departure
    metadata = dict(target.layout.meta)
    metadata["presentation_fingerprint"] = "stale"
    target.update_layout(meta=metadata)

    with pytest.raises(UnifiedPage5ArtifactError, match="metadata"):
        _build_artifacts(
            accepted_presentation,
            _xlsx_bytes(accepted_presentation, tmp_path),
            demand_figure=demand,
            departure_figure=departure,
        )


def test_changed_departure_time_with_unchanged_metadata_fails_closed(
    fast_render,
    tmp_path: Path,
    accepted_presentation,
) -> None:
    demand, departure = _stored_figures(accepted_presentation)
    original_metadata = dict(departure.layout.meta)
    changed_times = list(departure.data[0].x)
    changed_times[0] += 60
    departure.data[0].x = changed_times

    assert dict(departure.layout.meta) == original_metadata
    with pytest.raises(
        UnifiedPage5ArtifactError,
        match="contents do not match",
    ):
        _build_artifacts(
            accepted_presentation,
            _xlsx_bytes(accepted_presentation, tmp_path),
            demand_figure=demand,
            departure_figure=departure,
        )


def test_changed_c_customdata_with_unchanged_metadata_fails_closed(
    fast_render,
    tmp_path: Path,
    accepted_presentation,
) -> None:
    demand, departure = _stored_figures(accepted_presentation)
    original_metadata = dict(departure.layout.meta)
    c_trace = next(trace for trace in departure.data if str(trace.name).startswith("C ·"))
    changed_customdata = [list(row) for row in c_trace.customdata]
    changed_customdata[0][7] = "fabricated-source-b-trip"
    changed_customdata[0][12] = "fabricated-vehicle"
    c_trace.customdata = changed_customdata

    assert dict(departure.layout.meta) == original_metadata
    with pytest.raises(
        UnifiedPage5ArtifactError,
        match="contents do not match",
    ):
        _build_artifacts(
            accepted_presentation,
            _xlsx_bytes(accepted_presentation, tmp_path),
            demand_figure=demand,
            departure_figure=departure,
        )


def test_removed_b_lane_with_unchanged_metadata_fails_closed(
    fast_render,
    tmp_path: Path,
    accepted_presentation,
) -> None:
    demand, departure = _stored_figures(accepted_presentation)
    original_metadata = dict(departure.layout.meta)
    departure.data = tuple(
        trace for trace in departure.data if not str(trace.name).startswith("B ·")
    )

    assert dict(departure.layout.meta) == original_metadata
    with pytest.raises(
        UnifiedPage5ArtifactError,
        match="contents do not match",
    ):
        _build_artifacts(
            accepted_presentation,
            _xlsx_bytes(accepted_presentation, tmp_path),
            demand_figure=demand,
            departure_figure=departure,
        )


def test_added_c_lane_in_no_c_presentation_fails_closed(
    fast_render,
    tmp_path: Path,
    alpha_presentation,
) -> None:
    demand, departure = _stored_figures(alpha_presentation)
    original_metadata = dict(departure.layout.meta)
    departure.add_trace(
        go.Scatter(
            name="C · fabricated rejected candidate",
            x=[8 * 3600],
            y=["C · fabricated rejected candidate"],
            mode="markers",
            customdata=[["rejected-candidate-raw-fact"]],
        )
    )

    assert dict(departure.layout.meta) == original_metadata
    with pytest.raises(
        UnifiedPage5ArtifactError,
        match="contents do not match",
    ):
        _build_artifacts(
            alpha_presentation,
            _xlsx_bytes(alpha_presentation, tmp_path),
            demand_figure=demand,
            departure_figure=departure,
        )


def test_unchanged_canonical_stored_departure_figure_passes(
    fast_render,
    tmp_path: Path,
    accepted_presentation,
) -> None:
    demand, stored_departure = _stored_figures(accepted_presentation)
    artifacts = _build_artifacts(
        accepted_presentation,
        _xlsx_bytes(accepted_presentation, tmp_path),
        demand_figure=demand,
        departure_figure=stored_departure,
    )

    assert artifacts.departure_figure is not stored_departure
    assert artifacts.departure_figure.to_plotly_json() == stored_departure.to_plotly_json()


def test_returned_and_html_departure_figure_is_canonical_verified_rebuild(
    monkeypatch,
    tmp_path: Path,
    accepted_presentation,
) -> None:
    demand = build_unified_demand_supply_figure_v1(accepted_presentation)
    canonical_departure = build_unified_departure_figure_v1(accepted_presentation)
    stored_departure = go.Figure(canonical_departure)
    captured = {}

    monkeypatch.setattr(
        page5_artifacts,
        "build_unified_departure_figure_v1",
        lambda presentation: canonical_departure,
    )

    def capture_html(
        presentation,
        demand_supply_figure,
        departure_figure,
        *,
        selected_direction,
    ):
        captured["departure_figure"] = departure_figure
        return b"<html>canonical departure</html>"

    monkeypatch.setattr(page5_artifacts, "_build_html_bytes", capture_html)
    monkeypatch.setattr(page5_artifacts, "_build_png_bytes", lambda figure: b"png")
    artifacts = _build_artifacts(
        accepted_presentation,
        _xlsx_bytes(accepted_presentation, tmp_path),
        demand_figure=demand,
        departure_figure=stored_departure,
    )

    assert artifacts.departure_figure is canonical_departure
    assert captured["departure_figure"] is canonical_departure


def test_none_accepted_c_fingerprints_and_facts_align(
    fast_render,
    tmp_path: Path,
    alpha_presentation,
) -> None:
    artifacts = _build_artifacts(
        alpha_presentation,
        _xlsx_bytes(alpha_presentation, tmp_path),
    )

    assert artifacts.accepted_solution_fingerprint is None
    assert alpha_presentation.scenario("C") is None
    assert all(
        trace.name != "Số chuyến C được chấp nhận" for trace in artifacts.demand_supply_figure.data
    )
    assert all(not str(trace.name).startswith("C ·") for trace in artifacts.departure_figure.data)


def test_accepted_c_fingerprint_aligns_every_page5_artifact(
    fast_render,
    tmp_path: Path,
    accepted_presentation,
) -> None:
    artifacts = _build_artifacts(
        accepted_presentation,
        _xlsx_bytes(accepted_presentation, tmp_path),
        selected_direction="outbound",
    )
    scenario_c = accepted_presentation.scenario("C")

    assert scenario_c is not None
    assert artifacts.accepted_solution_fingerprint == scenario_c.source_fingerprint
    assert (
        dict(artifacts.demand_supply_figure.layout.meta)["accepted_solution_fingerprint"]
        == scenario_c.source_fingerprint
    )
    assert (
        dict(artifacts.departure_figure.layout.meta)["accepted_solution_fingerprint"]
        == scenario_c.source_fingerprint
    )


def test_artifact_builder_invokes_no_analysis_or_solver(
    fast_render,
    monkeypatch,
    tmp_path: Path,
    accepted_presentation,
) -> None:
    def forbidden(*args, **kwargs):
        raise AssertionError("execution boundary must not be invoked")

    monkeypatch.setattr(service, "run_analysis", forbidden)
    monkeypatch.setattr(
        optimization_service,
        "analyze_and_optimize_schedule_v1",
        forbidden,
    )

    _build_artifacts(
        accepted_presentation,
        _xlsx_bytes(accepted_presentation, tmp_path),
    )


def test_page5_sources_reference_no_execution_entrypoints() -> None:
    source = " ".join(
        path.read_text(encoding="utf-8")
        for path in (
            Path("src/bus_schedule_engine/unified_page5_artifacts.py"),
            Path("app_pages/05_xuat_file.py"),
        )
    )

    for prohibited in (
        "run_analysis(",
        "run_and_build_artifacts(",
        "analyze_and_optimize_schedule_v1(",
        "run_side_by_side_validation_v1(",
        "build_side_by_side_validation_report_v1(",
        "build_unified_presentation_v1(",
    ):
        assert prohibited not in source


def test_html_is_deterministic_offline_and_contains_required_authority_metadata(
    monkeypatch,
    tmp_path: Path,
    accepted_presentation,
) -> None:
    monkeypatch.setattr(page5_artifacts, "_build_png_bytes", lambda figure: b"png")
    content = _xlsx_bytes(accepted_presentation, tmp_path)

    first = _build_artifacts(accepted_presentation, content, selected_direction="outbound")
    second = _build_artifacts(accepted_presentation, content, selected_direction="outbound")
    html = first.html_bytes.decode("utf-8")

    assert first.html_bytes == second.html_bytes
    assert 'id="contract-v1-demand-supply"' in html
    assert 'id="contract-v1-departures"' in html
    assert html.count("plotly.js v") == 1
    assert "<script src=" not in html
    assert accepted_presentation.route_id in html
    assert accepted_presentation.presentation_mode in html
    assert accepted_presentation.presentation_fingerprint in html
    assert accepted_presentation.source_b_fingerprint in html
    assert accepted_presentation.accepted_solution_fingerprint in html
    assert "không phải phê duyệt khai thác" in html


def test_html_does_not_expose_a_local_source_path(
    monkeypatch,
    tmp_path: Path,
    accepted_presentation,
) -> None:
    monkeypatch.setattr(page5_artifacts, "_build_png_bytes", lambda figure: b"png")
    local_path = r"C:\runtime\temporary\private-input.xlsx"
    unsafe_route_name = "<script>alert('unsafe')</script>"
    presentation = _with_recomputed_fingerprint(
        replace(
            accepted_presentation,
            source_id=local_path,
            route_name=unsafe_route_name,
        )
    )

    artifacts = _build_artifacts(
        presentation,
        _xlsx_bytes(presentation, tmp_path),
    )
    html = artifacts.html_bytes.decode("utf-8")

    assert local_path not in html
    assert unsafe_route_name not in html
    assert "&lt;script&gt;alert(&#x27;unsafe&#x27;)&lt;/script&gt;" in html


def test_no_c_html_states_absence_without_c_timetable_facts(
    monkeypatch,
    tmp_path: Path,
    alpha_presentation,
) -> None:
    monkeypatch.setattr(page5_artifacts, "_build_png_bytes", lambda figure: b"png")
    artifacts = _build_artifacts(
        alpha_presentation,
        _xlsx_bytes(alpha_presentation, tmp_path),
    )
    html = artifacts.html_bytes.decode("utf-8")

    assert "không có Scenario C được chấp nhận" in html
    assert "Số chuyến C được chấp nhận" not in html
    assert "C ·" not in html


def test_png_renderer_receives_exact_selected_figure(
    monkeypatch,
    tmp_path: Path,
    accepted_presentation,
) -> None:
    captured = {}

    def render(figure, **kwargs):
        captured["figure"] = figure
        captured["kwargs"] = kwargs
        return b"deterministic-png"

    monkeypatch.setattr(page5_artifacts.pio, "to_image", render)
    artifacts = _build_artifacts(
        accepted_presentation,
        _xlsx_bytes(accepted_presentation, tmp_path),
        selected_direction="inbound",
    )

    assert artifacts.png_bytes == b"deterministic-png"
    assert captured["figure"] is artifacts.demand_supply_figure
    assert dict(captured["figure"].layout.meta)["displayed_direction"] == "inbound"
    assert captured["kwargs"] == {
        "format": "png",
        "width": 1600,
        "height": 900,
        "scale": 1,
        "validate": True,
        "engine": "kaleido",
    }


def test_png_failure_rejects_the_entire_bundle(
    monkeypatch,
    tmp_path: Path,
    accepted_presentation,
) -> None:
    def fail(*args, **kwargs):
        raise RuntimeError("renderer unavailable")

    monkeypatch.setattr(page5_artifacts.pio, "to_image", fail)
    with pytest.raises(UnifiedPage5ArtifactError, match="renderer unavailable"):
        _build_artifacts(
            accepted_presentation,
            _xlsx_bytes(accepted_presentation, tmp_path),
        )


def test_real_png_render_is_nonempty(
    tmp_path: Path,
    accepted_presentation,
) -> None:
    artifacts = _build_artifacts(
        accepted_presentation,
        _xlsx_bytes(accepted_presentation, tmp_path),
        selected_direction="outbound",
    )

    assert artifacts.png_bytes.startswith(b"\x89PNG\r\n\x1a\n")


def test_contract_filenames_are_fixed(
    fast_render,
    tmp_path: Path,
    accepted_presentation,
) -> None:
    artifacts = _build_artifacts(
        accepted_presentation,
        _xlsx_bytes(accepted_presentation, tmp_path),
    )

    assert artifacts.xlsx_filename == UNIFIED_PAGE5_XLSX_FILENAME
    assert artifacts.html_filename == UNIFIED_PAGE5_HTML_FILENAME
    assert artifacts.png_filename == UNIFIED_PAGE5_PNG_FILENAME


def test_rejected_candidate_timetable_is_absent_from_page5_artifacts(
    monkeypatch,
    tmp_path: Path,
) -> None:
    monkeypatch.setattr(page5_artifacts, "_build_png_bytes", lambda figure: b"png")
    presentation = build_unified_presentation_v1(*rejected_result_and_report())
    artifacts = _build_artifacts(
        presentation,
        _xlsx_bytes(presentation, tmp_path),
    )
    html = artifacts.html_bytes.decode("utf-8")

    assert "rejected-diagnostic-candidate" not in html
    assert all(not str(trace.name).startswith("C ·") for trace in artifacts.departure_figure.data)
    assert b"rejected-diagnostic-candidate" not in artifacts.xlsx_bytes


def test_combined_and_directional_demand_remain_exact(
    fast_render,
    tmp_path: Path,
    accepted_presentation,
) -> None:
    combined = build_unified_presentation_v1(*build_result_and_report(combined_demand=True))
    combined_artifacts = _build_artifacts(
        combined,
        _xlsx_bytes(combined, tmp_path, "combined.xlsx"),
        selected_direction="combined",
    )
    combined_demand = next(
        trace
        for trace in combined_artifacts.demand_supply_figure.data
        if trace.name == "Nhu cầu hành khách"
    )

    assert {row[1] for row in combined_demand.customdata} == {"combined"}
    assert available_unified_directions_v1(combined) == ("combined",)
    assert available_unified_directions_v1(accepted_presentation) == (
        "outbound",
        "inbound",
    )


def test_beta_gap_and_no_c_authority_are_preserved(
    fast_render,
    tmp_path: Path,
) -> None:
    presentation = build_unified_presentation_v1(
        *build_corpus_result_and_report("corpus_beta_46.json")
    )
    artifacts = _build_artifacts(
        presentation,
        _xlsx_bytes(presentation, tmp_path),
        selected_direction="outbound",
    )

    assert artifacts.accepted_solution_fingerprint is None
    assert presentation.scenario("C") is None
    assert any(
        gap.direction == "outbound"
        and gap.start_time_seconds == 17 * 3600
        and gap.end_time_seconds == 18 * 3600
        for gap in presentation.demand_gaps
    )
