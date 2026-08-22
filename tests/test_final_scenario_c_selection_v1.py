from __future__ import annotations

import hashlib
import importlib.util
import json
import unittest
from decimal import Decimal
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parents[1]
OUTPUT = ROOT / "outputs" / "final_scenario_c_selection_v1_20260821"
BRIDGE = (
    ROOT
    / "private_inputs"
    / "final_selection"
    / "final_selection_demand_authority_bridge_v1_20260821.json"
)
EXPECTED_BRIDGE_SHA256 = "77891c65713420c5bc9d8774b6964beefdcf221bf103f37dd32addbf0bcc00ea"


def _load_selector_module():
    path = ROOT / "src" / "bus_schedule_engine" / "contracts_v1" / "final_scenario_c_selection.py"
    spec = importlib.util.spec_from_file_location("final_scenario_c_selection_v1_test_target", path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"cannot load selector module from {path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


SELECTOR = _load_selector_module()


class FinalScenarioCSelectionV1Tests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.manifest = SELECTOR.build_final_selection_manifest(ROOT)

    def test_authority_bridge_hash_and_status(self) -> None:
        self.assertEqual(hashlib.sha256(BRIDGE.read_bytes()).hexdigest(), EXPECTED_BRIDGE_SHA256)
        bridge = json.loads(BRIDGE.read_text(encoding="utf-8"))
        self.assertEqual(bridge["status"], "TEMPORARY_AUTHORITATIVE_BRIDGE")

    def test_selection_is_deterministic(self) -> None:
        repeated = SELECTOR.build_final_selection_manifest(ROOT)

        def canonical(value: object) -> str:
            return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))

        self.assertEqual(canonical(repeated), canonical(self.manifest))

    def test_selected_pairs_and_operational_equivalence(self) -> None:
        expected = {
            "6": ("C2_CONSERVATIVE", "C1_DEMAND_FIT", "0.015029345"),
            "10": ("C3_BALANCED", "C2_CONSERVATIVE", "0.017868607"),
        }
        for route, (outbound, inbound, mismatch) in expected.items():
            result = self.manifest["checkpoint"]["routes"][route]
            self.assertEqual(result["selected_outbound_representative"], outbound)
            self.assertEqual(result["selected_inbound_representative"], inbound)
            self.assertEqual(result["combined_mismatch"], mismatch)
            if route == "6":
                self.assertEqual(
                    result["selected_outbound_provenance"], ["C2_CONSERVATIVE", "C3_BALANCED"]
                )
            else:
                self.assertEqual(
                    result["selected_inbound_provenance"], ["C2_CONSERVATIVE", "C3_BALANCED"]
                )

    def test_ineligible_and_dominated_pairs_are_never_selected(self) -> None:
        for route in ("6", "10"):
            rows = self.manifest["checkpoint"]["routes"][route]["decision_space"]
            selected = [item for item in rows if item.get("selection_status") == "SELECTED"]
            self.assertEqual(len(selected), 1)
            self.assertTrue(selected[0]["eligible"])
            self.assertEqual(selected[0]["pareto_status"], "PARETO_NONDOMINATED")
            self.assertFalse(
                any(
                    item.get("selection_status") == "SELECTED"
                    for item in rows
                    if not item["eligible"]
                )
            )
            self.assertFalse(
                any(
                    item.get("selection_status") == "SELECTED"
                    for item in rows
                    if item["pareto_status"] == "PARETO_DOMINATED"
                )
            )

    def test_compiler_departures_and_fleet_blocks_reconcile_exactly(self) -> None:
        for route in ("6", "10"):
            checkpoint = self.manifest["checkpoint"]["routes"][route]
            product = self.manifest["product_routes"][route]
            compiled_departures: set[tuple[str, int]] = set()
            for direction_name, direction in (
                ("outbound", "terminal_1_to_2"),
                ("inbound", "terminal_2_to_1"),
            ):
                compiler_path = (
                    ROOT / checkpoint["compiler_artifact_references"][direction_name]["path"]
                )
                compiler = json.loads(compiler_path.read_text(encoding="utf-8"))
                compiled_departures.update(
                    (direction, int(item["departure_minute"]))
                    for item in compiler["exact_departures"]
                )
            product_departures = {
                (
                    {"OUTBOUND": "terminal_1_to_2", "INBOUND": "terminal_2_to_1"}[
                        item["direction"]
                    ],
                    item["departure_minute"],
                )
                for item in product["scenario_c_trips"]
            }
            block_departures = {
                (item["direction"], item["departure_minute"]) for item in product["vehicle_blocks"]
            }
            self.assertEqual(product_departures, compiled_departures)
            self.assertEqual(block_departures, compiled_departures)
            self.assertEqual(len(product_departures), len(product["scenario_c_trips"]))
            product_arrivals = {
                (
                    {"OUTBOUND": "terminal_1_to_2", "INBOUND": "terminal_2_to_1"}[
                        item["direction"]
                    ],
                    item["departure_minute"],
                    item["arrival_minute"],
                )
                for item in product["scenario_c_trips"]
            }
            block_arrivals = {
                (item["direction"], item["departure_minute"], item["arrival_minute"])
                for item in product["vehicle_blocks"]
            }
            self.assertEqual(product_arrivals, block_arrivals)

    def test_trip_allocations_reconcile_by_direction(self) -> None:
        expected_per_direction = {"6": 78, "10": 51}
        for route, expected in expected_per_direction.items():
            product = self.manifest["product_routes"][route]
            for direction in ("OUTBOUND", "INBOUND"):
                allocations = [
                    item for item in product["demand_allocation"] if item["direction"] == direction
                ]
                self.assertEqual(
                    sum(item["scenario_c_trip_count"] for item in allocations), expected
                )
                self.assertEqual(
                    sum(item["scenario_b_trip_count"] for item in allocations), expected
                )
                observed_total = sum(
                    Decimal(str(item["observed_demand_share"])) for item in allocations
                )
                self.assertLessEqual(abs(observed_total - Decimal("1")), Decimal("0.0001"))

    def test_route_totals_fleet_limits_and_layovers(self) -> None:
        expected = {
            "6": {"direction": 78, "total": 156, "limit": 20},
            "10": {"direction": 51, "total": 102, "limit": 13},
        }
        for route, limits in expected.items():
            checkpoint = self.manifest["checkpoint"]["routes"][route]
            product = self.manifest["product_routes"][route]
            self.assertEqual(len(product["scenario_c_trips"]), limits["total"])
            for direction in ("OUTBOUND", "INBOUND"):
                self.assertEqual(
                    sum(item["direction"] == direction for item in product["scenario_c_trips"]),
                    limits["direction"],
                )
            self.assertLessEqual(checkpoint["fleet_c"], limits["limit"])
            self.assertGreaterEqual(product["minimum_layover_minutes"], 5)
            self.assertTrue(
                all(
                    item["next_trip_layover_minutes"] is None
                    or item["next_trip_layover_minutes"] >= 5
                    for item in product["vehicle_blocks"]
                )
            )

    def test_product_workbooks_contain_every_selected_trip_once(self) -> None:
        required_route_sheets = {
            "SUMMARY",
            "SCENARIO_B",
            "SCENARIO_C",
            "SERVICE_REGIMES",
            "DEMAND_ALLOCATION",
            "VEHICLE_BLOCKS",
            "COMPARISON",
            "DECISION_EVIDENCE",
            "CHARTS",
        }
        for route in ("6", "10"):
            workbook_path = OUTPUT / f"Route_{route}_Final_Scenario_C.xlsx"
            workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=False)
            self.assertEqual(set(workbook.sheetnames), required_route_sheets)
            sheet = workbook["SCENARIO_C"]
            rows = list(sheet.iter_rows(min_row=5, values_only=True))
            populated = [row for row in rows if row[0] is not None]
            expected_trips = self.manifest["product_routes"][route]["scenario_c_trips"]
            self.assertEqual(len(populated), len(expected_trips))
            self.assertEqual(len({(row[0], row[1]) for row in populated}), len(expected_trips))
            workbook.close()

        combined = openpyxl.load_workbook(
            OUTPUT / "Routes_6_10_Final_Scenario_C.xlsx", read_only=True
        )
        self.assertIn("OVERVIEW", combined.sheetnames)
        self.assertTrue(
            all(
                f"R{route}_{name}" in combined.sheetnames
                for route in ("6", "10")
                for name in required_route_sheets
            )
        )
        combined.close()

    def test_final_checkpoint_hashes_match_product_files(self) -> None:
        checkpoint_path = OUTPUT / "final_scenario_c_selection_v1.json"
        checkpoint = json.loads(checkpoint_path.read_text(encoding="utf-8"))
        for route in ("6", "10"):
            reference = checkpoint["routes"][route]["product_xlsx"]
            self.assertEqual(
                hashlib.sha256((ROOT / reference["path"]).read_bytes()).hexdigest(),
                reference["sha256"],
            )
        reference = checkpoint["combined_product_xlsx"]
        self.assertEqual(
            hashlib.sha256((ROOT / reference["path"]).read_bytes()).hexdigest(), reference["sha256"]
        )


if __name__ == "__main__":
    unittest.main()
