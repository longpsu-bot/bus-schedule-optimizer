from __future__ import annotations

import datetime as dt
import json
from pathlib import Path

import run_pr62_g_final_pilot as final
from openpyxl import load_workbook

_ROOT = Path(__file__).resolve().parents[1]
_EVIDENCE = _ROOT / final.OUTPUT_JSON


def _payload() -> dict:
    return json.loads(_EVIDENCE.read_text(encoding="utf-8"))


def _seconds(value: dt.time | float) -> int:
    if isinstance(value, dt.time):
        return value.hour * 3600 + value.minute * 60 + value.second
    return round(float(value) * 86400)


def test_selected_pairs_endpoints_trip_totals_and_fleets() -> None:
    payload = _payload()
    expected = {
        "6": (final.ROUTES["6"]["fingerprint"], 78, 78, 19),
        "10": (final.ROUTES["10"]["fingerprint"], 51, 51, 12),
    }
    for route_id, (fingerprint, out_count, in_count, fleet) in expected.items():
        route = payload["routes"][route_id]
        assert route["pair_fingerprint"] == fingerprint
        assert len(route["directions"]["outbound"]["exact_departures"]) == out_count
        assert len(route["directions"]["inbound"]["exact_departures"]) == in_count
        assert route["metrics"]["fleet_required"] == fleet
        for direction in ("outbound", "inbound"):
            departures = route["directions"][direction]["exact_departures"]
            assert (
                departures[0]
                == route["directions"][direction]["service_regimes"][0]["first_departure"]
            )
            assert (
                departures[-1]
                == route["directions"][direction]["service_regimes"][-1]["last_departure"]
            )
            assert all(value % 60 == 0 for value in departures)
            assert all(a < b for a, b in zip(departures, departures[1:], strict=False))


def test_route6_static_ten_minute_revalidation() -> None:
    route = _payload()["routes"]["6"]
    assert route["g0_robustness_classification"] == "BASELINE_TIMETABLE_ROBUST_AT_10"
    assert route["ten_minute_fleet_requirement"] == 20
    assert route["ten_minute_fleet_margin"] == 0
    assert route["departure_shifts_vs_baseline"] == 0


def test_required_sheets_timetable_times_and_fleet_coverage() -> None:
    payload = _payload()
    expected_sheets = {
        "6": [
            "Summary",
            "Timetable",
            "ServiceRegimes",
            "Demand_Comparison",
            "Comparison",
            "Fleet_Plan",
            "Layover_Robustness",
        ],
        "10": [
            "Summary",
            "Timetable",
            "ServiceRegimes",
            "Demand_Comparison",
            "Comparison",
            "Fleet_Plan",
        ],
    }
    for route_id in ("6", "10"):
        route = payload["routes"][route_id]
        path = _ROOT / route["workbook"]["path"]
        workbook = load_workbook(path, data_only=False, keep_links=True)
        assert workbook.sheetnames == expected_sheets[route_id]
        assert not workbook._external_links
        sheet = workbook["Timetable"]
        outbound = route["directions"]["outbound"]["exact_departures"]
        inbound = route["directions"]["inbound"]["exact_departures"]
        assert [
            _seconds(sheet.cell(row=index + 2, column=2).value) for index in range(len(outbound))
        ] == outbound
        assert [
            _seconds(sheet.cell(row=index + 2, column=7).value) for index in range(len(inbound))
        ] == inbound
        fleet = workbook["Fleet_Plan"]
        trips = [fleet.cell(row=row, column=2).value for row in range(2, fleet.max_row + 1)]
        expected_trips = {
            *(f"OUTBOUND-{index:03d}" for index in range(1, len(outbound) + 1)),
            *(f"INBOUND-{index:03d}" for index in range(1, len(inbound) + 1)),
        }
        assert len(trips) == len(expected_trips)
        assert set(trips) == expected_trips


def test_route6_layover_robustness_sheet_values() -> None:
    route = _payload()["routes"]["6"]
    workbook = load_workbook(_ROOT / route["workbook"]["path"], data_only=False)
    sheet = workbook["Layover_Robustness"]
    assert [sheet.cell(row=2, column=2).value, sheet.cell(row=2, column=3).value] == [70, 5]
    assert [sheet.cell(row=2, column=5).value, sheet.cell(row=2, column=6).value] == [19, 1]
    assert [sheet.cell(row=3, column=2).value, sheet.cell(row=3, column=3).value] == [70, 10]
    assert [sheet.cell(row=3, column=5).value, sheet.cell(row=3, column=6).value] == [20, 0]
    assert sheet.cell(row=2, column=7).value == sheet.cell(row=3, column=7).value == "No"
    assert sheet.cell(row=2, column=8).value == sheet.cell(row=3, column=8).value == 0


def test_logical_fingerprints_and_evidence_render_are_deterministic() -> None:
    payload = _payload()
    assert final._canonical(payload) == final._canonical(payload)
    assert final.render_markdown(payload) == final.render_markdown(payload)
    for route_id in ("6", "10"):
        route = payload["routes"][route_id]
        path = _ROOT / route["workbook"]["path"]
        assert (
            final.logical_workbook_fingerprint(path, route)
            == route["workbook"]["logical_fingerprint"]
        )
        assert route["workbook"]["logical_fingerprint_repeated_generation_equal"] is True
