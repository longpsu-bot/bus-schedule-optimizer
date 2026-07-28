from __future__ import annotations

from datetime import UTC, datetime

import pytest
from openpyxl import load_workbook

from bus_schedule_engine import run_side_by_side_validation_v1
from bus_schedule_engine.contracts_v1 import (
    DemandConfidence,
    InputSourceType,
    normalize_imported_workbook_v1,
)
from bus_schedule_engine.contracts_v1.terminal_occupancy import (
    TERMINAL_2_OCCUPANCY_CAPACITY_NOT_EVALUATED,
    TERMINAL_OCCUPANCY_CAPACITY_NOT_EVALUATED,
)
from bus_schedule_engine.excel_exporter import create_input_template
from bus_schedule_engine.fleet import assign_fleet
from bus_schedule_engine.importer import import_workbook
from bus_schedule_engine.input_authority import (
    AVAILABLE_FLEET_LIMIT_REQUIRED_FOR_OPTIMIZATION,
    DEMAND_CONFIDENCE_REQUIRED_FOR_OPTIMIZATION,
    DEMAND_RESPONSE_MODE_REQUIRED_FOR_OPTIMIZATION,
    DEMAND_SOURCE_TYPE_REQUIRED_FOR_OPTIMIZATION,
    OPERATING_DAY_TYPE_REQUIRED_FOR_OPTIMIZATION,
    SCENARIO_A_AVAILABLE_FLEET_LIMIT_REQUIRED_FOR_OPTIMIZATION,
    SCENARIO_A_OPERATING_DAY_TYPE_REQUIRED_FOR_OPTIMIZATION,
    WorkbookOptimizationAuthorityError,
    assess_workbook_input_readiness_v1,
    normalization_options_from_workbook_v1,
)

IMPORTED_AT = datetime(2026, 7, 28, 8, 0, tzinfo=UTC)


def _set_parameter(sheet, key: str, value: object | None) -> None:
    row = next(cell.row for cell in sheet["A"] if cell.value == key)
    sheet.cell(row, 2).value = value


def _remove_scenario_a(workbook) -> None:
    workbook.remove(workbook["THONG_SO_A"])
    workbook.remove(workbook["BIEU_DO_A"])


def _clear_demand(workbook) -> None:
    sheet = workbook["SAN_LUONG"]
    if sheet.max_row >= 4:
        sheet.delete_rows(4, sheet.max_row - 3)


def _blank_demand_authority(workbook) -> None:
    sheet = workbook["THONG_TIN_DU_LIEU"]
    for key in (
        "demand_source_type",
        "demand_confidence",
        "demand_response_mode",
    ):
        _set_parameter(sheet, key, None)


def test_blank_available_fleet_imports_and_blocks_strict_builder_without_inference(
    tmp_path,
) -> None:
    path = create_input_template(tmp_path / "blank-fleet.xlsx")
    workbook = load_workbook(path)
    _set_parameter(workbook["THONG_SO_B"], "available_fleet_limit", None)
    for row in range(4, workbook["BIEU_DO_B"].max_row + 1):
        workbook["BIEU_DO_B"].cell(row, 7).value = f"XE-{row:03d}"
    workbook.save(path)
    workbook.close()

    imported = import_workbook(path)
    original_trips = list(imported.trips_b)
    minimum_fleet = assign_fleet(imported.trips_b, imported.parameters_b).minimum_vehicles
    readiness = assess_workbook_input_readiness_v1(imported)

    assert imported.parameters_b.available_fleet_limit is None
    assert minimum_fleet >= 1
    assert readiness.import_ready is True
    assert readiness.optimization_ready is False
    assert readiness.missing_optimization_authority_codes == (
        AVAILABLE_FLEET_LIMIT_REQUIRED_FOR_OPTIMIZATION,
    )
    assert imported.trips_b == original_trips
    assert not hasattr(imported, "scenario_c")

    with pytest.raises(WorkbookOptimizationAuthorityError) as exc_info:
        normalization_options_from_workbook_v1(
            imported,
            source_id="blank-fleet",
            imported_at=IMPORTED_AT,
        )
    assert exc_info.value.codes == (AVAILABLE_FLEET_LIMIT_REQUIRED_FOR_OPTIMIZATION,)


def test_supplied_fleet_round_trips_and_minimum_fleet_remains_separate(tmp_path) -> None:
    path = create_input_template(tmp_path / "complete.xlsx")
    imported = import_workbook(path)
    readiness = assess_workbook_input_readiness_v1(imported)
    options = normalization_options_from_workbook_v1(
        imported,
        source_id="complete-template",
        imported_at=IMPORTED_AT,
    )
    normalized = normalize_imported_workbook_v1(imported, options)
    minimum_fleet = assign_fleet(imported.trips_b, imported.parameters_b).minimum_vehicles

    assert imported.parameters_b.available_fleet_limit == 8
    assert readiness.optimization_ready is True
    assert options.available_fleet_limit_b == 8
    assert normalized.scenario_b.available_fleet_limit == 8
    assert minimum_fleet < normalized.scenario_b.available_fleet_limit


def test_blank_optional_fleet_and_terminal_fields_do_not_block_optimization(
    tmp_path,
) -> None:
    path = create_input_template(tmp_path / "blank-optionals.xlsx")
    workbook = load_workbook(path)
    for sheet_name in ("THONG_SO_A", "THONG_SO_B"):
        _set_parameter(workbook[sheet_name], "approved_active_fleet", None)
    for key in (
        "terminal_1_max_occupancy_vehicles",
        "terminal_2_max_occupancy_vehicles",
    ):
        _set_parameter(workbook["THONG_SO_B"], key, None)
    workbook.save(path)
    workbook.close()

    imported = import_workbook(path)
    readiness = assess_workbook_input_readiness_v1(imported)
    options = normalization_options_from_workbook_v1(
        imported,
        source_id="blank-optionals",
        imported_at=IMPORTED_AT,
    )

    assert imported.parameters_b.approved_active_fleet is None
    assert imported.parameters_b.terminal_1_max_occupancy_vehicles is None
    assert imported.parameters_b.terminal_2_max_occupancy_vehicles is None
    assert readiness.optimization_ready is True
    assert readiness.optional_limitations == (TERMINAL_OCCUPANCY_CAPACITY_NOT_EVALUATED,)
    assert options.approved_active_fleet_b is None


def test_one_terminal_limit_reports_partial_evaluation(tmp_path) -> None:
    path = create_input_template(tmp_path / "partial-terminal.xlsx")
    workbook = load_workbook(path)
    _set_parameter(workbook["THONG_SO_B"], "terminal_2_max_occupancy_vehicles", None)
    workbook.save(path)
    workbook.close()

    imported = import_workbook(path)
    readiness = assess_workbook_input_readiness_v1(imported)

    assert readiness.optimization_ready is True
    assert readiness.optional_limitations == (TERMINAL_2_OCCUPANCY_CAPACITY_NOT_EVALUATED,)


def test_missing_scenario_a_does_not_block_scenario_b_readiness(tmp_path) -> None:
    path = create_input_template(tmp_path / "b-only.xlsx")
    workbook = load_workbook(path)
    _remove_scenario_a(workbook)
    workbook.save(path)
    workbook.close()

    imported = import_workbook(path)
    readiness = assess_workbook_input_readiness_v1(imported)

    assert imported.parameters_a is None
    assert readiness.optimization_ready is True


def test_present_scenario_a_requires_its_optimization_authority(tmp_path) -> None:
    path = create_input_template(tmp_path / "a-missing-authority.xlsx")
    workbook = load_workbook(path)
    _set_parameter(workbook["THONG_SO_A"], "available_fleet_limit", None)
    _set_parameter(workbook["THONG_SO_A"], "operating_day_type", None)
    workbook.save(path)
    workbook.close()

    readiness = assess_workbook_input_readiness_v1(import_workbook(path))

    assert readiness.missing_optimization_authority_codes == (
        SCENARIO_A_AVAILABLE_FLEET_LIMIT_REQUIRED_FOR_OPTIMIZATION,
        SCENARIO_A_OPERATING_DAY_TYPE_REQUIRED_FOR_OPTIMIZATION,
    )


def test_no_demand_does_not_require_demand_authority(tmp_path) -> None:
    path = create_input_template(tmp_path / "no-demand.xlsx")
    workbook = load_workbook(path)
    _clear_demand(workbook)
    _blank_demand_authority(workbook)
    workbook.save(path)
    workbook.close()

    imported = import_workbook(path)
    readiness = assess_workbook_input_readiness_v1(imported)

    assert imported.demand == []
    assert readiness.optimization_ready is True
    assert readiness.missing_optimization_authority_codes == ()


@pytest.mark.parametrize(
    ("key", "expected_code"),
    [
        ("demand_source_type", DEMAND_SOURCE_TYPE_REQUIRED_FOR_OPTIMIZATION),
        ("demand_confidence", DEMAND_CONFIDENCE_REQUIRED_FOR_OPTIMIZATION),
        ("demand_response_mode", DEMAND_RESPONSE_MODE_REQUIRED_FOR_OPTIMIZATION),
    ],
)
def test_demand_observations_require_declared_authority(
    tmp_path,
    key: str,
    expected_code: str,
) -> None:
    path = create_input_template(tmp_path / f"missing-{key}.xlsx")
    workbook = load_workbook(path)
    _set_parameter(workbook["THONG_TIN_DU_LIEU"], key, None)
    workbook.save(path)
    workbook.close()

    readiness = assess_workbook_input_readiness_v1(import_workbook(path))

    assert readiness.optimization_ready is False
    assert expected_code in readiness.missing_optimization_authority_codes


def test_declared_demand_confidence_is_preserved_without_upgrade(tmp_path) -> None:
    path = create_input_template(tmp_path / "low-confidence.xlsx")
    workbook = load_workbook(path)
    _set_parameter(workbook["THONG_TIN_DU_LIEU"], "demand_confidence", "low")
    workbook.save(path)
    workbook.close()

    imported = import_workbook(path)
    options = normalization_options_from_workbook_v1(
        imported,
        source_id="low-confidence",
        imported_at=IMPORTED_AT,
    )
    normalized = normalize_imported_workbook_v1(imported, options)

    assert imported.authority_metadata.demand_confidence == DemandConfidence.LOW
    assert options.demand_confidence == DemandConfidence.LOW
    assert {
        observation.demand_confidence for observation in normalized.observed_demand.observations
    } == {DemandConfidence.LOW}


def test_blank_optional_dataset_id_uses_runtime_source_identity_without_fake_suffix(
    tmp_path,
) -> None:
    path = create_input_template(tmp_path / "blank-dataset-id.xlsx")
    workbook = load_workbook(path)
    _set_parameter(workbook["THONG_TIN_DU_LIEU"], "demand_dataset_id", None)
    _set_parameter(workbook["THONG_TIN_DU_LIEU"], "source_notes", None)
    workbook.save(path)
    workbook.close()

    imported = import_workbook(path)
    original_metadata = imported.authority_metadata
    options = normalization_options_from_workbook_v1(
        imported,
        source_id="  runtime-source-identity  ",
        imported_at=IMPORTED_AT,
    )
    normalized = normalize_imported_workbook_v1(imported, options)

    assert imported.authority_metadata.demand_dataset_id is None
    assert imported.authority_metadata.source_notes is None
    assert imported.authority_metadata == original_metadata
    assert options.source_id == "runtime-source-identity"
    assert options.demand_dataset_id == "runtime-source-identity"
    assert options.imported_at is IMPORTED_AT
    assert options.source_type == InputSourceType.XLSX
    assert normalized.observed_demand.demand_dataset_id == "runtime-source-identity"
    assert not normalized.observed_demand.demand_dataset_id.endswith(":demand")


@pytest.mark.parametrize("source_id", ["", "   "])
def test_blank_runtime_source_id_is_rejected(
    tmp_path,
    source_id: str,
) -> None:
    imported = import_workbook(create_input_template(tmp_path / "complete.xlsx"))

    with pytest.raises(ValueError, match="source_id must be a non-empty string"):
        normalization_options_from_workbook_v1(
            imported,
            source_id=source_id,
            imported_at=IMPORTED_AT,
        )


@pytest.mark.parametrize("source_id", [None, 123])
def test_non_string_runtime_source_id_is_rejected(
    tmp_path,
    source_id: object,
) -> None:
    imported = import_workbook(create_input_template(tmp_path / "complete.xlsx"))

    with pytest.raises(TypeError, match="source_id must be a string"):
        normalization_options_from_workbook_v1(
            imported,
            source_id=source_id,  # type: ignore[arg-type]
            imported_at=IMPORTED_AT,
        )


def test_declared_dataset_id_takes_precedence_over_cleaned_runtime_source_id(
    tmp_path,
) -> None:
    path = create_input_template(tmp_path / "declared-dataset-id.xlsx")
    workbook = load_workbook(path)
    _set_parameter(workbook["THONG_TIN_DU_LIEU"], "demand_dataset_id", "  DEMAND-42  ")
    workbook.save(path)
    workbook.close()

    imported = import_workbook(path)
    original_metadata = imported.authority_metadata
    options = normalization_options_from_workbook_v1(
        imported,
        source_id="  upload-2026-07-28  ",
        imported_at=IMPORTED_AT,
    )

    assert imported.authority_metadata.demand_dataset_id == "DEMAND-42"
    assert imported.authority_metadata == original_metadata
    assert options.source_id == "upload-2026-07-28"
    assert options.demand_dataset_id == "DEMAND-42"


def test_old_blank_authority_workbook_imports_but_is_not_optimization_ready(
    tmp_path,
) -> None:
    path = create_input_template(tmp_path / "old-blank.xlsx")
    workbook = load_workbook(path)
    workbook.remove(workbook["THONG_TIN_DU_LIEU"])
    _set_parameter(workbook["THONG_SO_B"], "available_fleet_limit", None)
    _set_parameter(workbook["THONG_SO_B"], "operating_day_type", None)
    workbook.save(path)
    workbook.close()

    imported = import_workbook(path)
    readiness = assess_workbook_input_readiness_v1(imported)

    assert readiness.import_ready is True
    assert readiness.optimization_ready is False
    assert AVAILABLE_FLEET_LIMIT_REQUIRED_FOR_OPTIMIZATION in (
        readiness.missing_optimization_authority_codes
    )
    assert OPERATING_DAY_TYPE_REQUIRED_FOR_OPTIMIZATION in (
        readiness.missing_optimization_authority_codes
    )
    assert readiness.missing_optimization_authority_codes == tuple(
        sorted(readiness.missing_optimization_authority_codes)
    )
    with pytest.raises(WorkbookOptimizationAuthorityError) as exc_info:
        normalization_options_from_workbook_v1(
            imported,
            source_id="old-blank",
            imported_at=IMPORTED_AT,
        )
    assert exc_info.value.codes == readiness.missing_optimization_authority_codes


def test_generated_template_runs_side_by_side_without_blocking_source_discrepancy(
    tmp_path,
) -> None:
    imported = import_workbook(create_input_template(tmp_path / "complete.xlsx"))
    options = normalization_options_from_workbook_v1(
        imported,
        source_id="generated-template",
        imported_at=IMPORTED_AT,
    )

    report = run_side_by_side_validation_v1(imported, options)

    assert report.blocking_discrepancy_codes == ()
