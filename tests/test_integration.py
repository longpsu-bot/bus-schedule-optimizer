from __future__ import annotations

from dataclasses import replace

import pytest
from openpyxl import load_workbook

from bus_schedule_engine.excel_exporter import create_input_template
from bus_schedule_engine.fleet import assign_fleet
from bus_schedule_engine.importer import ImportedWorkbook, InputDataError, import_workbook
from bus_schedule_engine.models import Direction, ScenarioCStatus, Trip
from bus_schedule_engine.service import _inherit_active_vehicle_ids, run_analysis


def test_template_to_offline_analysis(tmp_path) -> None:
    template = create_input_template(tmp_path / "input.xlsx")
    imported = import_workbook(template)
    assert imported.parameters_a.runtime_options == (55, 65)
    assert imported.parameters_b.runtime_options == (55, 65)
    bundle = run_analysis(imported)
    assert {result.name for result in bundle.scenarios} >= {"A", "B", "C"}
    assert bundle.get("B").validation.passed
    assert bundle.get("C").strategy_id == "fixed_resource_redistribution"
    assert bundle.get("C").active_vehicle_count == bundle.get("B").active_vehicle_count
    assert bundle.get("C").fleet.minimum_vehicles <= bundle.get("C").active_vehicle_count
    assert len(bundle.get("C").trips) == len(bundle.get("B").trips)


def test_legacy_template_with_single_runtime_remains_supported(tmp_path) -> None:
    source = create_input_template(tmp_path / "legacy-runtime.xlsx")
    workbook = load_workbook(source)
    for sheet_name in ("THONG_SO_A", "THONG_SO_B"):
        sheet = workbook[sheet_name]
        row = next(cell.row for cell in sheet["A"] if cell.value == "allowed_trip_runtime_minutes")
        sheet.delete_rows(row)
    workbook.save(source)
    workbook.close()

    imported = import_workbook(source)

    assert imported.parameters_a.runtime_options == (65,)
    assert imported.parameters_b.runtime_options == (65,)


def test_template_keeps_runtime_range_as_text_and_recovers_decimal_comma(
    tmp_path,
) -> None:
    source = create_input_template(tmp_path / "decimal-comma-runtime.xlsx")
    workbook = load_workbook(source)
    for sheet_name in ("THONG_SO_A", "THONG_SO_B"):
        sheet = workbook[sheet_name]
        row = next(cell.row for cell in sheet["A"] if cell.value == "allowed_trip_runtime_minutes")
        assert sheet.cell(row, 2).number_format == "@"
        sheet.cell(row, 2).value = 55.65
    workbook.save(source)
    workbook.close()

    imported = import_workbook(source)

    assert imported.parameters_a.runtime_options == (55, 65)
    assert imported.parameters_b.runtime_options == (55, 65)


def test_vehicle_id_inheritance_reports_cardinality_mismatch(tmp_path) -> None:
    template = create_input_template(tmp_path / "input.xlsx")
    result_c = run_analysis(import_workbook(template)).get("C")
    result_c.active_vehicle_ids = ()

    with pytest.raises(InputDataError, match="Không thể ánh xạ đội xe Scenario C"):
        _inherit_active_vehicle_ids(result_c)


def test_b_only_workbook_runs_without_a_demand_or_configuration(tmp_path) -> None:
    source = create_input_template(tmp_path / "b-only.xlsx")
    workbook = load_workbook(source)
    for sheet_name in list(workbook.sheetnames):
        if sheet_name not in {"THONG_SO_B", "BIEU_DO_B"}:
            workbook.remove(workbook[sheet_name])
    workbook.save(source)
    workbook.close()

    imported = import_workbook(source)
    assert imported.parameters_a is None
    assert imported.trips_a == []
    assert imported.demand == []
    assert imported.configuration == {}

    bundle = run_analysis(imported)
    assert {result.name for result in bundle.scenarios} == {"B", "C"}
    assert bundle.get("C").generation_status == ScenarioCStatus.INSUFFICIENT_DATA


def test_infeasible_fallback_preserves_b_arrivals_and_does_not_expand_fleet(
    make_parameters,
) -> None:
    parameters = make_parameters(
        trip_runtime_minutes=60,
        minimum_layover_minutes=10,
        terminal_2_first_departure=6 * 3600 + 30 * 60,
        terminal_2_last_departure=7 * 3600 + 30 * 60,
    )
    definitions = (
        ("B-01", parameters.terminal_1_name, Direction.TERMINAL_1_TO_2, 6 * 3600),
        ("B-02", parameters.terminal_2_name, Direction.TERMINAL_2_TO_1, 6 * 3600 + 30 * 60),
        ("B-03", parameters.terminal_1_name, Direction.TERMINAL_1_TO_2, 7 * 3600),
        ("B-04", parameters.terminal_2_name, Direction.TERMINAL_2_TO_1, 7 * 3600 + 30 * 60),
    )
    trips_b = [
        Trip(
            scenario="B",
            trip_id=trip_id,
            departure_terminal=terminal,
            direction=direction,
            departure_seconds=departure,
            arrival_seconds=departure + 20 * 60,
            vehicle_id="XE-001",
        )
        for trip_id, terminal, direction, departure in definitions
    ]
    assert assign_fleet(trips_b, parameters).minimum_vehicles == 1
    imported = ImportedWorkbook(
        parameters_a=replace(parameters),
        trips_a=[replace(trip, scenario="A") for trip in trips_b],
        parameters_b=parameters,
        trips_b=trips_b,
        demand=[],
        configuration={},
    )

    bundle = run_analysis(imported)
    result_b = bundle.get("B")
    result_c = bundle.get("C")

    assert result_c.generation_status == ScenarioCStatus.INSUFFICIENT_DATA
    # Current MVP fallback intentionally preserves B unchanged. This is not the amended
    # Contract V1 default, where C derives its own minimum under B's available limit.
    assert result_c.fleet.minimum_vehicles == result_b.fleet.minimum_vehicles == 1
    assert [trip.arrival_seconds for trip in result_c.trips] == [
        trip.arrival_seconds for trip in result_b.trips
    ]
    assert result_c.active_vehicle_ids == result_b.active_vehicle_ids == ("XE-001",)
