from __future__ import annotations

import inspect
from copy import deepcopy
from dataclasses import FrozenInstanceError, fields, replace
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from typing import get_type_hints

import pytest
from openpyxl import load_workbook
from presentation_support import (
    rejected_result_and_report,
    small_fixed_resource_fixture,
)
from route_corpus_support import (
    imported_workbook_from_fixture,
    load_corpus_fixture,
    normalization_options_from_fixture,
)

import bus_schedule_engine
import bus_schedule_engine.application_pipeline as application
import bus_schedule_engine.optimization_service as optimization_service
import bus_schedule_engine.service as legacy_service
import bus_schedule_engine.side_by_side_validation as side_by_side
import bus_schedule_engine.ui_utils as ui_utils
from bus_schedule_engine.application_pipeline import (
    CONTRACT_V1_ARTIFACT_FAILED,
    CONTRACT_V1_SEMANTIC_INTEGRITY_MISMATCH,
    CONTRACT_V1_SOLVER_FAILED,
    UNIFIED_SHADOW_RUNTIME_FAILURE,
    ParallelApplicationRunV1,
    ParallelRuntimeStatusV1,
    UnifiedApplicationRunV1,
    UnifiedApplicationStatusV1,
    UnifiedRuntimeFailureV1,
    run_parallel_application_pipeline_v1,
    run_unified_application_pipeline_v1,
)
from bus_schedule_engine.contracts_v1 import ContractValidationError, GenerationResultStatus
from bus_schedule_engine.excel_exporter import create_input_template
from bus_schedule_engine.importer import (
    ImportedWorkbook,
    WorkbookAuthorityMetadata,
    import_workbook,
)
from bus_schedule_engine.input_authority import (
    AVAILABLE_FLEET_LIMIT_REQUIRED_FOR_OPTIMIZATION,
    DEMAND_CONFIDENCE_REQUIRED_FOR_OPTIMIZATION,
    DEMAND_RESPONSE_MODE_REQUIRED_FOR_OPTIMIZATION,
    DEMAND_SOURCE_TYPE_REQUIRED_FOR_OPTIMIZATION,
    OPERATING_DAY_TYPE_REQUIRED_FOR_OPTIMIZATION,
    WorkbookInputReadinessV1,
)
from bus_schedule_engine.optimization_service import (
    OptimizationExecutionErrorV1,
    OptimizationExecutionStageV1,
    analyze_and_optimize_schedule_v1,
)
from bus_schedule_engine.service import run_analysis
from bus_schedule_engine.unified_diagram import (
    build_unified_demand_supply_figure_v1,
    build_unified_departure_figure_v1,
)
from bus_schedule_engine.unified_presentation import (
    build_unified_presentation_v1,
)
from bus_schedule_engine.unified_result_exporter import (
    export_unified_result_workbook_v1,
    read_unified_export_metadata_v1,
)

IMPORTED_AT = datetime(2026, 7, 28, 8, 0, tzinfo=UTC)
SOURCE_ID = "streamlit-upload-sha256:" + "a" * 64


def _set_parameter(sheet, key: str, value: object | None) -> None:
    row = next(cell.row for cell in sheet["A"] if cell.value == key)
    sheet.cell(row, 2).value = value


def _template_import(tmp_path: Path, name: str = "input.xlsx") -> ImportedWorkbook:
    return import_workbook(create_input_template(tmp_path / name))


def _mutated_template_import(
    tmp_path: Path,
    name: str,
    mutate,
) -> ImportedWorkbook:
    path = create_input_template(tmp_path / name)
    workbook = load_workbook(path)
    mutate(workbook)
    workbook.save(path)
    workbook.close()
    return import_workbook(path)


def _legacy_artifacts(bundle) -> dict[str, bytes]:
    scenario_c = bundle.get("C")
    c_fingerprint = scenario_c.timetable_fingerprint if scenario_c is not None else ""
    return {
        "xlsx": b"legacy-xlsx",
        "comparison_xlsx": b"legacy-comparison-xlsx",
        "html": b"legacy-html",
        "png": b"legacy-png",
        "c_fingerprint": c_fingerprint.encode(),
    }


def _install_legacy_spy(
    monkeypatch: pytest.MonkeyPatch,
) -> dict[str, int]:
    calls = {"legacy": 0}

    def legacy_path(imported: ImportedWorkbook):
        calls["legacy"] += 1
        bundle = run_analysis(imported)
        return bundle, object(), _legacy_artifacts(bundle)

    monkeypatch.setattr(application, "run_and_build_artifacts", legacy_path)
    return calls


def _ready_from_options(imported: ImportedWorkbook, options) -> ImportedWorkbook:
    parameters_a = (
        replace(
            imported.parameters_a,
            available_fleet_limit=options.available_fleet_limit_a,
            approved_active_fleet=options.approved_active_fleet_a,
            operating_day_type=options.operating_day_type_a.value,
        )
        if imported.parameters_a is not None
        else None
    )
    return replace(
        imported,
        parameters_a=parameters_a,
        parameters_b=replace(
            imported.parameters_b,
            available_fleet_limit=options.available_fleet_limit_b,
            approved_active_fleet=options.approved_active_fleet_b,
            operating_day_type=options.operating_day_type_b.value,
            terminal_1_max_occupancy_vehicles=options.terminal_1_max_occupancy_vehicles_b,
            terminal_2_max_occupancy_vehicles=options.terminal_2_max_occupancy_vehicles_b,
        ),
        authority_metadata=WorkbookAuthorityMetadata(
            demand_dataset_id=options.demand_dataset_id,
            demand_source_type=options.demand_source_type,
            demand_confidence=options.demand_confidence,
            demand_response_mode=options.demand_response_mode,
            source_notes=options.source_notes,
        ),
    )


def _figure_metadata(run: ParallelApplicationRunV1) -> tuple[dict, dict]:
    return (
        dict(run.unified_demand_supply_figure.layout.meta),
        dict(run.unified_departure_figure.layout.meta),
    )


def test_public_application_api_and_model_shape() -> None:
    assert (
        bus_schedule_engine.run_parallel_application_pipeline_v1
        is run_parallel_application_pipeline_v1
    )
    assert bus_schedule_engine.ParallelApplicationRunV1 is ParallelApplicationRunV1
    assert ParallelApplicationRunV1.__dataclass_params__.frozen is True
    assert "__slots__" in ParallelApplicationRunV1.__dict__
    assert {field.name for field in fields(ParallelApplicationRunV1)} == {
        "status",
        "legacy_bundle",
        "legacy_figure",
        "legacy_artifacts",
        "input_readiness",
        "unified_result",
        "side_by_side_report",
        "unified_presentation",
        "unified_demand_supply_figure",
        "unified_departure_figure",
        "unified_xlsx_bytes",
        "source_id",
        "imported_at",
        "failure_code",
        "failure_message",
    }
    assert get_type_hints(ParallelApplicationRunV1)["input_readiness"] == (
        WorkbookInputReadinessV1 | None
    )
    assert list(inspect.signature(run_parallel_application_pipeline_v1).parameters) == [
        "imported",
        "source_id",
        "imported_at",
        "solver_choice",
    ]
    assert tuple(status.value for status in ParallelRuntimeStatusV1) == (
        "INPUT_NOT_READY",
        "PARALLEL_VALIDATION_COMPLETE",
        "UNIFIED_RUNTIME_FAILED",
    )


def test_complete_template_runs_legacy_once_and_unified_once(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    imported = _template_import(tmp_path)
    original = deepcopy(imported)
    calls = {"legacy": 0, "unified": 0}
    received: dict[str, ImportedWorkbook] = {}
    real_legacy = legacy_service.run_analysis
    real_unified = application.analyze_and_optimize_schedule_v1

    def legacy_spy(value):
        calls["legacy"] += 1
        received["legacy"] = value
        value.configuration["legacy_spy_mutation"] = True
        return real_legacy(value)

    def unified_spy(value, options, **kwargs):
        calls["unified"] += 1
        received["unified"] = value
        assert "legacy_spy_mutation" not in value.configuration
        result = real_unified(value, options, **kwargs)
        value.trips_b.clear()
        value.demand.clear()
        value.configuration["unified_spy_mutation"] = True
        return result

    monkeypatch.setattr(legacy_service, "run_analysis", legacy_spy)
    monkeypatch.setattr(application, "analyze_and_optimize_schedule_v1", unified_spy)

    run = run_parallel_application_pipeline_v1(
        imported,
        source_id=SOURCE_ID,
        imported_at=IMPORTED_AT,
    )

    assert calls == {"legacy": 1, "unified": 1}
    assert imported == original
    assert received["legacy"] is not imported
    assert received["unified"] is not imported
    assert received["legacy"] is not received["unified"]
    assert received["legacy"].trips_b is not received["unified"].trips_b
    assert "unified_spy_mutation" not in received["legacy"].configuration
    assert run.status == ParallelRuntimeStatusV1.PARALLEL_VALIDATION_COMPLETE
    assert run.input_readiness.optimization_ready is True
    result_b = run.legacy_bundle.get("B")
    assert result_b is not None
    assert result_b.trips
    assert result_b.trips is not imported.trips_b
    assert result_b.trips[0] is not imported.trips_b[0]
    assert run.unified_result is not None
    assert run.side_by_side_report is not None
    assert run.unified_presentation is not None
    assert run.unified_demand_supply_figure is not None
    assert run.unified_departure_figure is not None
    assert run.unified_xlsx_bytes
    assert run.source_id == SOURCE_ID
    assert run.imported_at is IMPORTED_AT
    with pytest.raises(FrozenInstanceError):
        run.status = ParallelRuntimeStatusV1.INPUT_NOT_READY


def test_legacy_artifacts_are_returned_unchanged_from_direct_path(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    imported = _template_import(tmp_path)
    expected_bundle, expected_figure, expected_artifacts = ui_utils.run_and_build_artifacts(
        imported
    )
    monkeypatch.setattr(
        application,
        "run_and_build_artifacts",
        lambda value: (expected_bundle, expected_figure, expected_artifacts),
    )

    run = run_parallel_application_pipeline_v1(
        imported,
        source_id=SOURCE_ID,
        imported_at=IMPORTED_AT,
    )

    assert run.legacy_bundle is expected_bundle
    assert run.legacy_figure is expected_figure
    assert run.legacy_artifacts is expected_artifacts
    assert run.legacy_artifacts == expected_artifacts


@pytest.mark.parametrize(
    ("name", "mutate", "expected_codes"),
    (
        (
            "blank-fleet.xlsx",
            lambda workbook: _set_parameter(workbook["THONG_SO_B"], "available_fleet_limit", None),
            (AVAILABLE_FLEET_LIMIT_REQUIRED_FOR_OPTIMIZATION,),
        ),
        (
            "blank-day.xlsx",
            lambda workbook: _set_parameter(workbook["THONG_SO_B"], "operating_day_type", None),
            (OPERATING_DAY_TYPE_REQUIRED_FOR_OPTIMIZATION,),
        ),
        (
            "blank-demand-authority.xlsx",
            lambda workbook: [
                _set_parameter(workbook["THONG_TIN_DU_LIEU"], key, None)
                for key in (
                    "demand_source_type",
                    "demand_confidence",
                    "demand_response_mode",
                )
            ],
            (
                DEMAND_CONFIDENCE_REQUIRED_FOR_OPTIMIZATION,
                DEMAND_RESPONSE_MODE_REQUIRED_FOR_OPTIMIZATION,
                DEMAND_SOURCE_TYPE_REQUIRED_FOR_OPTIMIZATION,
            ),
        ),
    ),
)
def test_input_not_ready_keeps_legacy_and_never_calls_unified(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    name: str,
    mutate,
    expected_codes: tuple[str, ...],
) -> None:
    imported = _mutated_template_import(tmp_path, name, mutate)
    calls = _install_legacy_spy(monkeypatch)

    def forbidden(*args, **kwargs):
        raise AssertionError("unified service must not run without input authority")

    monkeypatch.setattr(application, "analyze_and_optimize_schedule_v1", forbidden)
    run = run_parallel_application_pipeline_v1(
        imported,
        source_id=SOURCE_ID,
        imported_at=IMPORTED_AT,
    )

    assert calls == {"legacy": 1}
    assert run.status == ParallelRuntimeStatusV1.INPUT_NOT_READY
    assert run.legacy_bundle.get("B") is not None
    assert run.input_readiness.missing_optimization_authority_codes == expected_codes
    assert run.unified_result is None
    assert run.side_by_side_report is None
    assert run.unified_presentation is None
    assert run.unified_xlsx_bytes is None
    assert run.failure_code is None


def test_no_demand_may_complete_without_demand_authority(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def mutate(workbook) -> None:
        demand = workbook["SAN_LUONG"]
        demand.delete_rows(4, demand.max_row - 3)
        for key in (
            "demand_source_type",
            "demand_confidence",
            "demand_response_mode",
        ):
            _set_parameter(workbook["THONG_TIN_DU_LIEU"], key, None)

    imported = _mutated_template_import(tmp_path, "no-demand.xlsx", mutate)
    _install_legacy_spy(monkeypatch)
    run = run_parallel_application_pipeline_v1(
        imported,
        source_id=SOURCE_ID,
        imported_at=IMPORTED_AT,
    )

    assert imported.demand == []
    assert run.input_readiness.missing_optimization_authority_codes == ()
    assert run.status == ParallelRuntimeStatusV1.PARALLEL_VALIDATION_COMPLETE


def test_completed_artifacts_share_presentation_b_and_absent_c_fingerprints(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    imported = _template_import(tmp_path)
    _install_legacy_spy(monkeypatch)
    run = run_parallel_application_pipeline_v1(
        imported,
        source_id=SOURCE_ID,
        imported_at=IMPORTED_AT,
    )
    demand_meta, departure_meta = _figure_metadata(run)
    exported_path = tmp_path / "shadow.xlsx"
    exported_path.write_bytes(run.unified_xlsx_bytes)
    xlsx_meta = read_unified_export_metadata_v1(exported_path)
    presentation = run.unified_presentation

    assert {
        presentation.presentation_fingerprint,
        demand_meta["presentation_fingerprint"],
        departure_meta["presentation_fingerprint"],
        xlsx_meta.presentation_fingerprint,
    } == {presentation.presentation_fingerprint}
    assert {
        run.unified_result.normalized_inputs.scenario_b_fingerprint,
        presentation.source_b_fingerprint,
        demand_meta["source_b_fingerprint"],
        departure_meta["source_b_fingerprint"],
        xlsx_meta.b_fingerprint,
    } == {presentation.source_b_fingerprint}
    assert presentation.accepted_solution_fingerprint is None
    assert demand_meta["accepted_solution_fingerprint"] is None
    assert departure_meta["accepted_solution_fingerprint"] is None
    assert xlsx_meta.accepted_solution_fingerprint is None
    assert presentation.scenario("C") is None
    workbook = load_workbook(BytesIO(run.unified_xlsx_bytes), read_only=True)
    try:
        assert "C_BIEU_DO" not in workbook.sheetnames
    finally:
        workbook.close()


def test_accepted_synthetic_c_and_blocking_report_remain_available(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    imported, options = small_fixed_resource_fixture()
    imported = _ready_from_options(imported, options)
    _install_legacy_spy(monkeypatch)
    real_builder = application.build_side_by_side_validation_report_v1

    def blocking_builder(legacy_bundle, unified_result):
        report = real_builder(legacy_bundle, unified_result)
        return side_by_side._report(
            replace(report.legacy_snapshot, route_id="LEGACY-OTHER-ROUTE"),
            report.unified_snapshot,
        )

    monkeypatch.setattr(
        application,
        "build_side_by_side_validation_report_v1",
        blocking_builder,
    )
    run = run_parallel_application_pipeline_v1(
        imported,
        source_id=options.source_id,
        imported_at=options.imported_at,
    )
    demand_meta, departure_meta = _figure_metadata(run)

    assert run.status == ParallelRuntimeStatusV1.PARALLEL_VALIDATION_COMPLETE
    assert run.unified_presentation.cutover_blocked is True
    assert run.side_by_side_report.blocking_discrepancy_codes
    accepted = run.unified_presentation.accepted_solution_fingerprint
    assert accepted
    assert accepted == run.unified_result.recommended_outcome.solution.solution_fingerprint
    assert demand_meta["accepted_solution_fingerprint"] == accepted
    assert departure_meta["accepted_solution_fingerprint"] == accepted
    path = tmp_path / "accepted-shadow.xlsx"
    path.write_bytes(run.unified_xlsx_bytes)
    assert read_unified_export_metadata_v1(path).accepted_solution_fingerprint == accepted


@pytest.mark.parametrize(
    ("fixture_name", "expected_gap"),
    (
        ("corpus_alpha_80.json", None),
        ("corpus_beta_46.json", ("outbound", 17 * 3600, 18 * 3600)),
    ),
)
def test_alpha_and_beta_shadow_characterization(
    monkeypatch: pytest.MonkeyPatch,
    fixture_name: str,
    expected_gap: tuple[str, int, int] | None,
) -> None:
    fixture = load_corpus_fixture(fixture_name)
    options = normalization_options_from_fixture(fixture)
    imported = _ready_from_options(imported_workbook_from_fixture(fixture), options)
    _install_legacy_spy(monkeypatch)
    run = run_parallel_application_pipeline_v1(
        imported,
        source_id=options.source_id,
        imported_at=options.imported_at,
    )

    assert run.status == ParallelRuntimeStatusV1.PARALLEL_VALIDATION_COMPLETE
    assert run.unified_presentation.scenario("C") is None
    assert run.unified_presentation.accepted_solution_fingerprint is None
    assert run.unified_presentation.outcome.solver_attempted is False
    if expected_gap is None:
        assert run.legacy_bundle.get("C") is not None
        assert any(
            discrepancy.reason_code == "LEGACY_C_WITHOUT_UNIFIED_AUTHORITY"
            for discrepancy in run.unified_presentation.discrepancies
        )
    else:
        direction, start, end = expected_gap
        assert any(
            gap.direction == direction
            and gap.start_time_seconds == start
            and gap.end_time_seconds == end
            for gap in run.unified_presentation.demand_gaps
        )


def test_rejected_candidate_keeps_codes_but_no_candidate_timetable(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    imported, options = small_fixed_resource_fixture()
    imported = _ready_from_options(imported, options)
    rejected_result, _ = rejected_result_and_report()
    assert rejected_result.recommended_outcome.result_status == (
        GenerationResultStatus.CANDIDATE_REJECTED_BY_DOMAIN_VALIDATOR
    )
    _install_legacy_spy(monkeypatch)
    monkeypatch.setattr(
        application,
        "analyze_and_optimize_schedule_v1",
        lambda *args, **kwargs: rejected_result,
    )
    run = run_parallel_application_pipeline_v1(
        imported,
        source_id=options.source_id,
        imported_at=options.imported_at,
    )

    assert run.status == ParallelRuntimeStatusV1.PARALLEL_VALIDATION_COMPLETE
    assert run.unified_presentation.outcome.validator_rejection_codes == (
        "SYNTHETIC_DOMAIN_REJECTION",
    )
    assert run.unified_presentation.scenario("C") is None
    assert run.unified_presentation.accepted_solution_fingerprint is None
    workbook = load_workbook(BytesIO(run.unified_xlsx_bytes), read_only=True)
    try:
        assert "C_TRANG_THAI" in workbook.sheetnames
        assert "C_BIEU_DO" not in workbook.sheetnames
        all_text = " ".join(
            str(cell.value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None
        )
        assert "SYNTHETIC_DOMAIN_REJECTION" in all_text
        assert "rejected-diagnostic-candidate" not in all_text
    finally:
        workbook.close()


def test_unified_pipeline_treats_candidate_rejection_as_complete(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    imported, options = small_fixed_resource_fixture()
    imported = _ready_from_options(imported, options)
    rejected_result, _report = rejected_result_and_report()
    monkeypatch.setattr(
        application,
        "_analyze_normalized_and_optimize_schedule_v1",
        lambda *args, **kwargs: rejected_result,
    )

    run = run_unified_application_pipeline_v1(
        imported,
        source_id=options.source_id,
        imported_at=options.imported_at,
    )

    assert run.status == UnifiedApplicationStatusV1.COMPLETE
    assert run.failure is None
    assert run.unified_presentation is not None
    assert run.unified_presentation.outcome.heuristic_result_status == (
        GenerationResultStatus.CANDIDATE_REJECTED_BY_DOMAIN_VALIDATOR.value
    )
    assert run.unified_presentation.outcome.validator_rejection_codes == (
        "SYNTHETIC_DOMAIN_REJECTION",
    )
    assert run.unified_presentation.scenario("C") is None
    assert run.unified_presentation.accepted_solution_fingerprint is None
    assert run.unified_xlsx_bytes


def test_integrity_mismatch_becomes_shadow_failure_without_partial_artifacts(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    imported = _template_import(tmp_path)
    _install_legacy_spy(monkeypatch)
    real_builder = application.build_unified_departure_figure_v1

    def mismatched_figure(presentation):
        figure = real_builder(presentation)
        metadata = dict(figure.layout.meta)
        metadata["presentation_fingerprint"] = "misaligned"
        figure.update_layout(meta=metadata)
        return figure

    monkeypatch.setattr(
        application,
        "build_unified_departure_figure_v1",
        mismatched_figure,
    )
    run = run_parallel_application_pipeline_v1(
        imported,
        source_id=SOURCE_ID,
        imported_at=IMPORTED_AT,
    )

    assert run.status == ParallelRuntimeStatusV1.UNIFIED_RUNTIME_FAILED
    assert run.failure_code == UNIFIED_SHADOW_RUNTIME_FAILURE
    assert "fingerprints do not align" in run.failure_message
    assert run.legacy_bundle.get("B") is not None
    assert run.unified_result is None
    assert run.side_by_side_report is None
    assert run.unified_presentation is None
    assert run.unified_demand_supply_figure is None
    assert run.unified_departure_figure is None
    assert run.unified_xlsx_bytes is None


def test_unexpected_unified_failure_retains_exact_legacy_result(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    imported = _template_import(tmp_path)
    expected = (run_analysis(imported), object(), {"xlsx": b"legacy"})
    monkeypatch.setattr(application, "run_and_build_artifacts", lambda value: expected)

    def fail(*args, **kwargs):
        raise RuntimeError("synthetic unified failure")

    monkeypatch.setattr(application, "analyze_and_optimize_schedule_v1", fail)
    run = run_parallel_application_pipeline_v1(
        imported,
        source_id=SOURCE_ID,
        imported_at=IMPORTED_AT,
    )

    assert run.status == ParallelRuntimeStatusV1.UNIFIED_RUNTIME_FAILED
    assert run.legacy_bundle is expected[0]
    assert run.legacy_figure is expected[1]
    assert run.legacy_artifacts is expected[2]
    assert run.failure_code == UNIFIED_SHADOW_RUNTIME_FAILURE
    assert run.failure_message == "synthetic unified failure"


def test_unexpected_readiness_failure_retains_exact_legacy_result(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    imported = _template_import(tmp_path)
    expected = (run_analysis(imported), object(), {"xlsx": b"legacy"})
    monkeypatch.setattr(application, "run_and_build_artifacts", lambda value: expected)

    def fail_readiness(_imported):
        raise RuntimeError("synthetic readiness failure")

    def forbidden(*args, **kwargs):
        raise AssertionError("unified service must not run after readiness failure")

    monkeypatch.setattr(
        application,
        "assess_workbook_input_readiness_v1",
        fail_readiness,
    )
    monkeypatch.setattr(application, "analyze_and_optimize_schedule_v1", forbidden)
    run = run_parallel_application_pipeline_v1(
        imported,
        source_id=SOURCE_ID,
        imported_at=IMPORTED_AT,
    )

    assert run.status == ParallelRuntimeStatusV1.UNIFIED_RUNTIME_FAILED
    assert run.legacy_bundle is expected[0]
    assert run.legacy_figure is expected[1]
    assert run.legacy_artifacts is expected[2]
    assert run.input_readiness is None
    assert run.unified_result is None
    assert run.side_by_side_report is None
    assert run.unified_presentation is None
    assert run.unified_demand_supply_figure is None
    assert run.unified_departure_figure is None
    assert run.unified_xlsx_bytes is None
    assert run.failure_code == UNIFIED_SHADOW_RUNTIME_FAILURE
    assert run.failure_message == "synthetic readiness failure"


def test_unexpected_legacy_failure_still_blocks_the_ordinary_run(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    imported = _template_import(tmp_path)

    def fail(_imported):
        raise ValueError("legacy failure")

    monkeypatch.setattr(application, "run_and_build_artifacts", fail)
    with pytest.raises(ValueError, match="legacy failure"):
        run_parallel_application_pipeline_v1(
            imported,
            source_id=SOURCE_ID,
            imported_at=IMPORTED_AT,
        )


def test_presentation_figures_and_export_do_not_rerun_analysis_paths(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    imported, options = small_fixed_resource_fixture()
    legacy_bundle = run_analysis(imported)
    unified_result = analyze_and_optimize_schedule_v1(imported, options)
    report = side_by_side.build_side_by_side_validation_report_v1(
        legacy_bundle,
        unified_result,
    )

    def forbidden(*args, **kwargs):
        raise AssertionError("artifact projection must not rerun analysis")

    monkeypatch.setattr(side_by_side, "run_analysis", forbidden)
    monkeypatch.setattr(side_by_side, "analyze_and_optimize_schedule_v1", forbidden)
    monkeypatch.setattr(application, "analyze_and_optimize_schedule_v1", forbidden)

    presentation = build_unified_presentation_v1(unified_result, report)
    build_unified_demand_supply_figure_v1(presentation)
    build_unified_departure_figure_v1(presentation)
    export_unified_result_workbook_v1(presentation, tmp_path / "projected.xlsx")


def test_unified_application_public_model_shape_and_default_solver() -> None:
    assert bus_schedule_engine.run_unified_application_pipeline_v1 is (
        run_unified_application_pipeline_v1
    )
    assert UnifiedApplicationRunV1.__dataclass_params__.frozen is True
    assert "__slots__" in UnifiedApplicationRunV1.__dict__
    assert {field.name for field in fields(UnifiedApplicationRunV1)} == {
        "status",
        "input_readiness",
        "unified_result",
        "unified_presentation",
        "unified_demand_supply_figure",
        "unified_departure_figure",
        "unified_xlsx_bytes",
        "source_id",
        "imported_at",
        "failure",
        "trip_ridership_analysis",
        "trip_ridership_failure",
        "protected_service_floor_assessment",
        "protected_service_floor_failure",
        "protected_service_floor_enforcement_authority",
        "protected_service_floor_enforcement_failure",
    }
    assert UnifiedRuntimeFailureV1.__dataclass_params__.frozen is True
    assert tuple(status.value for status in UnifiedApplicationStatusV1) == (
        "INPUT_NOT_READY",
        "COMPLETE",
        "ARTIFACT_FAILED",
        "FAILED",
    )
    assert (
        inspect.signature(run_unified_application_pipeline_v1)
        .parameters["solver_choice"]
        .default.value
        == "HEURISTIC"
    )


def test_unified_pipeline_stops_at_readiness_without_any_analysis_or_artifact(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    imported = _template_import(tmp_path)
    imported = replace(
        imported,
        parameters_b=replace(
            imported.parameters_b,
            available_fleet_limit=None,
        ),
    )

    def forbidden(*args, **kwargs):
        raise AssertionError("readiness blocker must stop the pipeline")

    monkeypatch.setattr(application, "normalization_options_from_workbook_v1", forbidden)
    monkeypatch.setattr(
        application,
        "_analyze_normalized_and_optimize_schedule_v1",
        forbidden,
    )
    monkeypatch.setattr(application, "run_and_build_artifacts", forbidden)
    monkeypatch.setattr(
        application,
        "build_side_by_side_validation_report_v1",
        forbidden,
    )
    monkeypatch.setattr(application, "build_unified_demand_supply_figure_v1", forbidden)
    monkeypatch.setattr(application, "export_unified_result_workbook_v1", forbidden)

    run = run_unified_application_pipeline_v1(
        imported,
        source_id=SOURCE_ID,
        imported_at=IMPORTED_AT,
    )

    assert run.status == UnifiedApplicationStatusV1.INPUT_NOT_READY
    assert AVAILABLE_FLEET_LIMIT_REQUIRED_FOR_OPTIMIZATION in (
        run.input_readiness.missing_optimization_authority_codes
    )
    assert run.unified_result is None
    assert run.unified_presentation is None
    assert run.unified_xlsx_bytes is None
    assert run.failure is None


def test_unified_pipeline_completes_without_loading_or_running_legacy(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    imported = _template_import(tmp_path)
    unified_calls = 0
    real_unified = application._analyze_normalized_and_optimize_schedule_v1

    def unified_spy(*args, **kwargs):
        nonlocal unified_calls
        unified_calls += 1
        return real_unified(*args, **kwargs)

    def forbidden(*args, **kwargs):
        raise AssertionError("ordinary unified runtime must not execute the oracle")

    monkeypatch.setattr(
        application,
        "_analyze_normalized_and_optimize_schedule_v1",
        unified_spy,
    )
    monkeypatch.setattr(application, "run_and_build_artifacts", forbidden)
    monkeypatch.setattr(
        application,
        "build_side_by_side_validation_report_v1",
        forbidden,
    )

    run = run_unified_application_pipeline_v1(
        imported,
        source_id=SOURCE_ID,
        imported_at=IMPORTED_AT,
    )

    assert unified_calls == 1
    assert run.status == UnifiedApplicationStatusV1.COMPLETE
    assert run.unified_result is not None
    assert run.unified_presentation is not None
    assert run.unified_presentation.discrepancies == ()
    assert run.unified_presentation.validation_explanations == ()
    assert run.unified_presentation.validation_limitations == ()
    assert run.unified_xlsx_bytes
    assert run.failure is None


def test_solver_exception_is_staged_and_fails_closed(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    imported = _template_import(tmp_path)

    def fail(*args, **kwargs):
        original = RuntimeError("solver exploded")
        wrapped = OptimizationExecutionErrorV1(
            OptimizationExecutionStageV1.HEURISTIC_SOLVER,
            original,
        )
        raise wrapped from original

    monkeypatch.setattr(
        application,
        "_analyze_normalized_and_optimize_schedule_v1",
        fail,
    )
    run = run_unified_application_pipeline_v1(
        imported,
        source_id=SOURCE_ID,
        imported_at=IMPORTED_AT,
    )

    assert run.status == UnifiedApplicationStatusV1.FAILED
    assert run.failure is not None
    assert run.failure.code == CONTRACT_V1_SOLVER_FAILED
    assert run.failure.stage == "HEURISTIC_SOLVER"
    assert run.unified_result is None
    assert run.unified_presentation is None


def test_normalization_contract_validation_error_fails_at_normalization(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    imported = _template_import(tmp_path)
    issue_error = ContractValidationError(())

    def fail(*args, **kwargs):
        raise issue_error

    monkeypatch.setattr(
        application,
        "normalize_imported_workbook_v1",
        fail,
    )
    run = run_unified_application_pipeline_v1(
        imported,
        source_id=SOURCE_ID,
        imported_at=IMPORTED_AT,
    )

    assert run.status == UnifiedApplicationStatusV1.FAILED
    assert run.failure is not None
    assert run.failure.code == application.CONTRACT_V1_NORMALIZATION_FAILED
    assert run.failure.stage == OptimizationExecutionStageV1.NORMALIZATION.value
    assert run.unified_result is None
    assert run.unified_presentation is None
    assert run.unified_demand_supply_figure is None
    assert run.unified_departure_figure is None
    assert run.unified_xlsx_bytes is None


def test_evaluation_contract_validation_error_fails_at_evaluation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    imported = _template_import(tmp_path)
    issue_error = ContractValidationError(())

    def fail(*args, **kwargs):
        raise issue_error

    monkeypatch.setattr(
        optimization_service,
        "build_service_adjustment_evaluation_context_v1",
        fail,
    )
    run = run_unified_application_pipeline_v1(
        imported,
        source_id=SOURCE_ID,
        imported_at=IMPORTED_AT,
    )

    assert run.status == UnifiedApplicationStatusV1.FAILED
    assert run.failure is not None
    assert run.failure.code == application.CONTRACT_V1_APPLICATION_ERROR
    assert run.failure.stage == OptimizationExecutionStageV1.EVALUATION.value
    assert run.unified_result is None
    assert run.unified_presentation is None
    assert run.unified_demand_supply_figure is None
    assert run.unified_departure_figure is None
    assert run.unified_xlsx_bytes is None


def test_artifact_failure_retains_only_verified_result_and_presentation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    imported = _template_import(tmp_path)

    def fail(_presentation):
        raise RuntimeError("chart renderer unavailable")

    monkeypatch.setattr(application, "build_unified_demand_supply_figure_v1", fail)
    run = run_unified_application_pipeline_v1(
        imported,
        source_id=SOURCE_ID,
        imported_at=IMPORTED_AT,
    )

    assert run.status == UnifiedApplicationStatusV1.ARTIFACT_FAILED
    assert run.failure is not None
    assert run.failure.code == CONTRACT_V1_ARTIFACT_FAILED
    assert run.failure.stage == OptimizationExecutionStageV1.ARTIFACTS.value
    assert run.unified_result is not None
    assert run.unified_presentation is not None
    assert run.unified_demand_supply_figure is None
    assert run.unified_departure_figure is None
    assert run.unified_xlsx_bytes is None


def test_semantic_mismatch_exposes_no_analytical_result(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    imported = _template_import(tmp_path)

    def fail(_presentation):
        raise application.UnifiedPresentationConsistencyError("PRESENTATION_FINGERPRINT_MISMATCH")

    monkeypatch.setattr(
        application,
        "verify_unified_presentation_integrity_v1",
        fail,
    )
    run = run_unified_application_pipeline_v1(
        imported,
        source_id=SOURCE_ID,
        imported_at=IMPORTED_AT,
    )

    assert run.status == UnifiedApplicationStatusV1.FAILED
    assert run.failure is not None
    assert run.failure.code == CONTRACT_V1_SEMANTIC_INTEGRITY_MISMATCH
    assert run.failure.stage == OptimizationExecutionStageV1.PRESENTATION.value
    assert run.unified_result is None
    assert run.unified_presentation is None
    assert run.unified_xlsx_bytes is None


def test_failure_correlation_is_deterministic_and_local_paths_are_removed(
    caplog: pytest.LogCaptureFixture,
) -> None:
    readiness = WorkbookInputReadinessV1(
        import_ready=True,
        optimization_ready=True,
        blocking_import_codes=(),
        missing_optimization_authority_codes=(),
        optional_limitations=(),
    )
    kwargs = {
        "code": CONTRACT_V1_ARTIFACT_FAILED,
        "stage": OptimizationExecutionStageV1.ARTIFACTS.value,
        "exc": RuntimeError(r"failed at C:\Users\private\workbook.xlsx"),
        "retryable": True,
        "solver_choice": application.SolverChoice.HEURISTIC,
        "source_id": SOURCE_ID,
        "imported_at": IMPORTED_AT,
        "input_readiness": readiness,
    }
    first = application.build_unified_runtime_failure_v1(**kwargs)
    second = application.build_unified_runtime_failure_v1(**kwargs)

    assert first == second
    assert first.correlation_id.startswith("m5c2-")
    assert r"C:\Users" not in first.sanitized_message
    assert "[path]" in first.sanitized_message

    caplog.clear()
    sensitive = application.build_unified_runtime_failure_v1(
        **{
            **kwargs,
            "exc": RuntimeError("raw workbook rows: SECRET_PASSENGER_OBSERVATION"),
        }
    )
    assert "SECRET_PASSENGER_OBSERVATION" not in sensitive.sanitized_message
    assert "SECRET_PASSENGER_OBSERVATION" not in caplog.text


def test_import_error_sanitizer_retains_safe_location_but_not_cell_value() -> None:
    message = application.sanitize_import_error_message_v1(
        ValueError(
            r"BIEU_DO_B, dòng Excel 4: direction không hợp lệ: "
            r"SECRET_PASSENGER_ROW C:\Users\private\workbook.xlsx"
        )
    )

    assert "BIEU_DO_B" in message
    assert "direction" in message
    assert "SECRET_PASSENGER_ROW" not in message
    assert r"C:\Users" not in message
