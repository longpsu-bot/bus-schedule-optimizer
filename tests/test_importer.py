from __future__ import annotations

import pytest
from openpyxl import load_workbook

from bus_schedule_engine.contracts_v1 import (
    DemandConfidence,
    DemandResponseMode,
    DemandSourceType,
)
from bus_schedule_engine.excel_exporter import create_input_template
from bus_schedule_engine.importer import (
    InputDataError,
    WorkbookAuthorityMetadata,
    import_workbook,
)


def _set_parameter(sheet, key: str, value: object | None) -> None:
    row = next(cell.row for cell in sheet["A"] if cell.value == key)
    sheet.cell(row, 2).value = value


def test_old_workbook_without_authority_sheet_still_imports(tmp_path) -> None:
    path = create_input_template(tmp_path / "old.xlsx")
    workbook = load_workbook(path)
    workbook.remove(workbook["THONG_TIN_DU_LIEU"])
    workbook.save(path)
    workbook.close()

    imported = import_workbook(path)

    assert imported.authority_metadata == WorkbookAuthorityMetadata()


def test_blank_optimization_and_optional_parameters_remain_none(tmp_path) -> None:
    path = create_input_template(tmp_path / "blank-values.xlsx")
    workbook = load_workbook(path)
    for key in (
        "available_fleet_limit",
        "approved_active_fleet",
        "terminal_1_max_occupancy_vehicles",
        "terminal_2_max_occupancy_vehicles",
    ):
        _set_parameter(workbook["THONG_SO_B"], key, None)
    _set_parameter(workbook["THONG_SO_B"], "operating_day_type", None)
    workbook.save(path)
    workbook.close()

    imported = import_workbook(path)

    assert imported.parameters_b.available_fleet_limit is None
    assert imported.parameters_b.approved_active_fleet is None
    assert imported.parameters_b.operating_day_type is None
    assert imported.parameters_b.terminal_1_max_occupancy_vehicles is None
    assert imported.parameters_b.terminal_2_max_occupancy_vehicles is None


def test_authority_metadata_round_trips_declared_values(tmp_path) -> None:
    path = create_input_template(tmp_path / "metadata.xlsx")
    workbook = load_workbook(path)
    sheet = workbook["THONG_TIN_DU_LIEU"]
    _set_parameter(sheet, "demand_dataset_id", "DEMAND-42")
    _set_parameter(sheet, "demand_source_type", "apc")
    _set_parameter(sheet, "demand_confidence", "low")
    _set_parameter(sheet, "demand_response_mode", "calibrated")
    _set_parameter(sheet, "source_notes", "Declared by operator")
    workbook.save(path)
    workbook.close()

    imported = import_workbook(path)

    assert imported.authority_metadata == WorkbookAuthorityMetadata(
        demand_dataset_id="DEMAND-42",
        demand_source_type=DemandSourceType.APC,
        demand_confidence=DemandConfidence.LOW,
        demand_response_mode=DemandResponseMode.CALIBRATED,
        source_notes="Declared by operator",
    )


def test_blank_optional_metadata_remains_none(tmp_path) -> None:
    path = create_input_template(tmp_path / "blank-metadata.xlsx")
    workbook = load_workbook(path)
    sheet = workbook["THONG_TIN_DU_LIEU"]
    _set_parameter(sheet, "demand_dataset_id", None)
    _set_parameter(sheet, "source_notes", None)
    workbook.save(path)
    workbook.close()

    imported = import_workbook(path)

    assert imported.authority_metadata.demand_dataset_id is None
    assert imported.authority_metadata.source_notes is None


@pytest.mark.parametrize("value", [None, 0, -1, True, 60.5, "not-a-number"])
def test_scenario_b_vehicle_capacity_must_be_a_positive_integer(
    tmp_path,
    value: object,
) -> None:
    path = create_input_template(tmp_path / "invalid-b-capacity.xlsx")
    workbook = load_workbook(path)
    _set_parameter(workbook["THONG_SO_B"], "vehicle_capacity_passengers", value)
    workbook.save(path)
    workbook.close()

    with pytest.raises(InputDataError, match="vehicle_capacity_passengers"):
        import_workbook(path)


@pytest.mark.parametrize(("value", "expected"), [(73, 73), ("60", 60)])
def test_scenario_b_vehicle_capacity_round_trips_valid_integer_or_numeric_text(
    tmp_path,
    value: object,
    expected: int,
) -> None:
    path = create_input_template(tmp_path / "valid-b-capacity.xlsx")
    workbook = load_workbook(path)
    _set_parameter(workbook["THONG_SO_B"], "vehicle_capacity_passengers", value)
    workbook.save(path)
    workbook.close()

    imported = import_workbook(path)

    assert imported.parameters_b.vehicle_capacity_passengers == expected


def test_scenario_a_absent_does_not_require_its_vehicle_capacity(tmp_path) -> None:
    path = create_input_template(tmp_path / "no-a.xlsx")
    workbook = load_workbook(path)
    _set_parameter(workbook["THONG_SO_A"], "vehicle_capacity_passengers", None)
    workbook.remove(workbook["BIEU_DO_A"])
    workbook.save(path)
    workbook.close()

    imported = import_workbook(path)

    assert imported.parameters_a is None


@pytest.mark.parametrize("value", [None, 60.5])
def test_scenario_a_present_requires_valid_vehicle_capacity(
    tmp_path,
    value: object,
) -> None:
    path = create_input_template(tmp_path / "invalid-a-capacity.xlsx")
    workbook = load_workbook(path)
    _set_parameter(workbook["THONG_SO_A"], "vehicle_capacity_passengers", value)
    workbook.save(path)
    workbook.close()

    with pytest.raises(InputDataError, match="vehicle_capacity_passengers"):
        import_workbook(path)


def test_scenario_a_vehicle_capacity_round_trips(tmp_path) -> None:
    path = create_input_template(tmp_path / "valid-a-capacity.xlsx")
    workbook = load_workbook(path)
    _set_parameter(workbook["THONG_SO_A"], "vehicle_capacity_passengers", 71)
    workbook.save(path)
    workbook.close()

    imported = import_workbook(path)

    assert imported.parameters_a is not None
    assert imported.parameters_a.vehicle_capacity_passengers == 71


def test_legacy_runtime_is_used_when_preferred_runtime_is_blank(tmp_path) -> None:
    path = create_input_template(tmp_path / "legacy-runtime.xlsx")
    workbook = load_workbook(path)
    for sheet_name in ("THONG_SO_A", "THONG_SO_B"):
        _set_parameter(workbook[sheet_name], "allowed_trip_runtime_minutes", None)
        _set_parameter(workbook[sheet_name], "trip_runtime_minutes", 64)
    workbook.save(path)
    workbook.close()

    imported = import_workbook(path)

    assert imported.parameters_a is not None
    assert imported.parameters_a.runtime_options == (64,)
    assert imported.parameters_b.runtime_options == (64,)


def test_both_runtime_fields_blank_fail_import(tmp_path) -> None:
    path = create_input_template(tmp_path / "missing-runtime.xlsx")
    workbook = load_workbook(path)
    _set_parameter(workbook["THONG_SO_B"], "allowed_trip_runtime_minutes", None)
    _set_parameter(workbook["THONG_SO_B"], "trip_runtime_minutes", None)
    workbook.save(path)
    workbook.close()

    with pytest.raises(InputDataError, match="trip_runtime_minutes"):
        import_workbook(path)


def test_preferred_runtime_keeps_precedence_when_both_fields_are_valid(tmp_path) -> None:
    path = create_input_template(tmp_path / "runtime-precedence.xlsx")
    workbook = load_workbook(path)
    _set_parameter(workbook["THONG_SO_B"], "allowed_trip_runtime_minutes", "55,65")
    _set_parameter(workbook["THONG_SO_B"], "trip_runtime_minutes", 99)
    workbook.save(path)
    workbook.close()

    imported = import_workbook(path)

    assert imported.parameters_b.runtime_options == (55, 65)
    assert imported.parameters_b.default_trip_runtime_minutes == 65
