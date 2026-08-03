from __future__ import annotations

import json
from dataclasses import replace
from datetime import UTC, date, datetime, timedelta
from pathlib import Path

import pytest
from jsonschema import Draft202012Validator, FormatChecker
from openpyxl import load_workbook

from bus_schedule_engine.contracts_v1 import (
    ContractDirection,
    DemandConfidence,
    NormalizationError,
    NormalizationOptions,
    OperatingDayType,
    ScenarioBInput,
    demand_to_contract_dict,
    normalize_imported_workbook_v1,
    scenario_to_contract_dict,
    validate_scenario_input,
)
from bus_schedule_engine.excel_exporter import create_input_template
from bus_schedule_engine.importer import ImportedWorkbook, import_workbook
from bus_schedule_engine.models import DemandRecord, Direction, Trip, VolumeType

ROOT = Path(__file__).resolve().parents[1]
SCHEMA_DIR = ROOT / "contracts" / "v1"


def _schema_errors(payload: dict[str, object], schema_name: str) -> list[str]:
    schema = json.loads((SCHEMA_DIR / schema_name).read_text(encoding="utf-8"))
    validator = Draft202012Validator(schema, format_checker=FormatChecker())
    return [error.message for error in validator.iter_errors(payload)]


def _options(**overrides) -> NormalizationOptions:
    values = {
        "source_id": "fixture-workbook-sha256",
        "imported_at": datetime(2026, 7, 22, 8, 0, tzinfo=UTC),
        "operating_day_type_b": OperatingDayType.WEEKDAY,
        "available_fleet_limit_b": 3,
        "demand_confidence": DemandConfidence.MEDIUM,
    }
    values.update(overrides)
    return NormalizationOptions(**values)


def _imported_b_only(parameters, trips, demand=None) -> ImportedWorkbook:
    return ImportedWorkbook(
        parameters_a=None,
        trips_a=[],
        parameters_b=parameters,
        trips_b=trips,
        demand=list(demand or []),
        configuration={},
    )


def _imported_a_b(parameters, trips, demand=None) -> ImportedWorkbook:
    trips_a = [replace(item, scenario="A") for item in trips]
    return ImportedWorkbook(
        parameters_a=parameters,
        trips_a=trips_a,
        parameters_b=parameters,
        trips_b=trips,
        demand=list(demand or []),
        configuration={},
    )


def test_legacy_scenario_b_normalizes_and_matches_schema(make_parameters, make_valid_trips) -> None:
    parameters = make_parameters()
    trips = list(reversed(make_valid_trips(parameters)))

    bundle = normalize_imported_workbook_v1(_imported_b_only(parameters, trips), _options())

    assert isinstance(bundle.scenario_b, ScenarioBInput)
    assert bundle.scenario_b.trips_by_direction.outbound == 2
    assert bundle.scenario_b.trips_by_direction.inbound == 2
    assert [trip.trip_id for trip in bundle.scenario_b.exact_timetable] == [
        "T1",
        "T2",
        "T3",
        "T4",
    ]
    assert len(bundle.scenario_b_fingerprint) == 64
    assert (
        _schema_errors(
            scenario_to_contract_dict(bundle.scenario_b),
            "scenario_b_input.schema.json",
        )
        == []
    )


def test_scenario_runtime_is_default_while_exact_trip_runtimes_are_authoritative(
    make_parameters,
    make_valid_trips,
) -> None:
    parameters = make_parameters(
        trip_runtime_minutes=60,
        allowed_trip_runtime_minutes=(55, 65),
    )
    trips = make_valid_trips(parameters)
    trips[0] = replace(
        trips[0],
        arrival_seconds=trips[0].departure_seconds + 55 * 60,
    )
    trips[1] = replace(
        trips[1],
        arrival_seconds=trips[1].departure_seconds + 65 * 60,
    )

    bundle = normalize_imported_workbook_v1(
        _imported_b_only(parameters, trips),
        _options(),
    )

    assert bundle.scenario_b.trip_runtime_minutes == 60
    runtime_by_id = {
        trip.trip_id: trip.runtime_minutes for trip in bundle.scenario_b.exact_timetable
    }
    assert runtime_by_id["T1"] == 55
    assert runtime_by_id["T2"] == 65
    assert validate_scenario_input(bundle.scenario_b).passed


def test_explicit_legacy_runtime_outside_allowed_range_fails_closed(
    make_parameters,
    make_valid_trips,
) -> None:
    parameters = make_parameters(
        trip_runtime_minutes=60,
        allowed_trip_runtime_minutes=(55, 65),
    )
    trips = make_valid_trips(parameters)
    trips[0] = replace(
        trips[0],
        arrival_seconds=trips[0].departure_seconds + 66 * 60,
    )

    with pytest.raises(
        NormalizationError,
        match="TRIP_RUNTIME_OUTSIDE_ALLOWED_RANGE",
    ) as raised:
        normalize_imported_workbook_v1(
            _imported_b_only(parameters, trips),
            _options(),
        )

    assert raised.value.code == "TRIP_RUNTIME_OUTSIDE_ALLOWED_RANGE"


def test_missing_legacy_arrival_uses_maximum_configured_runtime_fallback(
    make_parameters,
    make_valid_trips,
) -> None:
    parameters = make_parameters(
        trip_runtime_minutes=60,
        allowed_trip_runtime_minutes=(55, 65),
    )
    trips = make_valid_trips(parameters)
    trips[0] = replace(trips[0], arrival_seconds=None)

    bundle = normalize_imported_workbook_v1(
        _imported_b_only(parameters, trips),
        _options(),
    )
    normalized = next(
        trip for trip in bundle.scenario_b.exact_timetable if trip.trip_id == trips[0].trip_id
    )

    assert bundle.scenario_b.trip_runtime_minutes == 60
    assert normalized.runtime_minutes == 65
    assert normalized.arrival_time == normalized.departure_time + 65 * 60


def test_normalization_refuses_to_infer_required_fleet(make_parameters, make_valid_trips) -> None:
    parameters = make_parameters()
    imported = _imported_b_only(parameters, make_valid_trips(parameters))

    with pytest.raises(NormalizationError, match="must not infer"):
        normalize_imported_workbook_v1(
            imported,
            _options(available_fleet_limit_b=None),
        )


def test_normalization_requires_explicit_operating_day(make_parameters, make_valid_trips) -> None:
    parameters = make_parameters()
    imported = _imported_b_only(parameters, make_valid_trips(parameters))

    with pytest.raises(NormalizationError, match="not inferred from dates"):
        normalize_imported_workbook_v1(
            imported,
            _options(operating_day_type_b=None),
        )


def test_scenario_fingerprint_ignores_import_timestamp(make_parameters, make_valid_trips) -> None:
    parameters = make_parameters()
    imported = _imported_b_only(parameters, make_valid_trips(parameters))
    first = normalize_imported_workbook_v1(imported, _options())
    second = normalize_imported_workbook_v1(
        imported,
        _options(imported_at=_options().imported_at + timedelta(hours=3)),
    )

    assert first.scenario_b_fingerprint == second.scenario_b_fingerprint


def test_combined_demand_is_preserved_and_average_day_is_exposed(
    make_parameters,
    make_valid_trips,
) -> None:
    parameters = make_parameters()
    demand = [
        DemandRecord(
            period_start=date(2026, 7, 1),
            period_end=date(2026, 7, 15),
            observation_days=15,
            block_start_seconds=6 * 3600,
            block_end_seconds=7 * 3600,
            direction=Direction.COMBINED,
            passenger_volume=150,
            volume_type=VolumeType.TOTAL_OBSERVATION_PERIOD,
        )
    ]
    bundle = normalize_imported_workbook_v1(
        _imported_a_b(parameters, make_valid_trips(parameters), demand),
        _options(
            operating_day_type_a=OperatingDayType.WEEKDAY,
            available_fleet_limit_a=3,
        ),
    )

    assert bundle.observed_demand is not None
    observation = bundle.observed_demand.observations[0]
    assert observation.direction == ContractDirection.COMBINED
    assert observation.average_daily_passenger_count(15) == 10
    assert (
        _schema_errors(
            demand_to_contract_dict(bundle.observed_demand),
            "observed_demand_input.schema.json",
        )
        == []
    )


def test_mixed_demand_periods_require_separate_datasets(make_parameters, make_valid_trips) -> None:
    parameters = make_parameters()
    records = [
        DemandRecord(
            period_start=date(2026, 7, 1),
            period_end=date(2026, 7, 15),
            observation_days=15,
            block_start_seconds=6 * 3600,
            block_end_seconds=7 * 3600,
            direction=Direction.COMBINED,
            passenger_volume=150,
            volume_type=VolumeType.TOTAL_OBSERVATION_PERIOD,
        ),
        DemandRecord(
            period_start=date(2026, 7, 16),
            period_end=date(2026, 7, 20),
            observation_days=5,
            block_start_seconds=7 * 3600,
            block_end_seconds=8 * 3600,
            direction=Direction.COMBINED,
            passenger_volume=60,
            volume_type=VolumeType.TOTAL_OBSERVATION_PERIOD,
        ),
    ]

    with pytest.raises(NormalizationError, match="multiple observation periods"):
        normalize_imported_workbook_v1(
            _imported_b_only(parameters, make_valid_trips(parameters), records),
            _options(),
        )


def test_validator_separates_approved_and_available_fleet(
    make_parameters,
    make_valid_trips,
) -> None:
    parameters = make_parameters()
    bundle = normalize_imported_workbook_v1(
        _imported_b_only(parameters, make_valid_trips(parameters)),
        _options(approved_active_fleet_b=2),
    )
    invalid = replace(bundle.scenario_b, approved_active_fleet=4)

    result = validate_scenario_input(invalid)

    assert "APPROVED_FLEET_EXCEEDS_AVAILABLE_LIMIT" in result.error_codes

    zero_fleet = replace(bundle.scenario_b, approved_active_fleet=0)
    zero_result = validate_scenario_input(zero_fleet)
    assert "INVALID_APPROVED_ACTIVE_FLEET" in zero_result.error_codes


def test_validator_rejects_naive_source_timestamp(make_parameters, make_valid_trips) -> None:
    parameters = make_parameters()
    imported = _imported_b_only(parameters, make_valid_trips(parameters))

    with pytest.raises(NormalizationError, match="SOURCE_IMPORTED_AT_NOT_TIMEZONE_AWARE"):
        normalize_imported_workbook_v1(
            imported,
            _options(imported_at=datetime(2026, 7, 22, 8, 0)),
        )


@pytest.mark.parametrize(
    ("workbook_day_type", "expected_day_type"),
    [
        ("saturday", OperatingDayType.SATURDAY),
        (" ALL_DAYS ", OperatingDayType.ALL_DAYS),
    ],
)
def test_optional_contract_fields_are_read_from_parameter_sheet(
    tmp_path,
    workbook_day_type: str,
    expected_day_type: OperatingDayType,
) -> None:
    source = create_input_template(tmp_path / "contract-input.xlsx")
    workbook = load_workbook(source)
    additions = {
        "available_fleet_limit": 8,
        "approved_active_fleet": 7,
        "operating_day_type": workbook_day_type,
    }
    for sheet_name in ("THONG_SO_A", "THONG_SO_B"):
        sheet = workbook[sheet_name]
        next_row = sheet.max_row + 1
        for offset, (key, value) in enumerate(additions.items()):
            sheet.cell(next_row + offset, 1).value = key
            sheet.cell(next_row + offset, 2).value = value
    workbook.save(source)
    workbook.close()

    imported = import_workbook(source)

    assert imported.parameters_b.available_fleet_limit == 8
    assert imported.parameters_b.approved_active_fleet == 7
    assert imported.parameters_b.operating_day_type == expected_day_type.value

    bundle = normalize_imported_workbook_v1(
        imported,
        NormalizationOptions(
            source_id="workbook-contract-metadata",
            imported_at=datetime(2026, 7, 22, 8, 0, tzinfo=UTC),
        ),
    )
    assert bundle.scenario_a is not None
    assert bundle.scenario_a.available_fleet_limit == 8
    assert bundle.scenario_a.operating_day_type == expected_day_type
    assert bundle.scenario_b.available_fleet_limit == 8
    assert bundle.scenario_b.approved_active_fleet == 7
    assert bundle.scenario_b.operating_day_type == expected_day_type
    serialized_b = scenario_to_contract_dict(bundle.scenario_b)
    assert serialized_b["operating_day_type"] == expected_day_type.value
    assert _schema_errors(serialized_b, "scenario_b_input.schema.json") == []


def test_normalization_does_not_mutate_legacy_trips(make_parameters, make_valid_trips) -> None:
    parameters = make_parameters()
    trips = make_valid_trips(parameters)
    snapshot = [
        Trip(
            scenario=item.scenario,
            trip_id=item.trip_id,
            departure_terminal=item.departure_terminal,
            direction=item.direction,
            departure_seconds=item.departure_seconds,
            arrival_seconds=item.arrival_seconds,
            vehicle_id=item.vehicle_id,
            vehicle_capacity_override=item.vehicle_capacity_override,
        )
        for item in trips
    ]

    normalize_imported_workbook_v1(_imported_b_only(parameters, trips), _options())

    assert trips == snapshot
