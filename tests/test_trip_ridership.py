from __future__ import annotations

from dataclasses import FrozenInstanceError, fields, replace
from datetime import UTC, date, datetime

import pytest
from openpyxl import load_workbook

import bus_schedule_engine.application_pipeline as application_pipeline
from bus_schedule_engine.application_pipeline import (
    UnifiedApplicationStatusV1,
    run_unified_application_pipeline_v1,
)
from bus_schedule_engine.contracts_v1.adapters import normalize_imported_workbook_v1
from bus_schedule_engine.contracts_v1.models import DemandResolutionType
from bus_schedule_engine.excel_exporter import create_input_template
from bus_schedule_engine.importer import ImportedWorkbook, InputDataError, import_workbook
from bus_schedule_engine.input_authority import normalization_options_from_workbook_v1
from bus_schedule_engine.models import (
    TripRidershipAnalysisV1,
    TripRidershipDatasetMetadataV1,
    TripRidershipDirectionV1,
    TripRidershipMatchMethodV1,
    TripRidershipMatchStatusV1,
    TripRidershipObservationV1,
)
from bus_schedule_engine.trip_ridership import (
    analyze_trip_ridership_v1,
    trip_ridership_analysis_is_current_v1,
    trip_ridership_input_fingerprint_v1,
)
from bus_schedule_engine.trip_ridership_codes import (
    AMBIGUOUS_TRIP_TIME_MATCH,
    DUPLICATE_OBSERVATION_FOR_TRIP_DATE,
    DUPLICATE_TRIP_RIDERSHIP_OBSERVATION_ID,
    EXPLICIT_SCHEDULED_TIME_MISMATCH,
    EXPLICIT_SCHEDULED_TRIP_ID_NOT_FOUND,
    EXPLICIT_TRIP_DIRECTION_MISMATCH,
    NO_TRIP_WITHIN_MATCH_TOLERANCE,
    TRIP_RIDERSHIP_ANALYSIS_FAILED,
    TRIP_RIDERSHIP_COMBINED_DIRECTION_NOT_ALLOWED,
    TRIP_RIDERSHIP_FORMULA_NOT_ALLOWED,
    TRIP_RIDERSHIP_MATCH_TOLERANCE_INVALID,
    TRIP_RIDERSHIP_METADATA_MISSING,
    TRIP_RIDERSHIP_OPERATING_DAY_TYPE_MISMATCH,
    TRIP_RIDERSHIP_PASSENGER_COUNT_INVALID,
    TRIP_RIDERSHIP_REFERENCE_MISSING,
)
from bus_schedule_engine.unified_result_exporter import (
    read_unified_export_metadata_bytes_v1,
)


def _set_metadata(workbook, key: str, value: object) -> None:
    sheet = workbook["THONG_TIN_SAN_LUONG_CHUYEN"]
    row = next(cell.row for cell in sheet["A"] if cell.value == key)
    sheet.cell(row, 2).value = value


def _write_observation(
    workbook,
    row: int = 4,
    *,
    observation_id: str = "OBS-001",
    service_date: date = date(2026, 7, 1),
    source_trip_id: str | None = None,
    scheduled_trip_id: str | None = "B-001",
    direction: str = "outbound",
    scheduled_departure_time: str | None = None,
    actual_departure_time: str | None = None,
    passenger_count: object = 20,
    vehicle_id: str | None = None,
    notes: str | None = None,
) -> None:
    values = [
        observation_id,
        service_date,
        source_trip_id,
        scheduled_trip_id,
        direction,
        scheduled_departure_time,
        actual_departure_time,
        passenger_count,
        vehicle_id,
        notes,
    ]
    sheet = workbook["SAN_LUONG_CHUYEN"]
    for column, value in enumerate(values, 1):
        sheet.cell(row, column).value = value


def _workbook_with_observations(tmp_path, rows: list[dict[str, object]]):
    path = create_input_template(tmp_path / "trip-ridership.xlsx")
    workbook = load_workbook(path)
    for index, row in enumerate(rows, 4):
        _write_observation(workbook, index, **row)
    workbook.save(path)
    workbook.close()
    return path


@pytest.fixture(scope="module")
def normalized_context(tmp_path_factory):
    path = create_input_template(tmp_path_factory.mktemp("trip-ridership") / "base.xlsx")
    imported = import_workbook(path)
    options = normalization_options_from_workbook_v1(
        imported,
        source_id="trip-ridership-tests",
        imported_at=datetime(2026, 7, 31, tzinfo=UTC),
    )
    normalized = normalize_imported_workbook_v1(imported, options)
    return imported, normalized.scenario_b


def _metadata(tolerance: int = 5) -> TripRidershipDatasetMetadataV1:
    return TripRidershipDatasetMetadataV1(
        dataset_id="TRIP-DATASET-1",
        source_type="manual_count",
        confidence="medium",
        observed_schedule_scenario="B",
        operating_day_type="weekday",
        match_tolerance_minutes=tolerance,
    )


def _observation(
    observation_id: str = "OBS-001",
    *,
    service_date: date = date(2026, 7, 1),
    source_trip_id: str | None = None,
    scheduled_trip_id: str | None = "B-001",
    direction: TripRidershipDirectionV1 = TripRidershipDirectionV1.OUTBOUND,
    scheduled_departure_seconds: int | None = None,
    actual_departure_seconds: int | None = None,
    passenger_count: int = 20,
    vehicle_id: str | None = None,
    notes: str | None = None,
) -> TripRidershipObservationV1:
    return TripRidershipObservationV1(
        observation_id=observation_id,
        service_date=service_date,
        source_trip_id=source_trip_id,
        scheduled_trip_id=scheduled_trip_id,
        direction=direction,
        scheduled_departure_seconds=scheduled_departure_seconds,
        actual_departure_seconds=actual_departure_seconds,
        passenger_count=passenger_count,
        vehicle_id=vehicle_id,
        notes=notes,
    )


def _analyze(
    normalized_context,
    observations: tuple[TripRidershipObservationV1, ...],
    *,
    tolerance: int = 5,
    imported_transform=None,
):
    imported, scenario_b = normalized_context
    with_observations = replace(
        imported,
        trip_ridership_metadata=_metadata(tolerance),
        trip_ridership_observations=observations,
    )
    if imported_transform is not None:
        with_observations = imported_transform(with_observations)
    return (
        with_observations,
        scenario_b,
        analyze_trip_ridership_v1(with_observations, scenario_b),
    )


def test_workbook_without_trip_sheets_imports_with_safe_defaults(tmp_path) -> None:
    path = create_input_template(tmp_path / "old.xlsx")
    workbook = load_workbook(path)
    del workbook["THONG_TIN_SAN_LUONG_CHUYEN"]
    del workbook["SAN_LUONG_CHUYEN"]
    workbook.save(path)
    workbook.close()

    imported = import_workbook(path)

    assert imported.trip_ridership_metadata is None
    assert imported.trip_ridership_observations == ()
    assert imported.demand


def test_empty_trip_sheet_and_metadata_only_are_not_provided(tmp_path) -> None:
    imported = import_workbook(create_input_template(tmp_path / "empty.xlsx"))

    assert imported.trip_ridership_metadata is None
    assert imported.trip_ridership_observations == ()


def test_trip_data_without_metadata_is_rejected(tmp_path) -> None:
    path = _workbook_with_observations(tmp_path, [{}])
    workbook = load_workbook(path)
    del workbook["THONG_TIN_SAN_LUONG_CHUYEN"]
    workbook.save(path)
    workbook.close()

    with pytest.raises(InputDataError, match=TRIP_RIDERSHIP_METADATA_MISSING):
        import_workbook(path)


def test_trip_sheet_required_columns_are_enforced(tmp_path) -> None:
    path = _workbook_with_observations(tmp_path, [{}])
    workbook = load_workbook(path)
    workbook["SAN_LUONG_CHUYEN"].cell(3, 8).value = "wrong_header"
    workbook.save(path)
    workbook.close()

    with pytest.raises(InputDataError, match="passenger_count"):
        import_workbook(path)


def test_duplicate_observation_ids_are_rejected(tmp_path) -> None:
    path = _workbook_with_observations(
        tmp_path,
        [
            {"observation_id": " DUPLICATE "},
            {"observation_id": "DUPLICATE", "service_date": date(2026, 7, 2)},
        ],
    )

    with pytest.raises(InputDataError, match=DUPLICATE_TRIP_RIDERSHIP_OBSERVATION_ID):
        import_workbook(path)


@pytest.mark.parametrize(
    ("passenger_count", "accepted"),
    ((0, True), (-1, False), (1.5, False)),
)
def test_passenger_count_validation(tmp_path, passenger_count, accepted) -> None:
    path = _workbook_with_observations(
        tmp_path,
        [{"passenger_count": passenger_count}],
    )

    if accepted:
        assert import_workbook(path).trip_ridership_observations[0].passenger_count == 0
    else:
        with pytest.raises(InputDataError, match=TRIP_RIDERSHIP_PASSENGER_COUNT_INVALID):
            import_workbook(path)


def test_combined_direction_and_missing_reference_are_rejected(tmp_path) -> None:
    combined = _workbook_with_observations(
        tmp_path,
        [{"direction": "combined"}],
    )
    with pytest.raises(InputDataError, match=TRIP_RIDERSHIP_COMBINED_DIRECTION_NOT_ALLOWED):
        import_workbook(combined)

    missing = _workbook_with_observations(
        tmp_path,
        [
            {
                "scheduled_trip_id": None,
                "scheduled_departure_time": None,
                "actual_departure_time": None,
            }
        ],
    )
    with pytest.raises(InputDataError, match=TRIP_RIDERSHIP_REFERENCE_MISSING):
        import_workbook(missing)


def test_operating_day_and_tolerance_metadata_are_strict(tmp_path) -> None:
    day_path = _workbook_with_observations(tmp_path, [{}])
    workbook = load_workbook(day_path)
    _set_metadata(workbook, "operating_day_type", "saturday")
    workbook.save(day_path)
    workbook.close()
    with pytest.raises(InputDataError, match=TRIP_RIDERSHIP_OPERATING_DAY_TYPE_MISMATCH):
        import_workbook(day_path)

    for tolerance in (-1, 31, 1.5):
        path = _workbook_with_observations(tmp_path, [{}])
        workbook = load_workbook(path)
        _set_metadata(workbook, "match_tolerance_minutes", tolerance)
        workbook.save(path)
        workbook.close()
        with pytest.raises(InputDataError, match=TRIP_RIDERSHIP_MATCH_TOLERANCE_INVALID):
            import_workbook(path)


@pytest.mark.parametrize(
    ("key", "value", "code"),
    (
        ("trip_ridership_dataset_id", "   ", "TRIP_RIDERSHIP_DATASET_ID_MISSING"),
        ("trip_ridership_source_type", "aggregate_report", "TRIP_RIDERSHIP_SOURCE_TYPE_INVALID"),
        ("trip_ridership_confidence", "invented", "TRIP_RIDERSHIP_CONFIDENCE_INVALID"),
        ("observed_schedule_scenario", "A", "TRIP_RIDERSHIP_SCENARIO_INVALID"),
    ),
)
def test_required_trip_metadata_vocabularies_are_enforced(
    tmp_path,
    key,
    value,
    code,
) -> None:
    path = _workbook_with_observations(tmp_path, [{}])
    workbook = load_workbook(path)
    _set_metadata(workbook, key, value)
    workbook.save(path)
    workbook.close()

    with pytest.raises(InputDataError, match=code):
        import_workbook(path)


def test_trip_formulas_are_not_evaluated_as_authority(tmp_path) -> None:
    path = _workbook_with_observations(tmp_path, [{}])
    workbook = load_workbook(path)
    workbook["SAN_LUONG_CHUYEN"].cell(4, 8).value = "=40+2"
    workbook.save(path)
    workbook.close()

    with pytest.raises(InputDataError, match=TRIP_RIDERSHIP_FORMULA_NOT_ALLOWED):
        import_workbook(path)


@pytest.mark.parametrize(
    "invalid_time",
    (-0.25, 1.0, 2.0, True, "24:00", "24:00:00"),
)
@pytest.mark.parametrize(
    "field_name",
    ("scheduled_departure_time", "actual_departure_time"),
)
def test_trip_times_must_remain_within_one_excel_service_day(
    tmp_path,
    invalid_time,
    field_name,
) -> None:
    path = _workbook_with_observations(
        tmp_path,
        [{"scheduled_trip_id": None, field_name: invalid_time}],
    )

    with pytest.raises(InputDataError, match=field_name):
        import_workbook(path)


@pytest.mark.parametrize(
    ("valid_time", "expected_seconds"),
    (
        ("00:00", 0),
        ("23:59", 23 * 3600 + 59 * 60),
        ("23:59:59", 24 * 3600 - 1),
    ),
)
@pytest.mark.parametrize(
    ("field_name", "attribute_name"),
    (
        ("scheduled_departure_time", "scheduled_departure_seconds"),
        ("actual_departure_time", "actual_departure_seconds"),
    ),
)
def test_trip_times_accept_the_supported_service_day_boundary(
    tmp_path,
    valid_time,
    expected_seconds,
    field_name,
    attribute_name,
) -> None:
    path = _workbook_with_observations(
        tmp_path,
        [{"scheduled_trip_id": None, field_name: valid_time}],
    )

    observation = import_workbook(path).trip_ridership_observations[0]

    assert getattr(observation, attribute_name) == expected_seconds


def test_explicit_trip_id_precedence_and_contradictions(normalized_context) -> None:
    _, _, exact = _analyze(normalized_context, (_observation(),))
    assert exact.match_rows[0].match_status == TripRidershipMatchStatusV1.MATCHED_EXACT
    assert exact.match_rows[0].matched_trip_id == "B-001"

    _, _, unknown = _analyze(
        normalized_context,
        (
            _observation(
                scheduled_trip_id="UNKNOWN",
                scheduled_departure_seconds=6 * 3600,
            ),
        ),
    )
    assert unknown.match_rows[0].match_status == TripRidershipMatchStatusV1.INVALID
    assert unknown.match_rows[0].matched_trip_id is None
    assert EXPLICIT_SCHEDULED_TRIP_ID_NOT_FOUND in unknown.match_rows[0].issue_codes

    _, _, direction = _analyze(
        normalized_context,
        (_observation(direction=TripRidershipDirectionV1.INBOUND),),
    )
    assert EXPLICIT_TRIP_DIRECTION_MISMATCH in direction.match_rows[0].issue_codes

    _, _, scheduled_time = _analyze(
        normalized_context,
        (_observation(scheduled_departure_seconds=6 * 3600 + 60),),
    )
    assert EXPLICIT_SCHEDULED_TIME_MISMATCH in scheduled_time.match_rows[0].issue_codes


def test_time_matching_is_nearest_inclusive_and_deterministic(normalized_context) -> None:
    observations = (
        _observation(
            "EXACT",
            scheduled_trip_id=None,
            scheduled_departure_seconds=6 * 3600,
        ),
        _observation(
            "NEAREST",
            service_date=date(2026, 7, 2),
            scheduled_trip_id=None,
            scheduled_departure_seconds=6 * 3600 + 4 * 60,
        ),
        _observation(
            "BOUNDARY",
            service_date=date(2026, 7, 3),
            scheduled_trip_id=None,
            scheduled_departure_seconds=6 * 3600 + 5 * 60,
        ),
    )
    _, _, analysis = _analyze(normalized_context, observations, tolerance=5)
    rows = {item.observation_id: item for item in analysis.match_rows}

    assert rows["EXACT"].match_status == TripRidershipMatchStatusV1.MATCHED_EXACT
    assert rows["NEAREST"].match_status == TripRidershipMatchStatusV1.MATCHED_WITHIN_TOLERANCE
    assert rows["NEAREST"].absolute_time_offset_seconds == 4 * 60
    assert rows["BOUNDARY"].match_status == TripRidershipMatchStatusV1.MATCHED_WITHIN_TOLERANCE


def test_equal_distance_is_ambiguous_and_no_candidate_is_unmatched(
    normalized_context,
) -> None:
    _, _, analysis = _analyze(
        normalized_context,
        (
            _observation(
                "AMBIGUOUS",
                scheduled_trip_id=None,
                scheduled_departure_seconds=6 * 3600 + 22 * 60 + 30,
            ),
            _observation(
                "UNMATCHED",
                service_date=date(2026, 7, 2),
                scheduled_trip_id=None,
                scheduled_departure_seconds=8 * 3600,
            ),
        ),
        tolerance=30,
    )
    rows = {item.observation_id: item for item in analysis.match_rows}

    assert rows["AMBIGUOUS"].match_status == TripRidershipMatchStatusV1.AMBIGUOUS
    assert AMBIGUOUS_TRIP_TIME_MATCH in rows["AMBIGUOUS"].issue_codes
    assert rows["UNMATCHED"].match_status == TripRidershipMatchStatusV1.UNMATCHED
    assert NO_TRIP_WITHIN_MATCH_TOLERANCE in rows["UNMATCHED"].issue_codes


def test_scheduled_time_precedes_actual_and_actual_is_fallback(normalized_context) -> None:
    _, _, analysis = _analyze(
        normalized_context,
        (
            _observation(
                "SCHEDULED",
                scheduled_trip_id=None,
                scheduled_departure_seconds=6 * 3600,
                actual_departure_seconds=10 * 3600 + 30 * 60,
            ),
            _observation(
                "ACTUAL",
                service_date=date(2026, 7, 2),
                scheduled_trip_id=None,
                actual_departure_seconds=6 * 3600 + 4 * 60,
            ),
        ),
    )
    rows = {item.observation_id: item for item in analysis.match_rows}

    assert rows["SCHEDULED"].match_method == TripRidershipMatchMethodV1.SCHEDULED_DEPARTURE_TIME
    assert rows["SCHEDULED"].matched_trip_id == "B-001"
    assert rows["ACTUAL"].match_method == TripRidershipMatchMethodV1.ACTUAL_DEPARTURE_TIME
    assert rows["ACTUAL"].matched_trip_id == "B-001"


def test_matching_never_crosses_direction(normalized_context) -> None:
    _, _, analysis = _analyze(
        normalized_context,
        (
            _observation(
                scheduled_trip_id=None,
                actual_departure_seconds=6 * 3600 + 15 * 60,
            ),
        ),
        tolerance=5,
    )

    assert analysis.match_rows[0].match_status == TripRidershipMatchStatusV1.UNMATCHED


def test_collisions_exclude_every_record_but_multiple_dates_are_valid(
    normalized_context,
) -> None:
    _, _, collided = _analyze(
        normalized_context,
        (_observation("A"), _observation("B", passenger_count=30)),
    )
    assert {item.match_status for item in collided.match_rows} == {
        TripRidershipMatchStatusV1.COLLISION
    }
    assert all(
        DUPLICATE_OBSERVATION_FOR_TRIP_DATE in item.issue_codes for item in collided.match_rows
    )
    assert collided.dataset_summary.usable_matched_records == 0
    trip = next(item for item in collided.trip_summaries if item.trip_id == "B-001")
    assert trip.observation_count == 0
    assert trip.passenger_mean is None

    _, _, multiple_dates = _analyze(
        normalized_context,
        (
            _observation("A"),
            _observation("B", service_date=date(2026, 7, 2), passenger_count=30),
        ),
    )
    trip = next(item for item in multiple_dates.trip_summaries if item.trip_id == "B-001")
    assert trip.observation_count == 2
    assert trip.distinct_observation_day_count == 2


def test_statistics_use_nearest_rank_capacity_override_and_real_zero(
    normalized_context,
) -> None:
    observations = tuple(
        _observation(
            f"OBS-{index}",
            service_date=date(2026, 7, index),
            passenger_count=count,
        )
        for index, count in enumerate((0, 20, 40, 60), 1)
    )

    def override_capacity(imported: ImportedWorkbook) -> ImportedWorkbook:
        trips = list(imported.trips_b)
        trips[0] = replace(trips[0], vehicle_capacity_override=50)
        return replace(imported, trips_b=trips)

    _, _, analysis = _analyze(
        normalized_context,
        observations,
        imported_transform=override_capacity,
    )
    trip = next(item for item in analysis.trip_summaries if item.trip_id == "B-001")

    assert trip.nominal_trip_capacity == 50
    assert trip.passenger_minimum == 0
    assert trip.passenger_maximum == 60
    assert trip.passenger_mean == 30
    assert trip.passenger_median == 30
    assert trip.passenger_p85 == 60
    assert trip.passenger_p90 == 60
    assert trip.mean_load_factor == pytest.approx(0.6)
    assert trip.days_at_or_above_target_load_factor == 1
    assert trip.days_above_maximum_load_factor == 1


def test_odd_median_and_nearest_rank_percentiles_are_deterministic(
    normalized_context,
) -> None:
    _, _, analysis = _analyze(
        normalized_context,
        tuple(
            _observation(
                f"ODD-{index}",
                service_date=date(2026, 7, index),
                passenger_count=count,
            )
            for index, count in enumerate((1, 3, 9), 1)
        ),
    )
    trip = next(item for item in analysis.trip_summaries if item.trip_id == "B-001")

    assert trip.passenger_median == 3
    assert trip.passenger_p85 == 9
    assert trip.passenger_p90 == 9


def test_missing_observations_remain_none_and_coverage_is_explicit(
    normalized_context,
) -> None:
    _, _, analysis = _analyze(
        normalized_context,
        (
            _observation("A", passenger_count=10),
            _observation(
                "B",
                service_date=date(2026, 7, 2),
                passenger_count=20,
            ),
        ),
    )
    summary = analysis.dataset_summary
    unobserved = next(item for item in analysis.trip_summaries if item.trip_id == "B-003")

    assert unobserved.observation_count == 0
    assert unobserved.passenger_minimum is None
    assert unobserved.passenger_mean is None
    assert unobserved.nominal_trip_capacity == 60
    assert summary.observed_matched_passengers == 30
    assert summary.scheduled_trip_coverage_rate == pytest.approx(1 / 24)
    assert summary.matched_trip_date_coverage_rate == pytest.approx(2 / (24 * 2))
    assert "not available" in summary.coverage_adjusted_interpretation


def test_current_analysis_requires_current_supplemental_input(normalized_context) -> None:
    imported, _scenario_b, analysis = _analyze(
        normalized_context,
        (
            _observation(
                source_trip_id="SOURCE-1",
                scheduled_departure_seconds=6 * 3600,
                actual_departure_seconds=6 * 3600 + 60,
                vehicle_id="VEHICLE-1",
            ),
        ),
    )

    assert analysis.trip_ridership_input_fingerprint != analysis.analysis_fingerprint
    assert trip_ridership_analysis_is_current_v1(
        analysis,
        imported,
        analysis.scenario_b_timetable_fingerprint,
    )


@pytest.mark.parametrize(
    "changed_fact",
    (
        "dataset_id",
        "source_type",
        "confidence",
        "observed_schedule_scenario",
        "operating_day_type",
        "match_tolerance_minutes",
        "observation_id",
        "service_date",
        "source_trip_id",
        "scheduled_trip_id",
        "direction",
        "scheduled_departure_seconds",
        "actual_departure_seconds",
        "passenger_count",
        "vehicle_id",
    ),
)
def test_same_scenario_b_with_changed_semantic_input_is_stale(
    normalized_context,
    changed_fact,
) -> None:
    imported, _scenario_b, analysis = _analyze(
        normalized_context,
        (
            _observation(
                source_trip_id="SOURCE-1",
                scheduled_departure_seconds=6 * 3600,
                actual_departure_seconds=6 * 3600 + 60,
                vehicle_id="VEHICLE-1",
            ),
        ),
    )
    metadata_changes = {
        "dataset_id": "TRIP-DATASET-2",
        "source_type": "apc",
        "confidence": "high",
        "observed_schedule_scenario": "A",
        "operating_day_type": "saturday",
        "match_tolerance_minutes": 6,
    }
    observation_changes = {
        "observation_id": "OBS-CHANGED",
        "service_date": date(2026, 7, 2),
        "source_trip_id": "SOURCE-2",
        "scheduled_trip_id": "B-003",
        "direction": TripRidershipDirectionV1.INBOUND,
        "scheduled_departure_seconds": 6 * 3600 + 60,
        "actual_departure_seconds": 6 * 3600 + 2 * 60,
        "passenger_count": 21,
        "vehicle_id": "VEHICLE-2",
    }
    if changed_fact in metadata_changes:
        assert imported.trip_ridership_metadata is not None
        changed_imported = replace(
            imported,
            trip_ridership_metadata=replace(
                imported.trip_ridership_metadata,
                **{changed_fact: metadata_changes[changed_fact]},
            ),
        )
    else:
        changed_imported = replace(
            imported,
            trip_ridership_observations=(
                replace(
                    imported.trip_ridership_observations[0],
                    **{changed_fact: observation_changes[changed_fact]},
                ),
            ),
        )

    current_fingerprint = trip_ridership_input_fingerprint_v1(
        changed_imported,
        analysis.scenario_b_timetable_fingerprint,
    )

    assert current_fingerprint != analysis.trip_ridership_input_fingerprint
    assert not trip_ridership_analysis_is_current_v1(
        analysis,
        changed_imported,
        analysis.scenario_b_timetable_fingerprint,
    )


def test_notes_and_row_order_do_not_make_current_analysis_stale(
    normalized_context,
) -> None:
    observations = (
        _observation(
            "OBS-A",
            source_trip_id="SOURCE-A",
            vehicle_id="VEHICLE-A",
            notes="first note",
        ),
        _observation(
            "OBS-B",
            service_date=date(2026, 7, 2),
            source_trip_id="SOURCE-B",
            vehicle_id="VEHICLE-B",
            notes="second note",
        ),
    )
    imported, _scenario_b, analysis = _analyze(normalized_context, observations)
    assert imported.trip_ridership_metadata is not None
    notes_changed = replace(
        imported,
        trip_ridership_metadata=replace(
            imported.trip_ridership_metadata,
            source_notes="new free-form dataset note",
        ),
        trip_ridership_observations=tuple(
            replace(item, notes=f"changed {item.observation_id}") for item in observations
        ),
    )
    reordered = replace(
        imported,
        trip_ridership_observations=tuple(reversed(observations)),
    )

    assert trip_ridership_input_fingerprint_v1(
        notes_changed,
        analysis.scenario_b_timetable_fingerprint,
    ) == trip_ridership_input_fingerprint_v1(
        reordered,
        analysis.scenario_b_timetable_fingerprint,
    )
    assert trip_ridership_analysis_is_current_v1(
        analysis,
        notes_changed,
        analysis.scenario_b_timetable_fingerprint,
    )
    assert trip_ridership_analysis_is_current_v1(
        analysis,
        reordered,
        analysis.scenario_b_timetable_fingerprint,
    )


def test_current_analysis_fails_closed_for_missing_input_changed_b_and_bad_integrity(
    normalized_context,
) -> None:
    imported, _scenario_b, analysis = _analyze(
        normalized_context,
        (_observation(),),
    )
    missing = replace(
        imported,
        trip_ridership_metadata=None,
        trip_ridership_observations=(),
    )
    fabricated = replace(
        analysis,
        analysis_fingerprint=analysis.trip_ridership_input_fingerprint,
    )

    assert (
        trip_ridership_input_fingerprint_v1(
            missing,
            analysis.scenario_b_timetable_fingerprint,
        )
        is None
    )
    assert not trip_ridership_analysis_is_current_v1(
        analysis,
        missing,
        analysis.scenario_b_timetable_fingerprint,
    )
    assert not trip_ridership_analysis_is_current_v1(
        analysis,
        imported,
        "0" * 64,
    )
    assert not trip_ridership_analysis_is_current_v1(
        fabricated,
        imported,
        analysis.scenario_b_timetable_fingerprint,
    )


def test_semantic_row_order_preserves_fingerprint_and_fact_changes_do_not(
    normalized_context,
) -> None:
    first = _observation("A", passenger_count=10)
    second = _observation(
        "B",
        service_date=date(2026, 7, 2),
        passenger_count=20,
    )
    _, _, ordered = _analyze(normalized_context, (first, second))
    _, _, reordered = _analyze(normalized_context, (second, first))
    _, _, changed_passengers = _analyze(
        normalized_context,
        (first, replace(second, passenger_count=21)),
    )
    _, _, changed_tolerance = _analyze(
        normalized_context,
        (first, second),
        tolerance=6,
    )

    assert ordered.analysis_fingerprint == reordered.analysis_fingerprint
    assert ordered.match_rows == reordered.match_rows
    assert ordered.analysis_fingerprint != changed_passengers.analysis_fingerprint
    assert ordered.analysis_fingerprint != changed_tolerance.analysis_fingerprint


def test_all_normalized_reference_facts_and_capacity_affect_fingerprint(
    normalized_context,
) -> None:
    base = _observation(
        scheduled_departure_seconds=6 * 3600,
        actual_departure_seconds=6 * 3600 + 60,
    )
    _, _, baseline = _analyze(normalized_context, (base,))
    variants = (
        replace(base, service_date=date(2026, 7, 2)),
        replace(base, direction=TripRidershipDirectionV1.INBOUND),
        replace(base, scheduled_trip_id="B-003", scheduled_departure_seconds=6 * 3600 + 45 * 60),
        replace(base, scheduled_departure_seconds=6 * 3600 + 60),
        replace(base, actual_departure_seconds=6 * 3600 + 2 * 60),
    )
    for variant in variants:
        _, _, changed = _analyze(normalized_context, (variant,))
        assert baseline.analysis_fingerprint != changed.analysis_fingerprint

    def override_capacity(imported: ImportedWorkbook) -> ImportedWorkbook:
        trips = list(imported.trips_b)
        trips[0] = replace(trips[0], vehicle_capacity_override=55)
        return replace(imported, trips_b=trips)

    _, _, changed_capacity = _analyze(
        normalized_context,
        (base,),
        imported_transform=override_capacity,
    )
    assert baseline.analysis_fingerprint != changed_capacity.analysis_fingerprint


def test_free_form_notes_are_excluded_from_analysis_fingerprint(
    normalized_context,
) -> None:
    imported, scenario_b, baseline = _analyze(
        normalized_context,
        (replace(_observation(), notes="first observation note"),),
    )
    changed_notes = replace(
        imported,
        trip_ridership_metadata=replace(
            imported.trip_ridership_metadata,
            source_notes="different dataset note",
        ),
        trip_ridership_observations=(
            replace(
                imported.trip_ridership_observations[0],
                notes="different observation note",
            ),
        ),
    )

    assert (
        analyze_trip_ridership_v1(changed_notes, scenario_b).analysis_fingerprint
        == baseline.analysis_fingerprint
    )


def test_scenario_b_and_collision_changes_change_fingerprint(normalized_context) -> None:
    imported, scenario_b, baseline = _analyze(
        normalized_context,
        (
            _observation("A"),
            _observation("B", service_date=date(2026, 7, 2)),
        ),
    )
    timetable = list(scenario_b.exact_timetable)
    timetable[0] = replace(timetable[0], departure_time=timetable[0].departure_time + 60)
    changed_b = replace(scenario_b, exact_timetable=tuple(timetable))
    changed_b_analysis = analyze_trip_ridership_v1(imported, changed_b)
    _, _, collision = _analyze(
        normalized_context,
        (_observation("A"), _observation("B")),
    )

    assert baseline.analysis_fingerprint != changed_b_analysis.analysis_fingerprint
    assert baseline.analysis_fingerprint != collision.analysis_fingerprint
    assert not trip_ridership_analysis_is_current_v1(
        baseline,
        imported,
        changed_b_analysis.scenario_b_timetable_fingerprint,
    )


def test_analysis_models_are_frozen_slotted_and_inputs_are_unchanged(
    normalized_context,
) -> None:
    imported, scenario_b = normalized_context
    observations = (_observation(),)
    with_observations = replace(
        imported,
        trip_ridership_metadata=_metadata(),
        trip_ridership_observations=observations,
    )
    original_trips = tuple(with_observations.trips_b)

    analysis = analyze_trip_ridership_v1(with_observations, scenario_b)

    assert "__slots__" in TripRidershipAnalysisV1.__dict__
    assert fields(analysis)
    with pytest.raises(FrozenInstanceError):
        analysis.dataset_id = "mutated"
    assert with_observations.trip_ridership_observations == observations
    assert tuple(with_observations.trips_b) == original_trips


@pytest.mark.parametrize(
    "metadata",
    (
        replace(_metadata(), observed_schedule_scenario="A"),
        replace(_metadata(), match_tolerance_minutes=-1),
        replace(_metadata(), match_tolerance_minutes=31),
    ),
)
def test_direct_analysis_rejects_invalid_policy_metadata(
    normalized_context,
    metadata,
) -> None:
    imported, scenario_b = normalized_context
    with_observations = replace(
        imported,
        trip_ridership_metadata=metadata,
        trip_ridership_observations=(_observation(),),
    )

    with pytest.raises(ValueError):
        analyze_trip_ridership_v1(with_observations, scenario_b)


def test_pipeline_contract_result_is_identical_with_supplemental_data(
    normalized_context,
    monkeypatch,
) -> None:
    imported, _scenario_b = normalized_context
    with_trip = replace(
        imported,
        trip_ridership_metadata=_metadata(),
        trip_ridership_observations=(_observation(),),
    )
    original_runner = application_pipeline.analyze_and_optimize_schedule_v1
    calls = 0

    def counted_runner(*args, **kwargs):
        nonlocal calls
        calls += 1
        return original_runner(*args, **kwargs)

    monkeypatch.setattr(
        application_pipeline,
        "analyze_and_optimize_schedule_v1",
        counted_runner,
    )
    imported_at = datetime(2026, 7, 31, tzinfo=UTC)
    baseline = run_unified_application_pipeline_v1(
        imported,
        source_id="pipeline-comparison",
        imported_at=imported_at,
    )
    supplemental = run_unified_application_pipeline_v1(
        with_trip,
        source_id="pipeline-comparison",
        imported_at=imported_at,
    )

    assert calls == 2
    assert baseline.status == supplemental.status == UnifiedApplicationStatusV1.COMPLETE
    assert replace(
        baseline.unified_result,
        protected_service_floor_enforcement_authority=None,
    ) == replace(
        supplemental.unified_result,
        protected_service_floor_enforcement_authority=None,
    )
    assert (
        not baseline.unified_result.protected_service_floor_enforcement_authority.protected_regimes
    )
    assert not supplemental.unified_result.protected_service_floor_enforcement_authority.protected_regimes
    assert baseline.unified_presentation == supplemental.unified_presentation
    assert read_unified_export_metadata_bytes_v1(
        baseline.unified_xlsx_bytes
    ) == read_unified_export_metadata_bytes_v1(supplemental.unified_xlsx_bytes)
    assert baseline.trip_ridership_analysis is None
    assert supplemental.trip_ridership_analysis is not None
    assert supplemental.trip_ridership_failure is None
    observed_demand = supplemental.unified_result.normalized_inputs.observed_demand
    assert observed_demand is not None
    assert all(
        item.source_resolution_type != DemandResolutionType.TRIP
        for item in observed_demand.observations
    )
    assert all(item.observation_id != "OBS-001" for item in observed_demand.observations)


def test_supplemental_failure_isolated_from_valid_contract_result(
    normalized_context,
    monkeypatch,
) -> None:
    imported, _scenario_b = normalized_context
    with_trip = replace(
        imported,
        trip_ridership_metadata=_metadata(),
        trip_ridership_observations=(_observation(),),
    )

    def fail_supplemental(*_args, **_kwargs):
        raise RuntimeError("synthetic raw observation should not be logged")

    monkeypatch.setattr(
        application_pipeline,
        "analyze_trip_ridership_v1",
        fail_supplemental,
    )
    monkeypatch.setattr(
        application_pipeline,
        "run_and_build_artifacts",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("legacy runtime must not run")
        ),
    )
    run = run_unified_application_pipeline_v1(
        with_trip,
        source_id="supplemental-failure",
        imported_at=datetime(2026, 7, 31, tzinfo=UTC),
    )

    assert run.status == UnifiedApplicationStatusV1.COMPLETE
    assert run.unified_result is not None
    assert run.unified_presentation is not None
    assert run.trip_ridership_analysis is None
    assert run.trip_ridership_failure is not None
    assert run.trip_ridership_failure.code == TRIP_RIDERSHIP_ANALYSIS_FAILED
    assert "synthetic raw observation" not in run.trip_ridership_failure.sanitized_message
