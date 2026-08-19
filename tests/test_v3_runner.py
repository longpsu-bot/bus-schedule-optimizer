from __future__ import annotations

import json
import os
import shutil
import zipfile
from datetime import date
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook

from bus_schedule_engine.contracts_v1 import (
    MultiPeriodDemandError,
    SolverPolicyV1,
    build_two_stage_uniform_request_v1,
)
from bus_schedule_engine.v3_result_exporter import (
    MATERIAL_PROFILE_SENSITIVITY,
    MINOR_PROFILE_SENSITIVITY,
    PROFILE_COMPARISON_ELIGIBLE,
    PROFILE_COMPARISON_INCONCLUSIVE,
    PROFILE_COMPARISON_INCONCLUSIVE_NO_COMPARABLE_C,
    PROFILE_COMPARISON_PARTIALLY_COMPARABLE,
    STABLE_ACROSS_PROFILES,
    _comparison_row,
    build_profile_comparison_v1,
)
from bus_schedule_engine.v3_runner import run_v3_profile_v1, write_deterministic_json
from bus_schedule_engine.v3_workbook import import_v3_multi_period_workbook_v1
from scripts.run_v3_two_stage import main


def _write_v3_workbook(path: Path) -> Path:
    workbook = Workbook()
    parameters = workbook.active
    parameters.title = "THONG_SO_B"
    parameters.append(["THÔNG SỐ SCENARIO B", None, None, None])
    parameters.append(["Tham số", "Giá trị", "Mức độ", "Diễn giải"])
    parameter_values = {
        "route_id": "SYNTH",
        "route_name": "Synthetic route",
        "route_type": "intra_provincial",
        "allowed_trip_runtime_minutes": "30",
        "trip_runtime_minutes": "30",
        "total_daily_trips": 6,
        "terminal_1_name": "Terminal 1",
        "terminal_1_first_departure": "06:00:00",
        "terminal_1_last_departure": "08:00:00",
        "terminal_2_name": "Terminal 2",
        "terminal_2_first_departure": "06:10:00",
        "terminal_2_last_departure": "08:10:00",
        "vehicle_capacity_passengers": 60,
        "minimum_layover_minutes": 5,
        "available_fleet_limit": 4,
        "operating_day_type": "all_days",
    }
    for key, value in parameter_values.items():
        parameters.append([key, value, "REQUIRED", ""])

    timetable = workbook.create_sheet("BIEU_DO_B")
    timetable.append(["SCENARIO B", None, None, None])
    timetable.append(["Synthetic", None, None, None])
    timetable.append(
        [
            "scenario",
            "trip_id",
            "departure_terminal",
            "direction",
            "departure_time",
            "arrival_time",
            "vehicle_id",
            "vehicle_capacity_override",
        ]
    )
    definitions = [
        ("O1", "Terminal 1", "terminal_1_to_2", "06:00:00", "06:30:00"),
        ("I1", "Terminal 2", "terminal_2_to_1", "06:10:00", "06:40:00"),
        ("O2", "Terminal 1", "terminal_1_to_2", "07:00:00", "07:30:00"),
        ("I2", "Terminal 2", "terminal_2_to_1", "07:10:00", "07:40:00"),
        ("O3", "Terminal 1", "terminal_1_to_2", "08:00:00", "08:30:00"),
        ("I3", "Terminal 2", "terminal_2_to_1", "08:10:00", "08:40:00"),
    ]
    for trip_id, terminal, direction, departure, arrival in definitions:
        timetable.append(["B", trip_id, terminal, direction, departure, arrival, None, None])

    catalog = workbook.create_sheet("PERIOD_CATALOG")
    catalog.append(
        [
            "period_id",
            "period_start",
            "period_end",
            "observation_days",
            "period_role",
            "status",
            "source_dataset_id",
            "notes",
        ]
    )
    catalog.append(["p1", date(2026, 3, 1), date(2026, 3, 2), 2, "CURRENT", "READY", "ds-p1", ""])
    catalog.append(["p2", date(2026, 4, 1), date(2026, 4, 4), 4, "CURRENT", "READY", "ds-p2", ""])

    demand = workbook.create_sheet("SAN_LUONG_MULTI_PERIOD")
    demand.append(
        [
            "period_id",
            "period_start",
            "period_end",
            "observation_days",
            "time_block_start",
            "time_block_end",
            "direction",
            "passenger_volume",
            "volume_type",
            "source_time_basis",
            "source_dataset_id",
        ]
    )
    for period_id, start_date, end_date, days, dataset, factor in (
        ("p1", date(2026, 3, 1), date(2026, 3, 2), 2, "ds-p1", 1),
        ("p2", date(2026, 4, 1), date(2026, 4, 4), 4, "ds-p2", 2),
    ):
        for direction in ("terminal_1_to_2", "terminal_2_to_1"):
            for block_index, (block_start, block_end) in enumerate(
                (
                    ("06:00:00", "07:00:00"),
                    ("07:00:00", "08:00:00"),
                    ("08:00:00", "08:30:00"),
                ),
                start=1,
            ):
                demand.append(
                    [
                        period_id,
                        start_date,
                        end_date,
                        days,
                        block_start,
                        block_end,
                        direction,
                        factor * block_index * 10,
                        "average_day",
                        "actual_departure_time",
                        dataset,
                    ]
                )

    profiles = workbook.create_sheet("DEMAND_PROFILE_CONFIG")
    profiles.append(
        [
            "profile_id",
            "included_period_ids",
            "aggregation_method",
            "period_weight",
            "authority_role",
            "status",
            "description",
        ]
    )
    profiles.append(
        ["stable", "p1,p2", "day_weighted_mean", "observation_days", "PRIMARY", "READY", ""]
    )
    profiles.append(
        ["current", "p2", "single_period", "observation_days", "SENSITIVITY", "READY", ""]
    )

    metadata = workbook.create_sheet("THONG_TIN_DU_LIEU")
    metadata.append(["DATA AUTHORITY", None, None, None])
    metadata.append(["Tham số", "Giá trị", "Mức độ", "Diễn giải"])
    for key, value in (
        ("demand_dataset_id", "synthetic-multi"),
        ("demand_source_type", "ticketing"),
        ("demand_confidence", "high"),
        ("demand_response_mode", "static"),
        ("default_demand_profile", "stable"),
        ("sensitivity_profiles", "current"),
    ):
        metadata.append([key, value, "REQUIRED", ""])

    guide = workbook.create_sheet("HUONG_DAN", 0)
    guide["A1"] = "Synthetic V3 fixture"
    workbook.save(path)
    return path


@pytest.fixture
def v3_workbook(tmp_path: Path) -> Path:
    return _write_v3_workbook(tmp_path / "synthetic_v3.xlsx")


def _comparison_run(
    profile_id: str,
    *,
    comparable_c: bool,
    maximum_service_gap: int = 10,
    maximum_shift: int = 4,
    total_shift: int = 12,
    final_acceptance_state: str = "FINAL_RECOMMENDED",
):
    quality_c = (
        {"maximum_positive_demand_service_gap_minutes": maximum_service_gap}
        if comparable_c
        else None
    )
    regimes = (
        [
            {
                "direction": "outbound",
                "regime_id": "R1",
                "start": "06:00:00",
                "end": "08:00:00",
                "uniform_headway_minutes": 30,
            }
        ]
        if comparable_c
        else []
    )
    selected = (
        {
            "allocation_by_demand_interval": [
                {
                    "block_id": "B1",
                    "direction": "outbound",
                    "start_time": "06:00:00",
                    "end_time": "08:00:00",
                    "trip_count": 3,
                }
            ]
        }
        if comparable_c
        else None
    )
    solution = SimpleNamespace() if comparable_c else None
    return SimpleNamespace(
        result=SimpleNamespace(
            candidate_outcome=(SimpleNamespace(solution=solution) if comparable_c else None)
        ),
        derivation=SimpleNamespace(
            profile=SimpleNamespace(
                profile_id=profile_id,
                profile_fingerprint=f"fingerprint-{profile_id}",
                derived_observations=(SimpleNamespace(),),
            )
        ),
        normalized_inputs=SimpleNamespace(scenario_b=SimpleNamespace(total_daily_trips=6)),
        payload={
            "route_id": "SYNTH",
            "route_name": "Synthetic route",
            "scenario_c_available": comparable_c,
            "final_acceptance_state": final_acceptance_state,
            "aggregate_native_status": "OPTIMAL" if comparable_c else "UNKNOWN",
            "stage_1": {"selected_allocation_plan": selected},
            "final_service_regimes": regimes,
            "quality": {
                "B": {"maximum_positive_demand_service_gap_minutes": 20},
                "C": quality_c,
            },
            "fleet": {
                "available": 4,
                "scenario_b_required": 3,
                "scenario_c_required": 3 if comparable_c else None,
            },
            "shift_metrics": {
                "shifted_trip_count": 2 if comparable_c else None,
                "total_shift_minutes": total_shift if comparable_c else None,
                "maximum_shift_minutes": maximum_shift if comparable_c else None,
            },
        },
    )


def test_v3_reader_loads_default_and_multiple_periods(v3_workbook: Path) -> None:
    imported = import_v3_multi_period_workbook_v1(v3_workbook)

    assert imported.base_workbook.parameters_b.route_id == "SYNTH"
    assert imported.multi_period_demand.default_profile_id == "stable"
    assert [item.period_id for item in imported.multi_period_demand.periods] == ["p1", "p2"]


def test_v3_reader_rejects_period_row_catalog_mismatch(v3_workbook: Path) -> None:
    workbook = load_workbook(v3_workbook)
    workbook["SAN_LUONG_MULTI_PERIOD"]["D2"] = 3
    workbook.save(v3_workbook)

    with pytest.raises(MultiPeriodDemandError) as exc_info:
        import_v3_multi_period_workbook_v1(v3_workbook)

    assert exc_info.value.code == "PERIOD_ROW_OBSERVATION_DAYS_MISMATCH"


def test_v3_reader_rejects_missing_observation_days(v3_workbook: Path) -> None:
    workbook = load_workbook(v3_workbook)
    workbook["PERIOD_CATALOG"]["D2"] = None
    workbook.save(v3_workbook)

    with pytest.raises(MultiPeriodDemandError) as exc_info:
        import_v3_multi_period_workbook_v1(v3_workbook)

    assert exc_info.value.code == "OBSERVATION_DAYS_MISSING"


@pytest.mark.parametrize("value", [31.5, 0, -1, True, float("nan"), float("inf"), "31.5"])
def test_v3_reader_rejects_malformed_observation_days(
    v3_workbook: Path,
    value: object,
) -> None:
    workbook = load_workbook(v3_workbook)
    workbook["PERIOD_CATALOG"]["D2"] = value
    workbook.save(v3_workbook)

    with pytest.raises(MultiPeriodDemandError):
        import_v3_multi_period_workbook_v1(v3_workbook)


@pytest.mark.parametrize("value", [-1, True, float("nan"), float("inf"), -float("inf"), "bad"])
def test_v3_reader_rejects_malformed_passenger_volume(
    v3_workbook: Path,
    value: object,
) -> None:
    workbook = load_workbook(v3_workbook)
    workbook["SAN_LUONG_MULTI_PERIOD"]["H2"] = value
    workbook.save(v3_workbook)

    with pytest.raises(MultiPeriodDemandError) as exc_info:
        import_v3_multi_period_workbook_v1(v3_workbook)

    assert exc_info.value.code == "PASSENGER_VOLUME_INVALID"


@pytest.mark.parametrize("value", [31, 31.0, "31"])
def test_v3_reader_accepts_integral_spreadsheet_observation_days(
    v3_workbook: Path,
    value: object,
) -> None:
    workbook = load_workbook(v3_workbook)
    workbook["PERIOD_CATALOG"]["D2"] = value
    for row in range(2, 8):
        workbook["SAN_LUONG_MULTI_PERIOD"].cell(row, 4, value)
    workbook.save(v3_workbook)

    imported = import_v3_multi_period_workbook_v1(v3_workbook)

    assert imported.multi_period_demand.periods[0].observation_days == 31


def test_default_profile_cli_creates_deterministic_json_and_required_xlsx(
    v3_workbook: Path,
    tmp_path: Path,
) -> None:
    output = tmp_path / "default"

    assert (
        main(
            [
                "--input",
                str(v3_workbook),
                "--output-dir",
                str(output),
                "--solve-budget-seconds",
                "0.001",
            ]
        )
        == 0
    )

    payload = json.loads((output / "result.json").read_text(encoding="utf-8"))
    assert payload["selected_profile"]["profile_id"] == "stable"
    assert payload["scenario_c_available"] is False
    assert payload["quality"]["C"] is None
    assert payload["fleet"]["scenario_b_required"] is not None
    assert payload["fleet"]["scenario_c_required"] is None
    assert payload["shift_metrics"] == {
        "maximum_shift_minutes": None,
        "shifted_trip_count": None,
        "total_shift_minutes": None,
    }
    result_workbook = load_workbook(output / "result.xlsx", read_only=False)
    assert result_workbook.sheetnames == [
        "SUMMARY",
        "DEMAND_PROFILE",
        "STAGE1_ALLOCATION",
        "REGIMES",
        "TIMETABLE_B",
        "TIMETABLE_C",
        "DIAGNOSTICS",
    ]
    assert len(result_workbook["DEMAND_PROFILE"]._charts) == 2
    demand_chart, service_chart = result_workbook["DEMAND_PROFILE"]._charts
    assert len(demand_chart.series) == 1
    assert len(service_chart.series) == 1
    with zipfile.ZipFile(output / "result.xlsx") as archive:
        chart_xml = b"".join(
            archive.read(name) for name in archive.namelist() if name.startswith("xl/charts/chart")
        )
    assert b"Passengers per block" in chart_xml
    assert b"Trips per block" in chart_xml
    assert b"scenario_c_service_count" not in chart_xml
    assert all(
        result_workbook["DEMAND_PROFILE"].cell(row, 6).value is None
        for row in range(4, result_workbook["DEMAND_PROFILE"].max_row + 1)
    )
    first = write_deterministic_json(tmp_path / "one.json", payload).read_bytes()
    second = write_deterministic_json(tmp_path / "two.json", payload).read_bytes()
    assert first == second


def test_explicit_profile_and_batch_profile_outputs(
    v3_workbook: Path,
    tmp_path: Path,
) -> None:
    single = tmp_path / "single"
    assert (
        main(
            [
                "--input",
                str(v3_workbook),
                "--profile",
                "current",
                "--output-dir",
                str(single),
                "--solve-budget-seconds",
                "0.001",
            ]
        )
        == 0
    )
    assert (
        json.loads((single / "result.json").read_text(encoding="utf-8"))["selected_profile"][
            "profile_id"
        ]
        == "current"
    )

    batch = tmp_path / "batch"
    assert (
        main(
            [
                "--input",
                str(v3_workbook),
                "--profiles",
                "stable,current",
                "--output-dir",
                str(batch),
                "--solve-budget-seconds",
                "0.001",
            ]
        )
        == 0
    )
    for profile_id in ("stable", "current"):
        assert (batch / profile_id / "result.json").is_file()
        assert (batch / profile_id / "result.xlsx").is_file()
    assert (batch / "profile_comparison.json").is_file()
    assert (batch / "profile_comparison.xlsx").is_file()
    comparison = json.loads((batch / "profile_comparison.json").read_text(encoding="utf-8"))
    assert comparison["comparison_eligibility"] == PROFILE_COMPARISON_INCONCLUSIVE
    assert comparison["stability_classification"] is None
    assert comparison["review_code"] == PROFILE_COMPARISON_INCONCLUSIVE_NO_COMPARABLE_C
    assert all(item["maximum_service_gap"] is None for item in comparison["profiles"])
    comparison_workbook = load_workbook(batch / "profile_comparison.xlsx", data_only=False)
    summary_values = {
        cell.value for row in comparison_workbook["PROFILE_COMPARISON"] for cell in row
    }
    assert PROFILE_COMPARISON_INCONCLUSIVE_NO_COMPARABLE_C in summary_values
    assert not any(
        isinstance(cell.value, str) and cell.value.startswith("#")
        for sheet in comparison_workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )
    assert not any(item.name == v3_workbook.name for item in batch.rglob("*"))


def test_profile_comparison_all_profiles_without_c_is_inconclusive() -> None:
    comparison = build_profile_comparison_v1(
        [
            _comparison_run("stable", comparable_c=False),
            _comparison_run("current", comparable_c=False),
        ]
    )

    assert comparison["comparison_eligibility"] == PROFILE_COMPARISON_INCONCLUSIVE
    assert comparison["stability_classification"] is None
    assert comparison["review_code"] == PROFILE_COMPARISON_INCONCLUSIVE_NO_COMPARABLE_C
    assert all(
        row["maximum_service_gap"] is None
        and row["fleet_required_c"] is None
        and row["shifted_trips"] is None
        and row["total_shift"] is None
        and row["maximum_shift"] is None
        for row in comparison["profiles"]
    )


def test_profile_comparison_mixed_c_availability_is_material() -> None:
    comparison = build_profile_comparison_v1(
        [
            _comparison_run("stable", comparable_c=True),
            _comparison_run("current", comparable_c=False),
        ]
    )

    assert comparison["comparison_eligibility"] == PROFILE_COMPARISON_PARTIALLY_COMPARABLE
    assert comparison["stability_classification"] == MATERIAL_PROFILE_SENSITIVITY


def test_profile_comparison_all_comparable_and_identical_is_stable() -> None:
    comparison = build_profile_comparison_v1(
        [
            _comparison_run("stable", comparable_c=True),
            _comparison_run("current", comparable_c=True),
        ]
    )

    assert comparison["comparison_eligibility"] == PROFILE_COMPARISON_ELIGIBLE
    assert comparison["stability_classification"] == STABLE_ACROSS_PROFILES
    assert _comparison_row(_comparison_run("stable", comparable_c=True))["fleet_required_c"] == 3


def test_profile_comparison_small_c_difference_is_minor() -> None:
    comparison = build_profile_comparison_v1(
        [
            _comparison_run("stable", comparable_c=True, maximum_service_gap=10),
            _comparison_run("current", comparable_c=True, maximum_service_gap=11),
        ]
    )

    assert comparison["stability_classification"] == MINOR_PROFILE_SENSITIVITY


def test_profile_comparison_material_c_difference_is_material() -> None:
    comparison = build_profile_comparison_v1(
        [
            _comparison_run("stable", comparable_c=True, maximum_service_gap=10),
            _comparison_run("current", comparable_c=True, maximum_service_gap=16),
        ]
    )

    assert comparison["stability_classification"] == MATERIAL_PROFILE_SENSITIVITY


def test_identical_workbook_content_is_path_and_mtime_independent(
    v3_workbook: Path,
    tmp_path: Path,
) -> None:
    copied = tmp_path / "renamed-identical-input.xlsx"
    shutil.copyfile(v3_workbook, copied)
    os.utime(copied, (1_700_000_000, 1_700_000_000))
    os.utime(v3_workbook, (1_800_000_000, 1_800_000_000))

    first = run_v3_profile_v1(v3_workbook, "stable", total_budget_seconds=0.001)
    second = run_v3_profile_v1(copied, "stable", total_budget_seconds=0.001)
    first_context, _ = build_two_stage_uniform_request_v1(
        first.normalized_inputs,
        first.b_evaluation,
        solver_policy=SolverPolicyV1(time_limit_seconds=0.001),
        demand_profile_fingerprint=first.derivation.profile.profile_fingerprint,
    )
    second_context, _ = build_two_stage_uniform_request_v1(
        second.normalized_inputs,
        second.b_evaluation,
        solver_policy=SolverPolicyV1(time_limit_seconds=0.001),
        demand_profile_fingerprint=second.derivation.profile.profile_fingerprint,
    )

    assert first.normalized_inputs.scenario_b_fingerprint == (
        second.normalized_inputs.scenario_b_fingerprint
    )
    assert first.derivation.profile.profile_fingerprint == (
        second.derivation.profile.profile_fingerprint
    )
    assert first.normalized_inputs.observed_demand_fingerprint == (
        second.normalized_inputs.observed_demand_fingerprint
    )
    assert first_context.problem.adapter_context_fingerprint == (
        second_context.problem.adapter_context_fingerprint
    )
    assert first_context.problem_fingerprint == second_context.problem_fingerprint
