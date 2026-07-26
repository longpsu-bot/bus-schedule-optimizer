"""Test-only constructors for the anonymized route corpus."""

from __future__ import annotations

import hashlib
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from bus_schedule_engine.contracts_v1 import (
    DemandConfidence,
    NormalizationOptions,
    OperatingDayType,
    normalize_imported_workbook_v1,
)
from bus_schedule_engine.importer import ImportedWorkbook
from bus_schedule_engine.models import (
    DemandRecord,
    Direction,
    RouteType,
    ScenarioParameters,
    Trip,
    VolumeType,
)

CORPUS_DIR = Path(__file__).parent / "fixtures" / "route_corpus" / "v1"
FIXTURE_FILES = ("corpus_alpha_80.json", "corpus_beta_46.json")


def load_corpus_fixture(filename: str) -> dict[str, Any]:
    if filename not in FIXTURE_FILES:
        raise ValueError(f"Unknown route corpus fixture: {filename}")
    return json.loads((CORPUS_DIR / filename).read_text(encoding="utf-8"))


def load_manifest() -> dict[str, Any]:
    return json.loads((CORPUS_DIR / "manifest.json").read_text(encoding="utf-8"))


def raw_trip_observations(fixture: dict[str, Any]) -> list[dict[str, Any]]:
    return fixture["demand_observations"]["raw_trip_observations"]["rows"]


def proxy_demand_blocks(fixture: dict[str, Any]) -> list[dict[str, Any]]:
    return fixture["demand_observations"]["departure_hour_proxy_v1"]["blocks"]


def _seconds(value: str) -> int:
    hour, minute = (int(part) for part in value.split(":"))
    return hour * 3600 + minute * 60


def _parameters(payload: dict[str, Any]) -> ScenarioParameters:
    return ScenarioParameters(
        route_id=payload["route_id"],
        route_name=payload["route_name"],
        route_type=RouteType(payload["route_type"]),
        trip_runtime_minutes=int(payload["trip_runtime_minutes"]),
        total_daily_trips=int(payload["total_daily_trips"]),
        terminal_1_name=payload["terminal_1_name"],
        terminal_1_first_departure=_seconds(payload["terminal_1_first_departure"]),
        terminal_1_last_departure=_seconds(payload["terminal_1_last_departure"]),
        terminal_2_name=payload["terminal_2_name"],
        terminal_2_first_departure=_seconds(payload["terminal_2_first_departure"]),
        terminal_2_last_departure=_seconds(payload["terminal_2_last_departure"]),
        vehicle_capacity_passengers=int(payload["vehicle_capacity_passengers"]),
        target_load_factor=float(payload["target_load_factor"]),
        maximum_load_factor=float(payload["maximum_load_factor"]),
        time_block_minutes=int(payload["time_block_minutes"]),
        minimum_layover_minutes=int(payload["minimum_layover_minutes"]),
        allowed_trip_runtime_minutes=tuple(
            int(value) for value in payload["allowed_trip_runtime_minutes"]
        ),
    )


def _trips(payload: list[dict[str, Any]]) -> list[Trip]:
    return [
        Trip(
            scenario=item["scenario"],
            trip_id=item["trip_id"],
            departure_terminal=item["departure_terminal"],
            direction=Direction(item["direction"]),
            departure_seconds=_seconds(item["departure_time"]),
            arrival_seconds=_seconds(item["arrival_time"]),
        )
        for item in payload
    ]


def _proxy_demand(payload: list[dict[str, Any]]) -> list[DemandRecord]:
    return [
        DemandRecord(
            period_start=date.fromisoformat(item["period_start"]),
            period_end=date.fromisoformat(item["period_end"]),
            observation_days=int(item["observation_days"]),
            block_start_seconds=_seconds(item["time_block_start"]),
            block_end_seconds=_seconds(item["time_block_end"]),
            direction=Direction(item["direction"]),
            passenger_volume=float(item["passenger_volume"]),
            volume_type=VolumeType(item["volume_type"]),
        )
        for item in payload
    ]


def imported_workbook_from_fixture(fixture: dict[str, Any]) -> ImportedWorkbook:
    """Construct the importer model with the proxy, never the overlapping raw rows."""
    return ImportedWorkbook(
        parameters_a=_parameters(fixture["scenario_a"]["parameters"]),
        trips_a=_trips(fixture["scenario_a"]["exact_trips"]),
        parameters_b=_parameters(fixture["scenario_b"]["parameters"]),
        trips_b=_trips(fixture["scenario_b"]["exact_trips"]),
        demand=_proxy_demand(proxy_demand_blocks(fixture)),
        configuration=dict(fixture["configuration"]),
    )


def normalization_options_from_fixture(
    fixture: dict[str, Any],
) -> NormalizationOptions:
    payload = fixture["normalization_options"]
    return NormalizationOptions(
        source_id=payload["source_id"],
        imported_at=datetime.fromisoformat(payload["imported_at"]),
        operating_day_type_a=OperatingDayType[payload["operating_day_type_a"]],
        operating_day_type_b=OperatingDayType[payload["operating_day_type_b"]],
        available_fleet_limit_a=int(payload["available_fleet_limit_a"]),
        available_fleet_limit_b=int(payload["available_fleet_limit_b"]),
        demand_dataset_id=payload["demand_dataset_id"],
        demand_confidence=DemandConfidence[payload["demand_confidence"]],
        source_notes=payload["source_notes"],
    )


def normalized_bundle_from_fixture(fixture: dict[str, Any]):
    return normalize_imported_workbook_v1(
        imported_workbook_from_fixture(fixture),
        normalization_options_from_fixture(fixture),
    )


def fact_fingerprint(fixture: dict[str, Any]) -> str:
    facts = {
        "demand_observations": fixture["demand_observations"],
        "scenario_a": fixture["scenario_a"],
        "scenario_b": fixture["scenario_b"],
    }
    canonical = json.dumps(
        facts,
        ensure_ascii=False,
        separators=(",", ":"),
        sort_keys=True,
    ).encode()
    return hashlib.sha256(canonical).hexdigest()


def _write_table(
    sheet: Worksheet,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
) -> None:
    sheet["A1"] = title
    sheet.append([])
    sheet.append(headers)
    for row in rows:
        sheet.append(row)


def _parameter_rows(parameters: dict[str, Any]) -> list[list[Any]]:
    keys = (
        "route_id",
        "route_name",
        "route_type",
        "trip_runtime_minutes",
        "allowed_trip_runtime_minutes",
        "total_daily_trips",
        "terminal_1_name",
        "terminal_1_first_departure",
        "terminal_1_last_departure",
        "terminal_2_name",
        "terminal_2_first_departure",
        "terminal_2_last_departure",
        "vehicle_capacity_passengers",
        "target_load_factor",
        "maximum_load_factor",
        "time_block_minutes",
        "minimum_layover_minutes",
    )
    rows: list[list[Any]] = []
    for key in keys:
        value = parameters[key]
        if key == "allowed_trip_runtime_minutes":
            options = list(value)
            value = str(options[0]) if len(options) == 1 else f"{min(options)},{max(options)}"
        rows.append([key, value, None])
    return rows


def render_sanitized_xlsx(
    fixture: dict[str, Any],
    destination: Path,
) -> Path:
    """Render a test-only importer workbook whose SAN_LUONG sheet is the proxy."""
    workbook = Workbook()
    guide = workbook.active
    guide.title = "HUONG_DAN"
    _write_table(
        guide,
        "Anonymized route corpus",
        ["field", "value", "notes"],
        [["fixture_id", fixture["fixture_id"], "Test-only sanitized workbook"]],
    )
    for scenario in ("A", "B"):
        payload = fixture[f"scenario_{scenario.lower()}"]
        parameter_sheet = workbook.create_sheet(f"THONG_SO_{scenario}")
        _write_table(
            parameter_sheet,
            f"Scenario {scenario} parameters",
            ["field", "value", "notes"],
            _parameter_rows(payload["parameters"]),
        )
        timetable_sheet = workbook.create_sheet(f"BIEU_DO_{scenario}")
        _write_table(
            timetable_sheet,
            f"Scenario {scenario} exact timetable",
            [
                "trip_id",
                "departure_terminal",
                "direction",
                "departure_time",
                "arrival_time",
                "vehicle_id",
                "vehicle_capacity_override",
            ],
            [
                [
                    trip["trip_id"],
                    trip["departure_terminal"],
                    trip["direction"],
                    trip["departure_time"],
                    trip["arrival_time"],
                    None,
                    None,
                ]
                for trip in payload["exact_trips"]
            ],
        )
    demand_sheet = workbook.create_sheet("SAN_LUONG")
    _write_table(
        demand_sheet,
        "LOW-confidence departure_hour_proxy_v1",
        [
            "period_start",
            "period_end",
            "observation_days",
            "time_block_start",
            "time_block_end",
            "direction",
            "passenger_volume",
            "volume_type",
        ],
        [
            [
                date.fromisoformat(block["period_start"]),
                date.fromisoformat(block["period_end"]),
                block["observation_days"],
                block["time_block_start"],
                block["time_block_end"],
                block["direction"],
                block["passenger_volume"],
                block["volume_type"],
            ]
            for block in proxy_demand_blocks(fixture)
        ],
    )
    configuration = workbook.create_sheet("CAU_HINH")
    _write_table(
        configuration,
        "Sanitized configuration",
        ["field", "value", "notes"],
        [[key, value, None] for key, value in sorted(fixture["configuration"].items())],
    )
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination
