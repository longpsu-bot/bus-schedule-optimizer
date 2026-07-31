from __future__ import annotations

from dataclasses import replace
from datetime import UTC, date, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook
from presentation_support import build_result_and_report, rejected_result_and_report
from streamlit.testing.v1 import AppTest

from bus_schedule_engine.application_pipeline import (
    CONTRACT_V1_ARTIFACT_FAILED,
    CONTRACT_V1_SOLVER_FAILED,
    WORKBOOK_IMPORT_INVALID,
    WORKBOOK_OPTIMIZATION_NOT_READY,
    UnifiedApplicationStatusV1,
    UnifiedRuntimeFailureV1,
    run_unified_application_pipeline_v1,
)
from bus_schedule_engine.excel_exporter import create_input_template
from bus_schedule_engine.importer import import_workbook
from bus_schedule_engine.input_authority import (
    AVAILABLE_FLEET_LIMIT_REQUIRED_FOR_OPTIMIZATION,
    WorkbookInputReadinessV1,
)
from bus_schedule_engine.unified_diagram import (
    build_unified_demand_supply_figure_v1,
    build_unified_departure_figure_v1,
)
from bus_schedule_engine.unified_presentation import (
    build_unified_application_presentation_v1,
)
from bus_schedule_engine.unified_result_exporter import (
    export_unified_result_workbook_v1,
)


def _readiness(*, ready: bool = True, missing: tuple[str, ...] = ()):
    return WorkbookInputReadinessV1(
        import_ready=True,
        optimization_ready=ready,
        blocking_import_codes=(),
        missing_optimization_authority_codes=missing,
        optional_limitations=(),
    )


def _failure(
    *,
    code: str,
    presentation=None,
) -> UnifiedRuntimeFailureV1:
    return UnifiedRuntimeFailureV1(
        code=code,
        stage="HEURISTIC_SOLVER" if code == CONTRACT_V1_SOLVER_FAILED else "ARTIFACT_CONSTRUCTION",
        correlation_id="m5c2-0123456789abcdef0123",
        sanitized_message="Synthetic bounded failure.",
        retryable=True,
        solver_choice="HEURISTIC",
        source_id=presentation.source_id if presentation is not None else "fixture-source",
        presentation_fingerprint=(
            presentation.presentation_fingerprint if presentation is not None else None
        ),
        b_fingerprint=(presentation.source_b_fingerprint if presentation is not None else None),
        accepted_solution_fingerprint=(
            presentation.accepted_solution_fingerprint if presentation is not None else None
        ),
    )


def _complete_state(tmp_path: Path, pair=None) -> dict[str, object]:
    result, _report = pair or build_result_and_report()
    presentation = build_unified_application_presentation_v1(result)
    xlsx_path = export_unified_result_workbook_v1(
        presentation,
        tmp_path / "contract-result.xlsx",
    )
    return {
        "unified_runtime_status": UnifiedApplicationStatusV1.COMPLETE,
        "workbook_input_readiness": _readiness(),
        "unified_optimization_result": result,
        "unified_presentation": presentation,
        "unified_demand_supply_figure": build_unified_demand_supply_figure_v1(presentation),
        "unified_departure_figure": build_unified_departure_figure_v1(presentation),
        "unified_download_artifacts": {
            "xlsx": xlsx_path.read_bytes(),
            "source_id": presentation.source_id,
            "presentation_fingerprint": presentation.presentation_fingerprint,
            "b_fingerprint": presentation.source_b_fingerprint,
            "accepted_solution_fingerprint": (presentation.accepted_solution_fingerprint),
        },
        "unified_runtime_failure": None,
    }


def _seed_page(app: AppTest, state: dict[str, object]) -> AppTest:
    for key, value in state.items():
        app.session_state[key] = value
    return app


def _trip_ridership_workbook(tmp_path: Path) -> Path:
    path = create_input_template(tmp_path / "trip-ridership-ui.xlsx")
    workbook = load_workbook(path)
    values = [
        "UI-OBS-001",
        date(2026, 7, 1),
        "UI-SOURCE-001",
        "B-001",
        "outbound",
        None,
        "06:03",
        42,
        None,
        "UI diagnostic sample",
    ]
    for column, value in enumerate(values, 1):
        workbook["SAN_LUONG_CHUYEN"].cell(4, column).value = value
    workbook.save(path)
    workbook.close()
    return path


def _trip_ridership_complete_state(tmp_path: Path) -> dict[str, object]:
    imported = import_workbook(_trip_ridership_workbook(tmp_path))
    run = run_unified_application_pipeline_v1(
        imported,
        source_id="trip-ridership-ui",
        imported_at=datetime(2026, 7, 31, tzinfo=UTC),
    )
    assert run.status == UnifiedApplicationStatusV1.COMPLETE
    assert run.unified_presentation is not None
    return {
        "unified_runtime_status": run.status,
        "workbook_input_readiness": run.input_readiness,
        "unified_optimization_result": run.unified_result,
        "unified_presentation": run.unified_presentation,
        "unified_demand_supply_figure": run.unified_demand_supply_figure,
        "unified_departure_figure": run.unified_departure_figure,
        "unified_download_artifacts": {
            "xlsx": run.unified_xlsx_bytes,
            "source_id": run.source_id,
            "presentation_fingerprint": run.unified_presentation.presentation_fingerprint,
            "b_fingerprint": run.unified_presentation.source_b_fingerprint,
            "accepted_solution_fingerprint": (
                run.unified_presentation.accepted_solution_fingerprint
            ),
        },
        "unified_runtime_failure": run.failure,
        "imported_workbook": imported,
        "trip_ridership_analysis": run.trip_ridership_analysis,
        "trip_ridership_failure": run.trip_ridership_failure,
        "protected_service_floor_assessment": (run.protected_service_floor_assessment),
        "protected_service_floor_failure": run.protected_service_floor_failure,
    }


def _assert_missing_state_key(app: AppTest, key: str) -> None:
    with pytest.raises(KeyError):
        _ = app.session_state[key]


def test_input_page_without_workbook_shows_starting_instruction() -> None:
    app = AppTest.from_file("app_pages/01_nhap_du_lieu.py")

    app.run(timeout=30)

    assert not app.exception
    assert app.info
    assert len(app.download_button) == 1
    assert app.download_button[0].label == "Tải template có dữ liệu minh họa"


def test_input_page_runs_unified_only_with_default_solver(tmp_path: Path) -> None:
    content = create_input_template(tmp_path / "input.xlsx").read_bytes()
    app = AppTest.from_file("app_pages/01_nhap_du_lieu.py")
    app.session_state["input_bytes"] = content

    app.run(timeout=30)
    runtime_input = next(
        item
        for item in app.text_input
        if item.label == "Khoảng thời gian hành trình cho phép (phút)"
    )
    runtime_input.set_value("45,65")
    submit = next(item for item in app.button if item.label == "Chạy kiểm tra và sinh phương án")
    submit.click().run(timeout=60)

    assert not app.exception
    assert app.session_state["imported_workbook"].parameters_b.runtime_options == (45, 65)
    assert app.session_state["unified_runtime_status"] == UnifiedApplicationStatusV1.COMPLETE
    assert app.session_state["unified_optimization_result"].solver_choice.value == "HEURISTIC"
    assert app.session_state["unified_presentation"] is not None
    assert app.session_state["unified_download_artifacts"]["xlsx"]
    for legacy_key in (
        "analysis_bundle",
        "diagram_figure",
        "download_artifacts",
        "scenario_c_fingerprint",
        "parallel_runtime_status",
        "side_by_side_validation_report",
    ):
        _assert_missing_state_key(app, legacy_key)


def test_page01_shows_trip_record_count_without_matching_before_submit(
    tmp_path: Path,
) -> None:
    content = _trip_ridership_workbook(tmp_path).read_bytes()
    app = AppTest.from_file("app_pages/01_nhap_du_lieu.py")
    app.session_state["input_bytes"] = content

    app.run(timeout=30)

    assert not app.exception
    metric = next(item for item in app.metric if item.label == "Quan sát sản lượng theo chuyến")
    assert metric.value == "1"
    markdown = "\n".join(item.value for item in app.markdown)
    assert "Thiết lập kế hoạch sàn dịch vụ bảo vệ 6A2A" in markdown
    assert any(
        "Regime chưa được phân loại trước khi gửi biểu mẫu" in item.value for item in app.caption
    )
    _assert_missing_state_key(app, "trip_ridership_analysis")
    _assert_missing_state_key(app, "trip_ridership_failure")
    _assert_missing_state_key(app, "protected_service_floor_assessment")
    _assert_missing_state_key(app, "protected_service_floor_failure")


def test_new_upload_removes_stale_legacy_and_unified_results(tmp_path: Path) -> None:
    first = create_input_template(tmp_path / "first.xlsx").read_bytes()
    second_path = create_input_template(tmp_path / "second.xlsx")
    workbook = load_workbook(second_path)
    sheet = workbook["THONG_SO_B"]
    row = next(cell.row for cell in sheet["A"] if cell.value == "route_name")
    sheet.cell(row, 2).value = "Replacement route"
    workbook.save(second_path)
    workbook.close()
    second = second_path.read_bytes()
    app = AppTest.from_file("app_pages/01_nhap_du_lieu.py")
    app.session_state["input_bytes"] = first
    for key in (
        "analysis_bundle",
        "diagram_figure",
        "download_artifacts",
        "scenario_c_fingerprint",
        "parallel_runtime_status",
        "side_by_side_validation_report",
    ):
        app.session_state[key] = object()
    for key in (
        "imported_workbook",
        "workbook_input_readiness",
        "unified_optimization_result",
        "unified_presentation",
        "unified_demand_supply_figure",
        "unified_departure_figure",
        "unified_download_artifacts",
        "unified_runtime_failure",
        "unified_runtime_status",
        "trip_ridership_analysis",
        "trip_ridership_failure",
        "protected_service_floor_assessment",
        "protected_service_floor_failure",
    ):
        app.session_state[key] = object()

    app.run(timeout=30)
    app.file_uploader[0].upload(
        "replacement.xlsx",
        second,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ).run(timeout=30)

    assert not app.exception
    for legacy_key in (
        "analysis_bundle",
        "diagram_figure",
        "download_artifacts",
        "scenario_c_fingerprint",
        "parallel_runtime_status",
        "side_by_side_validation_report",
    ):
        _assert_missing_state_key(app, legacy_key)
    assert app.session_state["unified_runtime_status"] is None
    assert app.session_state["unified_presentation"] is None
    assert app.session_state["trip_ridership_analysis"] is None
    assert app.session_state["trip_ridership_failure"] is None
    assert app.session_state["protected_service_floor_assessment"] is None
    assert app.session_state["protected_service_floor_failure"] is None


def test_import_invalid_is_stable_and_only_template_is_downloadable() -> None:
    app = AppTest.from_file("app_pages/01_nhap_du_lieu.py")
    app.session_state["input_bytes"] = b"not a workbook secret-row"

    app.run(timeout=30)

    assert not app.exception
    assert WORKBOOK_IMPORT_INVALID in app.error[0].value
    assert "secret-row" not in app.error[0].value
    assert [item.label for item in app.download_button] == ["Tải template có dữ liệu minh họa"]


def test_page01_not_ready_stores_only_readiness_and_exact_codes(
    tmp_path: Path,
) -> None:
    workbook_path = create_input_template(tmp_path / "not-ready.xlsx")
    workbook = load_workbook(workbook_path)
    sheet = workbook["THONG_SO_B"]
    row = next(cell.row for cell in sheet["A"] if cell.value == "available_fleet_limit")
    sheet.cell(row, 2).value = None
    workbook.save(workbook_path)
    workbook.close()
    app = AppTest.from_file("app_pages/01_nhap_du_lieu.py")
    app.session_state["input_bytes"] = workbook_path.read_bytes()

    app.run(timeout=30)
    submit = next(item for item in app.button if item.label == "Chạy kiểm tra và sinh phương án")
    submit.click().run(timeout=30)

    assert not app.exception
    assert app.session_state["unified_runtime_status"] == UnifiedApplicationStatusV1.INPUT_NOT_READY
    assert app.session_state["unified_optimization_result"] is None
    assert app.session_state["unified_presentation"] is None
    assert app.session_state["unified_download_artifacts"] is None
    warning = next(item for item in app.warning if WORKBOOK_OPTIMIZATION_NOT_READY in item.value)
    assert AVAILABLE_FLEET_LIMIT_REQUIRED_FOR_OPTIMIZATION in warning.value


@pytest.mark.parametrize(
    "page",
    (
        "app_pages/02_kiem_tra.py",
        "app_pages/03_nhu_cau.py",
        "app_pages/04_khuyen_nghi.py",
    ),
)
def test_pages_02_to_04_render_only_verified_unified_facts(
    page: str,
    tmp_path: Path,
) -> None:
    state = _complete_state(tmp_path)
    app = _seed_page(AppTest.from_file(page), state)

    app.run(timeout=30)

    assert not app.exception
    assert app.info
    assert app.info[0].value.startswith("Nguồn kết quả hiển thị: Contract V1.")
    assert app.dataframe
    assert all("pipeline legacy" not in item.value for item in app.warning)


def test_page03_renders_supplemental_match_quality_and_trip_summaries(
    tmp_path: Path,
) -> None:
    state = _trip_ridership_complete_state(tmp_path)
    app = _seed_page(AppTest.from_file("app_pages/03_nhu_cau.py"), state)

    app.run(timeout=30)

    assert not app.exception
    assert any(item.value == "Sản lượng theo từng chuyến" for item in app.subheader)
    assert any("chưa được sử dụng để sinh phương án C" in item.value for item in app.warning)
    labels = {item.label for item in app.metric}
    assert {
        "Số ngày quan sát",
        "Tỷ lệ ghép dùng được",
        "Tỷ lệ ghép chính xác",
        "Bao phủ chuyến B",
        "Bao phủ chuyến-ngày",
    }.issubset(labels)
    markdown = "\n".join(item.value for item in app.markdown)
    assert "Thống kê mô tả theo chuyến B" in markdown
    assert "Tổng hợp theo chiều" in markdown
    assert any(item.value == "Đánh giá regime cần bảo vệ" for item in app.subheader)
    assert any(
        "Kết quả 6A2A chỉ xác định regime đề xuất bảo vệ" in item.value for item in app.warning
    )
    assert "Ngưỡng chính sách 6A2A" in markdown
    assert "Preview sàn dịch vụ tương lai — chưa thực thi" in markdown


def test_page03_refuses_stale_supplemental_analysis(tmp_path: Path) -> None:
    state = _trip_ridership_complete_state(tmp_path)
    state["trip_ridership_analysis"] = replace(
        state["trip_ridership_analysis"],
        scenario_b_timetable_fingerprint="0" * 64,
    )
    app = _seed_page(AppTest.from_file("app_pages/03_nhu_cau.py"), state)

    app.run(timeout=30)

    assert not app.exception
    assert any("không khớp workbook hoặc Scenario B hiện tại" in item.value for item in app.warning)
    markdown = "\n".join(item.value for item in app.markdown)
    assert "Thống kê mô tả theo chuyến B" not in markdown


def test_page03_refuses_same_b_analysis_from_different_trip_dataset(
    tmp_path: Path,
) -> None:
    state = _trip_ridership_complete_state(tmp_path)
    imported = state["imported_workbook"]
    changed_observation = replace(
        imported.trip_ridership_observations[0],
        passenger_count=imported.trip_ridership_observations[0].passenger_count + 1,
    )
    state["imported_workbook"] = replace(
        imported,
        trip_ridership_observations=(changed_observation,),
    )
    app = _seed_page(AppTest.from_file("app_pages/03_nhu_cau.py"), state)

    app.run(timeout=30)

    assert not app.exception
    assert any("không khớp workbook hoặc Scenario B hiện tại" in item.value for item in app.warning)
    labels = {item.label for item in app.metric}
    assert "Số ngày quan sát" not in labels
    assert "Tỷ lệ ghép dùng được" not in labels
    markdown = "\n".join(item.value for item in app.markdown)
    assert "Thống kê mô tả theo chuyến B" not in markdown
    assert "Tổng hợp theo chiều" not in markdown
    assert "Bản ghi chẩn đoán bị loại" not in markdown


@pytest.mark.parametrize(
    "page",
    (
        "app_pages/02_kiem_tra.py",
        "app_pages/03_nhu_cau.py",
        "app_pages/04_khuyen_nghi.py",
        "app_pages/05_xuat_file.py",
    ),
)
def test_input_not_ready_pages_show_exact_codes_and_no_result(
    page: str,
) -> None:
    missing = ("AUTHORITY_ONE", "AUTHORITY_TWO")
    state = {
        "unified_runtime_status": UnifiedApplicationStatusV1.INPUT_NOT_READY,
        "workbook_input_readiness": _readiness(ready=False, missing=missing),
        "unified_optimization_result": None,
        "unified_presentation": None,
        "unified_demand_supply_figure": None,
        "unified_departure_figure": None,
        "unified_download_artifacts": None,
        "unified_runtime_failure": None,
    }
    app = _seed_page(AppTest.from_file(page), state)

    app.run(timeout=30)

    assert not app.exception
    assert WORKBOOK_OPTIMIZATION_NOT_READY in app.warning[0].value
    assert all(code in app.warning[0].value for code in missing)
    assert not app.dataframe
    assert not app.download_button


@pytest.mark.parametrize(
    "page",
    (
        "app_pages/02_kiem_tra.py",
        "app_pages/03_nhu_cau.py",
        "app_pages/04_khuyen_nghi.py",
        "app_pages/05_xuat_file.py",
    ),
)
def test_failed_runtime_pages_show_code_stage_and_correlation(page: str) -> None:
    failure = _failure(code=CONTRACT_V1_SOLVER_FAILED)
    state = {
        "unified_runtime_status": UnifiedApplicationStatusV1.FAILED,
        "workbook_input_readiness": _readiness(),
        "unified_optimization_result": None,
        "unified_presentation": None,
        "unified_demand_supply_figure": None,
        "unified_departure_figure": None,
        "unified_download_artifacts": None,
        "unified_runtime_failure": failure,
    }
    app = _seed_page(AppTest.from_file(page), state)

    app.run(timeout=30)

    assert not app.exception
    assert CONTRACT_V1_SOLVER_FAILED in app.error[0].value
    assert "HEURISTIC_SOLVER" in app.error[0].value
    assert failure.correlation_id in app.error[0].value
    assert not app.dataframe
    assert not app.download_button


def test_artifact_failure_keeps_pages_02_to_04_but_disables_page05(
    tmp_path: Path,
) -> None:
    state = _complete_state(tmp_path)
    presentation = state["unified_presentation"]
    state.update(
        {
            "unified_runtime_status": UnifiedApplicationStatusV1.ARTIFACT_FAILED,
            "unified_demand_supply_figure": None,
            "unified_departure_figure": None,
            "unified_download_artifacts": None,
            "unified_runtime_failure": _failure(
                code=CONTRACT_V1_ARTIFACT_FAILED,
                presentation=presentation,
            ),
        }
    )

    page02 = _seed_page(AppTest.from_file("app_pages/02_kiem_tra.py"), state)
    page02.run(timeout=30)
    assert not page02.exception
    assert CONTRACT_V1_ARTIFACT_FAILED in page02.warning[0].value
    assert page02.dataframe

    page05 = _seed_page(AppTest.from_file("app_pages/05_xuat_file.py"), state)
    page05.run(timeout=30)
    assert not page05.exception
    assert CONTRACT_V1_ARTIFACT_FAILED in page05.warning[0].value
    assert not page05.get("plotly_chart")
    assert not page05.download_button


def test_page05_complete_has_two_figures_and_exactly_three_contract_downloads(
    tmp_path: Path,
) -> None:
    state = _complete_state(tmp_path)
    app = _seed_page(AppTest.from_file("app_pages/05_xuat_file.py"), state)

    app.run(timeout=60)

    assert not app.exception
    assert len(app.get("plotly_chart")) == 2
    assert [item.label for item in app.download_button] == [
        "Workbook Contract V1",
        "Báo cáo biểu đồ Contract V1 (.html)",
        "Tổng quan đã chọn (.png)",
    ]
    source = Path("src/bus_schedule_engine/unified_page5_artifacts.py").read_text(encoding="utf-8")
    for filename in (
        "Bus_Schedule_Contract_V1_Result.xlsx",
        "Bus_Schedule_Contract_V1_Charts.html",
        "Bus_Schedule_Contract_V1_Overview.png",
    ):
        assert filename in source


def test_rejected_candidate_pages_never_show_scenario_c_timetable(
    tmp_path: Path,
) -> None:
    state = _complete_state(tmp_path, rejected_result_and_report())
    presentation = state["unified_presentation"]
    assert presentation.outcome.accepted_c_exists is False
    assert presentation.scenario("C") is None
    app = _seed_page(AppTest.from_file("app_pages/04_khuyen_nghi.py"), state)

    app.run(timeout=30)

    assert not app.exception
    combined = "\n".join(
        item.value
        for collection in (app.warning, app.error, app.success, app.markdown)
        for item in collection
        if isinstance(item.value, str)
    )
    assert "không tồn tại phương án c" in combined.lower()
    assert presentation.outcome.validator_rejection_codes
    assert all(code in combined for code in presentation.outcome.validator_rejection_codes)


def test_corrupted_complete_session_fails_closed_before_render(
    tmp_path: Path,
) -> None:
    state = _complete_state(tmp_path)
    presentation = state["unified_presentation"]
    state["unified_presentation"] = replace(
        presentation,
        presentation_fingerprint="0" * 64,
    )
    app = _seed_page(AppTest.from_file("app_pages/05_xuat_file.py"), state)

    app.run(timeout=30)

    assert not app.exception
    assert "CONTRACT_V1_SEMANTIC_INTEGRITY_MISMATCH" in app.error[0].value
    assert not app.get("plotly_chart")
    assert not app.download_button
