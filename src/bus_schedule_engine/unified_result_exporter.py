"""Editable XLSX export for validation-only unified presentations."""

from __future__ import annotations

import json
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .unified_presentation import UnifiedPresentationBundleV1

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
_WHITE_FONT = Font(color="FFFFFF", bold=True)
_BOLD_FONT = Font(bold=True)
_TIME_FORMAT = "[h]:mm"


@dataclass(frozen=True, slots=True)
class UnifiedExportMetadataV1:
    presentation_fingerprint: str
    b_fingerprint: str
    accepted_solution_fingerprint: str | None
    source_id: str
    presentation_mode: str
    cutover_blocked: bool


def _safe_value(value: object) -> object:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _json_text(value: object) -> str:
    return json.dumps(
        value,
        ensure_ascii=False,
        allow_nan=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def _excel_time(seconds: int | None) -> float | None:
    return None if seconds is None else seconds / 86_400


def _write_key_values(
    sheet: Worksheet,
    rows: Iterable[tuple[str, object]],
) -> None:
    sheet.append(["Trường", "Giá trị"])
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _WHITE_FONT
    for key, value in rows:
        sheet.append([key, _safe_value(value)])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:B{sheet.max_row}"
    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 90
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_table(
    sheet: Worksheet,
    headers: list[str],
    rows: Iterable[Iterable[object]],
    *,
    time_columns: tuple[int, ...] = (),
) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        sheet.append([_safe_value(value) for value in row])
    for column in time_columns:
        for row_index in range(2, sheet.max_row + 1):
            sheet.cell(row_index, column).number_format = _TIME_FORMAT
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(headers)).column_letter}{sheet.max_row}"
    for column_index, header in enumerate(headers, 1):
        width = min(max(len(header) + 3, 14), 44)
        sheet.column_dimensions[sheet.cell(1, column_index).column_letter].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _dimension_map(presentation: UnifiedPresentationBundleV1) -> dict[str, str]:
    return {item.dimension_name: item.status for item in presentation.dimensions}


def _overview_rows(
    presentation: UnifiedPresentationBundleV1,
) -> list[tuple[str, object]]:
    statuses = _dimension_map(presentation)
    terminal_statuses = dict(presentation.terminal_occupancy_terminal_statuses)
    terminal_limits = dict(presentation.terminal_occupancy_limits)
    return [
        ("presentation_mode", presentation.presentation_mode),
        ("route_id", presentation.route_id),
        ("route_name", presentation.route_name),
        ("terminal_1_name", presentation.terminal_1_name),
        ("terminal_2_name", presentation.terminal_2_name),
        ("source_id", presentation.source_id),
        ("imported_at", presentation.imported_at),
        ("b_disposition", presentation.outcome.b_disposition),
        ("adjustment_decision", presentation.outcome.adjustment_decision),
        ("selected_action", presentation.outcome.selected_action),
        ("solver_choice", presentation.outcome.solver_choice),
        ("solver_attempted", presentation.outcome.solver_attempted),
        ("b_technical_status", statuses["technical_feasibility"]),
        ("b_demand_status", statuses["demand_suitability"]),
        ("b_fleet_status", statuses["fleet_feasibility"]),
        ("terminal_occupancy_status", presentation.terminal_occupancy_status),
        ("terminal_1_occupancy_status", terminal_statuses.get("terminal_1")),
        ("terminal_2_occupancy_status", terminal_statuses.get("terminal_2")),
        ("terminal_1_occupancy_limit", terminal_limits.get("terminal_1")),
        ("terminal_2_occupancy_limit", terminal_limits.get("terminal_2")),
        ("accepted_c_exists", presentation.outcome.accepted_c_exists),
        ("accepted_c_authority", presentation.outcome.accepted_c_authority),
        ("cutover_blocked", presentation.cutover_blocked),
        ("requires_expert_review", presentation.requires_expert_review),
        (
            "blocking_discrepancy_codes",
            _json_text(presentation.blocking_discrepancy_codes),
        ),
        (
            "expert_review_required_codes",
            _json_text(presentation.expert_review_required_codes),
        ),
        ("b_fingerprint", presentation.source_b_fingerprint),
        ("accepted_c_fingerprint", presentation.accepted_solution_fingerprint),
        ("presentation_fingerprint", presentation.presentation_fingerprint),
    ]


def _write_a_schedule(
    workbook: Workbook,
    presentation: UnifiedPresentationBundleV1,
) -> None:
    scenario = presentation.scenario("A")
    if scenario is None:
        return
    sheet = workbook.create_sheet("A_BIEU_DO")
    _write_table(
        sheet,
        [
            "scenario_id",
            "trip_id",
            "direction",
            "departure_terminal",
            "arrival_terminal",
            "departure_time",
            "arrival_time",
            "runtime_seconds",
            "vehicle_assignment",
            "a_fingerprint",
        ],
        (
            (
                trip.scenario_id,
                trip.trip_id,
                trip.direction,
                trip.departure_terminal,
                trip.arrival_terminal,
                _excel_time(trip.departure_time_seconds),
                _excel_time(trip.arrival_time_seconds),
                trip.runtime_seconds,
                trip.vehicle_assignment,
                scenario.source_fingerprint,
            )
            for trip in scenario.trips
        ),
        time_columns=(6, 7),
    )


def _write_b_schedule(
    workbook: Workbook,
    presentation: UnifiedPresentationBundleV1,
) -> None:
    scenario = presentation.scenario("B")
    if scenario is None:
        raise ValueError("presentation must contain Scenario B")
    sheet = workbook.create_sheet("B_BIEU_DO")
    _write_table(
        sheet,
        [
            "scenario_id",
            "trip_id",
            "direction",
            "departure_terminal",
            "arrival_terminal",
            "departure_time",
            "arrival_time",
            "runtime_seconds",
            "vehicle_assignment",
            "b_fingerprint",
        ],
        (
            (
                trip.scenario_id,
                trip.trip_id,
                trip.direction,
                trip.departure_terminal,
                trip.arrival_terminal,
                _excel_time(trip.departure_time_seconds),
                _excel_time(trip.arrival_time_seconds),
                trip.runtime_seconds,
                trip.vehicle_assignment,
                presentation.source_b_fingerprint,
            )
            for trip in scenario.trips
        ),
        time_columns=(6, 7),
    )


def _write_c_schedule(
    workbook: Workbook,
    presentation: UnifiedPresentationBundleV1,
) -> None:
    scenario = presentation.scenario("C")
    if scenario is None:
        return
    sheet = workbook.create_sheet("C_BIEU_DO")
    _write_table(
        sheet,
        [
            "c_trip_id",
            "source_b_trip_id",
            "direction",
            "departure_terminal",
            "arrival_terminal",
            "b_departure_time",
            "c_departure_time",
            "arrival_time",
            "runtime_seconds",
            "shift_minutes",
            "vehicle",
            "regime_id",
            "change_reason",
            "accepted_solution_fingerprint",
        ],
        (
            (
                trip.trip_id,
                trip.source_b_trip_id,
                trip.direction,
                trip.departure_terminal,
                trip.arrival_terminal,
                _excel_time(trip.b_departure_time_seconds),
                _excel_time(trip.departure_time_seconds),
                _excel_time(trip.arrival_time_seconds),
                trip.runtime_seconds,
                trip.shift_minutes,
                trip.vehicle_assignment,
                trip.headway_regime_id,
                trip.change_reason,
                presentation.accepted_solution_fingerprint,
            )
            for trip in scenario.trips
        ),
        time_columns=(6, 7, 8),
    )


def _write_no_c_status(
    workbook: Workbook,
    presentation: UnifiedPresentationBundleV1,
) -> None:
    if presentation.scenario("C") is not None:
        return
    outcome = presentation.outcome
    sheet = workbook.create_sheet("C_TRANG_THAI")
    _write_key_values(
        sheet,
        [
            ("authoritative_c_statement", "Không có Scenario C có thẩm quyền."),
            ("selected_action", outcome.selected_action),
            ("solver_attempted", outcome.solver_attempted),
            ("heuristic_result_status", outcome.heuristic_result_status),
            ("ortools_result_status", outcome.ortools_result_status),
            ("heuristic_native_solver_status", outcome.heuristic_native_solver_status),
            ("ortools_native_solver_status", outcome.ortools_native_solver_status),
            ("validator_rejection_codes", _json_text(outcome.validator_rejection_codes)),
            ("explanations", _json_text(outcome.explanations)),
            ("limitations", _json_text(outcome.limitations)),
        ],
    )


def _write_b_c_comparison(
    workbook: Workbook,
    presentation: UnifiedPresentationBundleV1,
) -> None:
    scenario = presentation.scenario("C")
    if scenario is None:
        return
    sheet = workbook.create_sheet("SO_SANH_B_C")
    _write_table(
        sheet,
        [
            "source_b_trip_id",
            "c_trip_id",
            "direction",
            "b_departure_time",
            "c_departure_time",
            "shift_minutes",
            "change_reason",
        ],
        (
            (
                trip.source_b_trip_id,
                trip.trip_id,
                trip.direction,
                _excel_time(trip.b_departure_time_seconds),
                _excel_time(trip.departure_time_seconds),
                trip.shift_minutes,
                trip.change_reason,
            )
            for trip in scenario.trips
        ),
        time_columns=(4, 5),
    )


def _write_blocks(
    workbook: Workbook,
    presentation: UnifiedPresentationBundleV1,
) -> None:
    sheet = workbook.create_sheet("CUNG_CAU_BLOCK")
    _write_table(
        sheet,
        [
            "block_id",
            "direction",
            "block_start",
            "block_end",
            "passenger_demand",
            "confidence",
            "vehicle_capacity",
            "a_trip_count",
            "b_trip_count",
            "c_actual_trip_count",
            "required_trips_85",
            "required_trips_90",
            "b_nominal_capacity",
            "c_nominal_capacity",
            "b_load_factor",
            "c_load_factor",
            "b_shortage",
            "c_shortage",
            "b_status",
            "c_status",
            "b_allocation_reason",
            "c_allocation_reason",
        ],
        (
            (
                block.block_id,
                block.direction,
                _excel_time(block.block_start_seconds),
                _excel_time(block.block_end_seconds),
                block.passenger_demand,
                block.confidence,
                block.vehicle_capacity,
                block.a_trip_count,
                block.b_trip_count,
                block.c_actual_trip_count,
                block.required_trips_85,
                block.required_trips_90,
                block.b_nominal_capacity,
                block.c_nominal_capacity,
                block.b_load_factor,
                block.c_load_factor,
                block.b_shortage,
                block.c_shortage,
                block.b_status,
                block.c_status,
                block.allocation_reason,
                block.c_allocation_reason,
            )
            for block in presentation.blocks
        ),
        time_columns=(3, 4),
    )


def _write_dimensions(
    workbook: Workbook,
    presentation: UnifiedPresentationBundleV1,
) -> None:
    sheet = workbook.create_sheet("DANH_GIA_B")
    _write_table(
        sheet,
        [
            "dimension_name",
            "status",
            "confidence",
            "explanation",
            "issue_codes",
            "issue_severities",
            "issue_messages",
            "evidence",
        ],
        (
            (
                dimension.dimension_name,
                dimension.status,
                dimension.confidence,
                dimension.explanation,
                _json_text(dimension.issue_codes),
                _json_text(dimension.issue_severities),
                _json_text(dimension.issue_messages),
                _json_text(dimension.evidence),
            )
            for dimension in presentation.dimensions
        ),
    )


def _write_fleet(
    workbook: Workbook,
    presentation: UnifiedPresentationBundleV1,
) -> None:
    initial = presentation.initial_fleet
    if initial is None:
        return
    sheet = workbook.create_sheet("FLEET_C")
    summary = [
        ("terminal_1_initial", initial.terminal_1_vehicle_count),
        ("terminal_2_initial", initial.terminal_2_vehicle_count),
        ("positioning_mode", initial.positioning_mode),
        ("available_fleet_limit", initial.available_fleet_limit),
        ("approved_active_fleet", initial.approved_active_fleet),
        ("minimum_required_fleet", initial.minimum_required_fleet),
        ("fleet_margin", initial.fleet_margin),
        ("maximum_simultaneous_vehicle_use", initial.maximum_simultaneous_vehicle_use),
        ("fleet_feasibility_status", initial.fleet_feasibility_status),
    ]
    for row_number, (key, value) in enumerate(summary, 1):
        sheet.cell(row_number, 1, key).font = _BOLD_FONT
        sheet.cell(row_number, 2, _safe_value(value))
    header_row = len(summary) + 2
    headers = [
        "vehicle_id",
        "c_trip_id",
        "departure_terminal",
        "arrival_terminal",
        "departure_time",
        "arrival_time",
        "ready_time",
    ]
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(header_row, column, header)
        cell.fill = _HEADER_FILL
        cell.font = _WHITE_FONT
    for row_index, item in enumerate(presentation.fleet_assignments, header_row + 1):
        values = (
            item.vehicle_id,
            item.trip_id,
            item.departure_terminal,
            item.arrival_terminal,
            _excel_time(item.departure_time_seconds),
            _excel_time(item.arrival_time_seconds),
            _excel_time(item.ready_time_seconds),
        )
        for column, value in enumerate(values, 1):
            sheet.cell(row_index, column, _safe_value(value))
        for column in (5, 6, 7):
            sheet.cell(row_index, column).number_format = _TIME_FORMAT
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:G{max(header_row, sheet.max_row)}"
    for column in "ABCDEFG":
        sheet.column_dimensions[column].width = 24


def _write_headways(
    workbook: Workbook,
    presentation: UnifiedPresentationBundleV1,
) -> None:
    if presentation.scenario("C") is None:
        return
    sheet = workbook.create_sheet("HEADWAY_C")
    _write_table(
        sheet,
        [
            "regime_id",
            "direction",
            "start_time",
            "end_time",
            "covered_blocks",
            "trip_count",
            "target_service_rate",
            "target_headway",
            "actual_sequence",
            "transition_headways",
            "exceptional_headways",
            "boundary_reason",
            "regularity_status",
        ],
        (
            (
                regime.regime_id,
                regime.direction,
                _excel_time(regime.start_time_seconds),
                _excel_time(regime.end_time_seconds),
                _json_text(regime.covered_analysis_blocks),
                regime.trip_count,
                regime.target_service_rate,
                regime.target_headway,
                _json_text(regime.actual_headway_sequence),
                _json_text(regime.transition_headways),
                _json_text(regime.exceptional_headways),
                regime.boundary_reason,
                regime.regularity_status,
            )
            for regime in presentation.headway_regimes
        ),
        time_columns=(3, 4),
    )


def _write_solver(
    workbook: Workbook,
    presentation: UnifiedPresentationBundleV1,
) -> None:
    outcome = presentation.outcome
    sheet = workbook.create_sheet("SOLVER")
    _write_key_values(
        sheet,
        [
            ("solver_choice", outcome.solver_choice),
            ("solver_attempted", outcome.solver_attempted),
            ("heuristic_result_status", outcome.heuristic_result_status),
            ("ortools_result_status", outcome.ortools_result_status),
            ("heuristic_native_solver_status", outcome.heuristic_native_solver_status),
            ("ortools_native_solver_status", outcome.ortools_native_solver_status),
            ("validator_rejection_codes", _json_text(outcome.validator_rejection_codes)),
            (
                "comparison_objective_names",
                (
                    _json_text(outcome.comparison_objective_names)
                    if outcome.comparison_objective_names is not None
                    else None
                ),
            ),
            (
                "heuristic_objective_vector",
                (
                    _json_text(outcome.heuristic_objective_vector)
                    if outcome.heuristic_objective_vector is not None
                    else None
                ),
            ),
            (
                "ortools_objective_vector",
                (
                    _json_text(outcome.ortools_objective_vector)
                    if outcome.ortools_objective_vector is not None
                    else None
                ),
            ),
            ("recommended_solver", outcome.recommended_solver),
            ("comparison_reason", outcome.comparison_reason),
            ("accepted_c_exists", outcome.accepted_c_exists),
            ("accepted_c_authority", outcome.accepted_c_authority),
            ("accepted_solution_fingerprint", outcome.accepted_solution_fingerprint),
            ("accepted_outcome_fingerprint", outcome.accepted_outcome_fingerprint),
        ],
    )


def _write_side_by_side(
    workbook: Workbook,
    presentation: UnifiedPresentationBundleV1,
) -> None:
    sheet = workbook.create_sheet("DOI_CHIEU_5A1")
    _write_table(
        sheet,
        [
            "fact_code",
            "category",
            "rule",
            "legacy_value",
            "unified_value",
            "status",
            "disposition",
            "reason_code",
            "explanation",
        ],
        (
            (
                item.fact_code,
                item.category,
                item.comparison_rule,
                _json_text(item.legacy_value),
                _json_text(item.unified_value),
                item.comparison_status,
                item.disposition,
                item.reason_code,
                item.explanation,
            )
            for item in presentation.discrepancies
        ),
    )


def _write_limitations(
    workbook: Workbook,
    presentation: UnifiedPresentationBundleV1,
) -> None:
    rows: list[tuple[object, ...]] = []
    rows.extend(
        ("UNIFIED_EXPLANATION", None, None, None, text) for text in presentation.explanations
    )
    rows.extend(("UNIFIED_LIMITATION", None, None, None, text) for text in presentation.limitations)
    rows.extend(
        ("VALIDATION_EXPLANATION", None, None, None, text)
        for text in presentation.validation_explanations
    )
    rows.extend(
        ("VALIDATION_LIMITATION", None, None, None, text)
        for text in presentation.validation_limitations
    )
    rows.extend(
        (
            "DEMAND_GAP",
            gap.code,
            gap.direction,
            (f"{gap.start_time_seconds}-{gap.end_time_seconds}"),
            "Khoảng thời gian nhu cầu chưa được bao phủ.",
        )
        for gap in presentation.demand_gaps
    )
    rows.extend(
        (
            "TERMINAL_CAPACITY",
            code,
            None,
            None,
            "Trạng thái/giới hạn sức chứa bến được giữ nguyên từ kết quả trả về.",
        )
        for code in presentation.terminal_occupancy_issue_codes
    )
    rows.extend(
        (
            "VALIDATOR_REJECTION",
            code,
            None,
            None,
            "Mã từ chối của bộ kiểm tra miền; không có lịch ứng viên thô được xuất.",
        )
        for code in presentation.outcome.validator_rejection_codes
    )
    sheet = workbook.create_sheet("GIOI_HAN")
    _write_table(
        sheet,
        ["record_type", "code", "direction", "range_seconds", "text"],
        rows,
    )


def _write_fingerprints(
    workbook: Workbook,
    presentation: UnifiedPresentationBundleV1,
) -> None:
    sheet = workbook.create_sheet("FINGERPRINTS")
    _write_key_values(
        sheet,
        [
            ("source_id", presentation.source_id),
            ("normalized_b_fingerprint", presentation.source_b_fingerprint),
            (
                "accepted_solution_fingerprint",
                presentation.accepted_solution_fingerprint,
            ),
            ("accepted_outcome_fingerprint", presentation.accepted_outcome_fingerprint),
            ("presentation_fingerprint", presentation.presentation_fingerprint),
            ("presentation_mode", presentation.presentation_mode),
            ("cutover_blocked", presentation.cutover_blocked),
        ],
    )


def _source_path_matches(
    presentation: UnifiedPresentationBundleV1,
    target: Path,
) -> bool:
    source = Path(presentation.source_id)
    return source.suffix.lower() in {".xlsx", ".xlsm", ".xls"} and (
        source.resolve(strict=False) == target.resolve(strict=False)
    )


def export_unified_result_workbook_v1(
    presentation: UnifiedPresentationBundleV1,
    path: str | Path,
    *,
    overwrite: bool = False,
) -> Path:
    """Write a formula-free validation workbook aligned to one presentation."""
    if not isinstance(presentation, UnifiedPresentationBundleV1):
        raise TypeError("presentation must be a UnifiedPresentationBundleV1")
    target = Path(path)
    if _source_path_matches(presentation, target):
        raise ValueError("output path must not overwrite the presentation source workbook")
    if target.exists() and not overwrite:
        raise FileExistsError(target)
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    overview = workbook.active
    overview.title = "TONG_QUAN"
    _write_key_values(overview, _overview_rows(presentation))
    _write_a_schedule(workbook, presentation)
    _write_b_schedule(workbook, presentation)
    _write_c_schedule(workbook, presentation)
    _write_no_c_status(workbook, presentation)
    _write_b_c_comparison(workbook, presentation)
    _write_blocks(workbook, presentation)
    _write_dimensions(workbook, presentation)
    _write_fleet(workbook, presentation)
    _write_headways(workbook, presentation)
    _write_solver(workbook, presentation)
    _write_side_by_side(workbook, presentation)
    _write_limitations(workbook, presentation)
    _write_fingerprints(workbook, presentation)

    workbook.properties.title = "Milestone 5A2B Unified Presentation Validation"
    workbook.properties.subject = presentation.presentation_mode
    workbook.properties.keywords = presentation.presentation_fingerprint
    workbook.save(target)
    return target


def read_unified_export_metadata_v1(
    path: str | Path,
) -> UnifiedExportMetadataV1:
    """Read alignment metadata from a unified export; never use it as authority."""
    source = Path(path)
    workbook = load_workbook(source, read_only=True, data_only=False)
    try:
        if "FINGERPRINTS" not in workbook.sheetnames:
            raise ValueError("FINGERPRINTS sheet is missing")
        sheet = workbook["FINGERPRINTS"]
        values = {
            str(key): value
            for key, value in sheet.iter_rows(min_row=2, max_col=2, values_only=True)
            if key is not None
        }
    finally:
        workbook.close()
    required = {
        "presentation_fingerprint",
        "normalized_b_fingerprint",
        "source_id",
        "presentation_mode",
        "cutover_blocked",
    }
    missing = sorted(required - values.keys())
    if missing:
        raise ValueError(f"FINGERPRINTS metadata is incomplete: {missing}")
    blocked = values["cutover_blocked"]
    if not isinstance(blocked, bool):
        raise ValueError("cutover_blocked metadata must be a boolean")
    accepted = values.get("accepted_solution_fingerprint")
    return UnifiedExportMetadataV1(
        presentation_fingerprint=str(values["presentation_fingerprint"]),
        b_fingerprint=str(values["normalized_b_fingerprint"]),
        accepted_solution_fingerprint=(str(accepted) if accepted not in {None, ""} else None),
        source_id=str(values["source_id"]),
        presentation_mode=str(values["presentation_mode"]),
        cutover_blocked=blocked,
    )


__all__ = [
    "UnifiedExportMetadataV1",
    "export_unified_result_workbook_v1",
    "read_unified_export_metadata_v1",
]
