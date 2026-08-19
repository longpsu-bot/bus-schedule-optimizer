"""Reviewable XLSX and batch-sensitivity exports for the local V3 runner."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .contracts_v1 import ContractDirection
from .v3_runner import V3ProfileRunV1

PROFILE_COMPARISON_PROFILE_V1 = "v3_profile_comparison_v1"
PROFILE_COMPARISON_ELIGIBLE = "ELIGIBLE"
PROFILE_COMPARISON_INCONCLUSIVE = "INCONCLUSIVE"
PROFILE_COMPARISON_PARTIALLY_COMPARABLE = "PARTIALLY_COMPARABLE"
PROFILE_COMPARISON_INCONCLUSIVE_NO_COMPARABLE_C = "PROFILE_COMPARISON_INCONCLUSIVE_NO_COMPARABLE_C"
STABLE_ACROSS_PROFILES = "STABLE_ACROSS_PROFILES"
MINOR_PROFILE_SENSITIVITY = "MINOR_PROFILE_SENSITIVITY"
MATERIAL_PROFILE_SENSITIVITY = "MATERIAL_PROFILE_SENSITIVITY"
PROFILE_SENSITIVITY_REVIEW_REQUIRED = "PROFILE_SENSITIVITY_REVIEW_REQUIRED"

_NAVY = "17365D"
_BLUE = "2F75B5"
_LIGHT_BLUE = "D9EAF7"
_LIGHT_GRAY = "E7E6E6"
_WHITE = "FFFFFF"
_GREEN = "E2F0D9"
_AMBER = "FFF2CC"
_RED = "FCE4D6"
_THIN_GRAY = Side(style="thin", color="D9E2F3")


def _title(ws, text: str, end_column: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = ws.cell(1, 1, text)
    cell.fill = PatternFill("solid", fgColor=_NAVY)
    cell.font = Font(color=_WHITE, bold=True, size=14)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.sheet_view.showGridLines = False


def _header(ws, row: int, headers: list[str]) -> None:
    for column, value in enumerate(headers, start=1):
        cell = ws.cell(row, column, value)
        cell.fill = PatternFill("solid", fgColor=_BLUE)
        cell.font = Font(color=_WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=_THIN_GRAY)
    ws.row_dimensions[row].height = 28


def _section(ws, row: int, text: str, end_column: int = 2) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_column)
    cell = ws.cell(row, 1, text)
    cell.fill = PatternFill("solid", fgColor=_LIGHT_BLUE)
    cell.font = Font(color=_NAVY, bold=True)
    cell.alignment = Alignment(vertical="center")


def _autowidth(ws, maximum: int = 42) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        length = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=8,
        )
        ws.column_dimensions[letter].width = min(max(length + 2, 10), maximum)


def _excel_time(value: str | None):
    if value is None:
        return None
    return datetime.strptime(value, "%H:%M:%S").time()


def _set_time_columns(ws, columns: tuple[int, ...], start_row: int, end_row: int) -> None:
    for row in range(start_row, end_row + 1):
        for column in columns:
            ws.cell(row, column).number_format = "hh:mm:ss"


def _service_count(
    run: V3ProfileRunV1,
    direction: ContractDirection,
    start: int,
    end: int,
    *,
    scenario_c: bool,
) -> int | None:
    if scenario_c:
        outcome = run.result.candidate_outcome
        solution = outcome.solution if outcome is not None else None
        if solution is None:
            return None
        departures = (
            (item.direction, item.c_departure_time) for item in solution.c_exact_timetable
        )
    else:
        departures = (
            (item.direction, item.departure_time)
            for item in run.normalized_inputs.scenario_b.exact_timetable
        )
    return sum(
        start <= departure < end
        and (direction == ContractDirection.COMBINED or item_direction == direction)
        for item_direction, departure in departures
    )


def _write_summary(ws, run: V3ProfileRunV1) -> None:
    _title(ws, "V3 TWO-STAGE RUN SUMMARY", 4)
    profile = run.payload["selected_profile"]
    budget = run.payload["budget"]
    stage_1 = run.payload["stage_1"]
    stage_2 = run.payload["stage_2"]
    fleet = run.payload["fleet"]
    shifts = run.payload["shift_metrics"]
    quality = run.payload["quality"]
    rows = [
        ("Route", f"{run.payload['route_id']} — {run.payload['route_name']}"),
        ("Selected profile", profile["profile_id"]),
        ("Profile fingerprint", profile["profile_fingerprint"]),
        ("Included periods", ", ".join(profile["included_period_ids"])),
        ("Total observation days", profile["total_observation_days"]),
        ("Direction grain", profile["direction_grain"]),
        ("Scenario C available", run.payload["scenario_c_available"]),
        ("Final acceptance", run.payload["final_acceptance_state"]),
        ("Aggregate native status", run.payload["aggregate_native_status"]),
        ("Fleet available", fleet["available"]),
        ("Scenario B required fleet", fleet["scenario_b_required"]),
        ("Scenario C required fleet", fleet["scenario_c_required"]),
        ("Shifted trips", shifts["shifted_trip_count"]),
        ("Total shift minutes", shifts["total_shift_minutes"]),
        ("Maximum shift minutes", shifts["maximum_shift_minutes"]),
        (
            "Solve budget",
            f"{budget['consumed_seconds']:.3f} / {budget['total_seconds']:.3f} seconds; "
            f"exhausted={str(budget['exhausted']).lower()}",
        ),
        (
            "Stage 1 counts",
            f"candidates={stage_1['candidate_count']}; admitted={stage_1['admitted_count']}; "
            f"pruned={stage_1['pruned_count']}",
        ),
        ("Stage 2 attempts", stage_2["allocation_attempt_count"]),
    ]
    _header(ws, 3, ["Metric", "Value"])
    for row, values in enumerate(rows, start=4):
        ws.cell(row, 1, values[0]).font = Font(bold=True, color=_NAVY)
        ws.cell(row, 2, values[1])
    quality_row = len(rows) + 6
    _section(ws, quality_row, "B / C QUALITY VECTOR", 4)
    _header(ws, quality_row + 1, ["Metric", "Scenario B", "Scenario C", "C - B"])
    b_quality = quality["B"] or {}
    c_quality = quality["C"] or {}
    for offset, metric in enumerate(b_quality, start=quality_row + 2):
        b_value = b_quality[metric]
        c_value = c_quality.get(metric)
        ws.cell(offset, 1, metric)
        ws.cell(offset, 2, b_value)
        ws.cell(offset, 3, c_value)
        ws.cell(offset, 4, None if c_value is None else c_value - b_value)
    acceptance_row = next(
        row for row, values in enumerate(rows, start=4) if values[0] == "Final acceptance"
    )
    acceptance_cell = ws.cell(acceptance_row, 2)
    acceptance_cell.fill = PatternFill(
        "solid",
        fgColor=(
            _GREEN if run.payload["final_acceptance_state"] == "FINAL_RECOMMENDED" else _AMBER
        ),
    )
    ws.freeze_panes = "A4"
    _autowidth(ws)


def _write_demand_profile(ws, run: V3ProfileRunV1) -> None:
    _title(ws, "DERIVED DEMAND PROFILE AND SERVICE BY 30-MINUTE BLOCK", 12)
    headers = [
        "direction",
        "time_block_start",
        "time_block_end",
        "average_daily_passengers",
        "scenario_b_service_count",
        "scenario_c_service_count",
        "chart_label",
    ]
    _header(ws, 3, headers)
    for row_number, item in enumerate(
        run.derivation.profile.derived_observations,
        start=4,
    ):
        start_text = f"{item.interval_start // 3600:02d}:{item.interval_start % 3600 // 60:02d}:00"
        end_text = f"{item.interval_end // 3600:02d}:{item.interval_end % 3600 // 60:02d}:00"
        values = [
            item.direction.value,
            _excel_time(start_text),
            _excel_time(end_text),
            item.average_daily_passengers,
            _service_count(
                run,
                item.direction,
                item.interval_start,
                item.interval_end,
                scenario_c=False,
            ),
            _service_count(
                run,
                item.direction,
                item.interval_start,
                item.interval_end,
                scenario_c=True,
            ),
            f"{item.direction.value} {start_text[:5]}",
        ]
        for column, value in enumerate(values, start=1):
            ws.cell(row_number, column, value)
    end_row = 3 + len(run.derivation.profile.derived_observations)
    _set_time_columns(ws, (2, 3), 4, end_row)
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 24
    categories = Reference(ws, min_col=7, min_row=4, max_row=end_row)
    demand_chart = LineChart()
    demand_chart.title = "Demand profile (passengers per block)"
    demand_chart.y_axis.title = "Passengers per block"
    demand_chart.x_axis.title = "Direction and block start"
    demand_chart.height = 10
    demand_chart.width = 23
    demand_chart.add_data(
        Reference(ws, min_col=4, max_col=4, min_row=3, max_row=end_row),
        titles_from_data=True,
    )
    demand_chart.set_categories(categories)

    service_chart = LineChart()
    service_chart.title = (
        "Service count B vs C (trips per block)"
        if run.payload["scenario_c_available"]
        else "Scenario B service (trips per block); Scenario C not available"
    )
    service_chart.y_axis.title = "Trips per block"
    service_chart.x_axis.title = "Direction and block start"
    service_chart.height = 10
    service_chart.width = 23
    service_chart.add_data(
        Reference(ws, min_col=5, max_col=5, min_row=3, max_row=end_row),
        titles_from_data=True,
    )
    if run.payload["scenario_c_available"]:
        service_chart.add_data(
            Reference(ws, min_col=6, max_col=6, min_row=3, max_row=end_row),
            titles_from_data=True,
        )
    service_chart.set_categories(categories)
    demand_chart.legend.position = "b"
    service_chart.legend.position = "b"
    ws.add_chart(demand_chart, "I3")
    ws.add_chart(service_chart, "I24")
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:G{end_row}"
    _autowidth(ws, maximum=30)


def _write_stage_1(ws, run: V3ProfileRunV1) -> None:
    _title(ws, "STAGE 1 ALLOCATION", 12)
    headers = [
        "block_id",
        "direction",
        "start",
        "end",
        "trip_count",
        "source_b_trip_count",
        "protected_minimum_trip_count",
        "observed_passengers",
        "required_trips_90",
        "required_trips_85",
    ]
    _header(ws, 3, headers)
    selected = run.payload["stage_1"]["selected_allocation_plan"]
    blocks = selected["allocation_by_demand_interval"] if selected else []
    for row_number, item in enumerate(blocks, start=4):
        values = [
            item["block_id"],
            item["direction"],
            _excel_time(item["start_time"]),
            _excel_time(item["end_time"]),
            item["trip_count"],
            item["source_b_trip_count"],
            item["protected_minimum_trip_count"],
            item["observed_passengers"],
            item["required_trips_90"],
            item["required_trips_85"],
        ]
        for column, value in enumerate(values, start=1):
            ws.cell(row_number, column, value)
    end_row = max(4, 3 + len(blocks))
    _set_time_columns(ws, (3, 4), 4, end_row)
    ws.freeze_panes = "A4"
    if blocks:
        ws.auto_filter.ref = f"A3:J{end_row}"
    _autowidth(ws)


def _write_regimes(ws, run: V3ProfileRunV1) -> None:
    _title(ws, "FINAL V3 SERVICE REGIMES", 9)
    headers = [
        "direction",
        "regime_id",
        "start",
        "end",
        "trip_count",
        "uniform_headway_minutes",
        "boundary_reason",
        "is_final_service_tail",
    ]
    _header(ws, 3, headers)
    regimes = run.payload["final_service_regimes"]
    for row_number, item in enumerate(regimes, start=4):
        values = [
            item["direction"],
            item["regime_id"],
            _excel_time(item["start"]),
            _excel_time(item["end"]),
            item["trip_count"],
            item["uniform_headway_minutes"],
            item["boundary_reason"],
            item["is_final_service_tail"],
        ]
        for column, value in enumerate(values, start=1):
            ws.cell(row_number, column, value)
    end_row = max(4, 3 + len(regimes))
    _set_time_columns(ws, (3, 4), 4, end_row)
    ws.freeze_panes = "A4"
    if regimes:
        ws.auto_filter.ref = f"A3:H{end_row}"
    _autowidth(ws)


def _write_timetable_b(ws, run: V3ProfileRunV1) -> None:
    _title(ws, "SCENARIO B EXACT TIMETABLE", 9)
    headers = [
        "trip_id",
        "direction",
        "departure_terminal",
        "departure_time",
        "arrival_time",
        "runtime_minutes",
        "vehicle_assignment",
    ]
    _header(ws, 3, headers)
    for row_number, item in enumerate(
        run.normalized_inputs.scenario_b.exact_timetable,
        start=4,
    ):
        values = [
            item.trip_id,
            item.direction.value,
            item.departure_terminal.value,
            _excel_time(
                f"{item.departure_time // 3600:02d}:{item.departure_time % 3600 // 60:02d}:00"
            ),
            _excel_time(f"{item.arrival_time // 3600:02d}:{item.arrival_time % 3600 // 60:02d}:00"),
            item.runtime_minutes,
            item.vehicle_assignment,
        ]
        for column, value in enumerate(values, start=1):
            ws.cell(row_number, column, value)
    end_row = 3 + len(run.normalized_inputs.scenario_b.exact_timetable)
    _set_time_columns(ws, (4, 5), 4, end_row)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:G{end_row}"
    _autowidth(ws)


def _write_timetable_c(ws, run: V3ProfileRunV1) -> None:
    _title(ws, "SCENARIO C EXACT TIMETABLE AND B → C SHIFTS", 10)
    headers = [
        "c_trip_id",
        "source_b_trip_id",
        "direction",
        "departure_terminal",
        "b_departure_time",
        "c_departure_time",
        "arrival_time",
        "shift_minutes",
        "headway_regime_id",
        "vehicle_assignment",
    ]
    _header(ws, 3, headers)
    timetable = run.payload["timetable_c"]
    for row_number, item in enumerate(timetable, start=4):
        values = [
            item["c_trip_id"],
            item["source_b_trip_id"],
            item["direction"],
            item["departure_terminal"],
            _excel_time(item["b_departure_time"]),
            _excel_time(item["c_departure_time"]),
            _excel_time(item["arrival_time"]),
            item["shift_minutes"],
            item["headway_regime_id"],
            item["vehicle_assignment"],
        ]
        for column, value in enumerate(values, start=1):
            ws.cell(row_number, column, value)
    end_row = max(4, 3 + len(timetable))
    _set_time_columns(ws, (5, 6, 7), 4, end_row)
    ws.freeze_panes = "A4"
    if timetable:
        ws.auto_filter.ref = f"A3:J{end_row}"
    _autowidth(ws)


def _write_diagnostics(ws, run: V3ProfileRunV1) -> None:
    _title(ws, "DIAGNOSTICS, EXPLANATIONS, AND LIMITATIONS", 10)
    _header(
        ws,
        3,
        [
            "period_id",
            "direction",
            "average_daily_passengers",
            "peak_block_start",
            "peak_block_end",
            "peak_share",
            "compared_period_id",
            "maximum_shape_distance",
            "threshold",
            "structural_change_detected",
        ],
    )
    diagnostics = run.payload["period_diagnostics"]
    for row_number, item in enumerate(diagnostics, start=4):
        values = [
            item["period_id"],
            item["direction"],
            item["average_daily_passengers"],
            _excel_time(item["peak_block"]["start"]),
            _excel_time(item["peak_block"]["end"]),
            item["peak_share"],
            item["compared_period_id"],
            item["maximum_shape_distance"],
            item["shape_distance_threshold"],
            item["structural_change_detected"],
        ]
        for column, value in enumerate(values, start=1):
            ws.cell(row_number, column, value)
    end_row = 3 + len(diagnostics)
    _set_time_columns(ws, (4, 5), 4, end_row)
    next_row = end_row + 2
    for title, values in (
        ("DIAGNOSTIC CODES", run.payload["diagnostic_codes"]),
        ("EXPLANATIONS", run.payload["explanations"]),
        ("LIMITATIONS", run.payload["limitations"]),
    ):
        _section(ws, next_row, title, 10)
        next_row += 1
        if not values:
            ws.cell(next_row, 1, "None")
            next_row += 1
        else:
            for value in values:
                ws.merge_cells(
                    start_row=next_row,
                    start_column=1,
                    end_row=next_row,
                    end_column=10,
                )
                ws.cell(next_row, 1, value).alignment = Alignment(wrap_text=True)
                next_row += 1
        next_row += 1
    ws.freeze_panes = "A4"
    _autowidth(ws, maximum=34)


def export_v3_result_xlsx_v1(run: V3ProfileRunV1, path: str | Path) -> Path:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "SUMMARY"
    demand = workbook.create_sheet("DEMAND_PROFILE")
    stage_1 = workbook.create_sheet("STAGE1_ALLOCATION")
    regimes = workbook.create_sheet("REGIMES")
    timetable_b = workbook.create_sheet("TIMETABLE_B")
    timetable_c = workbook.create_sheet("TIMETABLE_C")
    diagnostics = workbook.create_sheet("DIAGNOSTICS")
    _write_summary(summary, run)
    _write_demand_profile(demand, run)
    _write_stage_1(stage_1, run)
    _write_regimes(regimes, run)
    _write_timetable_b(timetable_b, run)
    _write_timetable_c(timetable_c, run)
    _write_diagnostics(diagnostics, run)
    workbook.save(output)
    return output


def _c_comparison_eligibility(run: V3ProfileRunV1) -> tuple[bool, tuple[str, ...]]:
    outcome = run.result.candidate_outcome
    solution = outcome.solution if outcome is not None else None
    reasons: list[str] = []
    if solution is None:
        reasons.append("SCENARIO_C_SOLUTION_MISSING")
    if run.payload["quality"]["C"] is None:
        reasons.append("SCENARIO_C_QUALITY_MISSING")
    if run.derivation.profile.derived_observations and not run.payload["final_service_regimes"]:
        reasons.append("SCENARIO_C_REGIMES_MISSING")
    return not reasons, tuple(reasons)


def _comparison_row(run: V3ProfileRunV1) -> dict[str, object]:
    selected = run.payload["stage_1"]["selected_allocation_plan"]
    allocation = (
        [
            {
                "block_id": item["block_id"],
                "direction": item["direction"],
                "start": item["start_time"],
                "end": item["end_time"],
                "trip_count": item["trip_count"],
            }
            for item in selected["allocation_by_demand_interval"]
        ]
        if selected
        else []
    )
    regimes = run.payload["final_service_regimes"]
    c_quality = run.payload["quality"]["C"]
    c_comparable, unavailable_reasons = _c_comparison_eligibility(run)
    return {
        "profile_id": run.derivation.profile.profile_id,
        "profile_fingerprint": run.derivation.profile.profile_fingerprint,
        "scenario_c_available": run.payload["scenario_c_available"],
        "c_comparable": c_comparable,
        "c_comparison_unavailable_reasons": list(unavailable_reasons),
        "final_acceptance_state": run.payload["final_acceptance_state"],
        "aggregate_native_status": run.payload["aggregate_native_status"],
        "stage_1_allocation_by_block": allocation,
        "regime_count": len(regimes),
        "regime_boundaries": [
            {
                "direction": item["direction"],
                "regime_id": item["regime_id"],
                "start": item["start"],
                "end": item["end"],
            }
            for item in regimes
        ],
        "uniform_headways": [
            {
                "direction": item["direction"],
                "regime_id": item["regime_id"],
                "minutes": item["uniform_headway_minutes"],
            }
            for item in regimes
        ],
        "fleet_required_c": run.payload["fleet"]["scenario_c_required"],
        "maximum_service_gap": (
            c_quality["maximum_positive_demand_service_gap_minutes"]
            if c_quality is not None
            else None
        ),
        "shifted_trips": run.payload["shift_metrics"]["shifted_trip_count"],
        "total_shift": run.payload["shift_metrics"]["total_shift_minutes"],
        "maximum_shift": run.payload["shift_metrics"]["maximum_shift_minutes"],
    }


def build_profile_comparison_v1(runs: list[V3ProfileRunV1]) -> dict[str, object]:
    if not runs:
        raise ValueError("at least one profile run is required")
    rows = [_comparison_row(run) for run in runs]
    comparable = [bool(row["c_comparable"]) for row in rows]
    if not any(comparable):
        comparison_eligibility = PROFILE_COMPARISON_INCONCLUSIVE
        classification = None
        review_code = PROFILE_COMPARISON_INCONCLUSIVE_NO_COMPARABLE_C
    elif not all(comparable):
        comparison_eligibility = PROFILE_COMPARISON_PARTIALLY_COMPARABLE
        classification = MATERIAL_PROFILE_SENSITIVITY
        review_code = PROFILE_SENSITIVITY_REVIEW_REQUIRED
    else:
        comparison_eligibility = PROFILE_COMPARISON_ELIGIBLE
        review_code = None

    exact_fields = (
        "final_acceptance_state",
        "aggregate_native_status",
        "stage_1_allocation_by_block",
        "regime_count",
        "regime_boundaries",
        "uniform_headways",
        "fleet_required_c",
        "maximum_service_gap",
        "shifted_trips",
        "total_shift",
        "maximum_shift",
    )
    stable = all(all(row[field] == rows[0][field] for field in exact_fields) for row in rows[1:])
    material_fields = (
        "final_acceptance_state",
        "aggregate_native_status",
        "stage_1_allocation_by_block",
        "regime_count",
        "regime_boundaries",
        "uniform_headways",
        "fleet_required_c",
    )
    if all(comparable):
        material = any(
            any(row[field] != rows[0][field] for field in material_fields)
            or abs(float(row["maximum_service_gap"]) - float(rows[0]["maximum_service_gap"])) > 5
            or abs(float(row["maximum_shift"]) - float(rows[0]["maximum_shift"])) > 5
            or abs(float(row["total_shift"]) - float(rows[0]["total_shift"]))
            > run.normalized_inputs.scenario_b.total_daily_trips
            for row, run in zip(rows[1:], runs[1:], strict=True)
        )
        classification = (
            STABLE_ACROSS_PROFILES
            if stable
            else MATERIAL_PROFILE_SENSITIVITY
            if material
            else MINOR_PROFILE_SENSITIVITY
        )
        review_code = (
            PROFILE_SENSITIVITY_REVIEW_REQUIRED
            if classification == MATERIAL_PROFILE_SENSITIVITY
            else None
        )
    return {
        "comparison_profile": PROFILE_COMPARISON_PROFILE_V1,
        "route_id": runs[0].payload["route_id"],
        "route_name": runs[0].payload["route_name"],
        "comparison_eligibility": comparison_eligibility,
        "stability_classification": classification,
        "review_code": review_code,
        "classification_rule": {
            "eligibility": (
                "C comparison requires a genuine Scenario C solution, a C quality vector, and "
                "final service regimes for routes with measurable demand regimes."
            ),
            "inconclusive": (
                "If no profile has comparable C, stability classification is null and review "
                f"code is {PROFILE_COMPARISON_INCONCLUSIVE_NO_COMPARABLE_C}."
            ),
            "mixed_availability": (
                "If only some profiles have comparable C, the result is material profile "
                "sensitivity requiring review."
            ),
            "stable": "All reported status, allocation, regime, fleet, gap, and shift fields match.",
            "material": (
                "Any status, Stage 1 allocation, regime structure/headway, or fleet difference; "
                "or >5-minute maximum-gap/maximum-shift difference; or total-shift difference "
                "greater than one minute per daily trip."
            ),
            "minor": "Differences exist but none cross the material rule.",
            "selection_policy": "Classification never changes the configured primary profile.",
        },
        "profiles": rows,
    }


def export_profile_comparison_xlsx_v1(
    comparison: dict[str, object],
    path: str | Path,
) -> Path:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "PROFILE_COMPARISON"
    allocation = workbook.create_sheet("ALLOCATION")
    regimes = workbook.create_sheet("REGIMES")
    _title(summary, "V3 PROFILE SENSITIVITY COMPARISON", 12)
    summary_rows = [
        ("Route", f"{comparison['route_id']} — {comparison['route_name']}"),
        ("Comparison eligibility", comparison["comparison_eligibility"]),
        ("Stability classification", comparison["stability_classification"]),
        ("Review code", comparison["review_code"]),
        ("Eligibility rule", comparison["classification_rule"]["eligibility"]),
        ("Inconclusive rule", comparison["classification_rule"]["inconclusive"]),
        ("Material rule", comparison["classification_rule"]["material"]),
        ("Selection policy", comparison["classification_rule"]["selection_policy"]),
    ]
    _header(summary, 3, ["Metric", "Value"])
    for row, (label, value) in enumerate(summary_rows, start=4):
        summary.cell(row, 1, label).font = Font(bold=True, color=_NAVY)
        summary.cell(row, 2, value).alignment = Alignment(wrap_text=True)
        if value is not None and len(str(value)) > 60:
            summary.row_dimensions[row].height = 45
    table_row = 14
    headers = [
        "profile_id",
        "scenario_c_available",
        "c_comparable",
        "final_acceptance_state",
        "aggregate_native_status",
        "regime_count",
        "fleet_required_c",
        "maximum_service_gap",
        "shifted_trips",
        "total_shift",
        "maximum_shift",
    ]
    _header(summary, table_row, headers)
    for row_number, item in enumerate(comparison["profiles"], start=table_row + 1):
        for column, key in enumerate(headers, start=1):
            summary.cell(row_number, column, item[key])
    _title(allocation, "STAGE 1 ALLOCATION BY PROFILE", 7)
    _header(
        allocation,
        3,
        ["profile_id", "block_id", "direction", "start", "end", "trip_count"],
    )
    row_number = 4
    for profile in comparison["profiles"]:
        for item in profile["stage_1_allocation_by_block"]:
            values = [
                profile["profile_id"],
                item["block_id"],
                item["direction"],
                _excel_time(item["start"]),
                _excel_time(item["end"]),
                item["trip_count"],
            ]
            for column, value in enumerate(values, start=1):
                allocation.cell(row_number, column, value)
            row_number += 1
    _set_time_columns(allocation, (4, 5), 4, max(4, row_number - 1))
    _title(regimes, "REGIME BOUNDARIES AND UNIFORM HEADWAYS", 8)
    _header(
        regimes,
        3,
        ["profile_id", "direction", "regime_id", "start", "end", "uniform_headway_minutes"],
    )
    row_number = 4
    for profile in comparison["profiles"]:
        headway_lookup = {
            (item["direction"], item["regime_id"]): item["minutes"]
            for item in profile["uniform_headways"]
        }
        for item in profile["regime_boundaries"]:
            values = [
                profile["profile_id"],
                item["direction"],
                item["regime_id"],
                _excel_time(item["start"]),
                _excel_time(item["end"]),
                headway_lookup[(item["direction"], item["regime_id"])],
            ]
            for column, value in enumerate(values, start=1):
                regimes.cell(row_number, column, value)
            row_number += 1
    _set_time_columns(regimes, (4, 5), 4, max(4, row_number - 1))
    for ws in workbook.worksheets:
        ws.freeze_panes = "A4"
        _autowidth(ws, maximum=42)
    summary.column_dimensions["B"].width = 80
    status_cell = next(
        summary.cell(row, 2)
        for row in range(4, 4 + len(summary_rows))
        if summary.cell(row, 1).value == "Stability classification"
    )
    status_cell.fill = PatternFill(
        "solid",
        fgColor=(
            _GREEN
            if comparison["stability_classification"] == STABLE_ACROSS_PROFILES
            else _RED
            if comparison["stability_classification"] == MATERIAL_PROFILE_SENSITIVITY
            else _AMBER
        ),
    )
    workbook.save(output)
    return output


__all__ = [
    "MATERIAL_PROFILE_SENSITIVITY",
    "MINOR_PROFILE_SENSITIVITY",
    "PROFILE_COMPARISON_ELIGIBLE",
    "PROFILE_COMPARISON_INCONCLUSIVE",
    "PROFILE_COMPARISON_INCONCLUSIVE_NO_COMPARABLE_C",
    "PROFILE_COMPARISON_PARTIALLY_COMPARABLE",
    "PROFILE_COMPARISON_PROFILE_V1",
    "PROFILE_SENSITIVITY_REVIEW_REQUIRED",
    "STABLE_ACROSS_PROFILES",
    "build_profile_comparison_v1",
    "export_profile_comparison_xlsx_v1",
    "export_v3_result_xlsx_v1",
]
