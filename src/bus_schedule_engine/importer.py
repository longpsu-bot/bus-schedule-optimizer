from __future__ import annotations

import math
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from numbers import Integral, Real
from pathlib import Path
from typing import TYPE_CHECKING, BinaryIO

import pandas as pd
from openpyxl import load_workbook

from .models import (
    DemandRecord,
    Direction,
    RouteType,
    ScenarioParameters,
    Trip,
    TripRidershipDatasetMetadataV1,
    TripRidershipDirectionV1,
    TripRidershipObservationV1,
    VolumeType,
)
from .time_utils import parse_runtime_options, parse_time_to_seconds
from .trip_ridership_codes import (
    DUPLICATE_TRIP_RIDERSHIP_OBSERVATION_ID,
    TRIP_RIDERSHIP_COMBINED_DIRECTION_NOT_ALLOWED,
    TRIP_RIDERSHIP_CONFIDENCE_INVALID,
    TRIP_RIDERSHIP_DATASET_ID_MISSING,
    TRIP_RIDERSHIP_DIRECTION_INVALID,
    TRIP_RIDERSHIP_FORMULA_NOT_ALLOWED,
    TRIP_RIDERSHIP_MATCH_TOLERANCE_INVALID,
    TRIP_RIDERSHIP_METADATA_MISSING,
    TRIP_RIDERSHIP_OPERATING_DAY_TYPE_MISMATCH,
    TRIP_RIDERSHIP_PASSENGER_COUNT_INVALID,
    TRIP_RIDERSHIP_REFERENCE_MISSING,
    TRIP_RIDERSHIP_SCENARIO_INVALID,
    TRIP_RIDERSHIP_SOURCE_TYPE_INVALID,
)

if TYPE_CHECKING:
    from .contracts_v1.models import DemandConfidence, DemandResponseMode, DemandSourceType


class InputDataError(ValueError):
    """Blocking error in the input workbook."""


@dataclass(frozen=True, slots=True)
class WorkbookAuthorityMetadata:
    """Workbook-owned demand authority; runtime provenance is intentionally excluded."""

    demand_dataset_id: str | None = None
    demand_source_type: DemandSourceType | None = None
    demand_confidence: DemandConfidence | None = None
    demand_response_mode: DemandResponseMode | None = None
    source_notes: str | None = None


@dataclass(frozen=True)
class ImportedWorkbook:
    parameters_a: ScenarioParameters | None
    trips_a: list[Trip]
    parameters_b: ScenarioParameters
    trips_b: list[Trip]
    demand: list[DemandRecord]
    configuration: dict[str, object]
    authority_metadata: WorkbookAuthorityMetadata = field(default_factory=WorkbookAuthorityMetadata)
    trip_ridership_metadata: TripRidershipDatasetMetadataV1 | None = None
    trip_ridership_observations: tuple[TripRidershipObservationV1, ...] = ()


REQUIRED_SHEETS = {
    "THONG_SO_B",
    "BIEU_DO_B",
}

SUPPORTED_SHEETS = REQUIRED_SHEETS | {
    "HUONG_DAN",
    "THONG_SO_A",
    "BIEU_DO_A",
    "SAN_LUONG",
    "THONG_TIN_DU_LIEU",
    "SAN_LUONG_CHUYEN",
    "THONG_TIN_SAN_LUONG_CHUYEN",
    "CAU_HINH",
}

OPERATING_DAY_TYPES = {"weekday", "saturday", "sunday", "holiday", "special"}
TRIP_RIDERSHIP_SOURCE_TYPES = {"ticketing", "manual_count", "apc", "survey", "other"}
TRIP_RIDERSHIP_CONFIDENCE_VALUES = {"high", "medium", "low", "unknown"}
TRIP_RIDERSHIP_REQUIRED_COLUMNS = {
    "observation_id",
    "service_date",
    "source_trip_id",
    "scheduled_trip_id",
    "direction",
    "scheduled_departure_time",
    "actual_departure_time",
    "passenger_count",
    "vehicle_id",
    "notes",
}


def _materialize_workbook_source(
    source: str | Path | bytes | BinaryIO,
) -> str | Path | bytes:
    if isinstance(source, (str, Path, bytes)):
        return source
    original_position: int | None = None
    try:
        original_position = source.tell()
        source.seek(0)
    except (AttributeError, OSError):
        original_position = None
    content = source.read()
    if original_position is not None:
        source.seek(original_position)
    return content


def _trip_ridership_formula_cells(source: str | Path | bytes) -> tuple[str, ...]:
    workbook_source = BytesIO(source) if isinstance(source, bytes) else source
    workbook = load_workbook(workbook_source, read_only=True, data_only=False)
    try:
        coordinates = []
        for sheet_name in ("THONG_TIN_SAN_LUONG_CHUYEN", "SAN_LUONG_CHUYEN"):
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            coordinates.extend(
                f"{sheet_name}!{cell.coordinate}"
                for row in sheet.iter_rows()
                for cell in row
                if cell.data_type == "f"
            )
        return tuple(sorted(coordinates))
    finally:
        workbook.close()


def _clean(value: object) -> object | None:
    if value is None or (not isinstance(value, str) and pd.isna(value)):
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def _direction(value: object) -> Direction:
    text = str(_clean(value) or "").lower().replace(" ", "_")
    aliases = {
        "terminal_1_to_2": Direction.TERMINAL_1_TO_2,
        "t1_t2": Direction.TERMINAL_1_TO_2,
        "bến_1_đến_bến_2": Direction.TERMINAL_1_TO_2,
        "ben_1_den_ben_2": Direction.TERMINAL_1_TO_2,
        "terminal_2_to_1": Direction.TERMINAL_2_TO_1,
        "t2_t1": Direction.TERMINAL_2_TO_1,
        "bến_2_đến_bến_1": Direction.TERMINAL_2_TO_1,
        "ben_2_den_ben_1": Direction.TERMINAL_2_TO_1,
        "combined": Direction.COMBINED,
        "tổng_hai_chiều": Direction.COMBINED,
        "tong_hai_chieu": Direction.COMBINED,
    }
    try:
        return aliases[text]
    except KeyError as exc:
        raise InputDataError(f"Chiều không hợp lệ: {value}") from exc


def _date(value: object) -> date:
    cleaned = _clean(value)
    if isinstance(cleaned, datetime):
        return cleaned.date()
    if isinstance(cleaned, date):
        return cleaned
    try:
        return pd.to_datetime(cleaned, dayfirst=True).date()
    except (TypeError, ValueError) as exc:
        raise InputDataError(f"Ngày không hợp lệ: {value}") from exc


def _key_value_sheet(frame: pd.DataFrame, sheet_name: str) -> dict[str, object]:
    if frame.shape[1] < 2:
        raise InputDataError(f"Sheet {sheet_name} phải có ít nhất hai cột")
    output: dict[str, object] = {}
    for key, value in frame.iloc[:, :2].itertuples(index=False, name=None):
        cleaned_key = _clean(key)
        if cleaned_key is not None:
            output[str(cleaned_key)] = _clean(value)
    return output


def _required(mapping: dict[str, object], key: str, sheet_name: str) -> object:
    value = mapping.get(key)
    if value is None:
        raise InputDataError(f"Thiếu {key} trong sheet {sheet_name}")
    return value


def _optional_positive_integer(
    value: object | None,
    *,
    key: str,
    sheet_name: str,
    allow_numeric_text: bool = False,
) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        raise InputDataError(f"{key} trong {sheet_name} phải là số nguyên >= 1")
    if isinstance(value, Integral) or (
        isinstance(value, Real) and math.isfinite(float(value)) and float(value).is_integer()
    ):
        parsed = int(value)
    elif allow_numeric_text and isinstance(value, str):
        try:
            parsed = int(value.strip())
        except ValueError as exc:
            raise InputDataError(f"{key} trong {sheet_name} phải là số nguyên >= 1") from exc
    else:
        raise InputDataError(f"{key} trong {sheet_name} phải là số nguyên >= 1")
    if parsed < 1:
        raise InputDataError(f"{key} trong {sheet_name} phải là số nguyên >= 1")
    return parsed


def _required_positive_integer(
    value: object | None,
    *,
    key: str,
    sheet_name: str,
    allow_numeric_text: bool = False,
) -> int:
    parsed = _optional_positive_integer(
        value,
        key=key,
        sheet_name=sheet_name,
        allow_numeric_text=allow_numeric_text,
    )
    if parsed is None:
        raise InputDataError(f"Thiếu {key} trong sheet {sheet_name}")
    return parsed


def _parameters(frame: pd.DataFrame, scenario: str) -> ScenarioParameters:
    sheet_name = f"THONG_SO_{scenario}"
    values = _key_value_sheet(frame, sheet_name)
    try:
        route_type = RouteType(str(_required(values, "route_type", sheet_name)).strip())
    except ValueError as exc:
        raise InputDataError("route_type phải là intra_provincial hoặc inter_provincial") from exc
    capacity = _required_positive_integer(
        values.get("vehicle_capacity_passengers"),
        key="vehicle_capacity_passengers",
        sheet_name=sheet_name,
        allow_numeric_text=True,
    )
    total_daily_trips = _required_positive_integer(
        values.get("total_daily_trips"),
        key="total_daily_trips",
        sheet_name=sheet_name,
        allow_numeric_text=True,
    )
    layover_raw = values.get("minimum_layover_minutes")
    allowed_runtime_raw = values.get("allowed_trip_runtime_minutes")
    legacy_runtime_raw = values.get("trip_runtime_minutes")
    available_fleet_raw = values.get("available_fleet_limit")
    approved_active_fleet_raw = values.get("approved_active_fleet")
    operating_day_type_raw = values.get("operating_day_type")
    terminal_1_max_occupancy_raw = values.get("terminal_1_max_occupancy_vehicles")
    terminal_2_max_occupancy_raw = values.get("terminal_2_max_occupancy_vehicles")
    operating_day_type = (
        None if operating_day_type_raw is None else str(operating_day_type_raw).strip().lower()
    )
    if operating_day_type is not None and operating_day_type not in OPERATING_DAY_TYPES:
        allowed = ", ".join(sorted(OPERATING_DAY_TYPES))
        raise InputDataError(f"operating_day_type trong {sheet_name} phải là một trong: {allowed}")
    try:
        runtime_options = parse_runtime_options(
            allowed_runtime_raw
            if allowed_runtime_raw is not None
            else _required(values, "trip_runtime_minutes", sheet_name)
        )
    except ValueError as exc:
        raise InputDataError(
            f"allowed_trip_runtime_minutes trong {sheet_name} không hợp lệ: {exc}"
        ) from exc
    runtime_default = (
        max(runtime_options) if allowed_runtime_raw is not None else int(legacy_runtime_raw)
    )
    return ScenarioParameters(
        route_id=str(_required(values, "route_id", sheet_name)),
        route_name=str(_required(values, "route_name", sheet_name)),
        route_type=route_type,
        trip_runtime_minutes=runtime_default,
        total_daily_trips=total_daily_trips,
        terminal_1_name=str(_required(values, "terminal_1_name", sheet_name)),
        terminal_1_first_departure=parse_time_to_seconds(
            _required(values, "terminal_1_first_departure", sheet_name)
        ),
        terminal_1_last_departure=parse_time_to_seconds(
            _required(values, "terminal_1_last_departure", sheet_name)
        ),
        terminal_2_name=str(_required(values, "terminal_2_name", sheet_name)),
        terminal_2_first_departure=parse_time_to_seconds(
            _required(values, "terminal_2_first_departure", sheet_name)
        ),
        terminal_2_last_departure=parse_time_to_seconds(
            _required(values, "terminal_2_last_departure", sheet_name)
        ),
        vehicle_capacity_passengers=capacity,
        target_load_factor=float(values.get("target_load_factor") or 0.85),
        maximum_load_factor=float(values.get("maximum_load_factor") or 0.90),
        time_block_minutes=int(values.get("time_block_minutes") or 60),
        minimum_layover_minutes=None if layover_raw is None else int(layover_raw),
        allowed_trip_runtime_minutes=runtime_options,
        available_fleet_limit=_optional_positive_integer(
            available_fleet_raw,
            key="available_fleet_limit",
            sheet_name=sheet_name,
            allow_numeric_text=True,
        ),
        approved_active_fleet=_optional_positive_integer(
            approved_active_fleet_raw,
            key="approved_active_fleet",
            sheet_name=sheet_name,
            allow_numeric_text=True,
        ),
        operating_day_type=operating_day_type,
        terminal_1_max_occupancy_vehicles=(
            _optional_positive_integer(
                terminal_1_max_occupancy_raw,
                key="terminal_1_max_occupancy_vehicles",
                sheet_name=sheet_name,
            )
            if scenario == "B"
            else None
        ),
        terminal_2_max_occupancy_vehicles=(
            _optional_positive_integer(
                terminal_2_max_occupancy_raw,
                key="terminal_2_max_occupancy_vehicles",
                sheet_name=sheet_name,
            )
            if scenario == "B"
            else None
        ),
    )


def _authority_metadata(frame: pd.DataFrame) -> WorkbookAuthorityMetadata:
    from .contracts_v1.models import DemandConfidence, DemandResponseMode, DemandSourceType

    sheet_name = "THONG_TIN_DU_LIEU"
    values = _key_value_sheet(frame, sheet_name)

    def optional_enum(key: str, enum_type):
        raw = values.get(key)
        if raw is None:
            return None
        try:
            return enum_type(str(raw).strip())
        except ValueError as exc:
            allowed = ", ".join(item.value for item in enum_type)
            raise InputDataError(f"{key} trong {sheet_name} phải là một trong: {allowed}") from exc

    return WorkbookAuthorityMetadata(
        demand_dataset_id=(
            None
            if values.get("demand_dataset_id") is None
            else str(values["demand_dataset_id"]).strip()
        ),
        demand_source_type=optional_enum("demand_source_type", DemandSourceType),
        demand_confidence=optional_enum("demand_confidence", DemandConfidence),
        demand_response_mode=optional_enum("demand_response_mode", DemandResponseMode),
        source_notes=(
            None if values.get("source_notes") is None else str(values["source_notes"]).strip()
        ),
    )


def _trips(frame: pd.DataFrame, scenario: str) -> list[Trip]:
    required = {"trip_id", "departure_terminal", "direction", "departure_time"}
    missing = required - set(frame.columns)
    if missing:
        raise InputDataError(f"BIEU_DO_{scenario} thiếu cột: {', '.join(sorted(missing))}")
    trips: list[Trip] = []
    for row_number, row in frame.iterrows():
        if _clean(row.get("trip_id")) is None:
            continue
        try:
            arrival_raw = _clean(row.get("arrival_time"))
            override_raw = _clean(row.get("vehicle_capacity_override"))
            trips.append(
                Trip(
                    scenario=scenario,
                    trip_id=str(row["trip_id"]).strip(),
                    departure_terminal=str(row["departure_terminal"]).strip(),
                    direction=_direction(row["direction"]),
                    departure_seconds=parse_time_to_seconds(row["departure_time"]),
                    arrival_seconds=(
                        None if arrival_raw is None else parse_time_to_seconds(arrival_raw)
                    ),
                    vehicle_id=(
                        None
                        if _clean(row.get("vehicle_id")) is None
                        else str(row.get("vehicle_id")).strip()
                    ),
                    vehicle_capacity_override=(None if override_raw is None else int(override_raw)),
                )
            )
        except (TypeError, ValueError) as exc:
            raise InputDataError(f"BIEU_DO_{scenario}, dòng Excel {row_number + 2}: {exc}") from exc
    return trips


def _demand(frame: pd.DataFrame) -> list[DemandRecord]:
    required = {
        "period_start",
        "period_end",
        "observation_days",
        "time_block_start",
        "time_block_end",
        "direction",
        "passenger_volume",
        "volume_type",
    }
    missing = required - set(frame.columns)
    if missing:
        raise InputDataError(f"SAN_LUONG thiếu cột: {', '.join(sorted(missing))}")
    records: list[DemandRecord] = []
    for row_number, row in frame.iterrows():
        if _clean(row.get("time_block_start")) is None:
            continue
        try:
            records.append(
                DemandRecord(
                    period_start=_date(row["period_start"]),
                    period_end=_date(row["period_end"]),
                    observation_days=int(row["observation_days"]),
                    block_start_seconds=parse_time_to_seconds(row["time_block_start"]),
                    block_end_seconds=parse_time_to_seconds(row["time_block_end"]),
                    direction=_direction(row["direction"]),
                    passenger_volume=float(row["passenger_volume"]),
                    volume_type=VolumeType(str(row["volume_type"]).strip()),
                )
            )
        except (TypeError, ValueError) as exc:
            raise InputDataError(f"SAN_LUONG, dòng Excel {row_number + 2}: {exc}") from exc
    return records


def _trip_metadata_required(
    values: dict[str, object],
    key: str,
) -> object:
    value = values.get(key)
    if value is None:
        code = (
            TRIP_RIDERSHIP_DATASET_ID_MISSING
            if key == "trip_ridership_dataset_id"
            else TRIP_RIDERSHIP_METADATA_MISSING
        )
        raise InputDataError(f"{code}: thiếu {key} trong sheet THONG_TIN_SAN_LUONG_CHUYEN")
    return value


def _trip_ridership_metadata(
    frame: pd.DataFrame,
    parameters_b: ScenarioParameters,
) -> TripRidershipDatasetMetadataV1:
    values = _key_value_sheet(frame, "THONG_TIN_SAN_LUONG_CHUYEN")
    dataset_id = str(_trip_metadata_required(values, "trip_ridership_dataset_id")).strip()
    if not dataset_id:
        raise InputDataError(
            f"{TRIP_RIDERSHIP_DATASET_ID_MISSING}: trip_ridership_dataset_id không được để trống"
        )

    source_type = str(_trip_metadata_required(values, "trip_ridership_source_type")).strip().lower()
    if source_type not in TRIP_RIDERSHIP_SOURCE_TYPES:
        allowed = ", ".join(sorted(TRIP_RIDERSHIP_SOURCE_TYPES))
        raise InputDataError(
            f"{TRIP_RIDERSHIP_SOURCE_TYPE_INVALID}: "
            f"trip_ridership_source_type phải là một trong: {allowed}"
        )

    confidence = str(_trip_metadata_required(values, "trip_ridership_confidence")).strip().lower()
    if confidence not in TRIP_RIDERSHIP_CONFIDENCE_VALUES:
        allowed = ", ".join(sorted(TRIP_RIDERSHIP_CONFIDENCE_VALUES))
        raise InputDataError(
            f"{TRIP_RIDERSHIP_CONFIDENCE_INVALID}: "
            f"trip_ridership_confidence phải là một trong: {allowed}"
        )

    scenario = str(_trip_metadata_required(values, "observed_schedule_scenario")).strip().upper()
    if scenario != "B":
        raise InputDataError(
            f"{TRIP_RIDERSHIP_SCENARIO_INVALID}: observed_schedule_scenario "
            "phải là B trong Milestone 6A1"
        )

    operating_day_type = str(_trip_metadata_required(values, "operating_day_type")).strip().lower()
    if (
        operating_day_type not in OPERATING_DAY_TYPES
        or operating_day_type != parameters_b.operating_day_type
    ):
        raise InputDataError(
            f"{TRIP_RIDERSHIP_OPERATING_DAY_TYPE_MISMATCH}: operating_day_type "
            "phải trùng với Scenario B; không suy từ ngày lịch"
        )

    tolerance_raw = _trip_metadata_required(values, "match_tolerance_minutes")
    if isinstance(tolerance_raw, bool):
        tolerance: int | None = None
    elif isinstance(tolerance_raw, Integral) or (
        isinstance(tolerance_raw, Real)
        and math.isfinite(float(tolerance_raw))
        and float(tolerance_raw).is_integer()
    ):
        tolerance = int(tolerance_raw)
    elif isinstance(tolerance_raw, str):
        try:
            parsed = float(tolerance_raw.strip())
        except ValueError:
            tolerance = None
        else:
            tolerance = int(parsed) if math.isfinite(parsed) and parsed.is_integer() else None
    else:
        tolerance = None
    if tolerance is None or not 0 <= tolerance <= 30:
        raise InputDataError(
            f"{TRIP_RIDERSHIP_MATCH_TOLERANCE_INVALID}: "
            "match_tolerance_minutes phải là số nguyên từ 0 đến 30"
        )

    source_notes = _clean(values.get("source_notes"))
    return TripRidershipDatasetMetadataV1(
        dataset_id=dataset_id,
        source_type=source_type,
        confidence=confidence,
        observed_schedule_scenario=scenario,
        operating_day_type=operating_day_type,
        match_tolerance_minutes=tolerance,
        source_notes=None if source_notes is None else str(source_notes).strip(),
    )


def _trip_ridership_time(value: object, *, field_name: str) -> int | None:
    cleaned = _clean(value)
    if cleaned is None:
        return None
    if isinstance(cleaned, bool) or (isinstance(cleaned, Real) and not 0 <= float(cleaned) < 1):
        raise InputDataError(f"{field_name} nằm ngoài ngày dịch vụ được hỗ trợ")
    try:
        parsed = parse_time_to_seconds(cleaned)
    except (TypeError, ValueError) as exc:
        raise InputDataError(f"{field_name} không hợp lệ") from exc
    if not 0 <= parsed < 86400:
        raise InputDataError(f"{field_name} nằm ngoài ngày dịch vụ được hỗ trợ")
    return parsed


def _trip_ridership_passenger_count(value: object) -> int:
    cleaned = _clean(value)
    parsed: float | None
    if cleaned is None or isinstance(cleaned, bool):
        parsed = None
    elif isinstance(cleaned, Real):
        parsed = float(cleaned)
    elif isinstance(cleaned, str):
        try:
            parsed = float(cleaned)
        except ValueError:
            parsed = None
    else:
        parsed = None
    if parsed is None or not math.isfinite(parsed) or not parsed.is_integer() or parsed < 0:
        raise InputDataError(
            f"{TRIP_RIDERSHIP_PASSENGER_COUNT_INVALID}: passenger_count phải là số nguyên >= 0"
        )
    return int(parsed)


def _trip_ridership_observations(
    frame: pd.DataFrame,
) -> tuple[TripRidershipObservationV1, ...]:
    missing = TRIP_RIDERSHIP_REQUIRED_COLUMNS - set(frame.columns)
    if missing:
        raise InputDataError("SAN_LUONG_CHUYEN thiếu cột: " + ", ".join(sorted(missing)))

    observations: list[TripRidershipObservationV1] = []
    seen_ids: set[str] = set()
    for row_number, row in frame.iterrows():
        if all(_clean(value) is None for value in row.tolist()):
            continue
        try:
            observation_id_raw = _clean(row.get("observation_id"))
            service_date_raw = _clean(row.get("service_date"))
            direction_raw = _clean(row.get("direction"))
            if observation_id_raw is None:
                raise InputDataError("observation_id là bắt buộc")
            if service_date_raw is None:
                raise InputDataError("service_date là bắt buộc")
            if direction_raw is None:
                raise InputDataError(f"{TRIP_RIDERSHIP_DIRECTION_INVALID}: thiếu direction")

            observation_id = str(observation_id_raw).strip()
            if observation_id in seen_ids:
                raise InputDataError(
                    f"{DUPLICATE_TRIP_RIDERSHIP_OBSERVATION_ID}: observation_id phải duy nhất"
                )
            seen_ids.add(observation_id)

            direction_text = str(direction_raw).strip().lower()
            if direction_text == "combined":
                raise InputDataError(
                    f"{TRIP_RIDERSHIP_COMBINED_DIRECTION_NOT_ALLOWED}: "
                    "direction combined không được phép"
                )
            try:
                direction = TripRidershipDirectionV1(direction_text)
            except ValueError as exc:
                raise InputDataError(
                    f"{TRIP_RIDERSHIP_DIRECTION_INVALID}: direction phải là outbound hoặc inbound"
                ) from exc

            source_trip_id_raw = _clean(row.get("source_trip_id"))
            scheduled_trip_id_raw = _clean(row.get("scheduled_trip_id"))
            vehicle_id_raw = _clean(row.get("vehicle_id"))
            notes_raw = _clean(row.get("notes"))
            scheduled_departure_seconds = _trip_ridership_time(
                row.get("scheduled_departure_time"),
                field_name="scheduled_departure_time",
            )
            actual_departure_seconds = _trip_ridership_time(
                row.get("actual_departure_time"),
                field_name="actual_departure_time",
            )
            scheduled_trip_id = (
                None if scheduled_trip_id_raw is None else str(scheduled_trip_id_raw).strip()
            )
            if (
                scheduled_trip_id is None
                and scheduled_departure_seconds is None
                and actual_departure_seconds is None
            ):
                raise InputDataError(
                    f"{TRIP_RIDERSHIP_REFERENCE_MISSING}: phải có ít nhất một "
                    "trong scheduled_trip_id, scheduled_departure_time hoặc "
                    "actual_departure_time"
                )
            observations.append(
                TripRidershipObservationV1(
                    observation_id=observation_id,
                    service_date=_date(service_date_raw),
                    source_trip_id=(
                        None if source_trip_id_raw is None else str(source_trip_id_raw).strip()
                    ),
                    scheduled_trip_id=scheduled_trip_id,
                    direction=direction,
                    scheduled_departure_seconds=scheduled_departure_seconds,
                    actual_departure_seconds=actual_departure_seconds,
                    passenger_count=_trip_ridership_passenger_count(row.get("passenger_count")),
                    vehicle_id=(None if vehicle_id_raw is None else str(vehicle_id_raw).strip()),
                    notes=None if notes_raw is None else str(notes_raw).strip(),
                )
            )
        except (TypeError, ValueError) as exc:
            if isinstance(exc, InputDataError):
                message = str(exc)
            else:
                message = f"giá trị không hợp lệ: {exc}"
            raise InputDataError(
                f"SAN_LUONG_CHUYEN, dòng Excel {row_number + 4}: {message}"
            ) from exc
    return tuple(observations)


def import_workbook(source: str | Path | bytes | BinaryIO) -> ImportedWorkbook:
    materialized_source = _materialize_workbook_source(source)
    formula_cells = _trip_ridership_formula_cells(materialized_source)
    if formula_cells:
        raise InputDataError(
            f"{TRIP_RIDERSHIP_FORMULA_NOT_ALLOWED}: công thức không được dùng "
            "làm thẩm quyền dữ liệu chuyến"
        )
    workbook_source: str | Path | BytesIO
    workbook_source = (
        BytesIO(materialized_source)
        if isinstance(materialized_source, bytes)
        else materialized_source
    )
    with pd.ExcelFile(workbook_source, engine="openpyxl") as excel_file:
        missing_sheets = REQUIRED_SHEETS - set(excel_file.sheet_names)
        if missing_sheets:
            raise InputDataError(f"Workbook thiếu sheet: {', '.join(sorted(missing_sheets))}")
        sheets = {
            sheet_name: pd.read_excel(excel_file, sheet_name=sheet_name, header=2)
            for sheet_name in SUPPORTED_SHEETS & set(excel_file.sheet_names)
            if sheet_name != "HUONG_DAN"
        }
    configuration = (
        _key_value_sheet(sheets["CAU_HINH"], "CAU_HINH")
        if "CAU_HINH" in sheets and sheets["CAU_HINH"].shape[1] >= 2
        else {}
    )
    authority_metadata = (
        _authority_metadata(sheets["THONG_TIN_DU_LIEU"])
        if "THONG_TIN_DU_LIEU" in sheets
        else WorkbookAuthorityMetadata()
    )
    parameters_b = _parameters(sheets["THONG_SO_B"], "B")
    parameters_a: ScenarioParameters | None = None
    trips_a: list[Trip] = []
    schedule_a = sheets.get("BIEU_DO_A")
    if schedule_a is not None and not schedule_a.dropna(how="all").empty:
        if "trip_id" not in schedule_a.columns:
            raise InputDataError("BIEU_DO_A có dữ liệu nhưng thiếu cột trip_id")
        has_a_trips = any(_clean(value) is not None for value in schedule_a["trip_id"])
        if has_a_trips:
            if "THONG_SO_A" not in sheets:
                raise InputDataError(
                    "BIEU_DO_A có chuyến nhưng thiếu THONG_SO_A; "
                    "hãy bổ sung thông số A hoặc bỏ dữ liệu A để chạy chế độ chỉ có B"
                )
            parameters_a = _parameters(sheets["THONG_SO_A"], "A")
            trips_a = _trips(schedule_a, "A")
    demand_sheet = sheets.get("SAN_LUONG")
    demand = (
        _demand(demand_sheet)
        if demand_sheet is not None and not demand_sheet.dropna(how="all").empty
        else []
    )
    trip_ridership_sheet = sheets.get("SAN_LUONG_CHUYEN")
    has_trip_ridership_rows = (
        trip_ridership_sheet is not None and not trip_ridership_sheet.dropna(how="all").empty
    )
    if has_trip_ridership_rows and "THONG_TIN_SAN_LUONG_CHUYEN" not in sheets:
        raise InputDataError(
            f"{TRIP_RIDERSHIP_METADATA_MISSING}: SAN_LUONG_CHUYEN có dữ liệu "
            "nhưng thiếu sheet THONG_TIN_SAN_LUONG_CHUYEN"
        )
    trip_ridership_metadata = (
        _trip_ridership_metadata(
            sheets["THONG_TIN_SAN_LUONG_CHUYEN"],
            parameters_b,
        )
        if has_trip_ridership_rows
        else None
    )
    trip_ridership_observations = (
        _trip_ridership_observations(trip_ridership_sheet)
        if has_trip_ridership_rows and trip_ridership_sheet is not None
        else ()
    )
    return ImportedWorkbook(
        parameters_a=parameters_a,
        trips_a=trips_a,
        parameters_b=parameters_b,
        trips_b=_trips(sheets["BIEU_DO_B"], "B"),
        demand=demand,
        configuration=configuration,
        authority_metadata=authority_metadata,
        trip_ridership_metadata=trip_ridership_metadata,
        trip_ridership_observations=trip_ridership_observations,
    )
