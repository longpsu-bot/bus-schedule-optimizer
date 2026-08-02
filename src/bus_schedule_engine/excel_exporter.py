from __future__ import annotations

from datetime import date, time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .models import (
    Direction,
    RouteType,
    ScenarioParameters,
    Trip,
)

NAVY = "17324D"
BLUE = "2563EB"
TEAL = "0F766E"
GREEN = "DCFCE7"
AMBER = "FEF3C7"
RED = "FEE2E2"
LIGHT_BLUE = "E0F2FE"
LIGHT_GRAY = "F1F5F9"
WHITE = "FFFFFF"
TEXT = "1E293B"
MUTED = "64748B"
THIN_GRAY = Side(style="thin", color="CBD5E1")

REQUIRED_LABEL = "BẮT BUỘC"
REQUIRED_FOR_OPTIMIZATION_LABEL = "BẮT BUỘC ĐỂ TỐI ƯU"
REQUIRED_FOR_DEMAND_AND_OPTIMIZATION_LABEL = "BẮT BUỘC ĐỂ ĐÁNH GIÁ NHU CẦU/TỐI ƯU"
CONDITIONAL_DEMAND_LABEL = "BẮT BUỘC ĐỂ TỐI ƯU KHI CÓ SẢN LƯỢNG"
CONDITIONAL_TRIP_RIDERSHIP_LABEL = "BẮT BUỘC KHI CÓ SAN_LUONG_CHUYEN"
OPTIONAL_LABEL = "TÙY CHỌN"


def _title(ws, title: str, end_column: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = ws.cell(1, 1, title)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34
    ws.sheet_view.showGridLines = False


def _table_header(ws, row: int, headers: list[str]) -> None:
    for column, header in enumerate(headers, 1):
        cell = ws.cell(row, column, header)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(name="Aptos", bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color="0B4F4A"))
    ws.row_dimensions[row].height = 30


def _body_style(ws, start_row: int, end_row: int, end_column: int) -> None:
    if end_row < start_row:
        return
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=end_column):
        for cell in row:
            cell.font = Font(name="Aptos", size=10, color=TEXT)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN_GRAY)
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")


def _style_requirement_rows(ws, start_row: int, end_row: int, level_column: int) -> None:
    fills = {
        REQUIRED_LABEL: PatternFill("solid", fgColor=RED),
        REQUIRED_FOR_OPTIMIZATION_LABEL: PatternFill("solid", fgColor=AMBER),
        REQUIRED_FOR_DEMAND_AND_OPTIMIZATION_LABEL: PatternFill("solid", fgColor=AMBER),
        CONDITIONAL_DEMAND_LABEL: PatternFill("solid", fgColor=AMBER),
        CONDITIONAL_TRIP_RIDERSHIP_LABEL: PatternFill("solid", fgColor=AMBER),
        OPTIONAL_LABEL: PatternFill("solid", fgColor=LIGHT_BLUE),
    }
    for row in range(start_row, end_row + 1):
        level_cell = ws.cell(row, level_column)
        level_cell.fill = fills[str(level_cell.value)]
        level_cell.font = Font(name="Aptos", size=10, bold=True, color=TEXT)
        ws.cell(row, 1).font = Font(name="Aptos", size=10, bold=True, color=TEXT)


def _autowidth(ws, maximum: int = 42) -> None:
    for column_cells in ws.columns:
        column = get_column_letter(column_cells[0].column)
        length = 0
        for cell in column_cells:
            value = cell.value
            if value is not None:
                parts = str(value).splitlines() or [""]
                length = max(length, max(len(part) for part in parts))
        ws.column_dimensions[column].width = min(maximum, max(10, length + 2))


def _as_time(seconds: int) -> time:
    seconds %= 86400
    hour, remainder = divmod(seconds, 3600)
    minute, second = divmod(remainder, 60)
    return time(hour, minute, second)


def _sample_parameters() -> tuple[ScenarioParameters, ScenarioParameters]:
    common = {
        "route_id": "MVP-01",
        "route_name": "Tuyến mẫu Bến Trung Tâm – Bến Phía Đông",
        "route_type": RouteType.INTRA_PROVINCIAL,
        "trip_runtime_minutes": 65,
        "allowed_trip_runtime_minutes": (55, 65),
        "total_daily_trips": 24,
        "terminal_1_name": "Bến Trung Tâm",
        "terminal_1_first_departure": 6 * 3600,
        "terminal_1_last_departure": 18 * 3600,
        "terminal_2_name": "Bến Phía Đông",
        "terminal_2_first_departure": 6 * 3600 + 15 * 60,
        "terminal_2_last_departure": 18 * 3600 + 15 * 60,
        "vehicle_capacity_passengers": 60,
        "target_load_factor": 0.85,
        "maximum_load_factor": 0.90,
        "time_block_minutes": 60,
        "minimum_layover_minutes": 10,
    }
    parameters_a = ScenarioParameters(
        **common,
        available_fleet_limit=8,
        approved_active_fleet=4,
        operating_day_type="weekday",
    )
    parameters_b = ScenarioParameters(
        **common,
        available_fleet_limit=8,
        approved_active_fleet=4,
        operating_day_type="weekday",
        terminal_1_max_occupancy_vehicles=4,
        terminal_2_max_occupancy_vehicles=4,
    )
    return parameters_a, parameters_b


def _inclusive_times(first: int, last: int, count: int) -> list[int]:
    if count == 1:
        return [first]
    step = (last - first) / (count - 1)
    return [round(first + index * step) for index in range(count)]


def _sample_trips(
    scenario: str, parameters: ScenarioParameters, proposed: bool = False
) -> list[Trip]:
    if proposed:
        t1_hours = [
            "06:00",
            "06:45",
            "09:30",
            "10:30",
            "11:30",
            "12:30",
            "13:30",
            "14:30",
            "15:30",
            "16:30",
            "17:30",
            "18:00",
        ]
        t2_hours = [
            "06:15",
            "07:00",
            "09:45",
            "10:45",
            "11:45",
            "12:45",
            "13:45",
            "14:45",
            "15:45",
            "16:45",
            "17:45",
            "18:15",
        ]

        def parse(text: str) -> int:
            hour, minute = (int(part) for part in text.split(":"))
            return hour * 3600 + minute * 60

        times_1 = [parse(item) for item in t1_hours]
        times_2 = [parse(item) for item in t2_hours]
    else:
        times_1 = _inclusive_times(
            parameters.terminal_1_first_departure,
            parameters.terminal_1_last_departure,
            12,
        )
        times_2 = _inclusive_times(
            parameters.terminal_2_first_departure,
            parameters.terminal_2_last_departure,
            12,
        )
    trips = []
    for direction, terminal, times in (
        (Direction.TERMINAL_1_TO_2, parameters.terminal_1_name, times_1),
        (Direction.TERMINAL_2_TO_1, parameters.terminal_2_name, times_2),
    ):
        for departure in times:
            trips.append(
                Trip(
                    scenario=scenario,
                    trip_id="",
                    departure_terminal=terminal,
                    direction=direction,
                    departure_seconds=departure,
                    arrival_seconds=departure + parameters.default_trip_runtime_minutes * 60,
                )
            )
    ordered = sorted(trips, key=lambda item: (item.departure_seconds, item.direction.value))
    return [
        Trip(
            scenario=trip.scenario,
            trip_id=f"{scenario}-{index:03d}",
            departure_terminal=trip.departure_terminal,
            direction=trip.direction,
            departure_seconds=trip.departure_seconds,
            arrival_seconds=trip.arrival_seconds,
        )
        for index, trip in enumerate(ordered, 1)
    ]


def _write_parameter_sheet(ws, parameters: ScenarioParameters, scenario: str) -> None:
    _title(ws, f"THÔNG SỐ SCENARIO {scenario}", 4)
    _table_header(ws, 3, ["Tham số", "Giá trị", "Mức độ", "Diễn giải"])
    rows = [
        ("route_id", parameters.route_id, REQUIRED_LABEL, "Mã tuyến"),
        ("route_name", parameters.route_name, REQUIRED_LABEL, "Tên tuyến"),
        ("route_type", parameters.route_type.value, REQUIRED_LABEL, "Loại tuyến"),
        (
            "allowed_trip_runtime_minutes",
            parameters.runtime_options_text,
            REQUIRED_LABEL,
            "Định dạng runtime ưu tiên; có thể dùng trip_runtime_minutes cho file tương thích cũ.",
        ),
        (
            "trip_runtime_minutes",
            parameters.default_trip_runtime_minutes,
            OPTIONAL_LABEL,
            "Giá trị tương thích cũ; chỉ dùng khi allowed_trip_runtime_minutes để trống.",
        ),
        (
            "total_daily_trips",
            parameters.total_daily_trips,
            REQUIRED_LABEL,
            "Tổng lượt hai chiều/ngày",
        ),
        ("terminal_1_name", parameters.terminal_1_name, REQUIRED_LABEL, "Tên bến 1"),
        (
            "terminal_1_first_departure",
            _as_time(parameters.terminal_1_first_departure),
            REQUIRED_LABEL,
            "Giờ đầu bến 1",
        ),
        (
            "terminal_1_last_departure",
            _as_time(parameters.terminal_1_last_departure),
            REQUIRED_LABEL,
            "Giờ cuối bến 1",
        ),
        ("terminal_2_name", parameters.terminal_2_name, REQUIRED_LABEL, "Tên bến 2"),
        (
            "terminal_2_first_departure",
            _as_time(parameters.terminal_2_first_departure),
            REQUIRED_LABEL,
            "Giờ đầu bến 2",
        ),
        (
            "terminal_2_last_departure",
            _as_time(parameters.terminal_2_last_departure),
            REQUIRED_LABEL,
            "Giờ cuối bến 2",
        ),
        (
            "vehicle_capacity_passengers",
            parameters.vehicle_capacity_passengers,
            REQUIRED_FOR_DEMAND_AND_OPTIMIZATION_LABEL,
            "Có thể để trống để review biểu đồ; bắt buộc cho tải cung/cầu và tối ưu, không tự suy đoán",
        ),
        (
            "available_fleet_limit",
            parameters.available_fleet_limit,
            REQUIRED_FOR_OPTIMIZATION_LABEL,
            "Giới hạn đội xe cứng; không phải đội xe tối thiểu",
        ),
        (
            "operating_day_type",
            parameters.operating_day_type,
            REQUIRED_FOR_OPTIMIZATION_LABEL,
            "weekday, saturday, sunday, holiday hoặc special",
        ),
        (
            "approved_active_fleet",
            parameters.approved_active_fleet,
            OPTIONAL_LABEL,
            "Siêu dữ liệu quản trị; không thay thế available_fleet_limit",
        ),
        (
            "target_load_factor",
            parameters.target_load_factor,
            OPTIONAL_LABEL,
            "Mặc định 85%",
        ),
        (
            "maximum_load_factor",
            parameters.maximum_load_factor,
            OPTIONAL_LABEL,
            "Trần khuyến nghị 90%",
        ),
        (
            "time_block_minutes",
            parameters.time_block_minutes,
            OPTIONAL_LABEL,
            "Chỉ 30 hoặc 60",
        ),
        (
            "minimum_layover_minutes",
            parameters.effective_layover_minutes,
            OPTIONAL_LABEL,
            "Để trống để dùng mức tối thiểu theo loại tuyến",
        ),
    ]
    if scenario == "B":
        rows.extend(
            [
                (
                    "terminal_1_max_occupancy_vehicles",
                    parameters.terminal_1_max_occupancy_vehicles,
                    OPTIONAL_LABEL,
                    "Để trống nếu chưa có thẩm quyền sức chứa vật lý bến 1",
                ),
                (
                    "terminal_2_max_occupancy_vehicles",
                    parameters.terminal_2_max_occupancy_vehicles,
                    OPTIONAL_LABEL,
                    "Để trống nếu chưa có thẩm quyền sức chứa vật lý bến 2",
                ),
            ]
        )
    for row_index, row in enumerate(rows, 4):
        for column, value in enumerate(row, 1):
            ws.cell(row_index, column, value)
    _body_style(ws, 4, 3 + len(rows), 4)
    _style_requirement_rows(ws, 4, 3 + len(rows), 3)
    for row in range(4, 4 + len(rows)):
        key = ws.cell(row, 1).value
        if key and "departure" in key:
            ws.cell(row, 2).number_format = "HH:mm"
        if key and "load_factor" in key:
            ws.cell(row, 2).number_format = "0%"
    route_type_validation = DataValidation(
        type="list", formula1='"intra_provincial,inter_provincial"'
    )
    block_validation = DataValidation(type="list", formula1='"30,60"')
    operating_day_validation = DataValidation(
        type="list", formula1='"weekday,saturday,sunday,holiday,special"'
    )
    required_positive_integer = DataValidation(
        type="whole", operator="greaterThan", formula1="0", allow_blank=False
    )
    optional_positive_integer = DataValidation(
        type="whole", operator="greaterThan", formula1="0", allow_blank=True
    )
    load_factor_validation = DataValidation(
        type="decimal", operator="between", formula1="0.01", formula2="1"
    )
    ws.add_data_validation(route_type_validation)
    ws.add_data_validation(block_validation)
    ws.add_data_validation(operating_day_validation)
    ws.add_data_validation(required_positive_integer)
    ws.add_data_validation(optional_positive_integer)
    ws.add_data_validation(load_factor_validation)
    key_rows = {ws.cell(row, 1).value: row for row in range(4, 4 + len(rows))}
    runtime_options_cell = ws.cell(key_rows["allowed_trip_runtime_minutes"], 2)
    runtime_options_cell.number_format = "@"
    runtime_list_validation = DataValidation(
        type="custom",
        formula1=f"=ISTEXT(B{runtime_options_cell.row})",
        allow_blank=True,
    )
    runtime_list_validation.errorTitle = "Phải nhập khoảng dạng văn bản"
    runtime_list_validation.error = (
        "Nhập hai đầu mút như 55,65 hoặc 55;65; mọi số nguyên ở giữa đều được phép."
    )
    runtime_list_validation.promptTitle = "Khoảng runtime bao gồm hai đầu"
    runtime_list_validation.prompt = (
        "Ví dụ: 55,65. Excel dùng dấu phẩy thập phân có thể nhập 55;65."
    )
    runtime_list_validation.showErrorMessage = True
    runtime_list_validation.showInputMessage = True
    ws.add_data_validation(runtime_list_validation)
    runtime_list_validation.add(runtime_options_cell)
    route_type_validation.add(ws.cell(key_rows["route_type"], 2))
    block_validation.add(ws.cell(key_rows["time_block_minutes"], 2))
    operating_day_validation.add(ws.cell(key_rows["operating_day_type"], 2))
    required_positive_integer.add(ws.cell(key_rows["total_daily_trips"], 2))
    for key in (
        "trip_runtime_minutes",
        "vehicle_capacity_passengers",
        "minimum_layover_minutes",
        "available_fleet_limit",
        "approved_active_fleet",
    ):
        optional_positive_integer.add(ws.cell(key_rows[key], 2))
    if scenario == "B":
        optional_positive_integer.add(ws.cell(key_rows["terminal_1_max_occupancy_vehicles"], 2))
        optional_positive_integer.add(ws.cell(key_rows["terminal_2_max_occupancy_vehicles"], 2))
    for key in ("target_load_factor", "maximum_load_factor"):
        load_factor_validation.add(ws.cell(key_rows[key], 2))
    ws.freeze_panes = "A4"
    _autowidth(ws)


def _write_authority_metadata_sheet(ws) -> None:
    _title(ws, "THÔNG TIN THẨM QUYỀN DỮ LIỆU", 4)
    _table_header(ws, 3, ["Tham số", "Giá trị", "Mức độ", "Diễn giải"])
    rows = [
        (
            "timetable_authority_status",
            "proposed",
            OPTIONAL_LABEL,
            "approved_operational, current_operational, proposed hoặc unknown; chỉ giá trị khai báo rõ ràng mới được bảo toàn",
        ),
        (
            "timetable_authority_reference",
            "SYNTHETIC-SAMPLE",
            OPTIONAL_LABEL,
            "Số quyết định, tài liệu hoặc tham chiếu do người dùng khai báo; không dùng để tự suy trạng thái",
        ),
        (
            "timetable_effective_date",
            date(2026, 8, 1),
            OPTIONAL_LABEL,
            "Ngày hiệu lực của biểu đồ do nguồn khai báo; không thay thế loại ngày vận hành",
        ),
        (
            "demand_dataset_id",
            "MVP-01-DEMAND-SAMPLE",
            OPTIONAL_LABEL,
            "Mã bộ dữ liệu do người dùng khai báo; không tự tạo mã giả",
        ),
        (
            "demand_source_type",
            "manual_count",
            CONDITIONAL_DEMAND_LABEL,
            "ticketing, manual_count, apc, survey, aggregate_report hoặc other",
        ),
        (
            "demand_confidence",
            "medium",
            CONDITIONAL_DEMAND_LABEL,
            "Chất lượng bằng chứng do người dùng khai báo: high, medium, low hoặc unknown",
        ),
        (
            "demand_response_mode",
            "static",
            CONDITIONAL_DEMAND_LABEL,
            "static, elasticity_scenario hoặc calibrated",
        ),
        (
            "source_notes",
            "Dữ liệu minh họa, không phải giá trị vận hành đã phê duyệt.",
            OPTIONAL_LABEL,
            "Ghi chú nguồn; hệ thống không diễn giải thành quy tắc vận hành",
        ),
    ]
    for row_index, row in enumerate(rows, 4):
        for column, value in enumerate(row, 1):
            ws.cell(row_index, column, value)
    _body_style(ws, 4, 3 + len(rows), 4)
    _style_requirement_rows(ws, 4, 3 + len(rows), 3)

    source_type_validation = DataValidation(
        type="list",
        formula1='"ticketing,manual_count,apc,survey,aggregate_report,other"',
        allow_blank=True,
    )
    confidence_validation = DataValidation(
        type="list", formula1='"high,medium,low,unknown"', allow_blank=True
    )
    response_validation = DataValidation(
        type="list",
        formula1='"static,elasticity_scenario,calibrated"',
        allow_blank=True,
    )
    timetable_status_validation = DataValidation(
        type="list",
        formula1='"approved_operational,current_operational,proposed,unknown"',
        allow_blank=True,
    )
    ws.add_data_validation(timetable_status_validation)
    ws.add_data_validation(source_type_validation)
    ws.add_data_validation(confidence_validation)
    ws.add_data_validation(response_validation)
    key_rows = {ws.cell(row, 1).value: row for row in range(4, 4 + len(rows))}
    timetable_status_validation.add(ws.cell(key_rows["timetable_authority_status"], 2))
    ws.cell(key_rows["timetable_effective_date"], 2).number_format = "DD/MM/YYYY"
    source_type_validation.add(ws.cell(key_rows["demand_source_type"], 2))
    confidence_validation.add(ws.cell(key_rows["demand_confidence"], 2))
    response_validation.add(ws.cell(key_rows["demand_response_mode"], 2))
    ws.freeze_panes = "A4"
    _autowidth(ws, 58)


def _write_trip_ridership_metadata_sheet(ws) -> None:
    _title(ws, "THÔNG TIN BỘ DỮ LIỆU SẢN LƯỢNG THEO CHUYẾN", 4)
    _table_header(ws, 3, ["Tham số", "Giá trị", "Mức độ", "Diễn giải"])
    rows = [
        (
            "trip_ridership_dataset_id",
            "MVP-01-TRIP-RIDERSHIP-SAMPLE",
            CONDITIONAL_TRIP_RIDERSHIP_LABEL,
            "Mã bộ dữ liệu không được để trống; một bộ chỉ chứa một loại ngày vận hành",
        ),
        (
            "trip_ridership_source_type",
            "manual_count",
            CONDITIONAL_TRIP_RIDERSHIP_LABEL,
            "ticketing, manual_count, apc, survey hoặc other; không dùng báo cáo tuyến tổng hợp",
        ),
        (
            "trip_ridership_confidence",
            "medium",
            CONDITIONAL_TRIP_RIDERSHIP_LABEL,
            "high, medium, low hoặc unknown",
        ),
        (
            "observed_schedule_scenario",
            "B",
            CONDITIONAL_TRIP_RIDERSHIP_LABEL,
            "Milestone 6A1 chỉ cho phép đối chiếu với Scenario B",
        ),
        (
            "operating_day_type",
            "weekday",
            CONDITIONAL_TRIP_RIDERSHIP_LABEL,
            "Phải trùng THONG_SO_B; không suy từ service_date",
        ),
        (
            "match_tolerance_minutes",
            5,
            CONDITIONAL_TRIP_RIDERSHIP_LABEL,
            "Số nguyên từ 0 đến 30, tính cả đúng biên dung sai",
        ),
        (
            "source_notes",
            "Dữ liệu minh họa bổ sung; không phải thẩm quyền tối ưu trong 6A1.",
            OPTIONAL_LABEL,
            "Ghi chú nguồn tự do; không dùng trong fingerprint phân tích",
        ),
    ]
    for row_index, row in enumerate(rows, 4):
        for column, value in enumerate(row, 1):
            ws.cell(row_index, column, value)
    _body_style(ws, 4, 3 + len(rows), 4)
    _style_requirement_rows(ws, 4, 3 + len(rows), 3)

    source_type_validation = DataValidation(
        type="list",
        formula1='"ticketing,manual_count,apc,survey,other"',
        allow_blank=False,
    )
    confidence_validation = DataValidation(
        type="list", formula1='"high,medium,low,unknown"', allow_blank=False
    )
    scenario_validation = DataValidation(type="list", formula1='"B"', allow_blank=False)
    operating_day_validation = DataValidation(
        type="list",
        formula1='"weekday,saturday,sunday,holiday,special"',
        allow_blank=False,
    )
    tolerance_validation = DataValidation(
        type="whole",
        operator="between",
        formula1="0",
        formula2="30",
        allow_blank=False,
    )
    for validation in (
        source_type_validation,
        confidence_validation,
        scenario_validation,
        operating_day_validation,
        tolerance_validation,
    ):
        ws.add_data_validation(validation)
    key_rows = {ws.cell(row, 1).value: row for row in range(4, 4 + len(rows))}
    source_type_validation.add(ws.cell(key_rows["trip_ridership_source_type"], 2))
    confidence_validation.add(ws.cell(key_rows["trip_ridership_confidence"], 2))
    scenario_validation.add(ws.cell(key_rows["observed_schedule_scenario"], 2))
    operating_day_validation.add(ws.cell(key_rows["operating_day_type"], 2))
    tolerance_validation.add(ws.cell(key_rows["match_tolerance_minutes"], 2))
    ws.freeze_panes = "A4"
    _autowidth(ws, 62)


def _write_trip_ridership_sheet(ws, sample_trip: Trip) -> None:
    headers = [
        "observation_id",
        "service_date",
        "source_trip_id",
        "scheduled_trip_id",
        "direction",
        "scheduled_departure_time",
        "actual_departure_time",
        "passenger_count",
        "vehicle_id",
        "notes",
    ]
    _title(ws, "SẢN LƯỢNG HÀNH KHÁCH THEO TỪNG CHUYẾN (BỔ SUNG 6A1)", len(headers))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sample_note = ws.cell(
        2,
        1,
        "DÒNG MẪU (không được import): SAMPLE-TRIP-001 | 2026-07-01 | "
        f"{sample_trip.trip_id} | outbound | "
        f"{_as_time(sample_trip.departure_seconds).strftime('%H:%M')} | "
        f"{_as_time(sample_trip.departure_seconds + 3 * 60).strftime('%H:%M')} | 42. "
        "Nhập dữ liệu thật từ dòng 4.",
    )
    sample_note.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    sample_note.font = Font(name="Aptos", size=10, italic=True, color=TEXT)
    sample_note.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 34
    _table_header(ws, 3, headers)

    direction_validation = DataValidation(
        type="list", formula1='"outbound,inbound"', allow_blank=False
    )
    passenger_validation = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=False,
    )
    ws.add_data_validation(direction_validation)
    ws.add_data_validation(passenger_validation)
    direction_validation.add("E4:E1003")
    passenger_validation.add("H4:H1003")
    ws.auto_filter.ref = "A3:J3"
    ws.freeze_panes = "A4"
    _autowidth(ws, 48)


def _write_trip_input_sheet(ws, trips: list[Trip], parameters: ScenarioParameters) -> None:
    headers = [
        "scenario",
        "trip_id",
        "departure_terminal",
        "direction",
        "departure_time",
        "arrival_time",
        "vehicle_id",
        "vehicle_capacity_override",
    ]
    _title(ws, f"BIỂU ĐỒ SCENARIO {trips[0].scenario}", len(headers))
    _table_header(ws, 3, headers)
    for row_index, trip in enumerate(trips, 4):
        values = [
            trip.scenario,
            trip.trip_id,
            trip.departure_terminal,
            trip.direction.value,
            _as_time(trip.departure_seconds),
            _as_time(trip.resolved_arrival_seconds(parameters.default_trip_runtime_minutes)),
            None,
            None,
        ]
        for column, value in enumerate(values, 1):
            ws.cell(row_index, column, value)
        ws.cell(row_index, 5).number_format = "HH:mm"
        ws.cell(row_index, 6).number_format = "HH:mm"
    end_row = max(4, 3 + len(trips))
    _body_style(ws, 4, end_row, len(headers))
    direction_validation = DataValidation(type="list", formula1='"terminal_1_to_2,terminal_2_to_1"')
    terminal_validation = DataValidation(
        type="list",
        formula1=f'"{parameters.terminal_1_name},{parameters.terminal_2_name}"',
    )
    positive_integer = DataValidation(type="whole", operator="greaterThan", formula1="0")
    ws.add_data_validation(direction_validation)
    ws.add_data_validation(terminal_validation)
    ws.add_data_validation(positive_integer)
    direction_validation.add("D4:D1003")
    terminal_validation.add("C4:C1003")
    positive_integer.add("H4:H1003")
    ws.auto_filter.ref = f"A3:H{end_row}"
    ws.freeze_panes = "A4"
    _autowidth(ws)


def create_input_template(path: str | Path) -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    parameters_a, parameters_b = _sample_parameters()
    trips_a = _sample_trips("A", parameters_a)
    trips_b = _sample_trips("B", parameters_b, proposed=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    guide = workbook.create_sheet("HUONG_DAN")
    _title(guide, "HƯỚNG DẪN NHẬP DỮ LIỆU", 3)
    _table_header(guide, 3, ["Chủ đề", "Quy tắc", "Thao tác"])
    guide_rows = [
        (
            "Ba mức yêu cầu",
            "BẮT BUỘC để import; các trường cho đánh giá nhu cầu/tối ưu có thể để trống khi review biểu đồ; TÙY CHỌN không chặn review.",
            "Đọc nhãn Mức độ, không chỉ dựa vào màu ô.",
        ),
        (
            "Review biểu đồ trước tối ưu",
            "Thiếu sức chứa hoặc thẩm quyền tối ưu không chặn kiểm tra số chuyến, bến, thời gian, runtime, headway và chuỗi xe nguồn.",
            "Chạy data_authority_review trước; chỉ chạy real_route_review khi OPTIMIZATION sẵn sàng.",
        ),
        (
            "Thẩm quyền biểu đồ",
            "Chỉ timetable_authority_status = approved_operational mới được báo cáo là nguồn đã phê duyệt.",
            "Review kỹ thuật không cấp hoặc thu hồi phê duyệt bên ngoài; proposed/unknown không được tự nâng cấp.",
        ),
        (
            "Giới hạn đội xe",
            "Để trống available_fleet_limit vẫn import được nhưng chặn tối ưu authoritative.",
            "Khai báo đúng giới hạn cứng; đây không phải đội xe tối thiểu được tính toán.",
        ),
        (
            "Đội xe phê duyệt",
            "approved_active_fleet là siêu dữ liệu quản trị tùy chọn.",
            "Không dùng trường này thay available_fleet_limit.",
        ),
        (
            "Sức chứa bến",
            "Hai giới hạn bến đều tùy chọn và có thể để trống độc lập.",
            "Có một giới hạn thì chỉ đánh giá một phần; để trống cả hai thì không đánh giá sức chứa bến.",
        ),
        (
            "Thẩm quyền nhu cầu",
            "Khi SAN_LUONG có quan sát, nguồn, độ tin cậy và chế độ phản hồi là bắt buộc để tối ưu.",
            "demand_confidence là chất lượng bằng chứng do người dùng khai báo; hệ thống không tự nâng cấp.",
        ),
        (
            "Nhu cầu gộp",
            "direction = combined chỉ cho phép kết luận tổng hợp.",
            "Hệ thống không tự chia sản lượng gộp thành hai chiều.",
        ),
        (
            "Nguồn runtime",
            "source_id, imported_at và source_type do ứng dụng cung cấp, không nằm trong workbook.",
            "Không suy source_id từ mã tuyến hoặc tên tệp.",
        ),
        (
            "Scenario A",
            "Toàn bộ Scenario A là tùy chọn; chế độ chỉ có Scenario B vẫn được hỗ trợ.",
            "Nếu có A, fleet limit và operating day của A phải đủ trước tối ưu authoritative.",
        ),
        (
            "Runtime chuyến",
            "Phải khai báo allowed_trip_runtime_minutes hoặc trip_runtime_minutes; allowed_trip_runtime_minutes là định dạng được ưu tiên.",
            "trip_runtime_minutes chỉ dùng để tương thích file cũ; nếu Excel dùng dấu phẩy thập phân, nhập 55;65.",
        ),
        (
            "Dữ liệu mẫu",
            "Workbook có dữ liệu minh họa đầy đủ để chạy pipeline thống nhất.",
            "Thay thế các dòng mẫu bằng dữ liệu thật.",
        ),
        (
            "Sản lượng theo chuyến 6A1",
            "SAN_LUONG_CHUYEN là dữ liệu mô tả bổ sung đối chiếu Scenario B; chưa được dùng để sinh phương án C.",
            "Khai báo siêu dữ liệu riêng trong THONG_TIN_SAN_LUONG_CHUYEN và thay dòng SAMPLE bằng dữ liệu thực tế.",
        ),
        (
            "Quan sát thiếu",
            "Thiếu quan sát cho một chuyến-ngày không được hiểu là 0 hành khách.",
            "Không tự điền 0 và không ngoại suy các chuyến-ngày chưa được quan sát.",
        ),
        (
            "Ghép không an toàn",
            "Bản ghi mơ hồ, va chạm cùng chuyến-ngày, không ghép được hoặc mâu thuẫn bị loại khỏi thống kê hành khách.",
            "Tách bộ dữ liệu theo operating_day_type; dùng outbound/inbound và dung sai 0–30 phút.",
        ),
    ]
    for row_index, row in enumerate(guide_rows, 4):
        for column, value in enumerate(row, 1):
            guide.cell(row_index, column, value)
    _body_style(guide, 4, 3 + len(guide_rows), 3)
    guide.freeze_panes = "A4"
    _autowidth(guide, 58)

    _write_parameter_sheet(workbook.create_sheet("THONG_SO_A"), parameters_a, "A")
    _write_trip_input_sheet(workbook.create_sheet("BIEU_DO_A"), trips_a, parameters_a)
    _write_parameter_sheet(workbook.create_sheet("THONG_SO_B"), parameters_b, "B")
    _write_trip_input_sheet(workbook.create_sheet("BIEU_DO_B"), trips_b, parameters_b)
    _write_authority_metadata_sheet(workbook.create_sheet("THONG_TIN_DU_LIEU"))
    _write_trip_ridership_metadata_sheet(workbook.create_sheet("THONG_TIN_SAN_LUONG_CHUYEN"))
    _write_trip_ridership_sheet(
        workbook.create_sheet("SAN_LUONG_CHUYEN"),
        trips_b[0],
    )

    demand_sheet = workbook.create_sheet("SAN_LUONG")
    demand_headers = [
        "period_start",
        "period_end",
        "observation_days",
        "time_block_start",
        "time_block_end",
        "direction",
        "passenger_volume",
        "volume_type",
    ]
    _title(demand_sheet, "SẢN LƯỢNG HÀNH KHÁCH", len(demand_headers))
    _table_header(demand_sheet, 3, demand_headers)
    daily_profile = [40, 105, 95, 45, 35, 30, 35, 40, 50, 90, 105, 55, 25]
    row_index = 4
    for direction_index, direction in enumerate(
        (Direction.TERMINAL_1_TO_2, Direction.TERMINAL_2_TO_1)
    ):
        for block_index, daily_volume in enumerate(daily_profile):
            start_seconds = (6 + block_index) * 3600
            adjusted_daily = daily_volume * (0.92 if direction_index else 1)
            row = [
                date(2026, 7, 1),
                date(2026, 7, 15),
                15,
                _as_time(start_seconds),
                _as_time(start_seconds + 3600),
                direction.value,
                round(adjusted_daily * 15),
                "total_observation_period",
            ]
            for column, value in enumerate(row, 1):
                demand_sheet.cell(row_index, column, value)
            for column in (1, 2):
                demand_sheet.cell(row_index, column).number_format = "dd/mm/yyyy"
            for column in (4, 5):
                demand_sheet.cell(row_index, column).number_format = "HH:mm"
            row_index += 1
    _body_style(demand_sheet, 4, row_index - 1, len(demand_headers))
    direction_validation = DataValidation(
        type="list", formula1='"terminal_1_to_2,terminal_2_to_1,combined"'
    )
    volume_validation = DataValidation(
        type="list", formula1='"total_observation_period,average_day"'
    )
    positive_integer = DataValidation(type="whole", operator="greaterThan", formula1="0")
    demand_sheet.add_data_validation(direction_validation)
    demand_sheet.add_data_validation(volume_validation)
    demand_sheet.add_data_validation(positive_integer)
    direction_validation.add("F4:F1003")
    volume_validation.add("H4:H1003")
    positive_integer.add("C4:C1003")
    demand_sheet.auto_filter.ref = f"A3:H{row_index - 1}"
    demand_sheet.freeze_panes = "A4"
    _autowidth(demand_sheet)

    config = workbook.create_sheet("CAU_HINH")
    _title(config, "CẤU HÌNH MVP", 3)
    _table_header(config, 3, ["Tham số", "Giá trị", "Ghi chú"])
    config_rows = [
        ("final_service_block_minutes", 90, "Block cuối ngày"),
        ("diagram_time_direction", "left_to_right", "Giờ tăng từ trái sang phải trên trục X"),
        ("combined_direction_policy", "do_not_infer", "Không tự suy đoán theo chiều"),
        ("generator_mode", "deterministic", "Không dùng số ngẫu nhiên/AI"),
        ("scoring_config", "config/scoring.json", "Trọng số tập trung"),
        ("direction_trip_lock_mode", "fixed_by_direction", "Giữ tổng chuyến từng chiều của B"),
        ("lock_first_departures", True, "Khóa chuyến đầu mỗi bến"),
        ("lock_last_departures", True, "Khóa chuyến cuối mỗi bến"),
        ("headway_rounding_tolerance_minutes", 1, "Balanced rounding trong regime"),
        ("minimum_departures_per_normal_regime", 3, "Quy mô regime bình thường"),
        ("minimum_regime_duration_minutes", 60, "Thời lượng regime bình thường"),
        ("maximum_headway_regimes_per_direction", 6, "Giới hạn regime mỗi chiều"),
        ("maximum_transition_headways_per_boundary", 1, "Một headway chuyển tiếp mỗi ranh giới"),
        ("maximum_transition_deviation_minutes", 5, "Dung sai headway chuyển tiếp"),
        ("minimum_sustained_change_intervals", 2, "Không phản ứng với nhiễu một block"),
        ("minimum_material_headway_change_minutes", 5, "Biến đổi tần suất đáng kể"),
        ("minimum_material_service_rate_change_ratio", 0.15, "Biến đổi suất phục vụ đáng kể"),
        ("preferred_max_shift_per_trip_minutes", 15, "Mức dịch chuyển ưu tiên"),
        ("absolute_max_shift_per_trip_minutes", 30, "Giới hạn dịch chuyển tuyệt đối"),
        (
            "protected_service_floor_maximum_protected_b_headway_minutes",
            30,
            "6A2A: giới hạn headway B được đề xuất bảo vệ",
        ),
        (
            "protected_service_floor_headway_rounding_tolerance_minutes",
            1,
            "6A2A: dung sai balanced rounding",
        ),
        (
            "protected_service_floor_minimum_departures_per_regime",
            3,
            "6A2A: số chuyến tối thiểu trong regime",
        ),
        (
            "protected_service_floor_minimum_regime_duration_minutes",
            30,
            "6A2A: thời lượng regime tối thiểu",
        ),
        (
            "protected_service_floor_minimum_observed_days_per_trip",
            3,
            "6A2A: số ngày quan sát tối thiểu mỗi chuyến",
        ),
        (
            "protected_service_floor_minimum_regime_trip_coverage_rate",
            0.80,
            "6A2A: tỷ lệ chuyến regime đủ bao phủ",
        ),
        (
            "protected_service_floor_minimum_high_load_trip_share",
            0.67,
            "6A2A: tỷ lệ chuyến P85 đạt target",
        ),
        (
            "protected_service_floor_protected_load_statistic",
            "P85",
            "6A2A: thống kê tải được hỗ trợ",
        ),
        (
            "protected_service_floor_minimum_trip_ridership_confidence",
            "medium",
            "6A2A: độ tin cậy tối thiểu",
        ),
        (
            "protected_service_floor_future_service_window_boundary_tolerance_minutes",
            0,
            "6A2A preview: chưa áp dụng trong Scenario C",
        ),
        ("configuration_version", "scenario_c_regimes_v1", "Phiên bản cấu hình C"),
    ]
    for row_number, row in enumerate(config_rows, 4):
        for column, value in enumerate(row, 1):
            config.cell(row_number, column, value)
    _body_style(config, 4, 3 + len(config_rows), 3)
    _autowidth(config)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(output_path)
    return output_path
