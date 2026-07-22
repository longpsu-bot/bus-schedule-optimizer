from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .block_supply import SupplyPresentationStatus, build_block_supply_comparison
from .c_config import ScenarioCConfig
from .fingerprint import timetable_fingerprint
from .models import (
    AnalysisBundle,
    Direction,
    HeadwayType,
    ScenarioCStatus,
    ScenarioResult,
)
from .time_utils import block_label

NAVY = "17365D"
TEAL = "0F766E"
LIGHT = "E8F1F5"
RED = "FECACA"
AMBER = "FDE68A"
GREEN = "D1FAE5"
MUTED = "475569"


def _title(ws, text: str, columns: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    cell = ws.cell(1, 1, text)
    cell.font = Font(name="Aptos Display", size=16, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    ws.sheet_view.showGridLines = False


def _header(ws, row: int, headers: list[str]) -> None:
    for column, value in enumerate(headers, 1):
        cell = ws.cell(row, column, value)
        cell.font = Font(name="Aptos", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 34


def _body(ws, start: int, end: int, columns: int) -> None:
    if end < start:
        return
    for row in ws.iter_rows(min_row=start, max_row=end, max_col=columns):
        for cell in row:
            cell.font = Font(name="Aptos", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")


def _autowidth(ws, maximum: int = 44) -> None:
    for column_index, column_cells in enumerate(ws.columns, 1):
        letter = get_column_letter(column_index)
        width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[letter].width = min(maximum, max(10, width + 2))


def _time(seconds: int | None) -> float | None:
    return None if seconds is None else seconds / 86400


def _direction_name(result: ScenarioResult, direction: Direction) -> str:
    if direction == Direction.TERMINAL_1_TO_2:
        return f"{result.parameters.terminal_1_name} → {result.parameters.terminal_2_name}"
    if direction == Direction.TERMINAL_2_TO_1:
        return f"{result.parameters.terminal_2_name} → {result.parameters.terminal_1_name}"
    return "Tổng hai chiều"


def _block_for_trip(result: ScenarioResult, departure: int, direction: Direction):
    return next(
        (
            block
            for block in result.evaluation.blocks
            if block.direction in {direction, Direction.COMBINED}
            and block.block_start_seconds <= departure < block.block_end_seconds
        ),
        None,
    )


def _counts_by_direction(result: ScenarioResult) -> tuple[int, int]:
    return (
        sum(trip.direction == Direction.TERMINAL_1_TO_2 for trip in result.trips),
        sum(trip.direction == Direction.TERMINAL_2_TO_1 for trip in result.trips),
    )


def _write_rows(ws, start_row: int, rows: list[list[object]]) -> int:
    for row_number, values in enumerate(rows, start_row):
        for column, value in enumerate(values, 1):
            ws.cell(row_number, column, value)
    return start_row + len(rows) - 1


def _assert_consistency(result_b: ScenarioResult, result_c: ScenarioResult) -> None:
    current_c = timetable_fingerprint(result_c.trips)
    current_b = timetable_fingerprint(result_b.trips)
    if current_c != result_c.timetable_fingerprint:
        raise ValueError("Fingerprint C không khớp object dùng để export")
    if current_b != result_c.source_timetable_fingerprint:
        raise ValueError("Fingerprint nguồn B của C không khớp timetable B")
    if result_b.trips is result_c.trips:
        raise ValueError("Từ chối export vì B và C dùng chung list chuyến")
    if len(result_b.trips) != len(result_c.trips):
        raise ValueError("Từ chối export vì C không giữ tổng chuyến B")
    b_times = sorted((trip.direction.value, trip.departure_seconds) for trip in result_b.trips)
    c_times = sorted((trip.direction.value, trip.departure_seconds) for trip in result_c.trips)
    unchanged_statuses = {
        ScenarioCStatus.NO_BETTER_REDISTRIBUTION,
        ScenarioCStatus.INFEASIBLE_FIXED_RESOURCES,
        ScenarioCStatus.INSUFFICIENT_DATA,
    }
    if b_times == c_times and result_c.generation_status not in unchanged_statuses:
        raise ValueError("Từ chối export C trùng B khi không có trạng thái no-improvement rõ ràng")
    if len({trace.source_b_trip_id for trace in result_c.trip_traces}) != len(result_b.trips):
        raise ValueError("Từ chối export vì ánh xạ B → C không phải một-một")
    source_by_id = {trip.trip_id: trip for trip in result_b.trips}
    default_runtime = result_b.parameters.default_trip_runtime_minutes
    for trip_c in result_c.trips:
        trip_b = source_by_id.get(trip_c.source_b_trip_id)
        if trip_b is None:
            raise ValueError("Từ chối export vì C thiếu truy vết chuyến nguồn B")
        b_runtime = trip_b.resolved_arrival_seconds(default_runtime) - trip_b.departure_seconds
        c_runtime = trip_c.resolved_arrival_seconds(default_runtime) - trip_c.departure_seconds
        if c_runtime != b_runtime:
            raise ValueError(
                f"Từ chối export vì C đổi thời gian hành trình chuyến nguồn {trip_b.trip_id}"
            )


def export_bc_comparison(bundle: AnalysisBundle, path: str | Path) -> Path:
    result_b = bundle.get("B")
    result_c = bundle.get("C")
    if result_b is None or result_c is None:
        raise ValueError("Cần có cả Scenario B và Scenario C để xuất workbook so sánh")
    _assert_consistency(result_b, result_c)
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    _overview_sheet(workbook, result_b, result_c, bundle)
    _b_schedule_sheet(workbook, result_b)
    _c_schedule_sheet(workbook, result_c)
    _trip_comparison_sheet(workbook, result_c)
    _demand_comparison_sheet(workbook, result_b, result_c)
    _block_supply_sheet(workbook, bundle)
    _regime_sheet(workbook, result_c)
    _adjacent_headway_sheet(workbook, result_c)
    _fleet_sheet(workbook, result_c)
    _warning_sheet(workbook, result_b, result_c)
    _optimization_log_sheet(workbook, result_c)
    _config_sheet(workbook, result_c)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.save(output_path)
    return output_path


def _overview_sheet(
    workbook: Workbook,
    result_b: ScenarioResult,
    result_c: ScenarioResult,
    bundle: AnalysisBundle,
) -> None:
    ws = workbook.create_sheet("TONG_QUAN_B_C")
    _title(ws, "TỔNG QUAN SO SÁNH B VÀ C", 4)
    _header(ws, 3, ["Chỉ tiêu", "Scenario B", "Scenario C", "Ghi chú"])
    b_d1, b_d2 = _counts_by_direction(result_b)
    c_d1, c_d2 = _counts_by_direction(result_c)
    b_regularity = result_b.regularity
    c_regularity = result_c.regularity
    shifted = [trace for trace in result_c.trip_traces if trace.shift_minutes != 0]
    rows = [
        ["Mã tuyến", result_b.parameters.route_id, result_c.parameters.route_id, "Khóa từ B"],
        ["Tên tuyến", result_b.parameters.route_name, result_c.parameters.route_name, "Khóa từ B"],
        [
            "Trạng thái kết quả",
            "Baseline",
            result_c.generation_status.value if result_c.generation_status else "—",
            "",
        ],
        ["Tổng chuyến", len(result_b.trips), len(result_c.trips), "Phải bằng nhau"],
        ["Chuyến chiều 1", b_d1, c_d1, "fixed_by_direction"],
        ["Chuyến chiều 2", b_d2, c_d2, "fixed_by_direction"],
        [
            "Số xe hoạt động",
            result_b.active_vehicle_count,
            result_c.active_vehicle_count,
            "Phải bằng nhau",
        ],
        [
            "Số xe tối thiểu",
            result_b.fleet.minimum_vehicles,
            result_c.fleet.minimum_vehicles,
            "C không vượt số xe hoạt động B",
        ],
        ["Sức chứa xe", result_b.parameters.capacity, result_c.parameters.capacity, "Khóa từ B"],
        [
            "Thời gian hành trình cho phép (phút)",
            result_b.parameters.runtime_range_text,
            result_c.parameters.runtime_range_text,
            "Khóa từ B",
        ],
        [
            "Hệ số tải mục tiêu",
            result_b.parameters.target_load_factor,
            result_c.parameters.target_load_factor,
            "",
        ],
        [
            "Hệ số tải tối đa",
            result_b.parameters.maximum_load_factor,
            result_c.parameters.maximum_load_factor,
            "",
        ],
        [
            "Hệ số tải thực tế cao nhất",
            result_b.evaluation.maximum_load_factor,
            result_c.evaluation.maximum_load_factor,
            "",
        ],
        [
            "Khung trên 85%",
            result_b.evaluation.blocks_over_target,
            result_c.evaluation.blocks_over_target,
            "",
        ],
        [
            "Khung trên 90%",
            result_b.evaluation.blocks_over_maximum,
            result_c.evaluation.blocks_over_maximum,
            "",
        ],
        [
            "Khoảng cách phục vụ lớn nhất (phút)",
            b_regularity.maximum_service_gap if b_regularity else None,
            c_regularity.maximum_service_gap if c_regularity else None,
            "Timeline liên tục",
        ],
        [
            "Số chế độ giãn cách",
            b_regularity.number_of_headway_regimes if b_regularity else None,
            c_regularity.number_of_headway_regimes if c_regularity else None,
            "Hai chiều cộng lại",
        ],
        [
            "Giãn cách ngoại lệ",
            b_regularity.number_of_exceptional_headways if b_regularity else None,
            c_regularity.number_of_exceptional_headways if c_regularity else None,
            "",
        ],
        ["Số chuyến dịch chuyển", 0, len(shifted), ""],
        ["Tổng phút dịch chuyển", 0, sum(abs(trace.shift_minutes) for trace in shifted), ""],
        ["Kết luận", "Biểu đồ người dùng đề xuất", result_c.recommendation_reason, ""],
        ["Giới hạn còn lại", "", "; ".join(bundle.limitations) or "Không có", ""],
        [
            "Fingerprint B",
            result_b.timetable_fingerprint,
            result_c.source_timetable_fingerprint,
            "Phải trùng",
        ],
        ["Fingerprint C", "—", result_c.timetable_fingerprint, "UI = diagram = XLSX"],
    ]
    end = _write_rows(ws, 4, rows)
    for row in range(4, end + 1):
        if ws.cell(row, 1).value in {
            "Hệ số tải mục tiêu",
            "Hệ số tải tối đa",
            "Hệ số tải thực tế cao nhất",
        }:
            ws.cell(row, 2).number_format = "0.0%"
            ws.cell(row, 3).number_format = "0.0%"
    _body(ws, 4, end, 4)
    ws.freeze_panes = "A4"
    _autowidth(ws, 64)


def _b_schedule_sheet(workbook: Workbook, result: ScenarioResult) -> None:
    ws = workbook.create_sheet("B_BIEU_DO_GIO")
    headers = [
        "STT",
        "Trip ID B",
        "Chiều",
        "Bến xuất phát",
        "Giờ đi",
        "Giờ đến",
        "Xe khai báo",
        "Sức chứa",
    ]
    _title(ws, "B — BIỂU ĐỒ GIỜ ĐỀ XUẤT (BẤT BIẾN)", len(headers))
    _header(ws, 3, headers)
    rows = []
    for sequence, trip in enumerate(
        sorted(result.trips, key=lambda item: (item.departure_seconds, item.trip_id)), 1
    ):
        rows.append(
            [
                sequence,
                trip.trip_id,
                _direction_name(result, trip.direction),
                trip.departure_terminal,
                _time(trip.departure_seconds),
                _time(
                    trip.resolved_arrival_seconds(result.parameters.default_trip_runtime_minutes)
                ),
                trip.vehicle_id or "—",
                trip.vehicle_capacity_override or result.parameters.capacity,
            ]
        )
    end = _write_rows(ws, 4, rows)
    for row in range(4, end + 1):
        ws.cell(row, 5).number_format = "HH:mm"
        ws.cell(row, 6).number_format = "HH:mm"
    _body(ws, 4, end, len(headers))
    ws.auto_filter.ref = f"A3:H{end}"
    ws.freeze_panes = "A4"
    _autowidth(ws)


def _c_schedule_sheet(workbook: Workbook, result: ScenarioResult) -> None:
    ws = workbook.create_sheet("C_BIEU_DO_GIO")
    headers = [
        "STT",
        "Trip ID C",
        "Trip nguồn B",
        "Chiều",
        "Bến xuất phát",
        "Giờ đi",
        "Giờ đến",
        "Chuyến trước",
        "Giãn cách liền kề",
        "Regime ID",
        "Giãn cách mục tiêu",
        "Loại headway",
        "Xe gán",
        "Sức chứa",
        "Khung nhu cầu",
        "Nhu cầu",
        "Load factor",
        "Trạng thái",
    ]
    _title(ws, "C — TÁI PHÂN BỔ ỔN ĐỊNH THEO NHU CẦU", len(headers))
    _header(ws, 3, headers)
    assignment = {item.trip_id: item.vehicle_id for item in result.fleet.assignments}
    trace_by_trip = {trace.c_trip_id: trace for trace in result.trip_traces}
    regime_by_id = {regime.regime_id: regime for regime in result.headway_regimes}
    rows = []
    for direction in (Direction.TERMINAL_1_TO_2, Direction.TERMINAL_2_TO_1):
        ordered = sorted(
            (trip for trip in result.trips if trip.direction == direction),
            key=lambda item: (item.departure_seconds, item.trip_id),
        )
        for index, trip in enumerate(ordered):
            trace = trace_by_trip[trip.trip_id]
            block = _block_for_trip(result, trip.departure_seconds, direction)
            regime = regime_by_id.get(trace.headway_regime_id)
            rows.append(
                [
                    len(rows) + 1,
                    trip.trip_id,
                    trip.source_b_trip_id,
                    _direction_name(result, direction),
                    trip.departure_terminal,
                    _time(trip.departure_seconds),
                    _time(
                        trip.resolved_arrival_seconds(
                            result.parameters.default_trip_runtime_minutes
                        )
                    ),
                    ordered[index - 1].trip_id if index else "—",
                    trace.new_previous_headway,
                    trace.headway_regime_id,
                    regime.target_headway_minutes if regime else None,
                    trace.headway_type.value,
                    assignment.get(trip.trip_id, "—"),
                    trip.vehicle_capacity_override or result.parameters.capacity,
                    trace.new_demand_interval,
                    block.demand if block else None,
                    block.load_factor if block else None,
                    block.status.value if block else "Ngoài khung nhu cầu",
                ]
            )
    end = _write_rows(ws, 4, rows)
    for row in range(4, end + 1):
        for column in (6, 7):
            ws.cell(row, column).number_format = "HH:mm"
        ws.cell(row, 17).number_format = "0.0%"
    _body(ws, 4, end, len(headers))
    ws.auto_filter.ref = f"A3:R{end}"
    ws.freeze_panes = "A4"
    _autowidth(ws, 38)


def _trip_comparison_sheet(workbook: Workbook, result: ScenarioResult) -> None:
    ws = workbook.create_sheet("SO_SANH_CHUYEN_B_C")
    headers = [
        "Trip ID C",
        "Trip nguồn B",
        "Chiều",
        "Bến xuất phát",
        "Giờ B",
        "Giờ C",
        "Dịch chuyển (phút)",
        "Giữ/Dịch",
        "Headway trước B",
        "Headway trước C",
        "Headway sau B",
        "Headway sau C",
        "Khung nhu cầu B",
        "Khung nhu cầu C",
        "Regime ID",
        "Loại headway",
        "Lý do thay đổi",
        "Lý do ngoại lệ",
    ]
    _title(ws, "TRUY VẾT TỪNG CHUYẾN B → C", len(headers))
    _header(ws, 3, headers)
    rows = [
        [
            trace.c_trip_id,
            trace.source_b_trip_id,
            _direction_name(result, trace.direction),
            trace.departure_terminal,
            _time(trace.b_departure_seconds),
            _time(trace.c_departure_seconds),
            trace.shift_minutes,
            trace.retained_or_shifted,
            trace.original_previous_headway,
            trace.new_previous_headway,
            trace.original_next_headway,
            trace.new_next_headway,
            trace.original_demand_interval,
            trace.new_demand_interval,
            trace.headway_regime_id,
            trace.headway_type.value,
            trace.change_reason,
            trace.exception_reason,
        ]
        for trace in result.trip_traces
    ]
    end = _write_rows(ws, 4, rows)
    for row in range(4, end + 1):
        ws.cell(row, 5).number_format = "HH:mm"
        ws.cell(row, 6).number_format = "HH:mm"
        ws.cell(row, 7).number_format = "+0.0;-0.0;0.0"
    _body(ws, 4, end, len(headers))
    ws.auto_filter.ref = f"A3:R{end}"
    ws.freeze_panes = "A4"
    _autowidth(ws, 34)


def _demand_comparison_sheet(
    workbook: Workbook, result_b: ScenarioResult, result_c: ScenarioResult
) -> None:
    ws = workbook.create_sheet("SO_SANH_NHU_CAU_B_C")
    headers = [
        "Bắt đầu",
        "Kết thúc",
        "Chiều",
        "Nhu cầu",
        "Chuyến B",
        "Chuyến C",
        "Sức chứa B",
        "Sức chứa C",
        "Sức chứa 85% B",
        "Sức chứa 85% C",
        "Sức chứa 90% B",
        "Sức chứa 90% C",
        "Load factor B",
        "Load factor C",
        "Chuyến cần 85%",
        "Thiếu còn lại",
        "Cảnh báo B",
        "Cảnh báo C",
    ]
    _title(ws, "SO SÁNH NHU CẦU VÀ NĂNG LỰC B–C", len(headers))
    _header(ws, 3, headers)
    c_by_key = {
        (block.block_start_seconds, block.block_end_seconds, block.direction): block
        for block in result_c.evaluation.blocks
    }
    rows = []
    for block_b in result_b.evaluation.blocks:
        block_c = c_by_key[
            (block_b.block_start_seconds, block_b.block_end_seconds, block_b.direction)
        ]
        rows.append(
            [
                _time(block_b.block_start_seconds),
                _time(block_b.block_end_seconds),
                _direction_name(result_b, block_b.direction),
                block_b.demand,
                block_b.trips,
                block_c.trips,
                block_b.nominal_capacity,
                block_c.nominal_capacity,
                block_b.target_capacity,
                block_c.target_capacity,
                block_b.maximum_recommended_capacity,
                block_c.maximum_recommended_capacity,
                block_b.load_factor,
                block_c.load_factor,
                block_b.required_trips,
                max(0, -block_c.trip_gap_to_target),
                block_b.status.value,
                block_c.status.value,
            ]
        )
    end = _write_rows(ws, 4, rows)
    for row in range(4, end + 1):
        ws.cell(row, 1).number_format = "HH:mm"
        ws.cell(row, 2).number_format = "HH:mm"
        ws.cell(row, 13).number_format = "0.0%"
        ws.cell(row, 14).number_format = "0.0%"
    _body(ws, 4, end, len(headers))
    ws.auto_filter.ref = f"A3:R{end}"
    ws.freeze_panes = "A4"
    _autowidth(ws, 34)


def _block_supply_sheet(workbook: Workbook, bundle: AnalysisBundle) -> None:
    ws = workbook.create_sheet("CUNG_UNG_THEO_BLOCK")
    headers = [
        "Bắt đầu block",
        "Kết thúc block",
        "Chiều",
        "Nhu cầu hành khách",
        "Sức chứa phương tiện",
        "B — số chuyến",
        "C — số chuyến",
        "Số chuyến cần tại LF 85%",
        "Số chuyến tối thiểu tại LF 90%",
        "B — sức cung danh nghĩa",
        "C — sức cung danh nghĩa",
        "B — sức cung tại LF 85%",
        "C — sức cung tại LF 85%",
        "B — load factor",
        "C — load factor",
        "B — chênh lệch chuyến tới 85%",
        "C — chênh lệch chuyến tới 85%",
        "Trạng thái B",
        "Trạng thái C",
        "Mức độ tin cậy nhu cầu",
    ]
    _title(ws, "CUNG ỨNG CHUYẾN XUẤT BẾN THEO BLOCK", len(headers))
    _header(ws, 3, headers)
    result_b = bundle.get("B")
    if result_b is None:
        rows = []
    else:
        rows = [
            [
                _time(row.block_start_seconds),
                _time(row.block_end_seconds),
                _direction_name(result_b, row.direction),
                row.passenger_demand,
                row.vehicle_capacity,
                row.b_trip_count,
                row.c_trip_count,
                row.required_trips_85,
                row.minimum_trips_90,
                row.b_nominal_capacity,
                row.c_nominal_capacity,
                row.b_target_capacity,
                row.c_target_capacity,
                row.b_load_factor,
                row.c_load_factor,
                row.b_trip_gap_to_85,
                row.c_trip_gap_to_85,
                row.b_status.value,
                row.c_status.value,
                row.demand_confidence,
            ]
            for row in build_block_supply_comparison(bundle)
        ]
    end = _write_rows(ws, 4, rows)
    for row_number in range(4, end + 1):
        ws.cell(row_number, 1).number_format = "HH:mm"
        ws.cell(row_number, 2).number_format = "HH:mm"
        ws.cell(row_number, 14).number_format = "0.0%"
        ws.cell(row_number, 15).number_format = "0.0%"
        ws.cell(row_number, 16).number_format = "+0;-0;0"
        ws.cell(row_number, 17).number_format = "+0;-0;0"
    _body(ws, 4, end, len(headers))
    if end >= 4:
        status_range = f"R4:S{end}"
        critical = SupplyPresentationStatus.CRITICAL.value
        no_service = SupplyPresentationStatus.NO_SERVICE_WITH_DEMAND.value
        warning = SupplyPresentationStatus.WARNING.value
        ws.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[f'OR(R4="{critical}",R4="{no_service}")'],
                fill=PatternFill("solid", fgColor=RED),
            ),
        )
        ws.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[f'R4="{warning}"'],
                fill=PatternFill("solid", fgColor=AMBER),
            ),
        )
    ws.auto_filter.ref = f"A3:T{max(3, end)}"
    ws.freeze_panes = "A4"
    _autowidth(ws, 42)


def _regime_sheet(workbook: Workbook, result: ScenarioResult) -> None:
    ws = workbook.create_sheet("C_CHE_DO_GIAN_CACH")
    headers = [
        "Regime ID",
        "Chiều",
        "Bắt đầu",
        "Kết thúc",
        "Chuyến đầu",
        "Chuyến cuối",
        "Số chuyến",
        "Headway mục tiêu",
        "Dãy headway thực tế",
        "Min",
        "Max",
        "Mean",
        "Độ lệch chuẩn",
        "CV",
        "Lý do ranh giới",
        "Trạng thái ổn định",
    ]
    _title(ws, "C — CÁC CHẾ ĐỘ GIÃN CÁCH", len(headers))
    _header(ws, 3, headers)
    rows = [
        [
            regime.regime_id,
            _direction_name(result, regime.direction),
            _time(regime.start_seconds),
            _time(regime.end_seconds),
            regime.first_trip_id,
            regime.last_trip_id,
            regime.trip_count,
            regime.target_headway_minutes,
            ", ".join(f"{value:g}" for value in regime.actual_headway_sequence),
            regime.minimum_headway_minutes,
            regime.maximum_headway_minutes,
            regime.mean_headway_minutes,
            regime.standard_deviation_minutes,
            regime.coefficient_of_variation,
            regime.boundary_reason.value,
            regime.headway_status,
        ]
        for regime in result.headway_regimes
    ]
    end = _write_rows(ws, 4, rows)
    for row in range(4, end + 1):
        ws.cell(row, 3).number_format = "HH:mm"
        ws.cell(row, 4).number_format = "HH:mm"
        ws.cell(row, 14).number_format = "0.0%"
    _body(ws, 4, end, len(headers))
    ws.auto_filter.ref = f"A3:P{end}"
    ws.freeze_panes = "A4"
    _autowidth(ws, 42)


def _adjacent_headway_sheet(workbook: Workbook, result: ScenarioResult) -> None:
    ws = workbook.create_sheet("C_GIAN_CACH_LIEN_KE")
    headers = [
        "Chiều",
        "Chuyến trước",
        "Chuyến hiện tại",
        "Giờ trước",
        "Giờ hiện tại",
        "Headway",
        "Headway trước đó",
        "Chênh với headway trước",
        "Regime ID",
        "Loại headway",
        "Headway mục tiêu",
        "Độ lệch",
        "Cảnh báo",
        "Lý do",
    ]
    _title(ws, "C — GIÃN CÁCH LIỀN KỀ TRÊN TIMELINE LIÊN TỤC", len(headers))
    _header(ws, 3, headers)
    trace_by_trip = {trace.c_trip_id: trace for trace in result.trip_traces}
    regime_by_id = {regime.regime_id: regime for regime in result.headway_regimes}
    rows = []
    for direction in (Direction.TERMINAL_1_TO_2, Direction.TERMINAL_2_TO_1):
        ordered = sorted(
            (trip for trip in result.trips if trip.direction == direction),
            key=lambda item: (item.departure_seconds, item.trip_id),
        )
        previous_headway = None
        for previous, current in zip(ordered, ordered[1:], strict=False):
            trace = trace_by_trip[current.trip_id]
            regime = regime_by_id.get(trace.headway_regime_id)
            headway = (current.departure_seconds - previous.departure_seconds) / 60
            deviation = headway - regime.target_headway_minutes if regime else None
            warning = "" if trace.headway_type != HeadwayType.EXCEPTIONAL else "HEADWAY NGOẠI LỆ"
            rows.append(
                [
                    _direction_name(result, direction),
                    previous.trip_id,
                    current.trip_id,
                    _time(previous.departure_seconds),
                    _time(current.departure_seconds),
                    headway,
                    previous_headway,
                    None if previous_headway is None else headway - previous_headway,
                    trace.headway_regime_id,
                    trace.headway_type.value,
                    regime.target_headway_minutes if regime else None,
                    deviation,
                    warning,
                    trace.exception_reason or trace.change_reason,
                ]
            )
            previous_headway = headway
    end = _write_rows(ws, 4, rows)
    for row in range(4, end + 1):
        ws.cell(row, 4).number_format = "HH:mm"
        ws.cell(row, 5).number_format = "HH:mm"
    _body(ws, 4, end, len(headers))
    ws.auto_filter.ref = f"A3:N{end}"
    ws.freeze_panes = "A4"
    _autowidth(ws, 40)


def _fleet_sheet(workbook: Workbook, result: ScenarioResult) -> None:
    ws = workbook.create_sheet("C_PHAN_CONG_XE")
    headers = [
        "Xe",
        "Thứ tự chuyến",
        "Trip ID",
        "Bến xuất phát",
        "Giờ đi",
        "Giờ đến",
        "Sẵn sàng tiếp",
        "Quay đầu (phút)",
        "Bến tiếp theo",
        "Trạng thái khả thi",
    ]
    _title(ws, "C — PHÂN CÔNG XE", len(headers))
    _header(ws, 3, headers)
    sequence_by_vehicle: Counter[str] = Counter()
    rows = []
    for assignment in result.fleet.assignments:
        sequence_by_vehicle[assignment.vehicle_id] += 1
        rows.append(
            [
                assignment.vehicle_id,
                sequence_by_vehicle[assignment.vehicle_id],
                assignment.trip_id,
                assignment.departure_terminal,
                _time(assignment.departure_seconds),
                _time(assignment.arrival_seconds),
                _time(assignment.ready_seconds),
                result.parameters.effective_layover_minutes,
                assignment.arrival_terminal,
                "KHẢ THI",
            ]
        )
    assigned_ids = {assignment.vehicle_id for assignment in result.fleet.assignments}
    for vehicle_id in result.active_vehicle_ids:
        if vehicle_id in assigned_ids:
            continue
        rows.append(
            [
                vehicle_id,
                0,
                "—",
                "—",
                None,
                None,
                None,
                result.parameters.effective_layover_minutes,
                "—",
                "DỰ PHÒNG/NHÀN RỖI",
            ]
        )
    end = _write_rows(ws, 4, rows)
    for row in range(4, end + 1):
        for column in (5, 6, 7):
            ws.cell(row, column).number_format = "HH:mm"
    _body(ws, 4, end, len(headers))
    ws.auto_filter.ref = f"A3:J{end}"
    ws.freeze_panes = "A4"
    _autowidth(ws)


def _warning_sheet(workbook: Workbook, result_b: ScenarioResult, result_c: ScenarioResult) -> None:
    ws = workbook.create_sheet("C_CANH_BAO")
    headers = [
        "Mức độ",
        "Thời gian/Khung",
        "Chiều",
        "Loại vấn đề",
        "Giá trị B",
        "Giá trị C",
        "Giải thích",
        "Hành động còn lại",
    ]
    _title(ws, "C — CẢNH BÁO VÀ HẠN CHẾ CÒN LẠI", len(headers))
    _header(ws, 3, headers)
    rows: list[list[object]] = []
    c_by_key = {
        (block.block_start_seconds, block.block_end_seconds, block.direction): block
        for block in result_c.evaluation.blocks
    }
    for block_b in result_b.evaluation.blocks:
        block_c = c_by_key[
            (block_b.block_start_seconds, block_b.block_end_seconds, block_b.direction)
        ]
        if (
            block_c.load_factor is None
            or block_c.load_factor > result_c.parameters.target_load_factor
        ):
            rows.append(
                [
                    "ERROR"
                    if block_c.load_factor is None
                    or block_c.load_factor > result_c.parameters.maximum_load_factor
                    else "WARNING",
                    block_label(block_c.block_start_seconds, block_c.block_end_seconds),
                    _direction_name(result_c, block_c.direction),
                    "LOAD_FACTOR",
                    block_b.load_factor,
                    block_c.load_factor,
                    "C vẫn thiếu cung trong khung này."
                    if block_c.trip_gap_to_target < 0
                    else "C cần tiếp tục theo dõi.",
                    "Xem xét tăng nguồn lực ngoài phạm vi C"
                    if block_c.trip_gap_to_target < 0
                    else "Theo dõi nhu cầu thực tế",
                ]
            )
    if result_c.regularity and not result_c.regularity.gate_passed:
        rows.append(
            [
                "ERROR",
                "Toàn ngày",
                "Hai chiều",
                "REGULARITY_GATE",
                "—",
                "KHÔNG ĐẠT",
                ", ".join(result_c.regularity.gate_failures),
                "Không gắn nhãn khuyến nghị khai thác",
            ]
        )
    if not rows:
        rows.append(
            [
                "INFO",
                "Toàn ngày",
                "Hai chiều",
                "NONE",
                "—",
                "—",
                "Không còn cảnh báo nổi bật.",
                "Theo dõi vận hành",
            ]
        )
    end = _write_rows(ws, 4, rows)
    for row in range(4, end + 1):
        ws.cell(row, 5).number_format = "0.0%"
        ws.cell(row, 6).number_format = "0.0%"
    _body(ws, 4, end, len(headers))
    ws.conditional_formatting.add(
        f"A4:A{end}", FormulaRule(formula=['A4="ERROR"'], fill=PatternFill("solid", fgColor=RED))
    )
    ws.conditional_formatting.add(
        f"A4:A{end}",
        FormulaRule(formula=['A4="WARNING"'], fill=PatternFill("solid", fgColor=AMBER)),
    )
    ws.auto_filter.ref = f"A3:H{end}"
    ws.freeze_panes = "A4"
    _autowidth(ws, 52)


def _optimization_log_sheet(workbook: Workbook, result: ScenarioResult) -> None:
    ws = workbook.create_sheet("NHAT_KY_TOI_UU")
    _title(ws, "NHẬT KÝ TỐI ƯU SCENARIO C", 3)
    _header(ws, 3, ["Chỉ tiêu", "Giá trị", "Ghi chú"])
    log = result.optimization_log
    rows = [
        ["candidate_count", log.candidate_count if log else 0, ""],
        ["accepted_candidates", log.accepted_candidates if log else 0, ""],
        ["rejected_candidates", log.rejected_candidates if log else 0, ""],
        [
            "rejection_reason_counts",
            "; ".join(
                f"{key}={value}" for key, value in (log.rejection_reason_counts if log else ())
            ),
            "",
        ],
        ["objective_before", str(log.objective_before if log else ()), "Tuple ưu tiên từ B"],
        ["objective_after", str(log.objective_after if log else ()), "Tuple phương án C"],
        ["regularity_gate", log.regularity_gate_result if log else "—", ""],
        ["generation_status", log.generation_status.value if log else "—", ""],
        ["configuration_version", log.configuration_version if log else "—", ""],
        [
            "generation_timestamp",
            log.generation_timestamp if log else "—",
            "Mốc tham chiếu xác định từ dữ liệu để giữ tính lặp lại",
        ],
        ["fingerprint_C", result.timetable_fingerprint, "Phải trùng UI và diagram"],
    ]
    end = _write_rows(ws, 4, rows)
    _body(ws, 4, end, 3)
    ws.freeze_panes = "A4"
    _autowidth(ws, 68)


def _config_sheet(workbook: Workbook, result: ScenarioResult) -> None:
    ws = workbook.create_sheet("CAU_HINH_C")
    _title(ws, "CẤU HÌNH SINH SCENARIO C", 3)
    _header(ws, 3, ["Tham số", "Giá trị", "Ghi chú"])
    config = ScenarioCConfig.from_mapping(result.generation_config)
    rows = [
        ["trip_lock", "exact_total", "C = B"],
        ["direction_trip_lock_mode", config.direction_trip_lock_mode, "Mặc định khóa theo chiều"],
        ["fleet_lock", result.active_vehicle_count, "Số xe hoạt động kế thừa B"],
        ["lock_first_departures", config.lock_first_departures, ""],
        ["lock_last_departures", config.lock_last_departures, ""],
        [
            "demand_analysis_duration",
            result.parameters.time_block_minutes,
            "Chỉ dùng tổng hợp nhu cầu",
        ],
        ["maximum_headway_regimes_per_direction", config.maximum_headway_regimes_per_direction, ""],
        ["minimum_regime_duration_minutes", config.minimum_regime_duration_minutes, ""],
        ["minimum_departures_per_normal_regime", config.minimum_departures_per_normal_regime, ""],
        ["headway_rounding_tolerance_minutes", config.headway_rounding_tolerance_minutes, ""],
        ["maximum_transition_deviation_minutes", config.maximum_transition_deviation_minutes, ""],
        ["minimum_sustained_change_intervals", config.minimum_sustained_change_intervals, ""],
        [
            "minimum_material_headway_change_minutes",
            config.minimum_material_headway_change_minutes,
            "",
        ],
        [
            "minimum_material_service_rate_change_ratio",
            config.minimum_material_service_rate_change_ratio,
            "",
        ],
        ["preferred_max_shift_per_trip_minutes", config.preferred_max_shift_per_trip_minutes, ""],
        ["absolute_max_shift_per_trip_minutes", config.absolute_max_shift_per_trip_minutes, ""],
        ["target_load_factor", result.parameters.target_load_factor, ""],
        ["maximum_load_factor", result.parameters.maximum_load_factor, ""],
        [
            "minimum_turnaround_minutes",
            result.parameters.effective_layover_minutes,
            "Hard constraint",
        ],
    ]
    end = _write_rows(ws, 4, rows)
    for row in range(4, end + 1):
        if ws.cell(row, 1).value in {
            "target_load_factor",
            "maximum_load_factor",
            "minimum_material_service_rate_change_ratio",
        }:
            ws.cell(row, 2).number_format = "0.0%"
    _body(ws, 4, end, 3)
    ws.freeze_panes = "A4"
    _autowidth(ws, 54)


def exported_c_fingerprint(path: str | Path) -> str:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        ws = workbook["TONG_QUAN_B_C"]
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row[0] == "Fingerprint C":
                return str(row[2])
        raise ValueError("Workbook không chứa fingerprint C")
    finally:
        workbook.close()
