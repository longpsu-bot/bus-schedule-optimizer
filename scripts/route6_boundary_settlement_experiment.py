"""Review-only Route 6 boundary-settlement experiment for PR62-C0.

This module intentionally keeps private-workbook parsing and strict local arithmetic
outside production compiler/search policy.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import unicodedata
from collections import Counter
from collections.abc import Mapping, Sequence
from dataclasses import asdict, dataclass
from datetime import datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from bus_schedule_engine.clean_boundary_pilot import build_minimum_fleet_plan_v1
from bus_schedule_engine.service_plan_coordinator import (
    DemandBucketEvidenceV1,
    load_route_coordinator_inputs_v1,
)

EXPERIMENT_PROFILE = "pr62_c0_route6_boundary_settlement_experiment_v1"
REFERENCE_LABELS = ("CURRENT", "EXTERNAL_AI", "HUMAN_FINAL")
OUTPUT_JSON = Path("docs/engine/evidence/PR62_C0_ROUTE6_BOUNDARY_SETTLEMENT_EXPERIMENT.json")
OUTPUT_MARKDOWN = Path("docs/engine/evidence/PR62_C0_ROUTE6_BOUNDARY_SETTLEMENT_EXPERIMENT.md")
EXPECTED_DIRECT_ARTIFACTS = (
    "outputs/demand_regime_model_selection/route_6_demand_regimes.json",
    "outputs/demand_regime_trip_allocation/route_6_demand_regime_trip_allocations.json",
    "outputs/end_tail_settlement_v3/end_tail_settlement_pilot_report.json",
    "outputs/end_tail_settlement_v3/Route_6_EndTail_V3.xlsx",
)


@dataclass(frozen=True, slots=True)
class HeadwayRun:
    headway_minutes: int
    gap_start_index: int
    gap_count: int


@dataclass(frozen=True, slots=True)
class SettlementResidual:
    left_headway: int
    residual_gap: int
    right_headway: int
    residual_gap_index: int
    left_support_gaps: int
    right_support_gaps: int
    residual_start: int
    residual_end: int


@dataclass(frozen=True, slots=True)
class StrictLocalCandidate:
    family: str
    gaps: tuple[int, ...]
    left_headway: int
    bridge_headway: int | None
    right_headway: int
    left_gap_count: int
    bridge_gap_count: int
    right_gap_count: int
    departures: tuple[int, ...]


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _hhmm(seconds: int) -> str:
    minutes = seconds // 60
    return f"{minutes // 60:02d}:{minutes % 60:02d}"


def _normalise(value: str) -> str:
    return " ".join(unicodedata.normalize("NFC", value).casefold().split())


def _cell_time_seconds(value: Any, *, coordinate: str) -> int:
    if isinstance(value, datetime):
        value = value.time()
    if not isinstance(value, time):
        raise ValueError(f"{coordinate} is not an exact Excel time: {value!r}")
    if value.second or value.microsecond:
        raise ValueError(f"{coordinate} is not a whole-minute timestamp: {value!r}")
    return (value.hour * 60 + value.minute) * 60


def discover_reference_sheets(sheet_names: Sequence[str]) -> dict[str, str]:
    """Resolve only the three conceptual Route 6 reference sheets, failing closed."""

    rules = {
        "CURRENT": lambda name: "06" in name and "hiện hữu" in name,
        "EXTERNAL_AI": lambda name: "06" in name and name.endswith(" ai"),
        "HUMAN_FINAL": lambda name: "06" in name and name.endswith(" final"),
    }
    resolved: dict[str, str] = {}
    for label, predicate in rules.items():
        matches = [name for name in sheet_names if predicate(_normalise(name))]
        if len(matches) != 1:
            raise ValueError(f"expected exactly one {label} sheet, found {matches}")
        resolved[label] = matches[0]
    return resolved


def parse_route6_reference_workbook(path: Path) -> dict[str, Any]:
    """Parse the isolated, explicitly headed B/F departure columns in the reference."""

    workbook = load_workbook(path, read_only=True, data_only=False)
    sheet_map = discover_reference_sheets(workbook.sheetnames)
    references: dict[str, Any] = {}
    for label in REFERENCE_LABELS:
        sheet_name = sheet_map[label]
        sheet = workbook[sheet_name]
        outbound_header = str(sheet["B4"].value or "")
        inbound_header = str(sheet["F4"].value or "")
        if "đi" not in _normalise(outbound_header) or "chợ lớn" not in _normalise(outbound_header):
            raise ValueError(f"{sheet_name}!B4 does not explicitly identify outbound departures")
        if "đi" not in _normalise(inbound_header) or "nông lâm" not in _normalise(inbound_header):
            raise ValueError(f"{sheet_name}!F4 does not explicitly identify inbound departures")

        populated_rows = [
            row
            for row in range(5, sheet.max_row + 1)
            if isinstance(sheet[f"B{row}"].value, (datetime, time))
            or isinstance(sheet[f"F{row}"].value, (datetime, time))
        ]
        if not populated_rows or populated_rows != list(
            range(populated_rows[0], populated_rows[-1] + 1)
        ):
            raise ValueError(f"{sheet_name} departure rows are not one contiguous block")
        outbound = tuple(
            _cell_time_seconds(sheet[f"B{row}"].value, coordinate=f"{sheet_name}!B{row}")
            for row in populated_rows
        )
        inbound = tuple(
            _cell_time_seconds(sheet[f"F{row}"].value, coordinate=f"{sheet_name}!F{row}")
            for row in populated_rows
        )
        for direction, departures in (("outbound", outbound), ("inbound", inbound)):
            if len(departures) != 78:
                raise ValueError(
                    f"{sheet_name} {direction} has {len(departures)} departures, not 78"
                )
            if tuple(sorted(departures)) != departures or len(set(departures)) != len(departures):
                raise ValueError(f"{sheet_name} {direction} departures are not strictly increasing")
            if departures[0] != 4 * 3600 + 55 * 60 or departures[-1] != 21 * 3600:
                raise ValueError(
                    f"{sheet_name} {direction} endpoints materially disagree with 04:55-21:00"
                )
        references[label] = {
            "sheet_name": sheet_name,
            "layout": {
                "data_rows": [populated_rows[0], populated_rows[-1]],
                "outbound_departure_column": "B",
                "outbound_header": outbound_header,
                "inbound_departure_column": "F",
                "inbound_header": inbound_header,
            },
            "outbound": outbound,
            "inbound": inbound,
        }
    human_witness = tuple(
        value for value in references["HUMAN_FINAL"]["inbound"] if 17 * 3600 <= value <= 18 * 3600
    )
    expected = tuple(
        (hour * 60 + minute) * 60
        for hour, minute in ((17, 0), (17, 8), (17, 16), (17, 30), (17, 45), (18, 0))
    )
    if not all(value in human_witness for value in expected):
        raise ValueError(
            "Human Final inbound materially disagrees with the expected late-afternoon witness: "
            f"{[_hhmm(value) for value in human_witness]}"
        )
    return {
        "all_sheet_names": list(workbook.sheetnames),
        "reference_sheet_names": sheet_map,
        "references": references,
    }


def exact_headway_runs(departures: Sequence[int]) -> tuple[HeadwayRun, ...]:
    gaps = tuple(
        (right - left) // 60 for left, right in zip(departures, departures[1:], strict=False)
    )
    if any(
        right - left <= 0 or (right - left) % 60
        for left, right in zip(departures, departures[1:], strict=False)
    ):
        raise ValueError("departures must be strictly increasing whole-minute timestamps")
    runs: list[HeadwayRun] = []
    start = 0
    while start < len(gaps):
        end = start + 1
        while end < len(gaps) and gaps[end] == gaps[start]:
            end += 1
        runs.append(HeadwayRun(gaps[start], start, end - start))
        start = end
    return tuple(runs)


def detect_settlement_residuals(
    departures: Sequence[int], *, minimum_adjacent_support: int = 3
) -> tuple[SettlementResidual, ...]:
    runs = exact_headway_runs(departures)
    residuals: list[SettlementResidual] = []
    for left, middle, right in zip(runs, runs[1:], runs[2:], strict=False):
        if (
            middle.gap_count == 1
            and left.gap_count >= minimum_adjacent_support
            and right.gap_count >= minimum_adjacent_support
            and middle.headway_minutes not in {left.headway_minutes, right.headway_minutes}
        ):
            index = middle.gap_start_index
            residuals.append(
                SettlementResidual(
                    left_headway=left.headway_minutes,
                    residual_gap=middle.headway_minutes,
                    right_headway=right.headway_minutes,
                    residual_gap_index=index,
                    left_support_gaps=left.gap_count,
                    right_support_gaps=right.gap_count,
                    residual_start=departures[index],
                    residual_end=departures[index + 1],
                )
            )
    return tuple(residuals)


def select_route6_residual(residuals: Sequence[SettlementResidual]) -> SettlementResidual:
    late_afternoon = [
        item
        for item in residuals
        if 16 * 3600 <= item.residual_start <= 19 * 3600 and item.left_headway < item.right_headway
    ]
    if not late_afternoon:
        raise ValueError("no supported late-afternoon increasing-rhythm residual was detected")
    target = 17 * 3600 + 16 * 60
    return min(
        late_afternoon,
        key=lambda item: (
            abs(item.residual_start - target),
            item.residual_start,
            item.left_headway,
            item.residual_gap,
            item.right_headway,
        ),
    )


def enumerate_strict_local_candidates(
    *, left_anchor: int, right_anchor: int, gap_count: int, h_left: int, h_right: int
) -> tuple[StrictLocalCandidate, ...]:
    """Exhaustively enumerate the two authorized small local strict families."""

    duration_minutes = (right_anchor - left_anchor) // 60
    if right_anchor <= left_anchor or (right_anchor - left_anchor) % 60:
        raise ValueError("anchors must define a positive whole-minute duration")
    candidates: list[StrictLocalCandidate] = []
    seen: set[tuple[int, ...]] = set()

    def add(
        family: str,
        gaps: tuple[int, ...],
        *,
        h_bridge: int | None,
        n_left: int,
        n_bridge: int,
        n_right: int,
    ) -> None:
        if len(gaps) != gap_count or sum(gaps) != duration_minutes:
            return
        departures = [left_anchor]
        for gap in gaps:
            departures.append(departures[-1] + gap * 60)
        vector = tuple(departures)
        if vector[-1] != right_anchor or vector in seen:
            return
        seen.add(vector)
        candidates.append(
            StrictLocalCandidate(
                family=family,
                gaps=gaps,
                left_headway=h_left,
                bridge_headway=h_bridge,
                right_headway=h_right,
                left_gap_count=n_left,
                bridge_gap_count=n_bridge,
                right_gap_count=n_right,
                departures=vector,
            )
        )

    for n_left in range(1, gap_count):
        n_right = gap_count - n_left
        gaps = (h_left,) * n_left + (h_right,) * n_right
        add(
            "TWO_RHYTHM",
            gaps,
            h_bridge=None,
            n_left=n_left,
            n_bridge=0,
            n_right=n_right,
        )
    for h_bridge in range(h_left + 1, h_right):
        for n_left in range(1, gap_count - 2):
            for n_bridge in range(2, gap_count - n_left):
                n_right = gap_count - n_left - n_bridge
                if n_right < 1:
                    continue
                gaps = (h_left,) * n_left + (h_bridge,) * n_bridge + (h_right,) * n_right
                add(
                    "BRIDGE_REGIME",
                    gaps,
                    h_bridge=h_bridge,
                    n_left=n_left,
                    n_bridge=n_bridge,
                    n_right=n_right,
                )
    return tuple(candidates)


def _bucket_counts(
    departures: Sequence[int], buckets: Sequence[DemandBucketEvidenceV1]
) -> tuple[int, ...]:
    return tuple(
        sum(bucket.start <= departure < bucket.end for departure in departures)
        for bucket in buckets
    )


def _tail_clockface(departures: Sequence[int], runs: Sequence[HeadwayRun]) -> dict[str, Any]:
    stable_runs = [run for run in runs if run.gap_count >= 3]
    if not stable_runs:
        raise ValueError("no stable tail run with at least three equal headways")
    tail = stable_runs[-1]
    h = tail.headway_minutes
    tail_departures = tuple(
        departures[tail.gap_start_index : tail.gap_start_index + tail.gap_count + 1]
    )
    phases = Counter((departure // 60) % h for departure in tail_departures)
    dominant_phase, aligned = min(phases.items(), key=lambda item: (-item[1], item[0]))
    deviations = [
        min(
            abs((departure // 60) % h - dominant_phase),
            h - abs((departure // 60) % h - dominant_phase),
        )
        for departure in tail_departures
    ]
    return {
        "tail_headway_minutes": h,
        "tail_start": _hhmm(tail_departures[0]),
        "tail_departure_count": len(tail_departures),
        "tail_stable_gap_count": tail.gap_count,
        "dominant_modulo_phase_minutes": dominant_phase,
        "exact_alignment_proportion": aligned / len(tail_departures),
        "maximum_phase_deviation_minutes": max(deviations, default=0),
    }


def directional_metrics(
    departures: Sequence[int],
    *,
    demand_buckets: Sequence[DemandBucketEvidenceV1],
    current_bucket_counts: Sequence[int] | None = None,
) -> dict[str, Any]:
    runs = exact_headway_runs(departures)
    gaps = tuple(
        (right - left) // 60 for left, right in zip(departures, departures[1:], strict=False)
    )
    counts = _bucket_counts(departures, demand_buckets)
    total_demand = sum(item.observed_demand for item in demand_buckets)
    demand_shares = tuple(item.observed_demand / total_demand for item in demand_buckets)
    service_shares = tuple(count / len(departures) for count in counts)
    mismatch = sum(
        (service - demand) ** 2
        for service, demand in zip(service_shares, demand_shares, strict=True)
    )
    raw_jumps = tuple(
        abs(math.log(right / left)) for left, right in zip(gaps, gaps[1:], strict=False)
    )
    histogram = Counter(gaps)
    longest: dict[int, int] = {}
    for run in runs:
        longest[run.headway_minutes] = max(longest.get(run.headway_minutes, 0), run.gap_count)
    bucket_evidence = []
    for index, (bucket, count, demand_share, service_share) in enumerate(
        zip(demand_buckets, counts, demand_shares, service_shares, strict=True)
    ):
        current = count if current_bucket_counts is None else current_bucket_counts[index]
        bucket_evidence.append(
            {
                "start": _hhmm(bucket.start),
                "end": _hhmm(bucket.end),
                "observed_demand": bucket.observed_demand,
                "demand_share": demand_share,
                "service_count": count,
                "service_share": service_share,
                "movement_vs_current": count - current,
            }
        )
    return {
        "total_departures": len(departures),
        "first_departure": _hhmm(departures[0]),
        "last_departure": _hhmm(departures[-1]),
        "exact_headway_sequence_minutes": list(gaps),
        "unique_headways_minutes": sorted(histogram),
        "headway_histogram": {str(key): histogram[key] for key in sorted(histogram)},
        "equal_headway_run_count": len(runs),
        "singleton_gap_run_count": sum(run.gap_count == 1 for run in runs),
        "longest_run_by_headway": {str(key): longest[key] for key in sorted(longest)},
        "largest_adjacent_raw_frequency_jump": max(raw_jumps, default=0.0),
        "total_raw_frequency_variation": sum(raw_jumps),
        "immutable_demand_mismatch": mismatch,
        "demand_bucket_service": bucket_evidence,
        "tail_clockface": _tail_clockface(departures, runs),
    }


def pair_metrics(
    outbound: Sequence[int],
    inbound: Sequence[int],
    *,
    outbound_metrics: Mapping[str, Any],
    inbound_metrics: Mapping[str, Any],
    runtime_minutes: int,
    minimum_layover_minutes: int,
    fleet_ceiling: int,
    candidate_id: str,
) -> dict[str, Any]:
    plan = build_minimum_fleet_plan_v1(
        route_id="6",
        outbound_candidate_id=f"{candidate_id}-OUTBOUND",
        inbound_candidate_id=f"{candidate_id}-INBOUND",
        outbound_departures=outbound,
        inbound_departures=inbound,
        runtime_minutes=runtime_minutes,
        minimum_layover_minutes=minimum_layover_minutes,
    )
    excess_waits = [
        max(0, assignment.connection_layover_minutes - minimum_layover_minutes)
        for assignment in plan.assignments
        if assignment.connection_layover_minutes is not None
    ]
    return {
        "fleet_required": plan.fleet_requirement,
        "fleet_ceiling": fleet_ceiling,
        "fleet_feasible": plan.fleet_requirement <= fleet_ceiling,
        "minimum_connection_layover_minutes": plan.minimum_connection_layover_minutes,
        "total_excess_terminal_wait_minutes": sum(excess_waits),
        "maximum_excess_terminal_wait_minutes": max(excess_waits, default=0),
        "immutable_demand_mismatch": (
            outbound_metrics["immutable_demand_mismatch"]
            + inbound_metrics["immutable_demand_mismatch"]
        ),
        "largest_adjacent_raw_frequency_jump": max(
            outbound_metrics["largest_adjacent_raw_frequency_jump"],
            inbound_metrics["largest_adjacent_raw_frequency_jump"],
        ),
        "total_raw_frequency_variation": (
            outbound_metrics["total_raw_frequency_variation"]
            + inbound_metrics["total_raw_frequency_variation"]
        ),
    }


def _dominates(left: Sequence[float], right: Sequence[float]) -> bool:
    return all(a <= b for a, b in zip(left, right, strict=True)) and any(
        a < b for a, b in zip(left, right, strict=True)
    )


def _comparison_vector(candidate: Mapping[str, Any]) -> tuple[float, ...]:
    pair = candidate["pair_metrics"]
    tail = candidate["inbound_metrics"]["tail_clockface"]
    peak = candidate["peak_preservation"]
    return (
        pair["immutable_demand_mismatch"],
        float(pair["fleet_required"]),
        float(pair["total_excess_terminal_wait_minutes"]),
        float(pair["maximum_excess_terminal_wait_minutes"]),
        pair["largest_adjacent_raw_frequency_jump"],
        pair["total_raw_frequency_variation"],
        float(tail["maximum_phase_deviation_minutes"]),
        -float(peak["minutes_of_left_rhythm_preserved"]),
    )


def _pareto_frontier(candidates: Sequence[Mapping[str, Any]]) -> list[str]:
    retained = []
    for candidate in candidates:
        vector = _comparison_vector(candidate)
        if any(
            _dominates(_comparison_vector(other), vector)
            for other in candidates
            if other["candidate_id"] != candidate["candidate_id"]
        ):
            continue
        retained.append(candidate["candidate_id"])
    return sorted(retained)


def _strict_classification(
    human: Mapping[str, Any], strict_candidates: Sequence[Mapping[str, Any]]
) -> tuple[str, list[str]]:
    if not strict_candidates:
        return "NO_STRICT_ALTERNATIVE_EXISTS", []
    human_vector = _comparison_vector(human)
    dominating = [
        candidate["candidate_id"]
        for candidate in strict_candidates
        if _dominates(_comparison_vector(candidate), human_vector)
    ]
    if dominating:
        return "STRICT_DOMINATING_REFERENCE_EXISTS", sorted(dominating)
    if all(
        all(a <= b for a, b in zip(human_vector, _comparison_vector(candidate), strict=True))
        for candidate in strict_candidates
    ):
        return "HUMAN_FINAL_DOMINATES_ALL_STRICT", []
    return "TRADEOFF_INCONCLUSIVE", []


def _replace_local_window(
    departures: Sequence[int], *, local_start_index: int, local_departures: Sequence[int]
) -> tuple[int, ...]:
    end_index = local_start_index + len(local_departures)
    replaced = (
        tuple(departures[:local_start_index])
        + tuple(local_departures)
        + tuple(departures[end_index:])
    )
    if (
        len(replaced) != len(departures)
        or replaced[0] != departures[0]
        or replaced[-1] != departures[-1]
        or replaced[:local_start_index] != tuple(departures[:local_start_index])
        or replaced[end_index:] != tuple(departures[end_index:])
    ):
        raise AssertionError("strict candidate violated full-timetable invariants")
    return replaced


def _reference_metrics(
    parsed: Mapping[str, Any], context: Any
) -> tuple[dict[str, Any], dict[str, tuple[int, ...]]]:
    current_counts = {
        direction: _bucket_counts(
            parsed["references"]["CURRENT"][direction], context.demand_buckets[direction]
        )
        for direction in ("outbound", "inbound")
    }
    metrics: dict[str, Any] = {}
    for label in REFERENCE_LABELS:
        item = parsed["references"][label]
        directional = {
            direction: directional_metrics(
                item[direction],
                demand_buckets=context.demand_buckets[direction],
                current_bucket_counts=current_counts[direction],
            )
            for direction in ("outbound", "inbound")
        }
        metrics[label] = {
            "source_label": label,
            "sheet_name": item["sheet_name"],
            "outbound": directional["outbound"],
            "inbound": directional["inbound"],
            "pair": pair_metrics(
                item["outbound"],
                item["inbound"],
                outbound_metrics=directional["outbound"],
                inbound_metrics=directional["inbound"],
                runtime_minutes=context.runtime_minutes,
                minimum_layover_minutes=context.minimum_layover_minutes,
                fleet_ceiling=context.fleet_ceiling,
                candidate_id=label,
            ),
        }
    return metrics, current_counts


def _authority_evidence(
    repo_root: Path, authority_root: Path, coordinator_workbook_path: Path, context: Any
) -> dict[str, Any]:
    manifest = json.loads(
        (repo_root / "config/service_plan_coordinator_frozen_prior_v1.json").read_text(
            encoding="utf-8"
        )
    )
    artifacts = []
    for relative in EXPECTED_DIRECT_ARTIFACTS:
        path = authority_root / relative
        if not path.is_file():
            raise FileNotFoundError(f"missing direct frozen Route 6 authority artifact: {relative}")
        actual = _sha256(path)
        expected = manifest["sha256"].get(relative)
        if actual != expected:
            raise ValueError(f"frozen Route 6 authority artifact changed: {relative}")
        artifacts.append({"relative_path": relative, "sha256": actual})
    if (context.runtime_minutes, context.minimum_layover_minutes, context.fleet_ceiling) != (
        70,
        5,
        20,
    ):
        raise ValueError("Route 6 pilot authority is not runtime=70, layover=5, fleet=20")
    return {
        "loader": "load_route_coordinator_inputs_v1",
        "route_id": context.route_id,
        "runtime_minutes_each_direction": context.runtime_minutes,
        "minimum_layover_minutes": context.minimum_layover_minutes,
        "fleet_ceiling": context.fleet_ceiling,
        "immutable_demand_sha256": context.immutable_demand_sha256,
        "coordinator_input_workbook": {
            "basename": coordinator_workbook_path.name,
            "sha256": _sha256(coordinator_workbook_path),
        },
        "demand_mismatch_formula": "sum((service_share - demand_share)^2)",
        "fleet_authority": "build_minimum_fleet_plan_v1 exact timetable path cover",
        "direct_frozen_artifacts": artifacts,
    }


def run_experiment(
    *,
    repo_root: Path,
    workbook_path: Path,
    authority_root: Path,
    coordinator_workbook_path: Path,
) -> dict[str, Any]:
    parsed = parse_route6_reference_workbook(workbook_path)
    context, _ = load_route_coordinator_inputs_v1(
        repo_root=authority_root,
        route_id="6",
        workbook_path=coordinator_workbook_path,
    )
    authority = _authority_evidence(repo_root, authority_root, coordinator_workbook_path, context)
    references, current_counts = _reference_metrics(parsed, context)

    human_outbound = parsed["references"]["HUMAN_FINAL"]["outbound"]
    human_inbound = parsed["references"]["HUMAN_FINAL"]["inbound"]
    residuals = detect_settlement_residuals(human_inbound)
    selected = select_route6_residual(residuals)
    local_start_index = selected.residual_gap_index - 3
    local_human = tuple(human_inbound[local_start_index : local_start_index + 8])
    local_gaps = tuple(
        (right - left) // 60 for left, right in zip(local_human, local_human[1:], strict=False)
    )
    if local_gaps != (
        (selected.left_headway,) * 3 + (selected.residual_gap,) + (selected.right_headway,) * 3
    ):
        raise ValueError(
            f"selected residual does not support the required local window: {local_gaps}"
        )

    enumerated = enumerate_strict_local_candidates(
        left_anchor=local_human[0],
        right_anchor=local_human[-1],
        gap_count=len(local_gaps),
        h_left=selected.left_headway,
        h_right=selected.right_headway,
    )
    human_peak = {
        "last_departure_on_left_rhythm": _hhmm(human_inbound[selected.residual_gap_index]),
        "left_rhythm_gap_count_preserved": 3,
        "minutes_of_left_rhythm_preserved": 3 * selected.left_headway,
        "left_rhythm_ends_earlier_than_human_minutes": 0,
    }
    human_comparison = {
        "candidate_id": "HUMAN_FINAL",
        "pair_metrics": references["HUMAN_FINAL"]["pair"],
        "inbound_metrics": references["HUMAN_FINAL"]["inbound"],
        "peak_preservation": human_peak,
    }
    strict_payload: list[dict[str, Any]] = []
    family_ordinals: Counter[str] = Counter()
    for candidate in enumerated:
        family_ordinals[candidate.family] += 1
        family_tag = "A" if candidate.family == "TWO_RHYTHM" else "B"
        candidate_id = f"STRICT_{family_tag}_{family_ordinals[candidate.family]:03d}"
        inbound = _replace_local_window(
            human_inbound,
            local_start_index=local_start_index,
            local_departures=candidate.departures,
        )
        inbound_metrics = directional_metrics(
            inbound,
            demand_buckets=context.demand_buckets["inbound"],
            current_bucket_counts=current_counts["inbound"],
        )
        outbound_metrics = references["HUMAN_FINAL"]["outbound"]
        pair = pair_metrics(
            human_outbound,
            inbound,
            outbound_metrics=outbound_metrics,
            inbound_metrics=inbound_metrics,
            runtime_minutes=context.runtime_minutes,
            minimum_layover_minutes=context.minimum_layover_minutes,
            fleet_ceiling=context.fleet_ceiling,
            candidate_id=candidate_id,
        )
        shifts = [
            abs(candidate_time - human_time) // 60
            for candidate_time, human_time in zip(candidate.departures, local_human, strict=True)
        ]
        candidate_last_left = candidate.departures[candidate.left_gap_count]
        strict_payload.append(
            {
                "candidate_id": candidate_id,
                "family": candidate.family,
                "h_left": candidate.left_headway,
                "h_bridge": candidate.bridge_headway,
                "h_right": candidate.right_headway,
                "run_gap_counts": {
                    "left": candidate.left_gap_count,
                    "bridge": candidate.bridge_gap_count,
                    "right": candidate.right_gap_count,
                },
                "local_headways_minutes": list(candidate.gaps),
                "local_departures": [_hhmm(value) for value in candidate.departures],
                "sum_absolute_local_shift_minutes": sum(shifts),
                "maximum_absolute_local_shift_minutes": max(shifts),
                "peak_preservation": {
                    "last_departure_on_left_rhythm": _hhmm(candidate_last_left),
                    "left_rhythm_gap_count_preserved": candidate.left_gap_count,
                    "minutes_of_left_rhythm_preserved": (
                        candidate.left_gap_count * candidate.left_headway
                    ),
                    "left_rhythm_ends_earlier_than_human_minutes": (
                        human_inbound[selected.residual_gap_index] - candidate_last_left
                    )
                    // 60,
                },
                "full_timetable_invariants": {
                    "outbound_departures": len(human_outbound),
                    "inbound_departures": len(inbound),
                    "total_departures": len(human_outbound) + len(inbound),
                    "fixed_global_endpoints_preserved": True,
                    "human_final_outbound_unchanged": True,
                    "outside_local_inbound_window_unchanged": True,
                    "whole_minute_timestamps": all(value % 60 == 0 for value in inbound),
                },
                "strictness_verification": {
                    "no_singleton_settlement_gap": True,
                    "bridge_is_repeated_rhythm": (
                        candidate.bridge_headway is None or candidate.bridge_gap_count >= 2
                    ),
                    "local_arithmetic_verified": True,
                    "production_compiler_replay": "NOT_EVALUATED_EXPERIMENT_LOCAL_ARITHMETIC_SUFFICIENT",
                },
                "outbound_metrics": outbound_metrics,
                "inbound_metrics": inbound_metrics,
                "pair_metrics": pair,
            }
        )

    frontier = _pareto_frontier(strict_payload)
    classification, dominating = _strict_classification(human_comparison, strict_payload)
    family_counts = Counter(item.family for item in enumerated)
    payload = {
        "experiment_profile": EXPERIMENT_PROFILE,
        "experiment_status": "COMPLETED",
        "review_only": True,
        "source_workbook": {
            "basename": workbook_path.name,
            "sha256": _sha256(workbook_path),
            "all_sheet_names": parsed["all_sheet_names"],
            "reference_sheet_names": parsed["reference_sheet_names"],
        },
        "parse_audit": {
            label: {
                "sheet_name": parsed["references"][label]["sheet_name"],
                "layout": parsed["references"][label]["layout"],
                "outbound_count": len(parsed["references"][label]["outbound"]),
                "inbound_count": len(parsed["references"][label]["inbound"]),
                "total_count": (
                    len(parsed["references"][label]["outbound"])
                    + len(parsed["references"][label]["inbound"])
                ),
                "outbound_endpoints": [
                    _hhmm(parsed["references"][label]["outbound"][0]),
                    _hhmm(parsed["references"][label]["outbound"][-1]),
                ],
                "inbound_endpoints": [
                    _hhmm(parsed["references"][label]["inbound"][0]),
                    _hhmm(parsed["references"][label]["inbound"][-1]),
                ],
                "acceptance_oracles_passed": True,
            }
            for label in REFERENCE_LABELS
        },
        "route6_authority": authority,
        "reference_timetable_metrics": references,
        "detected_settlement_residual": {
            "selected": {
                **asdict(selected),
                "residual_start": _hhmm(selected.residual_start),
                "residual_end": _hhmm(selected.residual_end),
            },
            "all_supported_residuals": [
                {
                    **asdict(item),
                    "residual_start": _hhmm(item.residual_start),
                    "residual_end": _hhmm(item.residual_end),
                }
                for item in residuals
            ],
        },
        "local_arithmetic": {
            "left_anchor": _hhmm(local_human[0]),
            "right_anchor": _hhmm(local_human[-1]),
            "human_final_departures": [_hhmm(value) for value in local_human],
            "human_final_headways_minutes": list(local_gaps),
            "local_gap_count": len(local_gaps),
            "anchor_duration_minutes": (local_human[-1] - local_human[0]) // 60,
            "outside_window_must_remain_human_final": True,
        },
        "strict_enumeration": {
            "deterministic": True,
            "deduplicated_by_exact_departure_vector": True,
            "two_rhythm_candidate_count": family_counts["TWO_RHYTHM"],
            "bridge_regime_candidate_count": family_counts["BRIDGE_REGIME"],
            "total_candidate_count": len(strict_payload),
            "candidates": strict_payload,
        },
        "strict_pareto_frontier": {
            "candidate_ids": frontier,
            "dimensions": {
                "minimize": [
                    "immutable_demand_mismatch",
                    "fleet_required",
                    "total_excess_terminal_wait_minutes",
                    "maximum_excess_terminal_wait_minutes",
                    "largest_adjacent_raw_frequency_jump",
                    "total_raw_frequency_variation",
                    "tail_clockface_maximum_phase_deviation_minutes",
                ],
                "maximize": ["minutes_of_left_rhythm_preserved"],
            },
        },
        "human_final_comparison": {
            "human_final_peak_preservation": human_peak,
            "strict_candidates_dominating_human_final": dominating,
            "classification": classification,
            "policy_decision": "NONE_EVIDENCE_ONLY",
        },
        "evidence_classification": classification,
        "limitations": [
            "Single Route 6 private reference workbook; no cross-route generalization.",
            "External AI is a supplied reference only and has no project-engine lineage.",
            "Clockface descriptors are non-objective engine diagnostics in this milestone.",
            "Strict alternatives are exhaustive only within the two specified local families.",
            "Workbook-displayed runtimes are descriptive and do not alter the 70-minute authority.",
            "No production compiler, search, fleet-validation, or transport policy changed.",
        ],
        "policy_changes": {
            "compiler_policy_changed": False,
            "search_budget_changed": False,
            "transport_policy_decision_made": False,
        },
    }
    return payload


def render_markdown(payload: Mapping[str, Any]) -> str:
    source = payload["source_workbook"]
    residual = payload["detected_settlement_residual"]["selected"]
    arithmetic = payload["local_arithmetic"]
    enumeration = payload["strict_enumeration"]
    candidates = {item["candidate_id"]: item for item in enumeration["candidates"]}
    frontier_ids = payload["strict_pareto_frontier"]["candidate_ids"]
    lines = [
        "# PR62-C0 — Route 6 boundary-settlement experiment",
        "",
        "> Review-only evidence. No compiler, search, fleet-validator, or transport policy changed.",
        "",
        "## Input and authority",
        "",
        f"- Workbook: `{source['basename']}`",
        f"- SHA-256: `{source['sha256']}`",
        "- Sheets: "
        + ", ".join(
            f"{label} = `{name}`" for label, name in source["reference_sheet_names"].items()
        ),
        "- Primary fleet authority: 70 minutes each direction, 5 minutes minimum layover, "
        "20-vehicle ceiling.",
        "- Demand objective: `sum((service_share - demand_share)^2)` against the existing "
        "immutable Route 6 demand buckets.",
        "",
        "All three references parse as 78 outbound + 78 inbound departures, with fixed "
        "04:55 and 21:00 endpoints.",
        "",
        "## CURRENT → EXTERNAL_AI → HUMAN_FINAL benchmark",
        "",
        "The `EXTERNAL_AI` timetable is an external supplied reference and is not engine lineage.",
        "",
        "| Reference | Fleet | Pair mismatch | Out/In unique headways | Out/In runs | "
        "Out/In singleton runs | Max raw jump | Total raw variation | Out/In final tail | "
        "Out/In tail start | Out/In exact alignment | Out/In tail phase error |",
        "|---|---:|---:|---|---:|---:|---:|---:|---|---|---:|---:|",
    ]
    for label in REFERENCE_LABELS:
        reference = payload["reference_timetable_metrics"][label]
        out = reference["outbound"]
        inbound = reference["inbound"]
        pair = reference["pair"]
        lines.append(
            f"| {label} | {pair['fleet_required']}/{pair['fleet_ceiling']} | "
            f"{pair['immutable_demand_mismatch']:.10f} | "
            f"{out['unique_headways_minutes']} / {inbound['unique_headways_minutes']} | "
            f"{out['equal_headway_run_count']} / {inbound['equal_headway_run_count']} | "
            f"{out['singleton_gap_run_count']} / {inbound['singleton_gap_run_count']} | "
            f"{pair['largest_adjacent_raw_frequency_jump']:.6f} | "
            f"{pair['total_raw_frequency_variation']:.6f} | "
            f"{out['tail_clockface']['tail_headway_minutes']} / "
            f"{inbound['tail_clockface']['tail_headway_minutes']} min | "
            f"{out['tail_clockface']['tail_start']} / "
            f"{inbound['tail_clockface']['tail_start']} | "
            f"{out['tail_clockface']['exact_alignment_proportion']:.0%} / "
            f"{inbound['tail_clockface']['exact_alignment_proportion']:.0%} | "
            f"{out['tail_clockface']['maximum_phase_deviation_minutes']} / "
            f"{inbound['tail_clockface']['maximum_phase_deviation_minutes']} min |"
        )
    lines.extend(
        [
            "",
            "## Detected Human Final settlement",
            "",
            f"The selected inbound witness is `{residual['left_headway']} → "
            f"{residual['residual_gap']} → {residual['right_headway']}` minutes, with the "
            f"residual from {residual['residual_start']} to {residual['residual_end']}.",
            "",
            f"Local anchors are {arithmetic['left_anchor']} and {arithmetic['right_anchor']}: "
            f"`{arithmetic['human_final_headways_minutes']}` = "
            f"{arithmetic['anchor_duration_minutes']} minutes over "
            f"{arithmetic['local_gap_count']} gaps.",
            "",
            "## Strict local alternatives",
            "",
            f"- Two-rhythm candidates: {enumeration['two_rhythm_candidate_count']}",
            f"- Repeated bridge-regime candidates: {enumeration['bridge_regime_candidate_count']}",
            f"- Deduplicated total: {enumeration['total_candidate_count']}",
            "",
            "| Candidate | Family | Local gaps | Pair mismatch | Fleet | Total/max excess wait | "
            "Peak minutes | Peak ends earlier | Tail start/error | Pareto |",
            "|---|---|---|---:|---:|---:|---:|---:|---|---|",
        ]
    )
    for candidate in enumeration["candidates"]:
        pair = candidate["pair_metrics"]
        peak = candidate["peak_preservation"]
        tail = candidate["inbound_metrics"]["tail_clockface"]
        lines.append(
            f"| {candidate['candidate_id']} | {candidate['family']} | "
            f"`{candidate['local_headways_minutes']}` | "
            f"{pair['immutable_demand_mismatch']:.10f} | "
            f"{pair['fleet_required']}/{pair['fleet_ceiling']} | "
            f"{pair['total_excess_terminal_wait_minutes']}/"
            f"{pair['maximum_excess_terminal_wait_minutes']} | "
            f"{peak['minutes_of_left_rhythm_preserved']} | "
            f"{peak['left_rhythm_ends_earlier_than_human_minutes']} min | "
            f"{tail['tail_start']} / {tail['maximum_phase_deviation_minutes']} min | "
            f"{'yes' if candidate['candidate_id'] in frontier_ids else 'no'} |"
        )
    lines.extend(
        [
            "",
            "Pareto frontier: "
            + (", ".join(f"`{item}`" for item in frontier_ids) if frontier_ids else "none"),
            "",
            f"Evidence classification: **{payload['evidence_classification']}**.",
            "",
            "This classification is evidence only. It does not decide whether the clean-boundary "
            "rule should change.",
            "",
            "## Pareto candidate detail",
            "",
        ]
    )
    for candidate_id in frontier_ids:
        candidate = candidates[candidate_id]
        lines.extend(
            [
                f"### {candidate_id}",
                "",
                f"- Local departures: `{candidate['local_departures']}`",
                f"- Absolute local movement: {candidate['sum_absolute_local_shift_minutes']} "
                f"minutes total; {candidate['maximum_absolute_local_shift_minutes']} minutes max.",
                f"- Last departure on the {candidate['h_left']}-minute rhythm: "
                f"{candidate['peak_preservation']['last_departure_on_left_rhythm']}.",
                "",
            ]
        )
    lines.extend(
        [
            "## Limitations",
            "",
            *[f"- {item}" for item in payload["limitations"]],
            "",
        ]
    )
    return "\n".join(lines)


def _resolve_private_workbook(repo_root: Path) -> Path:
    exact = repo_root / "private/Route_6_Current_ExternalAI_HumanFinal.xlsx"
    if exact.is_file():
        return exact
    matches = sorted((repo_root / "private").glob("Route_6_Current_ExternalAI_HumanFinal*.xlsx"))
    if len(matches) != 1:
        raise ValueError(f"expected exactly one private Route 6 workbook, found {len(matches)}")
    return matches[0]


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--repo-root", type=Path, default=Path.cwd())
    parser.add_argument(
        "--authority-root",
        type=Path,
        default=None,
        help="Root containing the existing frozen outputs (defaults to --repo-root).",
    )
    parser.add_argument(
        "--coordinator-workbook",
        type=Path,
        default=None,
        help="Route 6 V3 workbook consumed by load_route_coordinator_inputs_v1.",
    )
    args = parser.parse_args()
    repo_root = args.repo_root.resolve()
    authority_root = (args.authority_root or repo_root).resolve()
    coordinator_workbook = (
        args.coordinator_workbook
        or repo_root / "Engine_Input_MST_6_V3_MultiPeriod_Mar-Jul_2026.xlsx"
    ).resolve()
    if not coordinator_workbook.is_file():
        raise FileNotFoundError(
            f"missing Route 6 coordinator workbook: {coordinator_workbook.name}"
        )
    workbook_path = _resolve_private_workbook(repo_root)
    payload = run_experiment(
        repo_root=repo_root,
        workbook_path=workbook_path,
        authority_root=authority_root,
        coordinator_workbook_path=coordinator_workbook,
    )
    json_path = repo_root / OUTPUT_JSON
    markdown_path = repo_root / OUTPUT_MARKDOWN
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    markdown_path.write_text(render_markdown(payload), encoding="utf-8", newline="\n")
    print(
        json.dumps(
            {
                "classification": payload["evidence_classification"],
                "json": str(OUTPUT_JSON),
                "markdown": str(OUTPUT_MARKDOWN),
                "strict_candidates": payload["strict_enumeration"]["total_candidate_count"],
            },
            sort_keys=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
